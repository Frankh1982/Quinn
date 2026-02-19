# -*- coding: utf-8 -*-
"""
Excel engine (extracted from server.py)

Hard rules:
- Must NOT import server.py (no circular imports).
- stdlib + openpyxl only.
"""

from __future__ import annotations

import io
import json
import os
import re
import time
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Callable


# ----------------------------
# Helpers
# ----------------------------

def now_iso() -> str:
    """
    Return current UTC time in ISO-8601 format with trailing 'Z'.
    Local helper to avoid importing server.py or project_store.
    """
    try:
        return time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
    except Exception:
        return ""

# ----------------------------
# Deliverables registry helper (C1)
# ----------------------------

def register_excel_deliverable(
    *,
    register_deliverable_fn,
    project_name: str,
    logical_name: str,
    rel_path: str,
    spec: Optional[Dict[str, Any]] = None,
) -> None:
    """
    C1: Register an explicitly-created Excel deliverable.
    This function does NOT import project_store (avoids circular imports).
    Caller injects register_deliverable_fn (single source of truth lives in project_store).
    """
    try:
        if not callable(register_deliverable_fn):
            return
        title = ""
        if isinstance(spec, dict):
            title = str(spec.get("name") or "").strip()
        if not title:
            title = str(logical_name or "").strip() or "Excel Deliverable"

        register_deliverable_fn(
            project_name,
            deliverable_type="xlsx",
            title=title,
            path=(rel_path or "").replace("\\", "/").strip(),
            source="excel_engine",
        )
    except Exception:
        return


def _is_cell_addr(s: str) -> bool:
    return bool(re.fullmatch(r"[A-Z]{1,3}[1-9][0-9]{0,6}", (s or "").strip().upper()))


def _is_a1_range(s: str) -> bool:
    return bool(re.fullmatch(r"[A-Z]{1,3}[1-9][0-9]{0,6}:[A-Z]{1,3}[1-9][0-9]{0,6}", (s or "").strip().upper()))


# ----------------------------
# openpyxl imports (isolated here)
# ----------------------------

try:
    from openpyxl import load_workbook
    from openpyxl import Workbook  # type: ignore
    from openpyxl.worksheet.datavalidation import DataValidation  # type: ignore
    from openpyxl.worksheet.table import Table, TableStyleInfo  # type: ignore
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # type: ignore
    from openpyxl.formatting.rule import FormulaRule, CellIsRule  # type: ignore
    from openpyxl.styles.differential import DifferentialStyle  # type: ignore
    from openpyxl.utils import range_boundaries  # type: ignore
except Exception:
    load_workbook = None
    Workbook = None  # type: ignore
    DataValidation = None  # type: ignore
    Table = None  # type: ignore
    TableStyleInfo = None  # type: ignore
    Font = None  # type: ignore
    PatternFill = None  # type: ignore
    Alignment = None  # type: ignore
    Border = None  # type: ignore
    Side = None  # type: ignore
    FormulaRule = None  # type: ignore
    CellIsRule = None  # type: ignore
    DifferentialStyle = None  # type: ignore
    range_boundaries = None  # type: ignore


# ----------------------------
# Excel structure + blueprint
# ----------------------------

def summarize_excel_structure(path: Path) -> Tuple[str, str]:
    if load_workbook is None:
        msg = "Excel detected but openpyxl is not installed, so I can't inspect it."
        return f"# Excel overview: {path.name}\n\n{msg}\n", msg

    try:
        wb = load_workbook(path, read_only=False, data_only=False)
    except Exception as e:
        msg = f"I tried to open this Excel workbook but hit an error: {e!r}"
        return f"# Excel overview: {path.name}\n\n{msg}\n", msg

    sheet_names = [ws.title for ws in wb.worksheets]
    md: List[str] = [f"# Excel overview: {path.name}", ""]
    md.append(f"- Sheets (total {len(sheet_names)}): " + ", ".join(sheet_names))
    md.append("")

    formula_cells = 0
    table_count = 0
    validation_count = 0

    try:
        defined_names = list(getattr(wb, "defined_names", {}).definedName)  # type: ignore[attr-defined]
        defined_name_count = len(defined_names)
    except Exception:
        defined_name_count = 0

    for ws in wb.worksheets:
        try:
            try:
                table_count += len(getattr(ws, "tables", {}) or {})
            except Exception:
                pass

            try:
                dv = getattr(ws, "data_validations", None)
                if dv is not None and getattr(dv, "dataValidation", None) is not None:
                    validation_count += len(dv.dataValidation)  # type: ignore[attr-defined]
            except Exception:
                pass

            max_cells = 8000
            inspected = 0
            dim = None
            try:
                dim = ws.calculate_dimension()
            except Exception:
                dim = None

            if dim and _is_a1_range(str(dim).upper()) and range_boundaries is not None:
                try:
                    min_col, min_row, max_col, max_row = range_boundaries(str(dim).upper())  # type: ignore[arg-type]
                except Exception:
                    min_col = min_row = 1
                    max_col = max_row = 1
            else:
                min_col = min_row = 1
                max_col = max_row = 50

            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    inspected += 1
                    if inspected > max_cells:
                        break
                    cell = ws.cell(row=r, column=c)
                    v = cell.value
                    if isinstance(v, str) and v.startswith("="):
                        formula_cells += 1
                if inspected > max_cells:
                    break
        except Exception:
            continue

    md.append("## Workbook signals")
    md.append(f"- Named ranges: {defined_name_count}")
    md.append(f"- Tables: {table_count}")
    md.append(f"- Data validations: {validation_count}")
    md.append(f"- Formula cells (bounded scan): {formula_cells}")
    md.append("")

    example_headers: List[str] = []
    for ws in wb.worksheets[:3]:
        md.append(f"## Sheet: {ws.title}")
        headers: List[str] = []
        first_row: List[str] = []
        try:
            rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
            if rows:
                headers = [str(v) for v in rows[0] if v is not None]
            if len(rows) > 1:
                first_row = [str(v) for v in rows[1] if v is not None]
        except Exception as e:
            md.append(f"- (Error reading rows: {e!r})")
            md.append("")
            continue

        if headers:
            md.append("- Header columns: " + ", ".join(headers[:12]))
            example_headers = headers
        else:
            md.append("- Header columns: (none detected)")
        if first_row:
            md.append("- Example data row: " + ", ".join(first_row[:12]))
        else:
            md.append("- Example data row: (no data in first rows)")
        md.append("")

    short_parts: List[str] = []
    if sheet_names:
        short_parts.append(f"It has {len(sheet_names)} sheet(s): " + ", ".join(sheet_names[:6]))
    short_parts.append(f"Signals: formulas~{formula_cells}, tables={table_count}, validations={validation_count}, names={defined_name_count}.")
    if example_headers:
        short_parts.append("Example columns: " + ", ".join(example_headers[:6]))
    short_summary = " ".join(short_parts) if short_parts else "Workbook opened, but I couldn't detect headers in the first rows."
    return "\n".join(md), short_summary


def extract_excel_blueprint(path: Path, *, max_formula_cells: int = 20000) -> Dict[str, Any]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is not installed (cannot extract blueprint).")

    # Two-pass load:
    # - wb_f: formulas (data_only=False)
    # - wb_v: cached values (data_only=True)
    wb_f = load_workbook(path, read_only=False, data_only=False)
    wb_v = load_workbook(path, read_only=False, data_only=True)

    bp: Dict[str, Any] = {
        "file": path.name,
        "sheets": [],
        "defined_names": [],
        "created_at": now_iso(),
        "limits": {"max_formula_cells": int(max_formula_cells)},
    }

    try:
        dn_list = list(getattr(wb_f, "defined_names", {}).definedName)  # type: ignore[attr-defined]
        for dn in dn_list:
            try:
                bp["defined_names"].append(
                    {
                        "name": str(getattr(dn, "name", "") or ""),
                        "attr_text": str(getattr(dn, "attr_text", "") or ""),
                        "localSheetId": getattr(dn, "localSheetId", None),
                    }
                )
            except Exception:
                continue
    except Exception:
        pass

    total_formulas = 0

    # Iterate worksheets by index so we can line up wb_f and wb_v sheets safely.
    for i_ws, ws_f in enumerate(wb_f.worksheets):
        ws_v = None
        try:
            ws_v = wb_v.worksheets[i_ws]
        except Exception:
            ws_v = None

        sh: Dict[str, Any] = {
            "title": ws_f.title,
            "tables": [],
            "data_validations": [],
            # Store formula + cached value (if available)
            "formula_cells": {},
        }

        # Tables
        try:
            tbls = getattr(ws_f, "tables", {}) or {}
            for tname, tbl in tbls.items():
                try:
                    sh["tables"].append(
                        {
                            "name": str(getattr(tbl, "name", tname) or tname),
                            "ref": str(getattr(tbl, "ref", "") or ""),
                            "style": str(getattr(getattr(tbl, "tableStyleInfo", None), "name", "") or ""),
                            "showRowStripes": bool(getattr(getattr(tbl, "tableStyleInfo", None), "showRowStripes", False)),
                        }
                    )
                except Exception:
                    continue
        except Exception:
            pass

        # Data validations
        try:
            dv = getattr(ws_f, "data_validations", None)
            dvs = getattr(dv, "dataValidation", None) if dv is not None else None
            if isinstance(dvs, list):
                for item in dvs:
                    try:
                        sh["data_validations"].append(
                            {
                                "type": str(getattr(item, "type", "") or ""),
                                "operator": str(getattr(item, "operator", "") or ""),
                                "formula1": getattr(item, "formula1", None),
                                "formula2": getattr(item, "formula2", None),
                                "allowBlank": bool(getattr(item, "allowBlank", False)),
                                "sqref": str(getattr(item, "sqref", "") or ""),
                                "promptTitle": str(getattr(item, "promptTitle", "") or ""),
                                "prompt": str(getattr(item, "prompt", "") or ""),
                                "errorTitle": str(getattr(item, "errorTitle", "") or ""),
                                "error": str(getattr(item, "error", "") or ""),
                            }
                        )
                    except Exception:
                        continue
        except Exception:
            pass

        # Formula scan (bounded) + cached values (if present)
        try:
            dim = None
            try:
                dim = ws_f.calculate_dimension()
            except Exception:
                dim = None

            if dim and _is_a1_range(str(dim).upper()) and range_boundaries is not None:
                try:
                    min_col, min_row, max_col, max_row = range_boundaries(str(dim).upper())  # type: ignore[arg-type]
                except Exception:
                    min_col = min_row = 1
                    max_col = max_row = 1
            else:
                min_col = min_row = 1
                max_col = max_row = 80

            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    if total_formulas >= max_formula_cells:
                        break
                    cell_f = ws_f.cell(row=r, column=c)
                    v_f = cell_f.value
                    if isinstance(v_f, str) and v_f.startswith("="):
                        cached = None
                        try:
                            if ws_v is not None:
                                cached = ws_v[cell_f.coordinate].value
                        except Exception:
                            cached = None

                        sh["formula_cells"][cell_f.coordinate] = {"f": v_f, "v": cached}
                        total_formulas += 1
                if total_formulas >= max_formula_cells:
                    break
        except Exception:
            pass

        bp["sheets"].append(sh)

    bp["counts"] = {
        "sheets": len(bp["sheets"]),
        "defined_names": len(bp["defined_names"]),
        "formula_cells": total_formulas,
        "tables": sum(len(s.get("tables") or []) for s in bp["sheets"]),
        "data_validations": sum(len(s.get("data_validations") or []) for s in bp["sheets"]),
    }
    return bp



# ----------------------------
# Spec validation + optional repair (caller supplies OpenAI callable)
# ----------------------------

def validate_excel_spec(spec: Dict[str, Any]) -> Dict[str, Any]:
    errors: List[str] = []
    warnings: List[str] = []

    if not isinstance(spec, dict):
        return {"ok": False, "errors": ["spec must be a JSON object"], "warnings": []}

    sheets = spec.get("sheets")
    if not isinstance(sheets, list) or not sheets:
        errors.append("spec.sheets must be a non-empty array")
        return {"ok": False, "errors": errors, "warnings": warnings}

    for i, sh in enumerate(sheets):
        if not isinstance(sh, dict):
            errors.append(f"sheets[{i}] must be an object")
            continue

        title = str(sh.get("title") or "").strip()
        if not title:
            errors.append(f"sheets[{i}].title is required")
        if len(title) > 31:
            warnings.append(f"sheets[{i}].title > 31 chars (Excel will truncate)")

        has_any = any(
            isinstance(sh.get(k), (dict, list)) and bool(sh.get(k))
            for k in ("cells", "tables", "dropdowns", "validations", "conditional_formats")
        )
        if not has_any:
            warnings.append(f"sheets[{i}] has no cells/tables/validations; will likely be empty")

        cells = sh.get("cells")
        if cells is not None:
            if not isinstance(cells, dict):
                errors.append(f"sheets[{i}].cells must be an object")
            else:
                for addr in cells.keys():
                    a = str(addr or "").strip().upper()
                    if not _is_cell_addr(a):
                        errors.append(f"sheets[{i}].cells has invalid address: {addr!r}")

        for key in ("dropdowns", "validations", "conditional_formats", "range_styles"):
            arr = sh.get(key)
            if arr is None:
                continue
            if not isinstance(arr, list):
                errors.append(f"sheets[{i}].{key} must be an array")
                continue
            for j, item in enumerate(arr):
                if not isinstance(item, dict):
                    errors.append(f"sheets[{i}].{key}[{j}] must be an object")
                    continue
                rng = str(item.get("range") or "").strip().upper()
                if not rng:
                    errors.append(f"sheets[{i}].{key}[{j}].range is required")
                    continue
                if not (_is_cell_addr(rng) or _is_a1_range(rng)):
                    errors.append(f"sheets[{i}].{key}[{j}].range invalid: {rng!r}")

    return {"ok": not errors, "errors": errors, "warnings": warnings}


def validate_excel_spec_master(spec: Dict[str, Any]) -> Dict[str, Any]:
    errors: List[str] = []
    warnings: List[str] = []

    if not isinstance(spec, dict):
        return {"ok": False, "errors": ["spec must be a JSON object"], "warnings": []}

    sheets = spec.get("sheets")
    if not isinstance(sheets, list) or not sheets:
        return {"ok": False, "errors": ["spec.sheets must be a non-empty array"], "warnings": []}

    has_master_sheet = False

    for i, sh in enumerate(sheets):
        if not isinstance(sh, dict):
            errors.append(f"sheets[{i}] must be an object")
            continue

        title = str(sh.get("title") or "").strip()
        if not title:
            errors.append(f"sheets[{i}].title is required")
        elif len(title) > 31:
            warnings.append(f"sheets[{i}].title > 31 chars (Excel will truncate)")

        cells = sh.get("cells")
        if cells is not None:
            if not isinstance(cells, dict):
                errors.append(f"sheets[{i}].cells must be an object")
            else:
                for addr in cells.keys():
                    a = str(addr or "").strip().upper()
                    if not _is_cell_addr(a):
                        errors.append(f"sheets[{i}].cells has invalid address: {addr!r}")

        for key in ("dropdowns", "validations", "conditional_formats", "range_styles"):
            arr = sh.get(key)
            if arr is None:
                continue
            if not isinstance(arr, list):
                errors.append(f"sheets[{i}].{key} must be an array")
                continue
            for j, item in enumerate(arr):
                if not isinstance(item, dict):
                    errors.append(f"sheets[{i}].{key}[{j}] must be an object")
                    continue
                rng = str(item.get("range") or "").strip().upper()
                if not rng:
                    errors.append(f"sheets[{i}].{key}[{j}].range is required")
                    continue
                if not (_is_cell_addr(rng) or _is_a1_range(rng)):
                    errors.append(f"sheets[{i}].{key}[{j}].range invalid: {rng!r}")

        tables = sh.get("tables") or []
        if tables is not None and not isinstance(tables, list):
            errors.append(f"sheets[{i}].tables must be an array")
            tables = []
        for j, t in enumerate(tables):
            if not isinstance(t, dict):
                errors.append(f"sheets[{i}].tables[{j}] must be an object")
                continue
            nm = str(t.get("name") or "").strip()
            ref = str(t.get("ref") or "").strip().upper()
            if not nm:
                errors.append(f"sheets[{i}].tables[{j}].name is required")
            if not ref:
                errors.append(f"sheets[{i}].tables[{j}].ref is required")
            elif not (_is_a1_range(ref)):
                errors.append(f"sheets[{i}].tables[{j}].ref invalid: {ref!r}")

        has_tables = isinstance(sh.get("tables"), list) and len(sh.get("tables") or []) > 0
        has_valids = (isinstance(sh.get("validations"), list) and len(sh.get("validations") or []) > 0) or (
            isinstance(sh.get("dropdowns"), list) and len(sh.get("dropdowns") or []) > 0
        )
        has_headers = isinstance(sh.get("header_rows"), list) and len(sh.get("header_rows") or []) > 0
        has_freeze = bool(str(sh.get("freeze_panes") or "").strip())
        if has_tables and has_valids and has_headers and has_freeze:
            has_master_sheet = True

    if not has_master_sheet:
        errors.append("No sheet meets master minimum: requires tables + (validations or dropdowns) + header_rows + freeze_panes on at least one sheet.")

    return {"ok": not errors, "errors": errors, "warnings": warnings}


def maybe_repair_excel_spec_once(
    spec: Dict[str, Any],
    report: Dict[str, Any],
    *,
    call_openai_chat_fn: Optional[Callable[[List[Dict[str, str]]], str]] = None,
) -> Tuple[Dict[str, Any], Dict[str, Any], bool]:
    if not isinstance(report, dict) or report.get("ok") is True:
        return spec, report, False
    if call_openai_chat_fn is None:
        return spec, report, False

    sys = (
        "You are an Excel workbook SPEC REPAIR tool.\n"
        "Fix the JSON spec so it will generate a usable professional workbook.\n"
        "Return ONLY valid JSON (no markdown, no tags, no commentary).\n"
        "Constraints:\n"
        "- Keep it compact (structure + formulas + validations; no huge data dumps).\n"
        "- Ensure at least one sheet has meaningful headers and a table range if appropriate.\n"
        "- All cell addresses and ranges must be valid A1 notation.\n"
    )
    user = json.dumps(
        {"errors": report.get("errors") or [], "warnings": report.get("warnings") or [], "spec": spec},
        ensure_ascii=False,
        indent=2,
    )

    try:
        raw = call_openai_chat_fn([{"role": "system", "content": sys}, {"role": "user", "content": user}])
        repaired = json.loads(raw)
        if not isinstance(repaired, dict):
            return spec, report, False
        new_report = validate_excel_spec(repaired)
        if new_report.get("ok") is True:
            return repaired, new_report, True
        return spec, report, False
    except Exception:
        return spec, report, False


def maybe_repair_excel_spec_once_master(
    spec: Dict[str, Any],
    report: Dict[str, Any],
    *,
    call_openai_chat_fn: Optional[Callable[[List[Dict[str, str]]], str]] = None,
) -> Tuple[Dict[str, Any], Dict[str, Any], bool]:
    if not isinstance(report, dict) or report.get("ok") is True:
        return spec, report, False
    if call_openai_chat_fn is None:
        return spec, report, False

    sys = (
        "You are an Excel SPEC REPAIR tool.\n"
        "Fix the JSON spec so it passes the master-quality validator.\n"
        "Return ONLY valid JSON (no markdown, no commentary).\n"
        "Must ensure at least one sheet has:\n"
        "- tables (valid A1 ref)\n"
        "- validations or dropdowns\n"
        "- header_rows\n"
        "- freeze_panes\n"
        "Use range_styles and conditional_formats to visually separate INPUT/CALC/OUTPUT.\n"
    )
    user = json.dumps({"report": report, "spec": spec}, ensure_ascii=False, indent=2)

    try:
        raw = call_openai_chat_fn([{"role": "system", "content": sys}, {"role": "user", "content": user}])
        repaired = json.loads(raw)
        if not isinstance(repaired, dict):
            return spec, report, False
        new_report = validate_excel_spec_master(repaired)
        if new_report.get("ok") is True:
            return repaired, new_report, True
        return spec, report, False
    except Exception:
        return spec, report, False


# ----------------------------
# Workbook generator (template-aware)
# ----------------------------

def generate_workbook_from_spec(spec: Dict[str, Any], *, project_raw_dir: Optional[Path] = None) -> bytes:
    if Workbook is None:
        raise RuntimeError("openpyxl Workbook is not available. Install openpyxl.")
    if Font is None or Alignment is None or PatternFill is None or Border is None or Side is None:
        raise RuntimeError("openpyxl style components not available. Upgrade/openpyxl install is incomplete.")

    def _mk_font(d: Dict[str, Any]) -> Font:
        kw: Dict[str, Any] = {}
        if "bold" in d:
            kw["bold"] = bool(d.get("bold"))
        if "italic" in d:
            kw["italic"] = bool(d.get("italic"))
        if "underline" in d:
            kw["underline"] = "single" if bool(d.get("underline")) else None
        if "size" in d:
            try:
                kw["size"] = float(d.get("size"))
            except Exception:
                pass
        if "name" in d and str(d.get("name") or "").strip():
            kw["name"] = str(d.get("name")).strip()
        if "color" in d and str(d.get("color") or "").strip():
            kw["color"] = str(d.get("color")).strip()
        return Font(**kw)

    def _mk_fill(d: Dict[str, Any]) -> PatternFill:
        color = str(d.get("color") or "").strip()
        ptype = str(d.get("patternType") or "solid").strip() or "solid"
        if not color:
            return PatternFill()
        return PatternFill(patternType=ptype, fgColor=color)

    def _mk_alignment(d: Dict[str, Any]) -> Alignment:
        kw: Dict[str, Any] = {}
        for k in ("horizontal", "vertical"):
            if str(d.get(k) or "").strip():
                kw[k] = str(d.get(k)).strip()
        if "wrap_text" in d:
            kw["wrap_text"] = bool(d.get("wrap_text"))
        if "text_rotation" in d:
            try:
                kw["text_rotation"] = int(d.get("text_rotation"))
            except Exception:
                pass
        if "shrink_to_fit" in d:
            kw["shrink_to_fit"] = bool(d.get("shrink_to_fit"))
        if "indent" in d:
            try:
                kw["indent"] = int(d.get("indent"))
            except Exception:
                pass
        return Alignment(**kw)

    def _mk_border(d: Dict[str, Any]) -> Border:
        style = str(d.get("style") or "thin").strip() or "thin"
        color = str(d.get("color") or "FF9CA3AF").strip() or "FF9CA3AF"
        side = Side(style=style, color=color)
        return Border(left=side, right=side, top=side, bottom=side)

    def _apply_style(cell, style: Dict[str, Any]) -> None:
        if not isinstance(style, dict) or not style:
            return
        if "font" in style and isinstance(style.get("font"), dict):
            cell.font = _mk_font(style["font"])
        if "fill" in style and isinstance(style.get("fill"), dict):
            cell.fill = _mk_fill(style["fill"])
        if "alignment" in style and isinstance(style.get("alignment"), dict):
            cell.alignment = _mk_alignment(style["alignment"])
        if "border" in style and isinstance(style.get("border"), dict):
            cell.border = _mk_border(style["border"])
        if "number_format" in style and style.get("number_format") is not None:
            cell.number_format = str(style.get("number_format"))

    def _apply_style_to_range(ws, rng: str, style: Dict[str, Any]) -> None:
        if not rng:
            return
        try:
            if _is_cell_addr(rng):
                _apply_style(ws[rng], style)
                return
            if range_boundaries is None:
                return
            min_col, min_row, max_col, max_row = range_boundaries(rng)
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    _apply_style(ws.cell(row=r, column=c), style)
        except Exception:
            return

    template_from_raw = str(spec.get("template_from_raw") or "").strip()
    auto_template = bool(spec.get("auto_template_latest_xlsx") or False)

    wb = None
    if (template_from_raw or auto_template) and project_raw_dir is not None:
        try:
            rawp = project_raw_dir

            template_path = None
            if template_from_raw:
                candidate = rawp / Path(template_from_raw).name
                if candidate.exists() and candidate.is_file():
                    template_path = candidate
            elif auto_template:
                xlsx_candidates = sorted(
                    [p for p in rawp.iterdir() if p.is_file() and p.suffix.lower() in (".xlsx", ".xlsm")],
                    key=lambda p: p.stat().st_mtime,
                    reverse=True,
                )
                if xlsx_candidates:
                    template_path = xlsx_candidates[0]

            if template_path is not None and load_workbook is not None:
                wb = load_workbook(template_path, read_only=False, data_only=False)
        except Exception:
            wb = None

    if wb is None:
        wb = Workbook()

    clones = spec.get("clone_sheets") or []
    if isinstance(clones, list) and clones:
        for item in clones:
            if not isinstance(item, dict):
                continue
            src = str(item.get("from") or "").strip()
            dst = str(item.get("to") or "").strip()
            if not src or not dst:
                continue
            try:
                if src in wb.sheetnames:
                    ws_src = wb[src]
                    ws_new = wb.copy_worksheet(ws_src)
                    ws_new.title = dst[:31]
            except Exception:
                continue

    if "sheets" in spec and isinstance(spec.get("sheets"), list) and spec["sheets"]:
        try:
            wb.remove(wb.active)
        except Exception:
            pass

    sheets = spec.get("sheets") or []
    if not isinstance(sheets, list) or not sheets:
        sheets = []

    for s in sheets:
        if not isinstance(s, dict):
            continue

        title = str(s.get("title") or "Sheet").strip()[:31] or "Sheet"
        if title in wb.sheetnames:
            ws = wb[title]
        else:
            ws = wb.create_sheet(title=title)

        if "gridlines" in s:
            try:
                ws.sheet_view.showGridLines = bool(s.get("gridlines"))
            except Exception:
                pass
        if "zoom" in s:
            try:
                ws.sheet_view.zoomScale = int(s.get("zoom"))
            except Exception:
                pass

        fp = str(s.get("freeze_panes") or "").strip().upper()
        if fp:
            try:
                ws.freeze_panes = fp
            except Exception:
                pass

        col_widths = s.get("col_widths") or {}
        if isinstance(col_widths, dict):
            for col, w in col_widths.items():
                c = str(col or "").strip().upper()
                try:
                    ws.column_dimensions[c].width = float(w)
                except Exception:
                    continue

        row_heights = s.get("row_heights") or {}
        if isinstance(row_heights, dict):
            for rr, h in row_heights.items():
                try:
                    r = int(rr)
                    ws.row_dimensions[r].height = float(h)
                except Exception:
                    continue

        cells = s.get("cells") or {}
        if isinstance(cells, dict):
            for addr, val in cells.items():
                a = str(addr or "").strip().upper()
                if not a:
                    continue
                cell = ws[a]
                if isinstance(val, dict):
                    if "f" in val and val.get("f") is not None:
                        cell.value = str(val.get("f"))
                    elif "v" in val:
                        cell.value = val.get("v")
                    else:
                        cell.value = json.dumps(val, ensure_ascii=False)

                    inline_style: Dict[str, Any] = {}
                    if "number_format" in val and val.get("number_format") is not None:
                        inline_style["number_format"] = str(val.get("number_format"))
                    if "style" in val and isinstance(val.get("style"), dict):
                        inline_style.update(val.get("style") or {})
                    if inline_style:
                        _apply_style(cell, inline_style)
                else:
                    cell.value = val

        header_rows = s.get("header_rows") or []
        if isinstance(header_rows, list):
            for rr in header_rows:
                try:
                    r = int(rr)
                except Exception:
                    continue
                if r <= 0:
                    continue
                max_c = max(26, int(ws.max_column or 26))
                for c in range(1, max_c + 1):
                    ws.cell(row=r, column=c).font = Font(bold=True)

        range_styles = s.get("range_styles") or []
        if isinstance(range_styles, list):
            for rs in range_styles:
                if not isinstance(rs, dict):
                    continue
                rng = str(rs.get("range") or "").strip().upper()
                style = rs.get("style") or {}
                if rng and isinstance(style, dict):
                    _apply_style_to_range(ws, rng, style)

        dropdowns = s.get("dropdowns") or []
        if dropdowns and DataValidation is None:
            raise RuntimeError("openpyxl DataValidation is not available. Upgrade/openpyxl install is incomplete.")
        if isinstance(dropdowns, list):
            for d in dropdowns:
                if not isinstance(d, dict):
                    continue
                rng = str(d.get("range") or "").strip().upper()
                if not rng:
                    continue
                allow_blank = bool(d.get("allow_blank", True))
                if "values" in d and isinstance(d.get("values"), list):
                    vals = [str(x) for x in d.get("values") if str(x).strip()]
                    joined = ",".join(vals)
                    formula1 = f"\"{joined}\""
                else:
                    formula1 = str(d.get("formula1") or "").strip()
                    if not formula1:
                        continue
                dv = DataValidation(type="list", formula1=formula1, allow_blank=allow_blank)
                ws.add_data_validation(dv)
                dv.add(rng)

        validations = s.get("validations") or []
        if validations and DataValidation is None:
            raise RuntimeError("openpyxl DataValidation is not available. Upgrade/openpyxl install is incomplete.")
        if isinstance(validations, list):
            for v in validations:
                if not isinstance(v, dict):
                    continue
                rng = str(v.get("range") or "").strip().upper()
                vtype = str(v.get("type") or "").strip()
                if not rng or not vtype:
                    continue
                allow_blank = bool(v.get("allow_blank", True))
                op = (str(v.get("operator") or "").strip() or None)
                f1 = v.get("formula1")
                f2 = v.get("formula2")

                formula1 = "" if f1 is None else str(f1)
                formula2 = "" if f2 is None else str(f2)

                dv = DataValidation(type=vtype, operator=op, formula1=formula1 or None, formula2=formula2 or None, allow_blank=allow_blank)
                if str(v.get("prompt_title") or "").strip():
                    dv.promptTitle = str(v.get("prompt_title")).strip()
                if str(v.get("prompt") or "").strip():
                    dv.prompt = str(v.get("prompt")).strip()
                if str(v.get("error_title") or "").strip():
                    dv.errorTitle = str(v.get("error_title")).strip()
                if str(v.get("error") or "").strip():
                    dv.error = str(v.get("error")).strip()
                if str(v.get("error_style") or "").strip():
                    dv.errorStyle = str(v.get("error_style")).strip()

                ws.add_data_validation(dv)
                dv.add(rng)

        cfs = s.get("conditional_formats") or []
        if cfs and (FormulaRule is None or CellIsRule is None or DifferentialStyle is None):
            raise RuntimeError("openpyxl conditional formatting support is not available.")
        if isinstance(cfs, list):
            for cf in cfs:
                if not isinstance(cf, dict):
                    continue
                rng = str(cf.get("range") or "").strip().upper()
                cftype = str(cf.get("type") or "").strip()
                if not rng or not cftype:
                    continue

                style = cf.get("style") or {}
                dxf = DifferentialStyle()
                if isinstance(style, dict):
                    if "font" in style and isinstance(style.get("font"), dict):
                        dxf.font = _mk_font(style["font"])
                    if "fill" in style and isinstance(style.get("fill"), dict):
                        dxf.fill = _mk_fill(style["fill"])

                if cftype == "cellIs":
                    op = str(cf.get("operator") or "").strip()
                    formulas = cf.get("formula") or []
                    if not isinstance(formulas, list) or not formulas:
                        continue
                    try:
                        rule = CellIsRule(operator=op, formula=[str(x) for x in formulas], dxf=dxf)
                        ws.conditional_formatting.add(rng, rule)
                    except Exception:
                        continue
                elif cftype == "formula":
                    formulas = cf.get("formula") or []
                    if not isinstance(formulas, list) or not formulas:
                        continue
                    try:
                        rule = FormulaRule(formula=[str(x) for x in formulas], dxf=dxf)
                        ws.conditional_formatting.add(rng, rule)
                    except Exception:
                        continue

        tables = s.get("tables") or []
        if tables and (Table is None or TableStyleInfo is None):
            raise RuntimeError("openpyxl Table support is not available.")
        if isinstance(tables, list):
            for t in tables:
                if not isinstance(t, dict):
                    continue
                nm = str(t.get("name") or "").strip()
                ref = str(t.get("ref") or "").strip().upper()
                if not nm or not ref:
                    continue
                tab = Table(displayName=nm, ref=ref)
                style_name = str(t.get("style") or "TableStyleMedium9").strip()
                show_row = bool(t.get("showRowStripes", True))
                show_col = bool(t.get("showColumnStripes", False))
                tab.tableStyleInfo = TableStyleInfo(
                    name=style_name,
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=show_row,
                    showColumnStripes=show_col,
                )
                ws.add_table(tab)

        prot = s.get("protection") or {}
        if isinstance(prot, dict) and bool(prot.get("enabled", False)):
            unlock_ranges = prot.get("unlock_ranges") or []
            lock_ranges = prot.get("lock_ranges") or []
            try:
                if range_boundaries is not None:
                    max_r = max(1, int(ws.max_row or 1))
                    max_c = max(1, int(ws.max_column or 1))
                    for r in range(1, max_r + 1):
                        for c in range(1, max_c + 1):
                            ws.cell(row=r, column=c).protection = ws.cell(row=r, column=c).protection.copy(locked=True)

                    for rng in lock_ranges:
                        rr = str(rng or "").strip().upper()
                        if not rr:
                            continue
                        if _is_cell_addr(rr):
                            ws[rr].protection = ws[rr].protection.copy(locked=True)
                        elif _is_a1_range(rr):
                            min_col, min_row, max_col, max_row = range_boundaries(rr)
                            for r in range(min_row, max_row + 1):
                                for c in range(min_col, max_col + 1):
                                    ws.cell(row=r, column=c).protection = ws.cell(row=r, column=c).protection.copy(locked=True)

                    for rng in unlock_ranges:
                        rr = str(rng or "").strip().upper()
                        if not rr:
                            continue
                        if _is_cell_addr(rr):
                            ws[rr].protection = ws[rr].protection.copy(locked=False)
                        elif _is_a1_range(rr):
                            min_col, min_row, max_col, max_row = range_boundaries(rr)
                            for r in range(min_row, max_row + 1):
                                for c in range(min_col, max_col + 1):
                                    ws.cell(row=r, column=c).protection = ws.cell(row=r, column=c).protection.copy(locked=False)
            except Exception:
                pass

            try:
                ws.protection.sheet = True
                pwd = str(prot.get("password") or "").strip()
                if pwd:
                    ws.protection.set_password(pwd)
            except Exception:
                pass

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
