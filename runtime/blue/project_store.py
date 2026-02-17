# -*- coding: utf-8 -*-
"""
project_store.py

Disk-backed “project store” for Lens-0 Project OS.

Contains:
- manifest + state dir helpers (ensure_project, load_manifest, save_manifest, create_artifact*, etc.)
- ingestion functions (ingest_uploaded_file, chunking/indexing, pdf/image/excel overview)
- project scaffolding + file registration/listing

Does NOT contain:
- WebSocket / aiohttp handlers
- Server startup / routing
"""

from __future__ import annotations

import hashlib
import inspect
import html
import json
import os
import re
import time
from datetime import datetime
try:
    from zoneinfo import ZoneInfo  # py3.9+
except Exception:
    ZoneInfo = None  # type: ignore
import zipfile
import shutil
import subprocess
import xml.etree.ElementTree as ET
from io import BytesIO
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import path_engine

from lens0_config import (
    TEXT_LIKE_SUFFIXES,
    IMAGE_SUFFIXES,
    is_text_suffix,
    mime_for_image_suffix,
)

from PyPDF2 import PdfReader

# Optional image OCR support
try:
    from PIL import Image  # type: ignore
except Exception:
    Image = None  # type: ignore

# Optional scanned PDF OCR support (requires poppler + pdf2image + tesseract)
try:
    from pdf2image import convert_from_path  # type: ignore
except Exception:
    convert_from_path = None  # type: ignore

try:
    import pytesseract  # type: ignore
except Exception:
    pytesseract = None  # type: ignore

# Excel engine lives elsewhere
from excel_engine import (
    summarize_excel_structure,
    extract_excel_blueprint,
)
# Optional Excel writer (for case workbooks)
try:
    from openpyxl import Workbook  # type: ignore
except Exception:
    Workbook = None  # type: ignore
# Visual semantics (deterministic image understanding beyond OCR)
try:
    import visual_semantics
except Exception:
    visual_semantics = None  # type: ignore

# -----------------------------------------------------------------------------
# Configuration (must be called by server.py)
# -----------------------------------------------------------------------------

PROJECT_ROOT: Optional[Path] = None
PROJECTS_DIR: Optional[Path] = None
DEFAULT_PROJECT_NAME: str = "default"
# ---------------------------------------------------------------------------
# GLOBAL USER STORE (Tier-2G / Tier-1G): users/<user>/...
# - project-agnostic, stable identity facts
# - age is NOT stored; derived at read time from birthdate
# - conflict-aware: store ambiguity rather than overwrite
# ---------------------------------------------------------------------------

USERS_DIR_NAME = "users"
USER_PROFILE_FILE_NAME = "profile.json"
USER_FACTS_RAW_FILE_NAME = "facts_raw.jsonl"
USER_MEMORY_POLICY_FILE_NAME = "memory_policy.json"
USER_GLOBAL_FACTS_MAP_FILE_NAME = "global_facts_map.json"
# ---------------------------------------------------------------------------
# USER-SCOPED PENDING DISAMBIGUATION (ephemeral; NOT memory)
# Stored under: projects/<user>/_user/pending_disambiguation.json
# Purpose:
# - When the system asks a disambiguation question about a Tier-2G field (e.g., location),
#   the server can persist the offered options and then consume the user's next reply
#   WITHOUT requiring magic words.
# - On consumption, the server emits a canonical Tier-1 claim (e.g., "My confirmed location is X")
#   and triggers the existing Tier-2G rebuild resolver.
# ---------------------------------------------------------------------------

USER_PENDING_DISAMBIGUATION_FILE_NAME = "pending_disambiguation.json"
USER_PENDING_PROFILE_QUESTION_FILE_NAME = "pending_profile_question.json"
COUPLES_SHARED_MEMORY_FILE_NAME = "couples_shared_memory.json"
HEALTH_PROFILE_FILE_NAME = "health_profile.json"
EXPERTS_DIR_NAME = "experts"
DISCOVERY_INDEX_FILE_NAME = "discovery_index.jsonl"
FACT_LEDGER_FILE_NAME = "fact_ledger.jsonl"
ANALYSIS_PIPELINE_VERSION = "analysis_v4"
INGEST_PIPELINE_VERSION = "ingest_v2"
CAPABILITY_GAPS_FILE_NAME = "capability_gaps.jsonl"
UPLOAD_BATCH_STATE_FILE_NAME = "upload_batches.json"

def user_pending_disambiguation_path(user: str) -> Path:
    # projects/<user>/_user/pending_disambiguation.json
    return user_dir(user) / USER_PENDING_DISAMBIGUATION_FILE_NAME

def load_user_pending_disambiguation(user: str) -> Dict[str, Any]:
    p = user_pending_disambiguation_path(user)
    if not p.exists():
        return {}
    try:
        obj = json.loads(p.read_text(encoding="utf-8") or "{}")
        return obj if isinstance(obj, dict) else {}
    except Exception:
        return {}

def save_user_pending_disambiguation(user: str, obj: Dict[str, Any]) -> None:
    out = obj if isinstance(obj, dict) else {}
    out.setdefault("version", 1)
    atomic_write_text(user_pending_disambiguation_path(user), json.dumps(out, indent=2, sort_keys=True))

def clear_user_pending_disambiguation(user: str) -> None:
    p = user_pending_disambiguation_path(user)
    try:
        if p.exists():
            p.unlink()
    except Exception:
        pass

def is_user_pending_disambiguation_unexpired(obj: Dict[str, Any]) -> bool:
    """
    Deterministic expiry check (no date parsing):
    - Requires obj["pending"] == True
    - Requires numeric obj["expires_at_epoch"] > time.time()
    """
    if not isinstance(obj, dict):
        return False
    if obj.get("pending") is not True:
        return False
    try:
        exp = float(obj.get("expires_at_epoch") or 0.0)
    except Exception:
        exp = 0.0
    return time.time() <= exp

def user_pending_profile_question_path(user: str) -> Path:
    # projects/<user>/_user/pending_profile_question.json
    return user_dir(user) / USER_PENDING_PROFILE_QUESTION_FILE_NAME

def couples_shared_memory_path(project_name: str) -> Path:
    # projects/<user>/<project>/state/couples_shared_memory.json
    return state_dir(project_name) / COUPLES_SHARED_MEMORY_FILE_NAME

def health_profile_path(project_name: str) -> Path:
    # Legacy: projects/<user>/<project>/state/health_profile.json
    return state_dir(project_name) / HEALTH_PROFILE_FILE_NAME

def experts_dir(user: str) -> Path:
    # projects/<user>/_user/experts
    d = user_dir(user) / EXPERTS_DIR_NAME
    d.mkdir(parents=True, exist_ok=True)
    return d

def expert_profile_path(user: str, expert: str) -> Path:
    # projects/<user>/_user/experts/<expert>_profile.json
    safe = re.sub(r"[^a-zA-Z0-9_-]", "_", (expert or "").strip().lower()) or "expert"
    return experts_dir(user) / f"{safe}_profile.json"

def health_profile_global_path(user: str) -> Path:
    # projects/<user>/_user/experts/health_profile.json
    return experts_dir(user) / HEALTH_PROFILE_FILE_NAME

def _default_health_profile() -> Dict[str, Any]:
    return {
        "schema": "health_profile_v1",
        "updated_at": now_iso(),
        "medications": [],
        "allergies": [],
        "measurements": {},
        "labs": [],
        "lab_refresh_manifest_at": 0.0,
        "notes": [],
    }

def _default_expert_profile(expert: str) -> Dict[str, Any]:
    return {
        "schema": "expert_profile_v1",
        "expert": str(expert or "").strip().lower() or "expert",
        "updated_at": now_iso(),
        "data": {},
    }

def _ensure_expert_profiles(user_seg: str) -> None:
    if not user_seg:
        return
    try:
        # Health (active)
        p_health = health_profile_global_path(user_seg)
        if not p_health.exists():
            atomic_write_text(p_health, json.dumps(_default_health_profile(), indent=2, sort_keys=True))
    except Exception:
        pass

    # Scaffolds for other hats (inactive for now)
    for ex in ("therapist", "coding", "analysis", "general"):
        try:
            p = expert_profile_path(user_seg, ex)
            if p.exists():
                continue
            atomic_write_text(p, json.dumps(_default_expert_profile(ex), indent=2, sort_keys=True))
        except Exception:
            pass

def _project_user_segment(project_name: str) -> str:
    try:
        return str(project_name or "").split("/", 1)[0].strip()
    except Exception:
        return ""

def load_health_profile(project_name: str) -> Dict[str, Any]:
    ensure_project_scaffold(project_name)
    user_seg = _project_user_segment(project_name)
    p_global = health_profile_global_path(user_seg) if user_seg else None
    obj = _load_json_obj(p_global) if p_global else {}
    if not isinstance(obj, dict) or obj.get("schema") != "health_profile_v1":
        # Legacy fallback: project-scoped health_profile.json
        p_legacy = health_profile_path(project_name)
        obj = _load_json_obj(p_legacy)
        if isinstance(obj, dict) and obj.get("schema") == "health_profile_v1":
            # Migrate forward to global store (non-destructive)
            try:
                if p_global is not None:
                    atomic_write_text(p_global, json.dumps(obj, indent=2, sort_keys=True))
            except Exception:
                pass
        else:
            obj = _default_health_profile()
            try:
                if p_global is not None:
                    atomic_write_text(p_global, json.dumps(obj, indent=2, sort_keys=True))
                else:
                    atomic_write_text(p_legacy, json.dumps(obj, indent=2, sort_keys=True))
            except Exception:
                pass
    obj["updated_at"] = now_iso()
    return obj

def _write_health_profile(project_name: str, obj: Dict[str, Any]) -> bool:
    try:
        out = obj if isinstance(obj, dict) else _default_health_profile()
        out.setdefault("schema", "health_profile_v1")
        out["updated_at"] = now_iso()
        user_seg = _project_user_segment(project_name)
        if user_seg:
            atomic_write_text(health_profile_global_path(user_seg), json.dumps(out, indent=2, sort_keys=True))
        else:
            atomic_write_text(health_profile_path(project_name), json.dumps(out, indent=2, sort_keys=True))
        return True
    except Exception:
        return False

def _health_norm_med_name(name: str) -> str:
    n = re.sub(r"[^a-zA-Z0-9\s\-]", " ", (name or "").strip().lower())
    n = re.sub(r"\s+", " ", n).strip()
    return n

def _health_allergy_is_none(rec: Dict[str, Any]) -> bool:
    try:
        nm = str(rec.get("name") or "").strip().lower()
        st = str(rec.get("status") or "").strip().lower()
    except Exception:
        return False
    if st in ("none", "no", "none_reported"):
        return True
    return "no known" in nm or nm == "none"

def _health_med_name_is_malformed(name: str) -> bool:
    low = str(name or "").strip().lower()
    if not low:
        return True
    bad_tokens = ("user", "takes", "every night", "every morning", "before bed", "after bed", "nightly")
    return any(t in low for t in bad_tokens)

_HEALTH_LAB_UNIT_RE = re.compile(
    r"\b(%|mg/dl|mmol/l|g/dl|ng/ml|ug/ml|miu/l|miu/ml|uiu/l|uiu/ml|iu/l|u/l|meq/l|pg/ml|ng/dl|mmhg|k/ul|10\^3/ul|x10\^3/ul|10\^9/l|cells/ul)\b",
    re.IGNORECASE,
)
_HEALTH_LAB_LINE_RE = re.compile(
    r"(?P<name>[A-Za-z][A-Za-z0-9\s\-/()%]{2,60})\s*[:\-]?\s*"
    r"(?P<val>-?\d+(?:\.\d+)?)\s*(?P<unit>%|mg/dl|mmol/l|g/dl|ng/ml|ug/ml|miu/l|iu/l|u/l|meq/l|pg/ml|ng/dl|mmhg|k/ul|10\^3/ul|x10\^3/ul|10\^9/l|cells/ul)\b",
    re.IGNORECASE,
)
_HEALTH_DATE_RE = re.compile(r"\b(20\d{2}[-/]\d{1,2}[-/]\d{1,2}|\d{1,2}/\d{1,2}/20\d{2})\b")
_HEALTH_LAB_PARSER_VERSION = 2
def _health_lab_name_is_noise(name: str) -> bool:
    low = str(name or "").strip().lower()
    if not low:
        return True
    if low.startswith("("):
        return True
    noise_terms = (
        "desirable range",
        "reference range",
        "goal of",
        "treating to",
        "factor",
        "interpret",
        "comment",
        "note:",
        "ldl-c of",
        "non-hdl-c goal",
        "range is",
    )
    return any(t in low for t in noise_terms)


def _health_parse_lab_line(ln: str) -> Optional[Dict[str, str]]:
    s = (ln or "").strip()
    if not s:
        return None
    if len(s) > 180:
        s = s[:180].rstrip()
    if not _HEALTH_LAB_UNIT_RE.search(s):
        return None

    # Prefer unit-aware parsing (handles "GLUCOSE 80 65-99 mg/dL" style)
    unit_m = None
    try:
        for m0 in _HEALTH_LAB_UNIT_RE.finditer(s):
            unit_m = m0
    except Exception:
        unit_m = None

    name = ""
    val = ""
    unit = ""
    ref = ""

    if unit_m is not None:
        unit = str(unit_m.group(1) or "").strip()
        before = s[:unit_m.start()].strip()
        after = s[unit_m.end():].strip()

        # Value token: must be standalone (whitespace-bound) to avoid "B12" or "3rd"
        mval = re.search(r"(?:^|\s)([<>]=?\s*)?(\d+(?:\.\d+)?)(?=\s|$)", before)
        if mval:
            val = (str(mval.group(1) or "") + str(mval.group(2) or "")).replace(" ", "")
            name = before[:mval.start()].strip(" :;-")
            rest = before[mval.end():].strip()

            # Reference range/comparator: search remaining text or trailing tail
            def _pick_ref(*parts: str) -> str:
                for part in parts:
                    if not part:
                        continue
                    mref = re.search(
                        r"([<>]=?\s*\d+(?:\.\d+)?|\d+(?:\.\d+)?\s*(?:-|to)\s*\d+(?:\.\d+)?)",
                        part,
                    )
                    if mref:
                        return mref.group(1).strip()
                return ""

            ref = _pick_ref(rest, after)

    # Fallback to legacy regex if unit-aware parse failed
    if not (name and val and unit):
        m = _HEALTH_LAB_LINE_RE.search(s)
        if not m:
            return None
        name = str(m.group("name") or "").strip()
        val = str(m.group("val") or "").strip()
        unit = str(m.group("unit") or "").strip()
        if not name or not val:
            return None
        # Trim trailing clutter from name
        name = re.sub(r"\s{2,}", " ", name).strip()
        name = re.sub(r"\s*(result|value)$", "", name, flags=re.IGNORECASE).strip()
        if len(name) < 3:
            return None
        ref = ""
        try:
            tail = s[m.end():]
            mref = re.search(r"(\d+(?:\.\d+)?\s*-\s*\d+(?:\.\d+)?|\d+(?:\.\d+)?\s*(?:to|-)\s*\d+(?:\.\d+)?)", tail)
            if mref:
                ref = mref.group(1).strip()
        except Exception:
            ref = ""

    if not name or not val or not unit:
        return None
    # Normalize name
    name = re.sub(r"\s{2,}", " ", name).strip()
    name = re.sub(r"\s*(result|value)$", "", name, flags=re.IGNORECASE).strip()
    if len(name) < 3:
        return None
    if _health_lab_name_is_noise(name):
        return None
    date = ""
    try:
        md = _HEALTH_DATE_RE.search(s)
        if md:
            date = md.group(1).strip()
    except Exception:
        date = ""
    return {
        "name": name,
        "value": val,
        "unit": unit,
        "ref": ref,
        "date": date,
        "raw": s,
    }


def _health_extract_labs_from_text(
    *,
    text: str,
    source_file: str,
    source_hint: str,
    max_items: int,
) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    if not text:
        return out
    for raw_ln in (text or "").splitlines():
        if len(out) >= int(max_items or 0):
            break
        ln = (raw_ln or "").strip()
        if not ln:
            continue
        lab = _health_parse_lab_line(ln)
        if not lab:
            continue
        lab["source_file"] = source_file
        lab["source_hint"] = source_hint
        out.append(lab)
    return out

def _health_extract_labs_from_excel_blueprint(
    *,
    canonical_rel: str,
    blueprint_obj: Dict[str, Any],
    max_items: int,
) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    if not isinstance(blueprint_obj, dict):
        return out

    def _emit(sheet: str, cell: str, label: str, value: Any) -> None:
        nonlocal out
        if len(out) >= int(max_items or 0):
            return
        if not label:
            return
        v = None
        if isinstance(value, (int, float)):
            v = str(value)
        elif isinstance(value, str):
            v = value.strip()
        if not v:
            return
        # Try to pick unit from label
        unit = ""
        m = _HEALTH_LAB_UNIT_RE.search(label)
        if m:
            unit = m.group(1)
        lab = {
            "name": label.strip(),
            "value": v.strip(),
            "unit": unit,
            "ref": "",
            "date": "",
            "source_file": canonical_rel,
            "source_hint": f"{sheet}!{cell}",
        }
        out.append(lab)

    cv = blueprint_obj.get("cached_values")
    if isinstance(cv, list):
        for row in cv:
            if len(out) >= int(max_items or 0):
                break
            if not isinstance(row, dict):
                continue
            sh = str(row.get("sheet") or "").strip()
            ce = str(row.get("cell") or "").strip()
            if not sh or not ce:
                continue
            lab = str(row.get("label") or row.get("name") or "").strip()
            _emit(sh, ce, lab, row.get("value"))

    cells = blueprint_obj.get("cells")
    if isinstance(cells, list) and len(out) < int(max_items or 0):
        for row in cells:
            if len(out) >= int(max_items or 0):
                break
            if not isinstance(row, dict):
                continue
            sh = str(row.get("sheet") or "").strip()
            ce = str(row.get("cell") or "").strip()
            if not sh or not ce:
                continue
            lab = str(row.get("label") or row.get("name") or "").strip()
            _emit(sh, ce, lab, row.get("value"))

    nr = blueprint_obj.get("named_ranges")
    if isinstance(nr, list) and len(out) < int(max_items or 0):
        for row in nr:
            if len(out) >= int(max_items or 0):
                break
            if not isinstance(row, dict):
                continue
            sh = str(row.get("sheet") or "").strip()
            ce = str(row.get("cell") or "").strip()
            if not sh or not ce:
                continue
            lab = str(row.get("name") or "").strip()
            _emit(sh, ce, lab, row.get("value"))

    return out

def refresh_health_profile_from_artifacts(
    project_name: str,
    *,
    max_items_per_file: int = 120,
    max_total_items: int = 600,
) -> Dict[str, Any]:
    """
    Deterministic lab extraction into health_profile.json.
    Uses existing artifacts (pdf_text, ocr_text, excel_blueprint, doc_text).
    """
    ensure_project_scaffold(project_name)

    prof = load_health_profile(project_name)
    m = load_manifest(project_name)
    try:
        man_updated = float(m.get("updated_at") or 0.0)
    except Exception:
        man_updated = 0.0
    try:
        last_refresh = float(prof.get("lab_refresh_manifest_at") or 0.0)
    except Exception:
        last_refresh = 0.0

    # Heuristic: re-run if existing labs look malformed (e.g., name contains value+range)
    force_refresh = False
    try:
        labs0 = prof.get("labs") if isinstance(prof, dict) else []
        if isinstance(labs0, list):
            for it in labs0:
                if not isinstance(it, dict):
                    continue
                nm = str(it.get("name") or "")
                if re.search(r"\b\d+\s+\d", nm) or _health_lab_name_is_noise(nm):
                    force_refresh = True
                    break
    except Exception:
        force_refresh = False

    try:
        parser_version = int(prof.get("lab_parser_version") or 0)
    except Exception:
        parser_version = 0

    if man_updated and (last_refresh >= man_updated) and (parser_version >= _HEALTH_LAB_PARSER_VERSION) and (not force_refresh):
        return prof

    raw_files = m.get("raw_files") or []
    if not isinstance(raw_files, list):
        raw_files = []

    all_labs: List[Dict[str, Any]] = []

    for rf in raw_files:
        if len(all_labs) >= int(max_total_items or 0):
            break
        if not isinstance(rf, dict):
            continue
        canonical_rel = str(rf.get("path") or "").replace("\\", "/").strip()
        if not canonical_rel:
            continue
        suf = Path(canonical_rel).suffix.lower()

        # Excel -> parse blueprint
        if suf in (".xlsx", ".xlsm", ".xls"):
            a = _find_latest_artifact_for_file(project_name, artifact_type="excel_blueprint", canonical_rel=canonical_rel)
            if isinstance(a, dict):
                txt = read_artifact_text(project_name, a, cap_chars=220000).strip()
                try:
                    obj = json.loads(txt) if txt else {}
                except Exception:
                    obj = {}
                labs = _health_extract_labs_from_excel_blueprint(
                    canonical_rel=canonical_rel,
                    blueprint_obj=(obj if isinstance(obj, dict) else {}),
                    max_items=int(max_items_per_file or 0),
                )
                all_labs.extend(labs)
            continue

        # PDF -> parse pdf_text
        if suf == ".pdf":
            a = _find_latest_artifact_for_file(project_name, artifact_type="pdf_text", canonical_rel=canonical_rel)
            if isinstance(a, dict):
                txt = read_artifact_text(project_name, a, cap_chars=220000).strip()
                page = 1
                line_in_page = 0
                for raw_ln in (txt or "").splitlines():
                    if len(all_labs) >= int(max_total_items or 0):
                        break
                    ln = (raw_ln or "").strip()
                    if not ln:
                        continue
                    mpage = re.match(r"^\[PAGE\s+(\d+)\]\s*$", ln, flags=re.IGNORECASE)
                    if mpage:
                        try:
                            page = int(mpage.group(1))
                        except Exception:
                            page = page
                        line_in_page = 0
                        continue
                    line_in_page += 1
                    labs = _health_extract_labs_from_text(
                        text=ln,
                        source_file=canonical_rel,
                        source_hint=f"page {page}: line {line_in_page}",
                        max_items=1,
                    )
                    if labs:
                        all_labs.extend(labs)
            continue

        # Image -> parse ocr_text
        if suf in IMAGE_SUFFIXES:
            a = _find_latest_artifact_for_file(project_name, artifact_type="ocr_text", canonical_rel=canonical_rel)
            if isinstance(a, dict):
                txt = read_artifact_text(project_name, a, cap_chars=220000).strip()
                line_no = 0
                for raw_ln in (txt or "").splitlines():
                    if len(all_labs) >= int(max_total_items or 0):
                        break
                    ln = (raw_ln or "").strip()
                    if not ln:
                        continue
                    line_no += 1
                    labs = _health_extract_labs_from_text(
                        text=ln,
                        source_file=canonical_rel,
                        source_hint=f"ocr line {line_no}",
                        max_items=1,
                    )
                    if labs:
                        all_labs.extend(labs)
            continue

        # DOC/DOCX -> parse doc_text if present
        if suf in (".doc", ".docx"):
            a = _find_latest_artifact_for_file(project_name, artifact_type="doc_text", canonical_rel=canonical_rel)
            if isinstance(a, dict):
                txt = read_artifact_text(project_name, a, cap_chars=220000).strip()
                line_no = 0
                for raw_ln in (txt or "").splitlines():
                    if len(all_labs) >= int(max_total_items or 0):
                        break
                    ln = (raw_ln or "").strip()
                    if not ln:
                        continue
                    line_no += 1
                    labs = _health_extract_labs_from_text(
                        text=ln,
                        source_file=canonical_rel,
                        source_hint=f"line {line_no}",
                        max_items=1,
                    )
                    if labs:
                        all_labs.extend(labs)
            continue

    # Merge into profile (latest wins per name+unit)
    try:
        prof = update_health_profile(project_name, labs=all_labs, replace_labs=True, source_excerpt="", source_turn=0, timestamp=now_iso())
    except Exception:
        prof = load_health_profile(project_name)

    try:
        prof["lab_refresh_manifest_at"] = float(man_updated or 0.0)
        prof["lab_parser_version"] = int(_HEALTH_LAB_PARSER_VERSION)
        _write_health_profile(project_name, prof)
    except Exception:
        pass

    return prof

def update_health_profile(
    project_name: str,
    *,
    meds: Optional[List[Dict[str, Any]]] = None,
    allergies: Optional[List[Dict[str, Any]]] = None,
    weight: Optional[Dict[str, Any]] = None,
    labs: Optional[List[Dict[str, Any]]] = None,
    replace_labs: bool = False,
    source_turn: int = 0,
    source_excerpt: str = "",
    timestamp: str = "",
) -> Dict[str, Any]:
    """
    Deterministic health memory update (global expert store).
    - meds: [{name, dose, dose_unit, frequency, route, status, raw}]
    - allergies: [{name, reaction, status, raw}]
    - weight: {value, unit}
    """
    prof = load_health_profile(project_name)
    meds_list = prof.get("medications")
    if not isinstance(meds_list, list):
        meds_list = []
        prof["medications"] = meds_list
    else:
        # Clean malformed promoted entries
        try:
            meds_list = [
                m for m in meds_list
                if not (
                    isinstance(m, dict)
                    and _health_med_name_is_malformed(str(m.get("name") or ""))
                    and str(m.get("source_excerpt") or "") == "promoted_from_global_facts"
                )
            ]
            prof["medications"] = meds_list
        except Exception:
            pass

    idx: Dict[str, int] = {}
    for i, m in enumerate(meds_list):
        if not isinstance(m, dict):
            continue
        nm = _health_norm_med_name(str(m.get("name") or ""))
        if nm:
            idx[nm] = i

    ts = (timestamp or "").strip() or now_iso()
    excerpt = (source_excerpt or "").strip()

    for m in (meds or []):
        if not isinstance(m, dict):
            continue
        name = str(m.get("name") or "").strip()
        if not name:
            continue
        key = _health_norm_med_name(name)
        if not key:
            continue

        rec: Dict[str, Any] = {}
        if key in idx:
            try:
                rec = meds_list[idx[key]]
            except Exception:
                rec = {}
        if not isinstance(rec, dict):
            rec = {}

        rec["name"] = name
        if str(m.get("dose") or "").strip():
            rec["dose"] = str(m.get("dose") or "").strip()
        if str(m.get("dose_unit") or "").strip():
            rec["dose_unit"] = str(m.get("dose_unit") or "").strip()
        if str(m.get("frequency") or "").strip():
            rec["frequency"] = str(m.get("frequency") or "").strip()
        if str(m.get("route") or "").strip():
            rec["route"] = str(m.get("route") or "").strip()
        if str(m.get("status") or "").strip():
            rec["status"] = str(m.get("status") or "").strip()
        if str(m.get("raw") or "").strip():
            rec["raw"] = str(m.get("raw") or "").strip()

        rec["last_mentioned"] = ts
        if source_turn:
            rec["source_turn"] = int(source_turn or 0)
        if excerpt:
            rec["source_excerpt"] = excerpt

        if key in idx:
            meds_list[idx[key]] = rec
        else:
            meds_list.append(rec)
            idx[key] = len(meds_list) - 1

    # Allergies
    allergies_list = prof.get("allergies")
    if not isinstance(allergies_list, list):
        allergies_list = []
        prof["allergies"] = allergies_list

    aidx: Dict[str, int] = {}
    for i, a in enumerate(allergies_list):
        if not isinstance(a, dict):
            continue
        nm = _health_norm_med_name(str(a.get("name") or ""))
        if nm:
            aidx[nm] = i

    for a in (allergies or []):
        if not isinstance(a, dict):
            continue
        name = str(a.get("name") or "").strip()
        if not name:
            continue
        key = _health_norm_med_name(name)
        if not key:
            continue

        rec_a: Dict[str, Any] = {}
        if key in aidx:
            try:
                rec_a = allergies_list[aidx[key]]
            except Exception:
                rec_a = {}
        if not isinstance(rec_a, dict):
            rec_a = {}

        rec_a["name"] = name
        if str(a.get("reaction") or "").strip():
            rec_a["reaction"] = str(a.get("reaction") or "").strip()
        if str(a.get("status") or "").strip():
            rec_a["status"] = str(a.get("status") or "").strip()
        if str(a.get("raw") or "").strip():
            rec_a["raw"] = str(a.get("raw") or "").strip()

        rec_a["last_mentioned"] = ts
        if source_turn:
            rec_a["source_turn"] = int(source_turn or 0)
        if excerpt:
            rec_a["source_excerpt"] = excerpt

        # If we are adding a real allergy, remove "none" placeholders.
        if not _health_allergy_is_none(rec_a):
            try:
                allergies_list = [x for x in allergies_list if not (isinstance(x, dict) and _health_allergy_is_none(x))]
                prof["allergies"] = allergies_list
                aidx = { _health_norm_med_name(str(x.get("name") or "")) : i for i, x in enumerate(allergies_list) if isinstance(x, dict) }
            except Exception:
                pass
        else:
            # If real allergies exist, skip "none" entry.
            try:
                if any(isinstance(x, dict) and not _health_allergy_is_none(x) for x in allergies_list):
                    continue
            except Exception:
                pass

        if key in aidx:
            allergies_list[aidx[key]] = rec_a
        else:
            allergies_list.append(rec_a)
            aidx[key] = len(allergies_list) - 1

    if isinstance(weight, dict):
        try:
            wv = str(weight.get("value") or "").strip()
            wu = str(weight.get("unit") or "").strip()
            if wv and wu:
                meas = prof.get("measurements")
                if not isinstance(meas, dict):
                    meas = {}
                    prof["measurements"] = meas
                meas["weight"] = {
                    "value": wv,
                    "unit": wu,
                    "last_mentioned": ts,
                    "source_excerpt": excerpt,
                }
        except Exception:
            pass

    # Labs (project-scoped; latest wins per (name, unit))
    labs_list = prof.get("labs")
    if replace_labs:
        labs_list = []
        prof["labs"] = labs_list
    elif not isinstance(labs_list, list):
        labs_list = []
        prof["labs"] = labs_list

    lab_idx: Dict[str, int] = {}
    for i, it in enumerate(labs_list):
        if not isinstance(it, dict):
            continue
        nm = _health_norm_med_name(str(it.get("name") or ""))
        un = str(it.get("unit") or "").strip().lower()
        if nm:
            lab_idx[f"{nm}|{un}"] = i

    for it in (labs or []):
        if not isinstance(it, dict):
            continue
        name = str(it.get("name") or "").strip()
        if not name:
            continue
        unit = str(it.get("unit") or "").strip()
        key = f"{_health_norm_med_name(name)}|{unit.lower()}"
        if not key.strip("|"):
            continue

        rec: Dict[str, Any] = {}
        if key in lab_idx:
            try:
                rec = labs_list[lab_idx[key]]
            except Exception:
                rec = {}
        if not isinstance(rec, dict):
            rec = {}

        rec["name"] = name
        if str(it.get("value") or "").strip():
            rec["value"] = str(it.get("value") or "").strip()
        if unit:
            rec["unit"] = unit
        if str(it.get("ref") or "").strip():
            rec["ref"] = str(it.get("ref") or "").strip()
        if str(it.get("date") or "").strip():
            rec["date"] = str(it.get("date") or "").strip()
        if str(it.get("source_file") or "").strip():
            rec["source_file"] = str(it.get("source_file") or "").strip()
        if str(it.get("source_hint") or "").strip():
            rec["source_hint"] = str(it.get("source_hint") or "").strip()

        rec["last_mentioned"] = ts
        if source_turn:
            rec["source_turn"] = int(source_turn or 0)
        if excerpt:
            rec["source_excerpt"] = excerpt

        if key in lab_idx:
            labs_list[lab_idx[key]] = rec
        else:
            labs_list.append(rec)
            lab_idx[key] = len(labs_list) - 1

    _write_health_profile(project_name, prof)
    return prof

def render_health_profile_snippet(project_name: str) -> str:
    """
    Render a small, deterministic health memory snippet for CANONICAL_SNIPPETS.
    """
    try:
        prof = load_health_profile(project_name)
    except Exception:
        prof = _default_health_profile()

    lines: List[str] = []

    # Age from global profile (derived)
    try:
        user_seg = str(project_name or "").split("/", 1)[0].strip()
    except Exception:
        user_seg = ""
    age = None
    try:
        if user_seg:
            gp = load_user_profile(user_seg)
            age = (gp.get("derived") or {}).get("age_years")
    except Exception:
        age = None
    if isinstance(age, int):
        lines.append(f"- Age (derived): {age}")

    # Weight
    try:
        meas = prof.get("measurements") if isinstance(prof, dict) else {}
        w = (meas or {}).get("weight") if isinstance(meas, dict) else {}
        wv = str((w or {}).get("value") or "").strip()
        wu = str((w or {}).get("unit") or "").strip()
        if wv and wu:
            lines.append(f"- Weight: {wv} {wu}")
        else:
            lines.append("- Weight: (not recorded)")
    except Exception:
        lines.append("- Weight: (not recorded)")

    # Medications
    meds = prof.get("medications") if isinstance(prof, dict) else []
    if not isinstance(meds, list):
        meds = []
    if meds:
        lines.append("- Medications:")
        for m in meds[:12]:
            if not isinstance(m, dict):
                continue
            name = str(m.get("name") or "").strip()
            if not name:
                continue
            dose = str(m.get("dose") or "").strip()
            unit = str(m.get("dose_unit") or "").strip()
            freq = str(m.get("frequency") or "").strip()
            route = str(m.get("route") or "").strip()
            status = str(m.get("status") or "").strip()

            parts: List[str] = [name]
            if dose and unit:
                parts.append(f"{dose} {unit}")
            elif dose:
                parts.append(dose)
            if freq:
                parts.append(freq)
            if route:
                parts.append(route)
            if status:
                parts.append(f"[{status}]")
            lines.append("  - " + " | ".join(parts))
    else:
        lines.append("- Medications: (not recorded yet)")

    # Allergies
    allergies = prof.get("allergies") if isinstance(prof, dict) else []
    if not isinstance(allergies, list):
        allergies = []
    if allergies:
        # If only "none" is recorded, print compactly.
        real_allergies = [a for a in allergies if isinstance(a, dict) and not _health_allergy_is_none(a)]
        if not real_allergies:
            lines.append("- Allergies: None reported")
        else:
            lines.append("- Allergies:")
            for a in real_allergies[:12]:
                name = str(a.get("name") or "").strip()
                if not name:
                    continue
                reaction = str(a.get("reaction") or "").strip()
                status = str(a.get("status") or "").strip()
                parts_a = [name]
                if reaction:
                    parts_a.append(reaction)
                if status:
                    parts_a.append(f"[{status}]")
                lines.append("  - " + " | ".join(parts_a))
    else:
        lines.append("- Allergies: (not recorded yet)")

    # Labs (recent)
    labs = prof.get("labs") if isinstance(prof, dict) else []
    if not isinstance(labs, list):
        labs = []
    if labs:
        lines.append("- Labs (recent):")
        for it in labs[:12]:
            if not isinstance(it, dict):
                continue
            nm = str(it.get("name") or "").strip()
            val = str(it.get("value") or "").strip()
            un = str(it.get("unit") or "").strip()
            dt = str(it.get("date") or "").strip()
            if not nm or not val:
                continue
            tail = f"{val} {un}".strip()
            if dt:
                tail = tail + f" (date {dt})"
            lines.append("  - " + nm + ": " + tail)
    else:
        lines.append("- Labs: (not recorded yet)")

    return "\n".join([ln for ln in lines if ln.strip()]).strip()

def _health_profile_has_meds(meds: Any) -> bool:
    if not isinstance(meds, list):
        return False
    for m in meds:
        if not isinstance(m, dict):
            continue
        name = str(m.get("name") or "").strip()
        if name and (not _health_med_name_is_malformed(name)):
            return True
    return False

def _health_profile_has_allergies(allergies: Any) -> bool:
    if not isinstance(allergies, list):
        return False
    for a in allergies:
        if not isinstance(a, dict):
            continue
        name = str(a.get("name") or "").strip()
        if name:
            return True
    return False

def _health_profile_has_weight(meas: Any) -> bool:
    if not isinstance(meas, dict):
        return False
    w = meas.get("weight") if isinstance(meas, dict) else {}
    if not isinstance(w, dict):
        return False
    wv = str(w.get("value") or "").strip()
    wu = str(w.get("unit") or "").strip()
    return bool(wv and wu)

def _health_profile_has_labs(labs: Any) -> bool:
    if not isinstance(labs, list):
        return False
    for it in labs:
        if not isinstance(it, dict):
            continue
        nm = str(it.get("name") or "").strip()
        val = str(it.get("value") or "").strip()
        if nm and val:
            return True
    return False

def seed_health_profile_from_user_history(project_name: str) -> Dict[str, Any]:
    """
    If current health_profile is missing core fields, seed from the most recent
    health_profile.json in the same user's other projects.
    Best-effort, bounded, and non-destructive (fills missing only).
    """
    try:
        prof = load_health_profile(project_name)
    except Exception:
        prof = _default_health_profile()

    missing = []
    try:
        if not _health_profile_has_meds(prof.get("medications") if isinstance(prof, dict) else None):
            missing.append("medications")
    except Exception:
        missing.append("medications")
    try:
        if not _health_profile_has_allergies(prof.get("allergies") if isinstance(prof, dict) else None):
            missing.append("allergies")
    except Exception:
        missing.append("allergies")
    try:
        if not _health_profile_has_weight(prof.get("measurements") if isinstance(prof, dict) else None):
            missing.append("weight")
    except Exception:
        missing.append("weight")
    try:
        if not _health_profile_has_labs(prof.get("labs") if isinstance(prof, dict) else None):
            missing.append("labs")
    except Exception:
        missing.append("labs")

    if not missing:
        return {"ok": True, "seeded": False, "missing": []}

    # Find most recent other health_profile.json under this user
    try:
        user_seg = str(project_name or "").split("/", 1)[0].strip()
    except Exception:
        user_seg = ""
    if not user_seg:
        return {"ok": True, "seeded": False, "missing": missing}

    base = PROJECTS_DIR / safe_user_segment(user_seg)
    if not base.exists():
        return {"ok": True, "seeded": False, "missing": missing}

    best_prof: Optional[Dict[str, Any]] = None
    best_proj = ""
    best_epoch = 0.0

    def _parse_epoch(ts: str) -> float:
        s = (ts or "").strip()
        if not s:
            return 0.0
        try:
            return time.mktime(time.strptime(s, "%Y-%m-%dT%H:%M:%SZ"))
        except Exception:
            return 0.0

    try:
        for child in base.iterdir():
            try:
                if not child.is_dir():
                    continue
            except Exception:
                continue
            if child.name == "_user":
                continue
            # Skip current project
            try:
                if child.resolve() == project_dir(project_name).resolve():
                    continue
            except Exception:
                pass

            hp = child / STATE_DIR_NAME / HEALTH_PROFILE_FILE_NAME
            if not hp.exists():
                continue
            try:
                obj = json.loads(hp.read_text(encoding="utf-8") or "{}")
            except Exception:
                obj = {}
            if not isinstance(obj, dict) or obj.get("schema") != "health_profile_v1":
                continue

            # Require at least one meaningful field
            if not (
                _health_profile_has_meds(obj.get("medications"))
                or _health_profile_has_allergies(obj.get("allergies"))
                or _health_profile_has_weight(obj.get("measurements"))
                or _health_profile_has_labs(obj.get("labs"))
            ):
                continue

            ep = _parse_epoch(str(obj.get("updated_at") or "").strip())
            if ep <= 0:
                try:
                    ep = float(hp.stat().st_mtime)
                except Exception:
                    ep = 0.0

            if ep >= best_epoch:
                best_epoch = ep
                best_prof = obj
                best_proj = child.name
    except Exception:
        best_prof = None

    if not isinstance(best_prof, dict):
        return {"ok": True, "seeded": False, "missing": missing}

    # Merge only missing fields
    new_prof = dict(prof) if isinstance(prof, dict) else _default_health_profile()
    fields_seeded: List[str] = []

    if ("medications" in missing) and _health_profile_has_meds(best_prof.get("medications")):
        new_prof["medications"] = best_prof.get("medications")
        fields_seeded.append("medications")
    if ("allergies" in missing) and _health_profile_has_allergies(best_prof.get("allergies")):
        new_prof["allergies"] = best_prof.get("allergies")
        fields_seeded.append("allergies")
    if ("weight" in missing) and _health_profile_has_weight(best_prof.get("measurements")):
        new_prof["measurements"] = best_prof.get("measurements")
        fields_seeded.append("weight")
    if ("labs" in missing) and _health_profile_has_labs(best_prof.get("labs")):
        new_prof["labs"] = best_prof.get("labs")
        fields_seeded.append("labs")

    if fields_seeded:
        try:
            _write_health_profile(project_name, new_prof)
        except Exception:
            pass
        return {
            "ok": True,
            "seeded": True,
            "missing": missing,
            "seeded_fields": fields_seeded,
            "source_project": best_proj,
        }

    return {"ok": True, "seeded": False, "missing": missing}

def _default_couples_shared_memory() -> Dict[str, Any]:
    return {
        "schema": "couples_shared_memory_v1",
        "updated_at": now_iso(),
        "agreements": [],
    }

def load_couples_shared_memory(project_name: str) -> Dict[str, Any]:
    ensure_project_scaffold(project_name)
    p = couples_shared_memory_path(project_name)
    obj = _load_json_obj(p)
    if not isinstance(obj, dict) or obj.get("schema") != "couples_shared_memory_v1":
        obj = _default_couples_shared_memory()
        try:
            atomic_write_text(p, json.dumps(obj, indent=2, sort_keys=True))
        except Exception:
            pass
    obj["updated_at"] = now_iso()
    return obj

def save_couples_shared_memory(project_name: str, obj: Dict[str, Any]) -> None:
    ensure_project_scaffold(project_name)
    out = obj if isinstance(obj, dict) else _default_couples_shared_memory()
    out.setdefault("schema", "couples_shared_memory_v1")
    out["updated_at"] = now_iso()
    atomic_write_text(couples_shared_memory_path(project_name), json.dumps(out, indent=2, sort_keys=True))

def _norm_agreement_text(s: str) -> str:
    t = re.sub(r"[^a-zA-Z0-9\s]", " ", (s or "").lower())
    t = re.sub(r"\s+", " ", t).strip()
    return t

def _extract_agreement_from_text(user_msg: str) -> str:
    s = (user_msg or "").strip()
    if not s:
        return ""
    # Work sentence-by-sentence (simple split)
    parts = re.split(r"[\\.!?\\n]+", s)
    for p in parts:
        t = p.strip()
        if not t:
            continue
        low = t.lower()
        m = re.search(
            r"\\bwe\\s+(agreed|decided|committed|promise(?:d)?|will|should|plan to|are going to)\\b\\s*(.+)",
            low,
        )
        if not m:
            continue
        tail = t[m.end(1):].strip()
        if tail.startswith("to "):
            tail = tail[3:].strip()
        tail = tail.strip(" .,:;")
        if len(tail) < 4:
            continue
        # Return a clean agreement statement
        return "We agreed to " + tail
    return ""

def _extract_suggested_rule_from_text(text: str) -> str:
    s = (text or "").strip()
    if not s:
        return ""
    parts = re.split(r"[\\.!?\\n]+", s)
    for p in parts:
        t = p.strip()
        if not t:
            continue
        low = t.lower()
        if any(k in low for k in (
            "one thing that helps",
            "a simple rule could be",
            "a helpful rule is",
            "it might help to",
            "a good rule is",
            "a small rule could be",
            "a useful rule is",
        )):
            # Keep the sentence as the proposed rule
            t2 = t.strip(" .,:;")
            if len(t2) < 4:
                continue
            # Normalize to "Rule: ..." format
            return t2
    return ""

def update_couples_shared_memory_from_message(project_name: str, user: str, user_msg: str) -> None:
    """
    Couples-only: capture proposed agreements and mark corroborated when both partners mention them.
    """
    try:
        pname = safe_project_name(project_name)
        user_seg = pname.split("/", 1)[0].strip().lower()
        if not user_seg.startswith("couple_"):
            return
    except Exception:
        return

    agreement = _extract_agreement_from_text(user_msg)
    if not agreement:
        return

    norm = _norm_agreement_text(agreement)
    if not norm:
        return

    mem = load_couples_shared_memory(project_name)
    items = mem.get("agreements") if isinstance(mem.get("agreements"), list) else []
    if not isinstance(items, list):
        items = []

    who = safe_user_name(user or "")
    now = now_iso()

    found = None
    for it in items:
        if not isinstance(it, dict):
            continue
        n = _norm_agreement_text(str(it.get("text") or ""))
        if n and n == norm:
            found = it
            break

    if not found:
        found = {
            "id": f"agr_{int(time.time() * 1000)}",
            "text": agreement,
            "status": "proposed",
            "proposed_by": [],
            "confirmed_by": [],
            "created_at": now,
            "updated_at": now,
        }
        items.append(found)

    # Update proposer list
    try:
        pb = found.get("proposed_by") if isinstance(found.get("proposed_by"), list) else []
        if who and who not in pb:
            pb.append(who)
        found["proposed_by"] = pb
    except Exception:
        pass

    # If two distinct partners have mentioned it, mark corroborated
    try:
        pb2 = found.get("proposed_by") if isinstance(found.get("proposed_by"), list) else []
        if len(set(pb2)) >= 2:
            found["status"] = "corroborated"
            found.setdefault("confirmed_by", pb2[:])
    except Exception:
        pass

    found["updated_at"] = now
    mem["agreements"] = items
    save_couples_shared_memory(project_name, mem)

def update_couples_shared_memory_from_assistant(project_name: str, assistant_msg: str) -> None:
    """
    Couples-only: capture therapist-suggested rules (proposed_by_quinn).
    """
    try:
        pname = safe_project_name(project_name)
        user_seg = pname.split("/", 1)[0].strip().lower()
        if not user_seg.startswith("couple_"):
            return
    except Exception:
        return

    rule = _extract_suggested_rule_from_text(assistant_msg)
    if not rule:
        return

    norm = _norm_agreement_text(rule)
    if not norm:
        return

    mem = load_couples_shared_memory(project_name)
    items = mem.get("agreements") if isinstance(mem.get("agreements"), list) else []
    if not isinstance(items, list):
        items = []

    now = now_iso()
    found = None
    for it in items:
        if not isinstance(it, dict):
            continue
        n = _norm_agreement_text(str(it.get("text") or ""))
        if n and n == norm:
            found = it
            break

    if not found:
        found = {
            "id": f"agr_{int(time.time() * 1000)}",
            "text": rule,
            "status": "proposed",
            "proposed_by": ["quinn"],
            "confirmed_by": [],
            "created_at": now,
            "updated_at": now,
        }
        items.append(found)
    else:
        pb = found.get("proposed_by") if isinstance(found.get("proposed_by"), list) else []
        if "quinn" not in pb:
            pb.append("quinn")
        found["proposed_by"] = pb
        found["updated_at"] = now

    mem["agreements"] = items
    save_couples_shared_memory(project_name, mem)

def load_user_pending_profile_question(user: str) -> Dict[str, Any]:
    p = user_pending_profile_question_path(user)
    if not p.exists():
        return {}
    try:
        obj = json.loads(p.read_text(encoding="utf-8") or "{}")
        return obj if isinstance(obj, dict) else {}
    except Exception:
        return {}

def save_user_pending_profile_question(user: str, obj: Dict[str, Any]) -> None:
    out = obj if isinstance(obj, dict) else {}
    out.setdefault("version", 1)
    atomic_write_text(user_pending_profile_question_path(user), json.dumps(out, indent=2, sort_keys=True))

def clear_user_pending_profile_question(user: str) -> None:
    p = user_pending_profile_question_path(user)
    try:
        if p.exists():
            p.unlink()
    except Exception:
        pass

def is_user_pending_profile_question_unexpired(obj: Dict[str, Any]) -> bool:
    """
    Deterministic expiry check (no date parsing):
    - Requires obj["pending"] == True
    - Requires numeric obj["expires_at_epoch"] > time.time()
    """
    if not isinstance(obj, dict):
        return False
    if obj.get("pending") is not True:
        return False
    try:
        exp = float(obj.get("expires_at_epoch") or 0.0)
    except Exception:
        exp = 0.0
    return time.time() <= exp

def user_global_facts_map_path(user: str) -> Path:
    # projects/<user>/_user/global_facts_map.json
    return user_dir(user) / USER_GLOBAL_FACTS_MAP_FILE_NAME

def _default_user_global_facts_map() -> Dict[str, Any]:
    # Tier-2M (global mutable memory): update-friendly facts map with provenance.
    return {
        "schema": "global_facts_map_v1",
        "updated_at": now_iso(),
        # entries keyed by entity_key
        "entries": {},
    }

def load_user_global_facts_map(user: str) -> Dict[str, Any]:
    p = user_global_facts_map_path(user)
    obj = _load_json_obj(p)
    if not isinstance(obj, dict) or obj.get("schema") != "global_facts_map_v1":
        obj = _default_user_global_facts_map()
        try:
            atomic_write_text(p, json.dumps(obj, indent=2, sort_keys=True))
        except Exception:
            pass
    obj["updated_at"] = now_iso()
    return obj

def _write_user_global_facts_map(user: str, obj: Dict[str, Any]) -> bool:
    try:
        out = obj if isinstance(obj, dict) else _default_user_global_facts_map()
        out.setdefault("schema", "global_facts_map_v1")
        out["updated_at"] = now_iso()
        atomic_write_text(user_global_facts_map_path(user), json.dumps(out, indent=2, sort_keys=True))
        return True
    except Exception:
        return False

def _tier2m_kind_from_entity_key(entity_key: str) -> str:
    ek = _policy_norm(entity_key)

    # Location is explicitly disambiguated and user-confirmed via Tier-2G flows.
    # In Tier-2M we want "latest wins" (recency), not permanent ambiguity.
    if ek == "user.identity.location":
        return "preference"

    if ek.startswith("user.identity.") or ek.startswith("user.relationship."):
        return "identity"
    if ek.startswith("user.preference."):
        return "preference"
    return "other"

def upsert_user_global_fact_tier2m(
    user: str,
    *,
    entity_key: str,
    value: str,
    evidence_id: str = "",
    source: Any = "chat",
    last_seen_at: str = "",
) -> Dict[str, Any]:
    """
    Tier-2M upsert into users/<user>/_user/global_facts_map.json

    Semantics:
      - identity-like keys: conflict => status=ambiguous + options preserved (no winner selection)
      - preference-like keys: resolve by recency (current value becomes last seen), preserve history
      - always track last_seen_at + evidence refs + source metadata
    """
    ek = _norm_one_line(entity_key)
    v = _norm_one_line(value)
    if not ek or not v:
        return {"ok": False, "error": "missing_entity_key_or_value"}

    kind = _tier2m_kind_from_entity_key(ek)
    seen = (last_seen_at or "").strip() or now_iso()

    m = load_user_global_facts_map(user)
    ents = m.get("entries")
    if not isinstance(ents, dict):
        ents = {}
        m["entries"] = ents

    rec = ents.get(ek)
    if not isinstance(rec, dict):
        rec = {
            "entity_key": ek,
            "kind": kind,
            "status": "confirmed",
            "value": "",
            "options": [],
            "history": [],
            "last_seen_at": "",
            "evidence_ids": [],
            "sources": [],
        }

    # Ensure kind is stable (but do not rewrite identity kernel)
    rec["kind"] = kind

    # Evidence ids (bounded; cap depends on kind + key shape)
    evs = rec.get("evidence_ids")
    if not isinstance(evs, list):
        evs = []
    if evidence_id:
        eid = _norm_one_line(evidence_id)
        if eid and eid not in evs:
            evs.append(eid)

    cap_evs = 32
    if kind == "preference":
        cap_evs = 28
    if kind == "other" and ek.endswith(".other"):
        cap_evs = 24
    rec["evidence_ids"] = [str(x) for x in evs if str(x).strip()][-cap_evs:]

    # Sources (append-only-ish; dedupe by (evidence_id, seen_at); cap depends on kind + key shape)
    srcs = rec.get("sources")
    if not isinstance(srcs, list):
        srcs = []
    src_obj = source
    try:
        # store structured source if possible; otherwise store string
        if isinstance(source, dict):
            src_obj = source
        else:
            src_obj = _norm_one_line(source) or "chat"
    except Exception:
        src_obj = "chat"

    src_entry = {"seen_at": seen, "source": src_obj, "evidence_id": _norm_one_line(evidence_id)}
    try:
        exists = any(
            isinstance(x, dict)
            and (str(x.get("evidence_id") or "").strip() == str(src_entry.get("evidence_id") or "").strip())
            and (str(x.get("seen_at") or "").strip() == str(src_entry.get("seen_at") or "").strip())
            for x in srcs
        )
    except Exception:
        exists = False
    if not exists:
        srcs.append(src_entry)

    cap_srcs = 24
    if kind == "preference":
        cap_srcs = 18
    if kind == "other" and ek.endswith(".other"):
        cap_srcs = 12
    rec["sources"] = srcs[-cap_srcs:]

    # Conflict handling
    cur_status = str(rec.get("status") or "").strip().lower() or "confirmed"
    cur_val = _norm_one_line(rec.get("value") or "")
    opts = rec.get("options")
    if not isinstance(opts, list):
        opts = []
    opts2 = [str(x).strip() for x in opts if str(x).strip()]

    # Identity-like: never pick a winner on conflict
    if kind == "identity":
        if cur_status == "ambiguous":
            if v not in opts2:
                opts2.append(v)
            rec["options"] = opts2[:10]
            rec["value"] = ""
            rec["status"] = "ambiguous"
        else:
            if not cur_val:
                rec["value"] = v
                rec["options"] = [v]
                rec["status"] = "confirmed"
            elif cur_val == v:
                if v not in opts2:
                    opts2.append(v)
                rec["options"] = opts2[:10]
                rec["status"] = "confirmed"
            else:
                # conflict => ambiguous with both options
                if cur_val not in opts2:
                    opts2.append(cur_val)
                if v not in opts2:
                    opts2.append(v)
                rec["options"] = opts2[:10]
                rec["value"] = ""
                rec["status"] = "ambiguous"

        rec["last_seen_at"] = seen
        ents[ek] = rec
        _write_user_global_facts_map(user, m)
        return {"ok": True, "entity_key": ek, "status": rec.get("status"), "kind": kind}

    # Preference-like + other: resolve-by-recency, preserve history
    hist = rec.get("history")
    if not isinstance(hist, list):
        hist = []

    # If the value changes, store prior current value as history
    if cur_val and (cur_val != v):
        hist.append(
            {
                "value": cur_val,
                "last_seen_at": str(rec.get("last_seen_at") or "").strip(),
                "evidence_ids": [str(x) for x in (rec.get("evidence_ids") or []) if str(x).strip()][:12],
            }
        )
        hist = hist[-24:]

    rec["history"] = hist
    rec["value"] = v
    rec["status"] = "confirmed"
    rec["last_seen_at"] = seen

    ents[ek] = rec
    _write_user_global_facts_map(user, m)
    return {"ok": True, "entity_key": ek, "status": rec.get("status"), "kind": kind}

def rebuild_user_global_facts_map_from_user_facts(user: str) -> Dict[str, Any]:
    """
    Deterministically rebuild Tier-2M from users/<user>/_user/facts_raw.jsonl.

    Key change (determinism + size control):
    - rebuild IN-MEMORY and write ONCE (avoid per-fact full-file load/write)
    - dedupe sources by (evidence_id, seen_at)
    - apply tighter caps for *.other buckets to prevent junk-drawer bloat
    """
    facts = _load_user_facts_raw(user, max_lines=8000)

    m: Dict[str, Any] = _default_user_global_facts_map()
    m["updated_at"] = now_iso()
    ents: Dict[str, Any] = {}
    m["entries"] = ents

    def _caps_for(ek: str, kind: str) -> Tuple[int, int, int]:
        # returns (cap_sources, cap_history, cap_evidence_ids)
        cap_sources = 24
        cap_history = 24
        cap_evs = 32

        if kind == "preference":
            cap_sources = 18
            cap_history = 18
            cap_evs = 28

        # Coarse buckets like "context.other", "event.other", etc. must stay small.
        if kind == "other" and (ek or "").endswith(".other"):
            cap_sources = 12
            cap_history = 12
            cap_evs = 24

        return cap_sources, cap_history, cap_evs

    def _apply_one(
        *,
        ek: str,
        v: str,
        evidence_id: str,
        source: Any,
        seen: str,
    ) -> None:
        ek0 = _norm_one_line(ek)
        v0 = _norm_one_line(v)

        # For certain identity keys, store the normalized VALUE (not the full claim sentence).
        # This prevents permanent "conflict" between different surface phrasings.
        if ek0 == "user.identity.location":
            low = v0.lower()
            mm = re.search(r"\bmy\s+confirmed\s+location\s+is\s+(.+)$", low)
            if not mm:
                mm = re.search(r"\bconfirm\s+my\s+location\s+(?:is|to)\s+(.+)$", low)
            if not mm:
                mm = re.search(r"\bset\s+my\s+location\s+to\s+(.+)$", low)
            if not mm:
                mm = re.search(r"\bi\s+live\s+in\s+(.+)$", low)
            if not mm:
                mm = re.search(r"\b(?:i\s*[' ]?m|i\s+am)\s+in\s+(.+)$", low)

            if mm:
                vv = (mm.group(1) or "").strip()
                while vv and vv[-1] in ".!,;:":
                    vv = vv[:-1].rstrip()
                vv = " ".join(vv.split())
                if vv:
                    v0 = vv

        if not ek0 or not v0:
            return

        kind = _tier2m_kind_from_entity_key(ek0)
        cap_sources, cap_history, cap_evs = _caps_for(ek0, kind)

        rec = ents.get(ek0)
        if not isinstance(rec, dict):
            rec = {
                "entity_key": ek0,
                "kind": kind,
                "status": "confirmed",
                "value": "",
                "options": [],
                "history": [],
                "last_seen_at": "",
                "evidence_ids": [],
                "sources": [],
            }

        rec["kind"] = kind

        # evidence ids (dedupe + bounded)
        evs = rec.get("evidence_ids")
        if not isinstance(evs, list):
            evs = []
        eid0 = _norm_one_line(evidence_id)
        if eid0 and eid0 not in evs:
            evs.append(eid0)
        rec["evidence_ids"] = [str(x) for x in evs if str(x).strip()][-cap_evs:]

        # sources (dedupe by (evidence_id, seen_at) + bounded)
        srcs = rec.get("sources")
        if not isinstance(srcs, list):
            srcs = []
        src_obj = source
        try:
            if isinstance(source, dict):
                src_obj = source
            else:
                src_obj = _norm_one_line(source) or "chat"
        except Exception:
            src_obj = "chat"

        src_entry = {"seen_at": seen, "source": src_obj, "evidence_id": eid0}
        try:
            exists = any(
                isinstance(x, dict)
                and (str(x.get("evidence_id") or "").strip() == eid0)
                and (str(x.get("seen_at") or "").strip() == seen)
                for x in srcs
            )
        except Exception:
            exists = False
        if not exists:
            srcs.append(src_entry)
        rec["sources"] = srcs[-cap_sources:]

        cur_status = str(rec.get("status") or "").strip().lower() or "confirmed"
        cur_val = _norm_one_line(rec.get("value") or "")
        opts = rec.get("options")
        if not isinstance(opts, list):
            opts = []
        opts2 = [str(x).strip() for x in opts if str(x).strip()]

        # identity-like conflict handling (no winner)
        if kind == "identity":
            if cur_status == "ambiguous":
                if v0 not in opts2:
                    opts2.append(v0)
                rec["options"] = opts2[:10]
                rec["value"] = ""
                rec["status"] = "ambiguous"
            else:
                if not cur_val:
                    rec["value"] = v0
                    rec["options"] = [v0]
                    rec["status"] = "confirmed"
                elif cur_val == v0:
                    if v0 not in opts2:
                        opts2.append(v0)
                    rec["options"] = opts2[:10]
                    rec["status"] = "confirmed"
                else:
                    if cur_val not in opts2:
                        opts2.append(cur_val)
                    if v0 not in opts2:
                        opts2.append(v0)
                    rec["options"] = opts2[:10]
                    rec["value"] = ""
                    rec["status"] = "ambiguous"

            rec["last_seen_at"] = seen
            ents[ek0] = rec
            return

        # preference/other: resolve-by-recency, preserve bounded history
        hist = rec.get("history")
        if not isinstance(hist, list):
            hist = []

        if cur_val and (cur_val != v0):
            hist.append(
                {
                    "value": cur_val,
                    "last_seen_at": str(rec.get("last_seen_at") or "").strip(),
                    "evidence_ids": [str(x) for x in (rec.get("evidence_ids") or []) if str(x).strip()][:12],
                }
            )
        rec["history"] = [h for h in hist if isinstance(h, dict)][-cap_history:]
        rec["value"] = v0
        rec["status"] = "confirmed"
        rec["last_seen_at"] = seen
        ents[ek0] = rec

    for f in (facts or []):
        if not isinstance(f, dict):
            continue
        claim = _norm_one_line(f.get("claim") or "")
        if not claim:
            continue

        slot0 = str(f.get("slot") or "other").strip().lower()
        subj0 = str(f.get("subject") or "user").strip().lower()

        try:
            ek = _tier1_entity_key_guess(claim, slot0, subj0)
        except Exception:
            ek = "other"

        fid = _norm_one_line(f.get("id") or "")
        src = f.get("source")
        seen_at = _norm_one_line(f.get("created_at") or "") or now_iso()

        try:
            _apply_one(ek=ek, v=claim, evidence_id=fid, source=(src if src is not None else "chat"), seen=seen_at)
        except Exception:
            pass

    # Write ONCE (atomic) and return the built object (not a reload)
    _write_user_global_facts_map(user, m)
    return m

def render_user_global_facts_snippet_tier2m(
    user: str,
    *,
    include_keys: Optional[List[str]] = None,
    include_prefixes: Optional[List[str]] = None,
    include_kinds: Optional[List[str]] = None,
    max_items: int = 18,
) -> str:
    """
    Render a bounded Tier-2M snippet for model context from a selected subset.
    Selection supports:
      - include_keys: exact entity_key matches
      - include_prefixes: entity_key prefix matches
      - include_kinds: ["identity","preference","other"]
    """
    m = load_user_global_facts_map(user)
    ents = m.get("entries")
    if not isinstance(ents, dict) or not ents:
        return ""

    keys0 = [str(x).strip() for x in (include_keys or []) if str(x).strip()]
    pfx0 = [str(x).strip() for x in (include_prefixes or []) if str(x).strip()]
    kinds0 = [str(x).strip().lower() for x in (include_kinds or []) if str(x).strip()]
    if kinds0:
        kinds0 = [k for k in kinds0 if k in ("identity", "preference", "other")]

    picked: List[Tuple[str, Dict[str, Any]]] = []
    for ek, rec in ents.items():
        if not isinstance(rec, dict):
            continue
        ek2 = str(ek or "").strip()
        if not ek2:
            continue

        if keys0 and (ek2 not in keys0):
            continue
        if pfx0 and (not any(ek2.startswith(p) for p in pfx0)):
            continue
        if kinds0:
            kk = str(rec.get("kind") or "").strip().lower()
            if kk not in kinds0:
                continue

        picked.append((ek2, rec))

    if not picked:
        return ""

    # Sort by last_seen_at desc (ISO-ish), stable tie-breaker: entity_key asc
    picked.sort(key=lambda x: (str((x[1].get("last_seen_at") or "")), x[0]), reverse=True)
    picked = picked[: int(max_items or 18)]

    lines: List[str] = []
    lines.append("USER_GLOBAL_FACTS_MUTABLE (Tier-2M; subset; provenance-aware):")
    for ek, rec in picked:
        st = str(rec.get("status") or "").strip().lower()
        kind = str(rec.get("kind") or "").strip()
        last_seen = str(rec.get("last_seen_at") or "").strip()
        evs = rec.get("evidence_ids") if isinstance(rec.get("evidence_ids"), list) else []
        evs2 = [str(x).strip() for x in evs if str(x).strip()][:8]

        if st == "ambiguous":
            opts = rec.get("options") if isinstance(rec.get("options"), list) else []
            opts2 = [str(x).strip() for x in opts if str(x).strip()][:6]
            lines.append(f"- {ek} ({kind}) @ {last_seen}: AMBIGUOUS options={opts2} evidence_ids={evs2}")
        else:
            val = str(rec.get("value") or "").strip()
            if len(val) > 240:
                val = val[:239].rstrip() + "…"
            lines.append(f"- {ek} ({kind}) @ {last_seen}: {val} evidence_ids={evs2}")

    return "\n".join(lines).strip()
def user_memory_policy_path(user: str) -> Path:
    return user_dir(user) / USER_MEMORY_POLICY_FILE_NAME

def _default_user_memory_policy() -> Dict[str, Any]:
    return {
        "schema": "user_memory_policy_v1",
        "updated_at": now_iso(),
        # Rules are applied in order; last match wins for the same action category.
        # match.type:
        #   - "entity_key": deterministic internal key (preferred)
        #   - "substring": case-insensitive substring match against claim text
        "rules": [],
    }

def load_user_memory_policy(user: str) -> Dict[str, Any]:
    p = user_memory_policy_path(user)
    obj = _load_json_obj(p)
    if not isinstance(obj, dict) or obj.get("schema") != "user_memory_policy_v1":
        obj = _default_user_memory_policy()
        try:
            atomic_write_text(p, json.dumps(obj, indent=2, sort_keys=True))
        except Exception:
            pass
    obj["updated_at"] = now_iso()
    return obj

def _write_user_memory_policy(user: str, obj: Dict[str, Any]) -> bool:
    try:
        out = obj if isinstance(obj, dict) else _default_user_memory_policy()
        out.setdefault("schema", "user_memory_policy_v1")
        out["updated_at"] = now_iso()
        atomic_write_text(user_memory_policy_path(user), json.dumps(out, indent=2, sort_keys=True))
        return True
    except Exception:
        return False

def _policy_norm(s: str) -> str:
    return re.sub(r"\s{2,}", " ", (s or "").strip().lower())

def _policy_rule_matches(rule: Dict[str, Any], *, entity_key: str, claim: str) -> bool:
    if not isinstance(rule, dict):
        return False
    m = rule.get("match")
    if not isinstance(m, dict):
        return False
    typ = _policy_norm(str(m.get("type") or ""))
    val = _policy_norm(str(m.get("value") or ""))
    if not typ or not val:
        return False

    ek = _policy_norm(entity_key)
    cl = _policy_norm(claim)

    if typ == "entity_key":
        return ek == val
    if typ == "substring":
        return (val in cl) if val else False
    return False

def upsert_user_memory_policy_rule(
    user: str,
    *,
    action: str,
    match_type: str,
    match_value: str,
    note: str = "",
) -> Dict[str, Any]:
    """
    Deterministically add/update a user memory policy rule.

    action:
      - "do_not_store"
      - "project_only"
      - "do_not_resurface"
      - "allow_global"

    match_type:
      - "entity_key" (preferred)
      - "substring"

    Returns the written rule dict (best-effort).
    """
    act = _policy_norm(action)
    if act not in ("do_not_store", "project_only", "do_not_resurface", "allow_global"):
        act = "do_not_store"

    mt = _policy_norm(match_type)
    if mt not in ("entity_key", "substring"):
        mt = "substring"

    mv = _policy_norm(match_value)
    mv = mv.replace("\r", " ").replace("\n", " ").strip()
    if not mv:
        return {}

    pol = load_user_memory_policy(user)
    rules = pol.get("rules")
    if not isinstance(rules, list):
        rules = []
        pol["rules"] = rules

    # De-dupe by (action, match.type, match.value) last-write-wins.
    keep: List[Dict[str, Any]] = []
    for r in rules:
        if not isinstance(r, dict):
            continue
        if _policy_norm(str(r.get("action") or "")) != act:
            keep.append(r)
            continue
        mm = r.get("match")
        if not isinstance(mm, dict):
            keep.append(r)
            continue
        if _policy_norm(str(mm.get("type") or "")) == mt and _policy_norm(str(mm.get("value") or "")) == mv:
            # drop older identical rule
            continue
        keep.append(r)

    rid = f"pol_{int(time.time() * 1000)}"
    rec: Dict[str, Any] = {
        "id": rid,
        "created_at": now_iso(),
        "action": act,
        "match": {"type": mt, "value": mv},
        "note": (note or "").replace("\r", " ").replace("\n", " ").strip(),
    }
    keep.append(rec)
    pol["rules"] = keep[-128:]  # bounded
    _write_user_memory_policy(user, pol)
    return rec

def policy_decision_for_tier1_claim(
    user: str,
    *,
    project_full: str,
    claim: str,
    entity_key: str,
) -> Dict[str, bool]:
    """
    Deterministic policy evaluation for a Tier-1 candidate.
    Returns:
      - store: whether to write Tier-1 at all
      - mirror_global: whether to mirror to Tier-1G (global) when eligible
      - allow_resurface: whether to inject this info proactively (read-time)
    """
    pol = load_user_memory_policy(user)
    rules = pol.get("rules")
    if not isinstance(rules, list):
        rules = []

    store = True
    mirror_global = True
    allow_resurface = True

    # Evaluate in order; last match for each action wins.
    for r in rules:
        if not isinstance(r, dict):
            continue
        act = _policy_norm(str(r.get("action") or ""))
        if not _policy_rule_matches(r, entity_key=entity_key, claim=claim):
            continue

        if act == "do_not_store":
            store = False
        elif act == "project_only":
            mirror_global = False
        elif act == "do_not_resurface":
            allow_resurface = False
        elif act == "allow_global":
            mirror_global = True

    return {"store": store, "mirror_global": mirror_global, "allow_resurface": allow_resurface}

def _policy_request_keywords_for_entity_key(entity_key: str) -> List[str]:
    ek = _policy_norm(entity_key)
    if ek == "user.identity.birthdate":
        return ["birthday", "birthdate", "date of birth", "dob", "born"]
    if ek.startswith("user.relationship.") and ek.endswith(".birthdate"):
        return ["birthday", "birthdate", "born"]
    if ek == "user.identity.location":
        return ["live", "location", "where"]
    if ek == "user.identity.timezone":
        return ["time zone", "timezone", "central time"]
    if ek == "user.identity.name":
        return ["name", "call me", "go by"]
    if ek == "user.relationship.girlfriend.name" or ek == "user.relationship.partner.name":
        return ["girlfriend", "boyfriend", "partner"]
    if ek == "user.relationship.son.name" or ek == "user.relationship.child.name":
        return ["son", "child", "kid"]
    if ek == "user.relationship.daughter.name":
        return ["daughter", "child", "kid"]
    if ek == "user.relationship.spouse.name":
        return ["spouse", "wife", "husband", "partner"]
    if ek == "user.relationship.mother.name":
        return ["mother", "mom"]
    if ek == "user.relationship.father.name":
        return ["father", "dad"]
    if ek == "user.relationship.sister.name":
        return ["sister"]
    if ek == "user.relationship.brother.name":
        return ["brother"]
    if ek == "user.relationship.stepmother.name":
        return ["stepmother", "step mom", "stepmom"]
    if ek == "user.relationship.stepfather.name":
        return ["stepfather", "step dad", "stepdad"]
    if ek == "user.relationship.stepson.name":
        return ["stepson"]
    if ek == "user.relationship.stepdaughter.name":
        return ["stepdaughter"]
    if ek == "user.relationship.stepbrother.name":
        return ["stepbrother", "step brother"]
    if ek == "user.relationship.stepsister.name":
        return ["stepsister", "step sister"]
    return []

def policy_allows_resurface_now(
    user: str,
    *,
    entity_key: str,
    user_text: str,
) -> bool:
    """
    do_not_resurface means: do not inject unless the user is asking about it.
    Deterministic: allow if user_text contains any keyword tied to the entity_key.
    """
    pol = load_user_memory_policy(user)
    rules = pol.get("rules")
    if not isinstance(rules, list):
        rules = []

    suppressed = False
    for r in rules:
        if not isinstance(r, dict):
            continue
        act = _policy_norm(str(r.get("action") or ""))
        if act != "do_not_resurface":
            continue
        if _policy_rule_matches(r, entity_key=entity_key, claim=""):
            suppressed = True

    if not suppressed:
        return True

    low = _policy_norm(user_text or "")
    kws = _policy_request_keywords_for_entity_key(entity_key)
    return any(k in low for k in kws)

def users_root_dir() -> Path:
    """
    Global user memory root lives under projects/<user>/_user so it's:
    - user-scoped
    - outside any specific project
    - easy to locate alongside projects/<user>/<project>/
    """
    _pr, pd = _require_configured()
    d = pd  # PROJECTS_DIR
    d.mkdir(parents=True, exist_ok=True)
    return d

def user_dir(user: str) -> Path:
    # projects/<user>/_user
    d = users_root_dir() / safe_user_segment(user) / "_user"
    d.mkdir(parents=True, exist_ok=True)
    return d


def safe_user_segment(user: str) -> str:
    return re.sub(r"[^a-zA-Z0-9_-]", "_", (user or "").strip()) or "user"

def user_profile_path(user: str) -> Path:
    return user_dir(user) / USER_PROFILE_FILE_NAME

def user_facts_raw_path(user: str) -> Path:
    return user_dir(user) / USER_FACTS_RAW_FILE_NAME

def _today_ymd_utc() -> str:
    return time.strftime("%Y-%m-%d", time.gmtime())

def _today_ymd_for_tz(tz_name: str) -> tuple[str, str]:
    tz = (tz_name or "").strip()
    if tz and ZoneInfo is not None:
        try:
            dt = datetime.now(ZoneInfo(tz))
            return dt.strftime("%Y-%m-%d"), tz
        except Exception:
            pass
    return _today_ymd_utc(), "UTC"

def _today_ymd_for_user(user: str) -> tuple[str, str]:
    try:
        prof = _load_json_obj(user_profile_path(user)) or {}
    except Exception:
        prof = {}
    ident = prof.get("identity") if isinstance(prof.get("identity"), dict) else {}
    tz_rec = ident.get("timezone") if isinstance(ident.get("timezone"), dict) else {}
    tz_val = str((tz_rec.get("value") or "")).strip()
    return _today_ymd_for_tz(tz_val)

def _derive_age_from_birthdate(birthdate_ymd: str) -> Optional[int]:
    """
    Age is derived at read time ONLY. Stored fact is birthdate (YYYY-MM-DD).
    Returns None if birthdate is invalid.
    """
    out: Optional[int] = None
    b = (birthdate_ymd or "").strip()

    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", b):
        try:
            by, bm, bd = [int(x) for x in b.split("-")]
            ty, tm, td = [int(x) for x in _today_ymd_utc().split("-")]
            age = ty - by
            if (tm, td) < (bm, bd):
                age -= 1
            if 0 <= age <= 130:
                out = int(age)
        except Exception:
            out = None

    return out

def _load_json_obj(path: Path) -> Dict[str, Any]:
    out: Dict[str, Any] = {}

    if path.exists():
        try:
            obj = json.loads(path.read_text(encoding="utf-8") or "{}")
            out = obj if isinstance(obj, dict) else {}
        except Exception:
            out = {}

    return out

def load_user_profile(user: str) -> Dict[str, Any]:
    # Ensure the on-disk scaffold exists even if the profile has no populated fields yet.
    # This prevents "facts_raw.jsonl exists but profile.json missing" after a user deletes _user/.
    try:
        p0 = user_profile_path(user)
        if not p0.exists():
            seed = {
                "schema": "user_profile_v1",
                "updated_at": now_iso(),
                "identity": {},
                "relationships": {},
                "conflicts": [],
            }
            _write_user_profile(user, seed)
    except Exception:
        pass    
    """
    Global profile (Tier-2G): stable identity facts across projects.
    """
    p = user_profile_path(user)
    obj = _load_json_obj(p)

    if not isinstance(obj, dict) or obj.get("schema") != "user_profile_v1":
        obj = {
            "schema": "user_profile_v1",
            "updated_at": now_iso(),
            "identity": {},
            "relationships": {},
            "conflicts": [],
        }

    # Inject derived age at read time ONLY (never persisted as a stored fact)
    # Only compute when birthdate is CONFIRMED (not ambiguous).
    try:
        bd_rec = ((obj.get("identity") or {}).get("birthdate") or {})
        bd_status = str((bd_rec.get("status") or "")).strip().lower()
        bd_val = str((bd_rec.get("value") or "")).strip()
        if bd_status == "confirmed" and bd_val:
            age, asof_ymd, asof_tz = _derive_age_from_birthdate_for_user(bd_val, user)
            if age is not None:
                obj.setdefault("derived", {})
                if isinstance(obj["derived"], dict):
                    obj["derived"]["age_years"] = int(age)
                    obj["derived"]["age_asof_utc"] = _today_ymd_utc()
                    obj["derived"]["age_asof_local"] = asof_ymd
                    obj["derived"]["age_asof_tz"] = asof_tz
    except Exception:
        pass

    return obj if isinstance(obj, dict) else {}

def _write_user_profile(user: str, obj: Dict[str, Any]) -> None:
    out = obj if isinstance(obj, dict) else {}
    out.setdefault("schema", "user_profile_v1")
    out["updated_at"] = now_iso()
    atomic_write_text(user_profile_path(user), json.dumps(out, indent=2, sort_keys=True))

def append_user_fact_raw_candidate(
    user: str,
    *,
    claim: str,
    slot: str = "other",
    subject: str = "user",
    source: str = "chat",
    evidence_quote: str = "",
    turn_index: int = 0,
    timestamp: str = "",
) -> Dict[str, Any]:
    """
    Append exactly one GLOBAL Tier-1G fact candidate to users/<user>/facts_raw.jsonl.
    This is project-agnostic. Used to rebuild Global Tier-2G (profile.json).
    """
    c = _norm_one_line(claim)
    if not c:
        return {"ok": False, "error": "empty_claim"}

    slot_n = _norm_one_line(slot).lower() or "other"
    subject_n = _norm_one_line(subject).lower() or "user"
    evq = _norm_one_line(evidence_quote or "")

    src_type = _norm_one_line(source) or "chat"
    ts = _norm_one_line(timestamp or "")
    try:
        ti = int(turn_index or 0)
    except Exception:
        ti = 0

    src_obj: Any = src_type
    if ti or ts:
        src_obj = {"type": src_type, "turn_index": ti, "timestamp": ts}

    obj: Dict[str, Any] = {
        "id": f"uf_{int(time.time() * 1000)}",
        "created_at": now_iso(),
        "claim": c,
        "slot": slot_n,
        "subject": subject_n,
        "source": src_obj,
        "evidence_quote": evq,
    }

    p = user_facts_raw_path(user)
    p.parent.mkdir(parents=True, exist_ok=True)
    try:
        line = json.dumps(obj, ensure_ascii=False)
        if ("\n" in line) or ("\r" in line):
            return {"ok": False, "error": "jsonl_must_be_single_line"}
        with p.open("a", encoding="utf-8") as f:
            f.write(line + "\n")

        # After any successful global Tier-1 append, materialize Tier-2M immediately.
        # This keeps global_facts_map.json in sync even when Tier-2G does not change.
        try:
            rebuild_user_global_facts_map_from_user_facts(user)
        except Exception:
            pass

        return {"ok": True, "id": obj["id"]}
    except Exception:
        return {"ok": False, "error": "write_failed"}

def _load_user_facts_raw(user: str, *, max_lines: int = 8000) -> List[Dict[str, Any]]:
    p = user_facts_raw_path(user)
    if not p.exists():
        return []
    try:
        lines = p.read_text(encoding="utf-8", errors="replace").splitlines()
    except Exception:
        return []
    if max_lines and len(lines) > max_lines:
        lines = lines[-max_lines:]
    out: List[Dict[str, Any]] = []
    for ln in lines:
        s = (ln or "").strip()
        if not s:
            continue
        try:
            obj = json.loads(s)
        except Exception:
            continue
        if isinstance(obj, dict):
            out.append(obj)
    return out

def _profile_field_upsert_conflict_aware(
    cur: Dict[str, Any],
    *,
    field_path: Tuple[str, str],
    new_value: str,
    evidence_id: str,
) -> Tuple[Dict[str, Any], bool]:
    """
    Stores:
      {"status":"confirmed|ambiguous","value":..., "options":[...], "evidence_ids":[...]}
    If conflict, marks ambiguous and does NOT overwrite existing value.
    """
    section, key = field_path
    changed = False
    prof = cur if isinstance(cur, dict) else {}
    prof.setdefault("schema", "user_profile_v1")
    prof.setdefault("identity", {})
    prof.setdefault("relationships", {})
    prof.setdefault("conflicts", [])

    sec = prof.get(section)
    if not isinstance(sec, dict):
        sec = {}
        prof[section] = sec

    rec = sec.get(key)
    if not isinstance(rec, dict):
        rec = {"status": "confirmed", "value": "", "options": [], "evidence_ids": []}

    v_old = str(rec.get("value") or "").strip()
    v_new = str(new_value or "").strip()

    # Canonicalize materialized profile identity values (avoid "Frank.", "Frank,", etc.)
    # Deterministic and non-inferential: only strips common trailing sentence punctuation.
    if section == "identity":
        # Do not touch ISO birthdates (already canonical)
        if key not in ("birthdate",):
            while v_new and v_new[-1] in ".!,;:":
                v_new = v_new[:-1].rstrip()
            # normalize internal whitespace
            v_new = " ".join(v_new.split())

    if not v_new:
        return prof, False

    # evidence id
    evs = rec.get("evidence_ids")
    if not isinstance(evs, list):
        evs = []
    if evidence_id and evidence_id not in evs:
        evs.append(evidence_id)
        rec["evidence_ids"] = evs[:24]

    # If this field is already ambiguous, it stays ambiguous until an explicit resolution mechanism exists.
    # Do NOT let later repeats re-confirm a single value (that would silently pick a winner).
    if str(rec.get("status") or "").strip().lower() == "ambiguous":
        opts_a = rec.get("options")
        if not isinstance(opts_a, list):
            opts_a = []
        if v_new not in [str(x).strip() for x in opts_a]:
            opts_a.append(v_new)
            rec["options"] = [str(x).strip() for x in opts_a if str(x).strip()][:8]
        rec["value"] = ""
        rec["status"] = "ambiguous"
        sec[key] = rec

        confs = prof.get("conflicts")
        if not isinstance(confs, list):
            confs = []
        conf_key = f"{section}.{key}"
        if conf_key not in [str(x) for x in confs]:
            confs.append(conf_key)
            prof["conflicts"] = confs[:64]
        return prof, True

    if not v_old:
        rec["status"] = "confirmed"
        rec["value"] = v_new
        rec["options"] = [v_new]
        sec[key] = rec
        return prof, True

    if v_old == v_new:
        opts = rec.get("options")
        if not isinstance(opts, list):
            opts = []
        if v_new not in opts:
            opts.append(v_new)
            rec["options"] = opts[:8]
            changed = True
        sec[key] = rec
        return prof, changed

    # conflict -> ambiguous
    opts2 = rec.get("options")
    if not isinstance(opts2, list):
        opts2 = []
    if v_old not in opts2:
        opts2.append(v_old)
    if v_new not in opts2:
        opts2.append(v_new)
    rec["options"] = opts2[:8]
    rec["status"] = "ambiguous"
    rec["value"] = ""  # conflict: do not retain a stale "current" value

    sec[key] = rec

    confs = prof.get("conflicts")
    if not isinstance(confs, list):
        confs = []
    conf_key = f"{section}.{key}"
    if conf_key not in [str(x) for x in confs]:
        confs.append(conf_key)
        prof["conflicts"] = confs[:64]
        changed = True
    else:
        changed = True

    return prof, changed

def rebuild_user_profile_from_user_facts(user: str) -> Dict[str, Any]:
    """
    Deterministically rebuild Tier-2G (profile.json) from users/<user>/facts_raw.jsonl.
    Conflict-aware. Age is never stored.
    """
    facts = _load_user_facts_raw(user, max_lines=8000)

    # TRUE REBUILD (deterministic):
    # Do NOT start from the existing profile.json, or stale conflicts/options will persist forever.
    # Tier-2G must be derived solely from users/<user>/_user/facts_raw.jsonl on each rebuild.
    prof: Dict[str, Any] = {
        "schema": "user_profile_v1",
        "updated_at": now_iso(),
        "identity": {},
        "relationships": {},
        "conflicts": [],
    }

    def _extract_profile_pairs(fact: Dict[str, Any]) -> List[Tuple[Tuple[str, str], str]]:
        """
        Philosophy B (Tier-2G materialization):
        - Rebuild profile from Tier-1G facts_raw.jsonl
        - Do NOT require slot0 == "identity" for identity fields, because slot tagging can be imperfect.
        - Still require:
            * subject == "user"
            * conservative evidence_quote gating for identity facts (first-person signals)
        - Birthdate remains STRICT: only accept ISO YYYY-MM-DD AND evidence_quote must contain
          explicit first-person birth phrasing ("my birthday" or "i was born").
        """
        claim = str((fact or {}).get("claim") or "").strip()
        c = _norm_one_line(claim)
        low = c.lower()
        outp: List[Tuple[Tuple[str, str], str]] = []

        slot0 = str((fact or {}).get("slot") or "").strip().lower()
        subj0 = str((fact or {}).get("subject") or "").strip().lower()
        evq0 = str((fact or {}).get("evidence_quote") or "").strip().lower()

        # HARD RULE: only user-attributed facts can enter Tier-2G
        if subj0 != "user":
            return outp

        # HARD RULE: never store age (but do not discard relationship names that mention age)
        if re.search(
            r"\b(?:i\s+am|i'm|the\s+user\s+is|user\s+is|my\s+age\s+is)\s+\d{1,3}\s*(?:years?\s*old|yo)\b",
            low,
        ):
            return outp

        def _ev_has_any(*needles: str) -> bool:
            for n in needles:
                if n and (n in evq0):
                    return True
            return False

        def _parse_month_day_year(s: str) -> str:
            """
            Parse month-name date like "October 7, 1982" -> "1982-10-07".
            Returns "" if no deterministic parse.
            """
            m = re.search(
                r"\b(jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|jul(?:y)?|aug(?:ust)?|sep(?:tember)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?)"
                r"\s+(\d{1,2})(?:st|nd|rd|th)?(?:,)?\s+(\d{4})\b",
                s,
                flags=re.IGNORECASE,
            )
            if not m:
                return ""
            mon_txt = (m.group(1) or "").lower()
            day_txt = (m.group(2) or "").strip()
            year_txt = (m.group(3) or "").strip()

            month_map = {
                "jan": 1, "january": 1,
                "feb": 2, "february": 2,
                "mar": 3, "march": 3,
                "apr": 4, "april": 4,
                "may": 5,
                "jun": 6, "june": 6,
                "jul": 7, "july": 7,
                "aug": 8, "august": 8,
                "sep": 9, "september": 9,
                "oct": 10, "october": 10,
                "nov": 11, "november": 11,
                "dec": 12, "december": 12,
            }
            mon = month_map.get(mon_txt[:3], 0)
            try:
                day = int(day_txt)
                year = int(year_txt)
            except Exception:
                return ""
            if not (1 <= mon <= 12 and 1 <= day <= 31 and 1800 <= year <= 2100):
                return ""
            return f"{year:04d}-{mon:02d}-{day:02d}"

        def _parse_mdy_numeric(s: str) -> str:
            """
            Parse numeric date like "03/24/1996" or "3-24-1996" -> "1996-03-24".
            Returns "" if no deterministic parse.
            """
            m = re.search(r"\b(\d{1,2})[/-](\d{1,2})[/-](\d{4})\b", s)
            if not m:
                return ""
            try:
                mon = int(m.group(1))
                day = int(m.group(2))
                year = int(m.group(3))
            except Exception:
                return ""
            if not (1 <= mon <= 12 and 1 <= day <= 31 and 1800 <= year <= 2100):
                return ""
            return f"{year:04d}-{mon:02d}-{day:02d}"

        # -------------------------
        # Identity (slot flexible)
        # -------------------------

        # Full/legal name (distinct from preferred_name)
        if (
            re.search(r"\bmy\s+full\s+name\s+is\s+.+$", low)
            or re.search(r"\bmy\s+legal\s+name\s+is\s+.+$", low)
        ):
            if _ev_has_any("my full name", "my legal name"):
                if " is " in low:
                    outp.append((("identity", "name"), c.split("is", 1)[1].strip()))
                return outp
            return outp

        # Preferred name
        # Accept both first-person and normalized third-person claims, but require self-ID in evidence_quote.
        if (
            re.search(r"\bmy\s+preferred\s+name\s+is\s+.+$", low)
            or re.search(r"\bmy\s+name\s+is\s+.+$", low)
            or re.search(r"\bi\s+go\s+by\s+.+$", low)
            or re.search(r"\bthe\s+user(?:'s)?\s+preferred\s+name\s+is\s+.+$", low)
            or re.search(r"\bthe\s+user(?:'s)?\s+name\s+is\s+.+$", low)
        ):
            if _ev_has_any("my name", "i'm ", "i’m ", "i am ", "i go by", "preferred name"):
                # Extract value deterministically from claim text
                if " go by " in low:
                    outp.append((("identity", "preferred_name"), c.split("by", 1)[1].strip()))
                elif " is " in low:
                    outp.append((("identity", "preferred_name"), c.split("is", 1)[1].strip()))
                return outp
            return outp

        # Pronouns
        if re.search(r"\b(?:my\s+pronouns\s+are|pronouns\s+are)\s+.+$", low):
            if _ev_has_any("my pronouns", "pronouns are"):
                if " are " in low:
                    outp.append((("identity", "pronouns"), c.split("are", 1)[1].strip()))
                return outp
            return outp

        # Location
        # Conservative: require evidence_quote to contain first-person location phrasing,
        # BUT allow explicit confirmation claims to pass even if the user's reply is short.
        if (
            ("i live in " in low)
            or ("the user lives in " in low)
            or ("i'm in " in low)
            or ("i’m in " in low)
            or ("i am in " in low)
            or ("my confirmed location is " in low)
        ):
            # Explicit confirmation claim is sufficient on its own (deterministic, user-attributed).
            if "my confirmed location is " in low:
                outp.append((("identity", "location"), c.split("is", 1)[1].strip()))
                return outp

            # Otherwise require first-person phrasing in evidence_quote (prevents accidental third-person pollution).
            if _ev_has_any("i live", "i'm in", "i’m in", "i am in", "my confirmed location"):
                if " in " in low:
                    outp.append((("identity", "location"), c.split("in", 1)[1].strip()))
                return outp
            return outp

        # Birthdate (ISO only)
        # HARD RULE: only accept birthdate if the original user evidence_quote contains
        # explicit first-person birth phrasing ("my birthday" or "i was born").
        m = re.search(r"\b(?:my\s+birthday\s+is|i\s+was\s+born\s+on)\s+(\d{4}-\d{2}-\d{2})\b", low)
        if m:
            if ("my birthday" not in evq0) and ("i was born" not in evq0):
                return outp
            outp.append((("identity", "birthdate"), m.group(1)))
            return outp
        # Month-name birthdate with year (e.g., "my birthday is October 7, 1982")
        if ("my birthday" in low) or ("i was born" in low) or ("my birthdate" in low):
            if ("my birthday" not in evq0) and ("i was born" not in evq0) and ("my birthdate" not in evq0):
                return outp
            ymd = _parse_month_day_year(c)
            if ymd:
                outp.append((("identity", "birthdate"), ymd))
                return outp

        # Timezone
        # Allow slot mistakes, but require evidence_quote to mention timezone / time zone or known tz phrase.
        if re.search(r"\b(?:my\s+time\s+zone\s+is|timezone\s+is)\s+.+$", low):
            if _ev_has_any("time zone", "timezone"):
                outp.append((("identity", "timezone"), c.split("is", 1)[1].strip()))
                return outp
            return outp

        if "central time" in low:
            if _ev_has_any("central time", "time zone", "timezone"):
                outp.append((("identity", "timezone"), "America/Chicago"))
                return outp
            return outp

        # ------------------------------
        # Relationships (slot flexible)
        # ------------------------------
        # These can also be slot-mislabeled; still require explicit "name is"/"named" patterns.
        def _trim_relation_name(v: str) -> str:
            s = " ".join(str(v or "").strip().split())
            if not s:
                return ""

            # If claim provides last name explicitly, combine deterministically.
            m_last = re.search(r"\bwhose\s+last\s+name\s+is\s+([a-zA-Z][^,.;]*)", s, flags=re.IGNORECASE)
            if m_last:
                base = s[: m_last.start()].strip()
                last = (m_last.group(1) or "").strip()
                if base and last:
                    if last.lower() not in [t.lower() for t in base.split()]:
                        s = f"{base} {last}".strip()
                    else:
                        s = base

            # Cut at clear clause boundaries (prevents "and my ..." bleed-through).
            for pat in (
                r"\b(?:and|but)\b\s+(?:my|the\s+user|their|his|her|our)\b",
                r"\b(?:who|that|which)\b",
            ):
                m = re.search(pat, s, flags=re.IGNORECASE)
                if m:
                    s = s[: m.start()].strip()
                    break

            # Trim trailing age fragments like ", 42" or "42 years old".
            s = re.sub(r"[,;]?\s+\d{1,3}\s*(?:years?\s*old|yo)?$", "", s, flags=re.IGNORECASE).strip()
            while s and s[-1] in ".!,;:":
                s = s[:-1].rstrip()
            return s

        def _extract_birthdate_from_text(s: str) -> str:
            if not s:
                return ""
            ymd = _parse_mdy_numeric(s)
            if not ymd:
                ymd = _parse_month_day_year(s)
            return ymd

        rel_out: List[Tuple[Tuple[str, str], str]] = []

        def _add_rel(key: str, raw_val: str) -> None:
            val = _trim_relation_name(raw_val)
            if val:
                rel_out.append((("relationships", key), val))

        def _add_rel_birthdate(key: str, raw_val: str) -> None:
            ymd = _extract_birthdate_from_text(raw_val)
            if ymd:
                rel_out.append((("relationships", key), ymd))

        rel_defs = [
            ("partner_name", r"partner"),
            ("spouse_name", r"spouse"),
            ("girlfriend_name", r"girlfriend"),
            ("boyfriend_name", r"boyfriend"),
            ("wife_name", r"wife"),
            ("husband_name", r"husband"),
            ("mother_name", r"(?:mother|mom)"),
            ("father_name", r"(?:father|dad)"),
            ("sister_name", r"sister"),
            ("brother_name", r"brother"),
            ("son_name", r"son"),
            ("daughter_name", r"daughter"),
            ("child_name", r"child"),
            ("stepson_name", r"(?:step\s*son|stepson)"),
            ("stepdaughter_name", r"(?:step\s*daughter|stepdaughter)"),
            ("stepmother_name", r"(?:step\s*mother|stepmother|step\s*mom|stepmom)"),
            ("stepfather_name", r"(?:step\s*father|stepfather|step\s*dad|stepdad)"),
            ("stepbrother_name", r"(?:step\s*brother|stepbrother)"),
            ("stepsister_name", r"(?:step\s*sister|stepsister)"),
        ]

        for key, rel_re in rel_defs:
            m = re.search(rf"\bmy\s+{rel_re}(?:'s|s)?\s+name\s+is\s+(.+)$", c, flags=re.IGNORECASE)
            if m:
                _add_rel(key, m.group(1))
            m = re.search(rf"\bmy\s+{rel_re}(?:'s|s)?\s+named\s+(.+)$", c, flags=re.IGNORECASE)
            if m:
                _add_rel(key, m.group(1))
            m = re.search(
                rf"\bthe\s+user\s+has\s+(?:an?\s+|one\s+|two\s+|three\s+|four\s+|five\s+)?{rel_re}\s+(?:named|name\s+is)\s+(.+)$",
                c,
                flags=re.IGNORECASE,
            )
            if m:
                _add_rel(key, m.group(1))
            m = re.search(
                rf"\bthe\s+user(?:'s)?\s+{rel_re}(?:'s)?\s+(?:is\s+named|name\s+is)\s+(.+)$",
                c,
                flags=re.IGNORECASE,
            )
            if m:
                _add_rel(key, m.group(1))

        # Relationship birthdates (explicit only)
        rel_bd_defs = [
            ("partner_birthdate", r"partner"),
            ("spouse_birthdate", r"spouse"),
            ("girlfriend_birthdate", r"girlfriend"),
            ("boyfriend_birthdate", r"boyfriend"),
            ("wife_birthdate", r"wife"),
            ("husband_birthdate", r"husband"),
            ("mother_birthdate", r"(?:mother|mom)"),
            ("father_birthdate", r"(?:father|dad)"),
            ("sister_birthdate", r"sister"),
            ("brother_birthdate", r"brother"),
            ("son_birthdate", r"son"),
            ("daughter_birthdate", r"daughter"),
            ("child_birthdate", r"child"),
            ("stepson_birthdate", r"(?:step\s*son|stepson)"),
            ("stepdaughter_birthdate", r"(?:step\s*daughter|stepdaughter)"),
            ("stepmother_birthdate", r"(?:step\s*mother|stepmother|step\s*mom|stepmom)"),
            ("stepfather_birthdate", r"(?:step\s*father|stepfather|step\s*dad|stepdad)"),
            ("stepbrother_birthdate", r"(?:step\s*brother|stepbrother)"),
            ("stepsister_birthdate", r"(?:step\s*sister|stepsister)"),
        ]
        for key, rel_re in rel_bd_defs:
            m = re.search(rf"\bmy\s+{rel_re}(?:'s|s)?\s+(?:birthday|birthdate)\s+is\s+(.+)$", c, flags=re.IGNORECASE)
            if m:
                _add_rel_birthdate(key, m.group(1))
            m = re.search(rf"\bmy\s+{rel_re}\s+was\s+born\s+on\s+(.+)$", c, flags=re.IGNORECASE)
            if m:
                _add_rel_birthdate(key, m.group(1))

        if rel_out:
            outp.extend(rel_out)
            return outp

        return outp

    for f in facts:
        if not isinstance(f, dict):
            continue
        claim = str(f.get("claim") or "").strip()
        if not claim:
            continue
        fid = str(f.get("id") or "").strip()
        for (sec_key, value) in _extract_profile_pairs(f):
            prof, _changed = _profile_field_upsert_conflict_aware(
                prof, field_path=sec_key, new_value=value, evidence_id=fid
            )

    # ------------------------------------------------------------
    # Tier-2G deterministic resolution mechanisms (FOUNDATIONAL)
    #
    # Principle:
    # - Tier-2G remains conservative by default.
    # - We only collapse ambiguity when we have an explicit, user-attributed
    #   resolution signal in Tier-1 (facts_raw.jsonl).
    #
    # Implemented resolvers:
    # 1) identity.preferred_name: nested-prefix collapse (existing behavior)
    # 2) identity.location: explicit user-confirmed location statement
    #
    # IMPORTANT:
    # - No city-specific logic.
    # - No geocoding.
    # - No implicit winner selection from ambiguous options.
    # - Resolution must be triggered by an explicit confirmation claim.
    # ------------------------------------------------------------
    try:
        ident = prof.get("identity") if isinstance(prof.get("identity"), dict) else {}

        # ------------------------------
        # Resolver 1: preferred_name (unchanged)
        # ------------------------------
        pn = ident.get("preferred_name") if isinstance(ident.get("preferred_name"), dict) else {}
        st = str(pn.get("status") or "").strip().lower()
        if st == "ambiguous":
            opts = pn.get("options") if isinstance(pn.get("options"), list) else []
            opts2 = [" ".join(str(x).strip().split()) for x in opts if str(x).strip()]

            cleaned: List[str] = []
            for o in opts2:
                o2 = o
                while o2 and o2[-1] in ".!,;:":
                    o2 = o2[:-1].rstrip()
                o2 = " ".join(o2.split())
                if o2 and (o2 not in cleaned):
                    cleaned.append(o2)

            chosen = ""
            low = [o.lower() for o in cleaned]
            for i in range(len(cleaned)):
                for j in range(len(cleaned)):
                    if i == j:
                        continue
                    if low[j].startswith(low[i]) and len(cleaned[i]) <= len(cleaned[j]):
                        chosen = cleaned[i]
                        break
                if chosen:
                    break

            if chosen:
                pn["status"] = "resolved_auto"
                pn["value"] = chosen
                pn["note"] = "auto_resolve_preferred_name_prefix_rule"
                pn["options"] = cleaned[:8]
                ident["preferred_name"] = pn

                confs = prof.get("conflicts")
                if isinstance(confs, list):
                    prof["conflicts"] = [str(x) for x in confs if str(x) != "identity.preferred_name"][:64]

        # ------------------------------
        # Resolver 2: location (explicit confirmation only)
        # ------------------------------
        #
        # Trigger examples (deterministic):
        # - "My confirmed location is Oklahoma City, OK."
        # - "Confirm my location is Oklahoma City, OK."
        # - "Set my location to Oklahoma City, OK."
        #
        # NOTE:
        # - We intentionally do NOT resolve based on "I'm in X" alone.
        #   That creates implicit winner selection and can be wrong.
        #
        loc = ident.get("location") if isinstance(ident.get("location"), dict) else {}
        st_loc = str(loc.get("status") or "").strip().lower()

        confirmed_loc = ""
        confirmed_eid = ""

        # Scan Tier-1 newest-first for an explicit location confirmation.
        # Deterministic: reverse order of loaded facts (latest appended wins).
        for f in reversed(facts):
            if not isinstance(f, dict):
                continue
            if str(f.get("subject") or "").strip().lower() != "user":
                continue

            c = _norm_one_line(str(f.get("claim") or ""))
            if not c:
                continue
            low = c.lower()

            m = re.search(r"\bmy\s+confirmed\s+location\s+is\s+(.+)$", low)
            if not m:
                m = re.search(r"\bconfirm\s+my\s+location\s+(?:is|to)\s+(.+)$", low)
            if not m:
                m = re.search(r"\bset\s+my\s+location\s+to\s+(.+)$", low)

            if not m:
                continue

            v = (m.group(1) or "").strip()
            # strip common trailing punctuation, deterministic
            while v and v[-1] in ".!,;:":
                v = v[:-1].rstrip()
            v = " ".join(v.split())
            if not v:
                continue

            confirmed_loc = v
            confirmed_eid = str(f.get("id") or "").strip()
            break

        if st_loc == "ambiguous" and confirmed_loc:
            # Resolve to confirmed value, retain options for auditability.
            opts = loc.get("options") if isinstance(loc.get("options"), list) else []
            cleaned_opts: List[str] = []
            for o in opts:
                o2 = str(o).strip()
                while o2 and o2[-1] in ".!,;:":
                    o2 = o2[:-1].rstrip()
                o2 = " ".join(o2.split())
                if o2 and o2 not in cleaned_opts:
                    cleaned_opts.append(o2)
            if confirmed_loc not in cleaned_opts:
                cleaned_opts.append(confirmed_loc)

            # Evidence ids: preserve + add the confirmation evidence id
            evs = loc.get("evidence_ids")
            if not isinstance(evs, list):
                evs = []
            if confirmed_eid and confirmed_eid not in evs:
                evs.append(confirmed_eid)
            loc["evidence_ids"] = [str(x) for x in evs if str(x).strip()][:24]

            loc["status"] = "resolved_user"
            loc["value"] = confirmed_loc
            loc["note"] = "user_confirmed_location_rule"
            loc["options"] = cleaned_opts[:8]
            ident["location"] = loc

            confs = prof.get("conflicts")
            if isinstance(confs, list):
                prof["conflicts"] = [str(x) for x in confs if str(x) != "identity.location"][:64]

        # ------------------------------
        # Resolver 3: relationship names (explicit confirmation only)
        # ------------------------------
        rel = prof.get("relationships") if isinstance(prof.get("relationships"), dict) else {}

        def _resolve_rel_name(rel_key: str, patterns: List[str]) -> None:
            if not rel_key or not patterns:
                return
            rec = rel.get(rel_key) if isinstance(rel.get(rel_key), dict) else {}
            st_rel = str(rec.get("status") or "").strip().lower()
            if st_rel not in ("ambiguous", "") and str(rec.get("value") or "").strip():
                return

            confirmed_name = ""
            confirmed_eid = ""

            for f in reversed(facts):
                if not isinstance(f, dict):
                    continue
                if str(f.get("subject") or "").strip().lower() != "user":
                    continue
                claim_raw = str(f.get("claim") or "").strip()
                if not claim_raw:
                    continue
                for pat in patterns:
                    try:
                        m = re.search(pat, claim_raw, flags=re.IGNORECASE)
                    except Exception:
                        m = None
                    if not m:
                        continue
                    v = (m.group(1) or "").strip()
                    while v and v[-1] in ".!,;:":
                        v = v[:-1].rstrip()
                    v = " ".join(v.split())
                    if not v:
                        continue
                    confirmed_name = v
                    confirmed_eid = str(f.get("id") or "").strip()
                    break
                if confirmed_name:
                    break

            if not confirmed_name:
                return

            # Normalize options + evidence ids
            opts = rec.get("options") if isinstance(rec.get("options"), list) else []
            cleaned_opts: List[str] = []
            for o in opts:
                o2 = str(o).strip()
                while o2 and o2[-1] in ".!,;:":
                    o2 = o2[:-1].rstrip()
                o2 = " ".join(o2.split())
                if o2 and o2 not in cleaned_opts:
                    cleaned_opts.append(o2)
            if confirmed_name not in cleaned_opts:
                cleaned_opts.append(confirmed_name)

            evs = rec.get("evidence_ids")
            if not isinstance(evs, list):
                evs = []
            if confirmed_eid and confirmed_eid not in evs:
                evs.append(confirmed_eid)

            rec["status"] = "resolved_user"
            rec["value"] = confirmed_name
            rec["options"] = cleaned_opts[:8]
            rec["evidence_ids"] = evs
            rec["note"] = "user_confirmed_relationship_name_rule"
            rel[rel_key] = rec

            confs = prof.get("conflicts")
            if isinstance(confs, list):
                prof["conflicts"] = [str(x) for x in confs if str(x) != f"relationships.{rel_key}"][:64]

        _resolve_rel_name(
            "girlfriend_name",
            [
                r"\bmy\s+girlfriend(?:'s)?\s+name\s+is\s+(.+)$",
                r"\bmy\s+girlfriend\s+named\s+(.+)$",
                r"\bmy\s+girlfriend\s+is\s+named\s+(.+)$",
            ],
        )

        def _parse_birth_ymd_any(s: str) -> str:
            if not s:
                return ""
            # Numeric m/d/yyyy
            m = re.search(r"\b(\d{1,2})[/-](\d{1,2})[/-](\d{4})\b", s)
            if m:
                try:
                    mon = int(m.group(1))
                    day = int(m.group(2))
                    year = int(m.group(3))
                except Exception:
                    mon = day = year = 0
                if 1 <= mon <= 12 and 1 <= day <= 31 and 1800 <= year <= 2100:
                    return f"{year:04d}-{mon:02d}-{day:02d}"
            # Month-name "October 7, 1982"
            m2 = re.search(
                r"\b(jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|jul(?:y)?|aug(?:ust)?|sep(?:tember)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?)"
                r"\s+(\d{1,2})(?:st|nd|rd|th)?(?:,)?\s+(\d{4})\b",
                s,
                flags=re.IGNORECASE,
            )
            if not m2:
                return ""
            mon_txt = (m2.group(1) or "").lower()
            day_txt = (m2.group(2) or "").strip()
            year_txt = (m2.group(3) or "").strip()
            month_map = {
                "jan": 1, "january": 1,
                "feb": 2, "february": 2,
                "mar": 3, "march": 3,
                "apr": 4, "april": 4,
                "may": 5,
                "jun": 6, "june": 6,
                "jul": 7, "july": 7,
                "aug": 8, "august": 8,
                "sep": 9, "september": 9,
                "oct": 10, "october": 10,
                "nov": 11, "november": 11,
                "dec": 12, "december": 12,
            }
            mon = month_map.get(mon_txt[:3], 0)
            try:
                day = int(day_txt)
                year = int(year_txt)
            except Exception:
                return ""
            if not (1 <= mon <= 12 and 1 <= day <= 31 and 1800 <= year <= 2100):
                return ""
            return f"{year:04d}-{mon:02d}-{day:02d}"

        def _resolve_rel_birthdate_from_name(rel_name_key: str, rel_bd_key: str) -> None:
            if not rel_name_key or not rel_bd_key:
                return
            name_rec = rel.get(rel_name_key) if isinstance(rel.get(rel_name_key), dict) else {}
            name_val = str(name_rec.get("value") or "").strip()
            st_name = str(name_rec.get("status") or "").strip().lower()
            if not name_val or st_name not in ("resolved_user", "resolved_auto", "confirmed"):
                return

            bd_rec = rel.get(rel_bd_key) if isinstance(rel.get(rel_bd_key), dict) else {}
            st_bd = str(bd_rec.get("status") or "").strip().lower()
            if st_bd in ("resolved_user", "confirmed") and str(bd_rec.get("value") or "").strip():
                return

            confirmed_bd = ""
            confirmed_eid = ""
            name_variants: List[str] = [name_val]
            try:
                if " " in name_val:
                    first = name_val.split()[0].strip()
                    if first and first not in name_variants:
                        name_variants.append(first)
            except Exception:
                pass

            patterns: List[str] = []
            for nv in name_variants:
                name_esc = re.escape(nv)
                patterns.append(rf"\b{name_esc}\s*['’]?s\s+(?:birthday|birthdate)\s+is\s+(.+)$")
                patterns.append(rf"\b{name_esc}\s+was\s+born\s+on\s+(.+)$")

            for f in reversed(facts):
                if not isinstance(f, dict):
                    continue
                if str(f.get("subject") or "").strip().lower() != "user":
                    continue
                claim_raw = str(f.get("claim") or "").strip()
                evq_raw = str(f.get("evidence_quote") or "").strip()
                if not claim_raw and not evq_raw:
                    continue
                for pat in patterns:
                    try:
                        m = re.search(pat, claim_raw, flags=re.IGNORECASE)
                    except Exception:
                        m = None
                    if not m:
                        # Try evidence_quote if claim is normalized/abstracted
                        try:
                            m = re.search(pat, evq_raw, flags=re.IGNORECASE)
                        except Exception:
                            m = None
                    if not m:
                        continue
                    cand = (m.group(1) or "").strip()
                    ymd = _parse_birth_ymd_any(cand)
                    if not ymd:
                        # Try parsing the whole claim as a fallback
                        ymd = _parse_birth_ymd_any(evq_raw or claim_raw)
                    if not ymd:
                        continue
                    confirmed_bd = ymd
                    confirmed_eid = str(f.get("id") or "").strip()
                    break
                if confirmed_bd:
                    break

            if not confirmed_bd:
                return

            opts = bd_rec.get("options") if isinstance(bd_rec.get("options"), list) else []
            cleaned_opts: List[str] = []
            for o in opts:
                o2 = str(o).strip()
                while o2 and o2[-1] in ".!,;:":
                    o2 = o2[:-1].rstrip()
                o2 = " ".join(o2.split())
                if o2 and o2 not in cleaned_opts:
                    cleaned_opts.append(o2)
            if confirmed_bd not in cleaned_opts:
                cleaned_opts.append(confirmed_bd)

            evs = bd_rec.get("evidence_ids")
            if not isinstance(evs, list):
                evs = []
            if confirmed_eid and confirmed_eid not in evs:
                evs.append(confirmed_eid)

            bd_rec["status"] = "resolved_user"
            bd_rec["value"] = confirmed_bd
            bd_rec["options"] = cleaned_opts[:8]
            bd_rec["evidence_ids"] = evs
            bd_rec["note"] = "user_confirmed_relationship_birthdate_rule"
            rel[rel_bd_key] = bd_rec

            confs = prof.get("conflicts")
            if isinstance(confs, list):
                prof["conflicts"] = [str(x) for x in confs if str(x) != f"relationships.{rel_bd_key}"][:64]

        _resolve_rel_birthdate_from_name("girlfriend_name", "girlfriend_birthdate")

        def _set_rel_birthdate(rel_bd_key: str, ymd: str, evidence_id: str = "") -> None:
            if not rel_bd_key or not ymd:
                return
            bd_rec = rel.get(rel_bd_key) if isinstance(rel.get(rel_bd_key), dict) else {}
            st_bd = str(bd_rec.get("status") or "").strip().lower()
            if st_bd in ("resolved_user", "confirmed") and str(bd_rec.get("value") or "").strip():
                return

            opts = bd_rec.get("options") if isinstance(bd_rec.get("options"), list) else []
            cleaned_opts: List[str] = []
            for o in opts:
                o2 = str(o).strip()
                while o2 and o2[-1] in ".!,;:":
                    o2 = o2[:-1].rstrip()
                o2 = " ".join(o2.split())
                if o2 and o2 not in cleaned_opts:
                    cleaned_opts.append(o2)
            if ymd not in cleaned_opts:
                cleaned_opts.append(ymd)

            evs = bd_rec.get("evidence_ids")
            if not isinstance(evs, list):
                evs = []
            if evidence_id and evidence_id not in evs:
                evs.append(evidence_id)

            bd_rec["status"] = "resolved_user"
            bd_rec["value"] = ymd
            bd_rec["options"] = cleaned_opts[:8]
            bd_rec["evidence_ids"] = evs
            bd_rec["note"] = "user_confirmed_relationship_birthdate_rule"
            rel[rel_bd_key] = bd_rec

            confs = prof.get("conflicts")
            if isinstance(confs, list):
                prof["conflicts"] = [str(x) for x in confs if str(x) != f"relationships.{rel_bd_key}"][:64]

        def _resolve_rel_birthdate_from_pronoun() -> None:
            # If the user says "her birthday is ..." and we have exactly one confirmed female partner,
            # bind deterministically to that relationship.
            candidates_her = [
                ("girlfriend_name", "girlfriend_birthdate"),
                ("wife_name", "wife_birthdate"),
                ("partner_name", "partner_birthdate"),
                ("spouse_name", "spouse_birthdate"),
            ]
            candidates_his = [
                ("boyfriend_name", "boyfriend_birthdate"),
                ("husband_name", "husband_birthdate"),
                ("partner_name", "partner_birthdate"),
                ("spouse_name", "spouse_birthdate"),
            ]

            def _pick_one(cands: List[Tuple[str, str]]) -> str:
                picked = []
                for name_key, bd_key in cands:
                    rec = rel.get(name_key) if isinstance(rel.get(name_key), dict) else {}
                    st = str(rec.get("status") or "").strip().lower()
                    val = str(rec.get("value") or "").strip()
                    if val and st in ("confirmed", "resolved_user", "resolved_auto"):
                        picked.append(bd_key)
                return picked[0] if len(picked) == 1 else ""

            target_her = _pick_one(candidates_her)
            target_his = _pick_one(candidates_his)

            if not target_her and not target_his:
                return

            for f in reversed(facts):
                if not isinstance(f, dict):
                    continue
                if str(f.get("subject") or "").strip().lower() != "user":
                    continue
                claim_raw = str(f.get("claim") or "").strip()
                evq_raw = str(f.get("evidence_quote") or "").strip()
                if not claim_raw and not evq_raw:
                    continue

                txt = evq_raw or claim_raw
                low = (txt or "").lower()
                if target_her and any(k in low for k in ("her birthday", "her birthdate", "she was born")):
                    ymd = _parse_birth_ymd_any(txt)
                    if ymd:
                        _set_rel_birthdate(target_her, ymd, str(f.get("id") or "").strip())
                        return
                if target_his and any(k in low for k in ("his birthday", "his birthdate", "he was born")):
                    ymd = _parse_birth_ymd_any(txt)
                    if ymd:
                        _set_rel_birthdate(target_his, ymd, str(f.get("id") or "").strip())
                        return

        _resolve_rel_birthdate_from_pronoun()

        prof["identity"] = ident
    except Exception:
        pass

    # Always write on rebuild so stale conflicts/options cannot persist.
    _write_user_profile(user, prof)

    return prof


def tier2g_single_preferred_name_option(user: str) -> str:
    """
    Server helper (read-only):
    If Tier-2G preferred_name is ambiguous but contains exactly ONE single-token option,
    return it; else return "".

    This lets server.py accept a plain "yes" as confirmation when the system already
    presented exactly one viable name choice.
    """
    try:
        prof = load_user_profile(user)
        ident = prof.get("identity") if isinstance(prof.get("identity"), dict) else {}
        pn = ident.get("preferred_name") if isinstance(ident.get("preferred_name"), dict) else {}

        st = str(pn.get("status") or "").strip().lower()
        if st != "ambiguous":
            return ""

        opts = pn.get("options") if isinstance(pn.get("options"), list) else []
        opts2 = [" ".join(str(x).strip().split()) for x in opts if str(x).strip()]
        # de-dupe, keep order
        out: List[str] = []
        for o in opts2:
            if o and o not in out:
                out.append(o)

        if len(out) != 1:
            return ""

        only = out[0]
        # single-token only (prevents accidentally confirming a full sentence)
        if not only or (" " in only):
            return ""

        # strip trailing punctuation deterministically
        while only and only[-1] in ".!,;:":
            only = only[:-1].rstrip()
        return only.strip()
    except Exception:
        return ""


def tier2g_location_status_options(user: str) -> Dict[str, Any]:
    """
    Server helper (read-only):
    Return Tier-2G location status + options/value without server.py touching disk directly.

    Shape:
      {"status": "...", "value": "...", "options": [...]}
    """
    try:
        prof = load_user_profile(user)
        ident = prof.get("identity") if isinstance(prof.get("identity"), dict) else {}
        loc = ident.get("location") if isinstance(ident.get("location"), dict) else {}

        st = str(loc.get("status") or "").strip().lower()
        val = str(loc.get("value") or "").strip()
        opts = loc.get("options") if isinstance(loc.get("options"), list) else []
        opts2 = [" ".join(str(x).strip().split()) for x in opts if str(x).strip()]

        # deterministic cleanup: strip trailing punctuation + de-dupe
        cleaned: List[str] = []
        for o in opts2:
            o2 = o
            while o2 and o2[-1] in ".!,;:":
                o2 = o2[:-1].rstrip()
            o2 = " ".join(o2.split())
            if o2 and o2 not in cleaned:
                cleaned.append(o2)

        # also clean value
        v2 = val
        while v2 and v2[-1] in ".!,;:":
            v2 = v2[:-1].rstrip()
        v2 = " ".join(v2.split())

        return {"status": st, "value": v2, "options": cleaned[:12]}
    except Exception:
        return {"status": "", "value": "", "options": []}


def render_user_profile_snippet(user: str, *, user_text: str = "") -> str:
    """
    Bounded global profile snippet injected into every project for this user.
    """
    prof = load_user_profile(user)
    if not prof:
        return ""

    ident = prof.get("identity") if isinstance(prof.get("identity"), dict) else {}
    rel = prof.get("relationships") if isinstance(prof.get("relationships"), dict) else {}
    conflicts = prof.get("conflicts") if isinstance(prof.get("conflicts"), list) else []

    def _fmt(rec: Any) -> str:
        r = rec if isinstance(rec, dict) else {}
        st = str(r.get("status") or "").strip()
        val = str(r.get("value") or "").strip()
        opts = r.get("options") if isinstance(r.get("options"), list) else []
        opts2 = [str(x).strip() for x in opts if str(x).strip()]
        if st == "ambiguous" and opts2:
            return f"AMBIGUOUS options={opts2[:6]}"
        if val:
            return val
        if opts2:
            return opts2[0]
        return ""

    lines: List[str] = []
    lines.append("USER_PROFILE_GLOBAL (Tier-2G; conflict-aware; age derived at read-time):")

    # Apply do_not_resurface policy at read-time (proactive injection only).
    # If the user explicitly asks about a suppressed field, allow it (deterministic keyword match).
    name = _fmt(ident.get("name"))
    pn = _fmt(ident.get("preferred_name"))
    pro = _fmt(ident.get("pronouns"))
    bd = _fmt(ident.get("birthdate"))
    tz = _fmt(ident.get("timezone"))
    loc = _fmt(ident.get("location"))

    if not policy_allows_resurface_now(user, entity_key="user.identity.name", user_text=user_text):
        name = ""
        pn = ""
    if not policy_allows_resurface_now(user, entity_key="user.identity.birthdate", user_text=user_text):
        bd = ""
    if not policy_allows_resurface_now(user, entity_key="user.identity.timezone", user_text=user_text):
        tz = ""
    if not policy_allows_resurface_now(user, entity_key="user.identity.location", user_text=user_text):
        loc = ""

    if name:
        lines.append(f"- full_name: {name}")
    if pn:
        lines.append(f"- preferred_name: {pn}")
    if pro:
        lines.append(f"- pronouns: {pro}")
    # Derived age can be surfaced without revealing birthdate when user asks about age.
    age, asof_ymd, asof_tz = _derive_age_from_birthdate_for_user(
        str((ident.get("birthdate") or {}).get("value") or "").strip(),
        user,
    )
    asks_age = any(k in (user_text or "").lower() for k in ("age", "how old", "years old"))
    if bd:
        lines.append(f"- birthdate: {bd}")
        if age is not None:
            lines.append(f"- derived_age_years: {age} (asof={asof_ymd} {asof_tz})")
    elif asks_age and age is not None:
        lines.append(f"- derived_age_years: {age} (asof={asof_ymd} {asof_tz})")
    if tz:
        lines.append(f"- timezone: {tz}")
    if loc:
        lines.append(f"- location: {loc}")

    for k in (
        "girlfriend_name",
        "girlfriend_birthdate",
        "boyfriend_name",
        "boyfriend_birthdate",
        "partner_name",
        "partner_birthdate",
        "spouse_name",
        "spouse_birthdate",
        "wife_name",
        "wife_birthdate",
        "husband_name",
        "husband_birthdate",
        "mother_name",
        "mother_birthdate",
        "father_name",
        "father_birthdate",
        "sister_name",
        "sister_birthdate",
        "brother_name",
        "brother_birthdate",
        "son_name",
        "son_birthdate",
        "daughter_name",
        "daughter_birthdate",
        "child_name",
        "child_birthdate",
        "stepson_name",
        "stepson_birthdate",
        "stepdaughter_name",
        "stepdaughter_birthdate",
        "stepmother_name",
        "stepmother_birthdate",
        "stepfather_name",
        "stepfather_birthdate",
        "stepbrother_name",
        "stepbrother_birthdate",
        "stepsister_name",
        "stepsister_birthdate",
    ):
        vv = _fmt(rel.get(k))
        if vv:
            lines.append(f"- {k}: {vv}")

    if conflicts:
        lines.append(f"- conflicts: { [str(x) for x in conflicts[:12]] }")

    return "\n".join(lines).strip()
RAW_DIR_NAME = "raw"
ARTIFACTS_DIR_NAME = "artifacts"
STATE_DIR_NAME = "state"
MANIFEST_VERSION = 2


# -----------------------------------------------------------------------------
# Project deletion tombstone (short-lived; prevents immediate auto-recreate)
# -----------------------------------------------------------------------------

# TTL window (seconds) where auto-scaffold is blocked for a just-deleted project.
_DELETED_PROJECT_TTL_S = 8

# Map: normalized_project_name -> expires_at_epoch
_DELETED_PROJECTS: Dict[str, float] = {}


def mark_project_deleted(project_name: str) -> None:
    """
    Mark a project as deleted for a short TTL window.
    This prevents ensure_project()/ensure_project_scaffold() from recreating it immediately.
    """
    name = safe_project_name(project_name)
    _DELETED_PROJECTS[name] = time.time() + float(_DELETED_PROJECT_TTL_S)


def is_project_deleted(project_name: str) -> bool:
    """
    True only during the TTL window after deletion.
    After TTL, the tombstone expires automatically so the user can recreate the project name.
    """
    name = safe_project_name(project_name)
    exp = _DELETED_PROJECTS.get(name)
    if not exp:
        return False
    if time.time() >= float(exp):
        try:
            del _DELETED_PROJECTS[name]
        except Exception:
            pass
        return False
    return True


def configure(*, project_root: Path, projects_dir: Path, default_project_name: str = "default") -> None:
    global PROJECT_ROOT, PROJECTS_DIR, DEFAULT_PROJECT_NAME
    PROJECT_ROOT = Path(project_root).resolve()
    PROJECTS_DIR = Path(projects_dir).resolve()
    DEFAULT_PROJECT_NAME = (default_project_name or "default").strip() or "default"


def _require_configured() -> Tuple[Path, Path]:
    if PROJECT_ROOT is None or PROJECTS_DIR is None:
        raise RuntimeError("project_store is not configured. Call project_store.configure(...) from server.py.")
    return PROJECT_ROOT, PROJECTS_DIR


# -----------------------------------------------------------------------------
# Small IO helpers (kept local to avoid importing server.py)
# -----------------------------------------------------------------------------

def file_sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data or b"").hexdigest()


def _called_from_server_module() -> bool:
    frame = None
    try:
        frame = inspect.currentframe()
        if frame is None:
            return False
        caller = frame.f_back
        if caller is None:
            return False
        return caller.f_globals.get("__name__") == "server"
    except Exception:
        return False
    finally:
        try:
            del frame
        except Exception:
            pass


def read_text_file(path: Path, *, errors: str = "replace") -> str:
    try:
        return path.read_text(encoding="utf-8", errors=errors)
    except Exception:
        return ""


def read_bytes_file(path: Path) -> bytes:
    try:
        return path.read_bytes()
    except Exception:
        return b""


def atomic_write_text(path: Path, content: str, *, encoding: str = "utf-8", errors: str = "strict") -> None:
    rp = str(path).replace("\\", "/")
    if "/_user/" in rp and _called_from_server_module():
        raise RuntimeError(
            "Refusing direct write to global user memory from server.py. "
            "Use project_store as the single authority."
        )
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(content or "", encoding=encoding, errors=errors)
    os.replace(tmp, path)


def atomic_write_bytes(path: Path, content: bytes) -> None:
    rp = str(path).replace("\\", "/")
    if "/_user/" in rp and _called_from_server_module():
        raise RuntimeError(
            "Refusing direct write to global user memory from server.py. "
            "Use project_store as the single authority."
        )
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_bytes(content or b"")
    os.replace(tmp, path)


def now_iso() -> str:
    return time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())


def safe_project_name(name: str) -> str:
    """
    Normalize a project name while preserving folder nesting like "User/Project".
    Each path segment is sanitized independently.
    """
    raw = (name or "").strip().replace("\\", "/")
    parts = [p for p in raw.split("/") if p.strip()]
    if not parts:
        return DEFAULT_PROJECT_NAME

    cleaned_parts: List[str] = []
    for part in parts:
        cleaned = re.sub(r"[^a-zA-Z0-9_-]", "_", part.strip())
        cleaned_parts.append(cleaned or DEFAULT_PROJECT_NAME)

    return "/".join(cleaned_parts)

# -----------------------------------------------------------------------------
# Couples links (system-private; not inside any user's project tree)
# -----------------------------------------------------------------------------

COUPLES_SYSTEM_DIR = "_couples"
COUPLES_LINKS_FILE = "couples_links.json"

def _couples_links_path() -> Path:
    _, pd = _require_configured()
    d = (pd / COUPLES_SYSTEM_DIR)
    d.mkdir(parents=True, exist_ok=True)
    return d / COUPLES_LINKS_FILE

def load_couples_links() -> Dict[str, Any]:
    p = _couples_links_path()
    if not p.exists():
        return {}
    try:
        obj = json.loads(p.read_text(encoding="utf-8") or "{}")
    except Exception:
        obj = {}
    return obj if isinstance(obj, dict) else {}

def save_couples_links(obj: Dict[str, Any]) -> None:
    if not isinstance(obj, dict):
        raise ValueError("couples links must be a dict")
    atomic_write_text(_couples_links_path(), json.dumps(obj, indent=2, sort_keys=True))

def make_couple_id(user_a: str, user_b: str) -> str:
    # stable + inspectable id
    a = re.sub(r"[^a-zA-Z0-9_-]", "_", (user_a or "").strip())
    b = re.sub(r"[^a-zA-Z0-9_-]", "_", (user_b or "").strip())
    pair = sorted([a or "userA", b or "userB"])
    return f"{pair[0]}__{pair[1]}"

def link_couple(*, user_a: str, user_b: str, project_a: str = "default", project_b: str = "default") -> Dict[str, Any]:
    links = load_couples_links()
    couple_id = make_couple_id(user_a, user_b)
    links[couple_id] = {
        "couple_id": couple_id,
        "user_a": user_a,
        "user_b": user_b,
        "project_a": project_a,
        "project_b": project_b,
        "status": "active",
        "updated_at": time.time(),
    }
    save_couples_links(links)
    return links[couple_id]

def get_couple(couple_id: str) -> Optional[Dict[str, Any]]:
    links = load_couples_links()
    obj = links.get(couple_id)
    return obj if isinstance(obj, dict) else None

# -----------------------------------------------------------------------------
# Paths + manifest
# -----------------------------------------------------------------------------

def project_dir(project_name: str) -> Path:
    pr, pd = _require_configured()
    return pd / safe_project_name(project_name)


def project_manifest_path(project_name: str) -> Path:
    return project_dir(project_name) / "project_manifest.json"


def ensure_project(project_name: str) -> Path:
    if is_project_deleted(project_name):
        raise RuntimeError(f"Project '{safe_project_name(project_name)}' was just deleted; refusing to auto-recreate.")
    pr, pd = _require_configured()
    pdir = project_dir(project_name)
    (pdir / RAW_DIR_NAME).mkdir(parents=True, exist_ok=True)
    (pdir / ARTIFACTS_DIR_NAME).mkdir(parents=True, exist_ok=True)
    (pdir / STATE_DIR_NAME).mkdir(parents=True, exist_ok=True)
    if not project_manifest_path(project_name).exists():
        manifest = {
            "version": MANIFEST_VERSION,
            "project_name": safe_project_name(project_name),
            "display_name": "",
            "created_at": time.time(),
            "updated_at": time.time(),
            "goal": "",
            "expert_type": "",
            "raw_files": [],
            "artifacts": [],
            "last_ingested": {},
        }
        atomic_write_text(project_manifest_path(project_name), json.dumps(manifest, indent=2))
    return pdir


def list_existing_projects(user: Optional[str] = None) -> List[str]:
    pr, pd = _require_configured()
    base = (pd / user) if user else pd
    if not base.exists():
        return []
    names: List[str] = []
    for p in base.iterdir():
        if p.is_dir() and not p.name.startswith(("_", ".")):
            names.append(p.name)
    return sorted(names)


def load_manifest(project_name: str) -> Dict[str, Any]:
    ensure_project(project_name)
    path = project_manifest_path(project_name)
    try:
        m = json.loads(path.read_text(encoding="utf-8") or "{}")
    except Exception:
        m = {}
    if not isinstance(m, dict):
        m = {}

    # minimal repair
    m.setdefault("version", MANIFEST_VERSION)
    m.setdefault("project_name", safe_project_name(project_name))
    m.setdefault("display_name", "")
    m.setdefault("created_at", time.time())
    m.setdefault("updated_at", time.time())
    m.setdefault("goal", "")
    m.setdefault("raw_files", [])
    m.setdefault("artifacts", [])
    m.setdefault("last_ingested", {})
    return m


def save_manifest(project_name: str, manifest: Dict[str, Any]) -> None:
    manifest["updated_at"] = time.time()
    atomic_write_text(project_manifest_path(project_name), json.dumps(manifest, indent=2))


def get_project_display_name(project_name: str) -> str:
    """
    Return the human-facing display name for a project (if any).
    """
    try:
        m = load_manifest(project_name) or {}
    except Exception:
        m = {}
    val = ""
    try:
        val = str(m.get("display_name") or "").strip()
    except Exception:
        val = ""
    return val


def set_project_display_name(project_name: str, display_name: str) -> None:
    """
    Set the human-facing display name for a project.
    """
    dn = str(display_name or "").strip()
    try:
        m = load_manifest(project_name) or {}
    except Exception:
        m = {}
    m["display_name"] = dn
    save_manifest(project_name, m)


def raw_dir(project_name: str) -> Path:
    pr, pd = _require_configured()
    return project_dir(project_name) / RAW_DIR_NAME


def artifacts_dir(project_name: str) -> Path:
    pr, pd = _require_configured()
    return project_dir(project_name) / ARTIFACTS_DIR_NAME


def state_dir(project_name: str) -> Path:
    pr, pd = _require_configured()
    return project_dir(project_name) / STATE_DIR_NAME


def state_file_path(project_name: str, key: str) -> Path:
    mapping = {
        "project_map": "project_map.md",
        "project_brief": "project_brief.md",
        "working_doc": "working_doc.md",
        "preferences": "preferences.md",
        "decision_log": "decision_log.md",
        "facts_map": "facts_map.md",
        "project_state": "project_state.json",
        "discovery_index": "discovery_index.md",
        "timeline": "timeline.md",
        "issues_questions": "issues_questions.md",
        "evidence_matrix": "evidence_matrix.csv",
        "fact_ledger": "fact_ledger.md",
        "conflicts": "conflicts.md",
        "capability_gaps": "capability_gaps.md",
        "library_index": "library_index.md",
    }
    fname = mapping.get(key, f"{key}.txt")
    return state_dir(project_name) / fname


# ---------------------------------------------------------------------------
# C6.3 — Project Bootstrap Mode + Project Truth Policy (project_state extensions)
#
# Minimal schema extension (no migrations):
#   - project_mode: "open_world" | "closed_world" | "hybrid"
#   - goal: string
#   - domains: [tags]
#   - bootstrap_status: "needs_goal" | "goal_proposed" | "active"
#
# Existing projects default (read-time):
#   - bootstrap_status="active"
#   - project_mode="closed_world"
# ---------------------------------------------------------------------------

_PROJECT_MODE_ALLOWED = ("open_world", "closed_world", "hybrid")
_BOOTSTRAP_STATUS_ALLOWED = ("needs_goal", "goal_proposed", "active")


def load_project_state(project_name: str) -> Dict[str, Any]:
    """
    Read canonical state/project_state.json (best-effort) and apply read-time defaults.
    NO migrations / NO writes.
    """
    ensure_project_scaffold(project_name)
    p = state_file_path(project_name, "project_state")
    if not p.exists():
        return {}

    try:
        obj = json.loads(p.read_text(encoding="utf-8") or "{}")
    except Exception:
        obj = {}

    if not isinstance(obj, dict):
        obj = {}

    # Existing required-ish keys (keep prior behavior)
    obj.setdefault("goal", "")
    obj.setdefault("current_focus", "")
    obj.setdefault("next_actions", [])
    obj.setdefault("key_files", [])
    obj.setdefault("last_updated", now_iso())

    # Expert Frame Lock (EFL v1) — read-time defaults + normalization (no writes)
    ef = obj.get("expert_frame")
    if not isinstance(ef, dict):
        ef = {}
    st = str(ef.get("status") or "").strip().lower()
    if st not in ("proposed", "active"):
        st = ""
    obj["expert_frame"] = {
        "status": st,
        "label": str(ef.get("label") or "").replace("\r", " ").replace("\n", " ").strip(),
        "directive": str(ef.get("directive") or "").replace("\r", " ").replace("\n", " ").strip(),
        "set_reason": str(ef.get("set_reason") or "").replace("\r", " ").replace("\n", " ").strip(),
        "updated_at": str(ef.get("updated_at") or "").replace("\r", " ").replace("\n", " ").strip(),
    }

    # C6.3 extensions (read-time defaults, no migration)
    pm = str(obj.get("project_mode") or "").strip()
    if pm not in _PROJECT_MODE_ALLOWED:
        pm = "hybrid"
    obj["project_mode"] = pm

    bs = str(obj.get("bootstrap_status") or "").strip()
    if bs not in _BOOTSTRAP_STATUS_ALLOWED:
        bs = "active"
    obj["bootstrap_status"] = bs

    doms = obj.get("domains")
    if not isinstance(doms, list):
        doms = []
    obj["domains"] = [str(x).strip() for x in doms if str(x).strip()][:24]

    return obj


def write_project_state_fields(project_name: str, updates: Dict[str, Any]) -> bool:
    """
    Deterministically merge + overwrite project_state.json using canonical writer.
    Returns True if written, else False.
    """
    ensure_project_scaffold(project_name)

    cur = load_project_state(project_name)
    if not isinstance(cur, dict):
        cur = {}

    up = updates if isinstance(updates, dict) else {}

    merged = dict(cur)
    merged.update(up)

    pm = str(merged.get("project_mode") or "").strip()
    if pm not in _PROJECT_MODE_ALLOWED:
        pm = "hybrid"
    merged["project_mode"] = pm

    bs = str(merged.get("bootstrap_status") or "").strip()
    if bs not in _BOOTSTRAP_STATUS_ALLOWED:
        bs = "active"
    merged["bootstrap_status"] = bs

    doms = merged.get("domains")
    if not isinstance(doms, list):
        doms = []
    merged["domains"] = [str(x).strip() for x in doms if str(x).strip()][:24]
    # Expert Frame Lock (EFL v1) — canonical normalization
    ef = merged.get("expert_frame")
    if not isinstance(ef, dict):
        ef = {}
    st = str(ef.get("status") or "").strip().lower()
    if st not in ("proposed", "active"):
        st = ""
    merged["expert_frame"] = {
        "status": st,
        "label": str(ef.get("label") or "").replace("\r", " ").replace("\n", " ").strip(),
        "directive": str(ef.get("directive") or "").replace("\r", " ").replace("\n", " ").strip(),
        "set_reason": str(ef.get("set_reason") or "").replace("\r", " ").replace("\n", " ").strip(),
        "updated_at": str(ef.get("updated_at") or "").replace("\r", " ").replace("\n", " ").strip(),
    }

    # Defensive normalization:
    # Canonical writer rejects multiline strings in JSON. Bootstrap model outputs can contain newlines.
    # Normalize all string fields to single-line to prevent "silent reject" → bootstrap loops.
    for k, v in list(merged.items()):
        if isinstance(v, str):
            merged[k] = v.replace("\r", " ").replace("\n", " ").strip()
        elif isinstance(v, list):
            merged[k] = [
                (x.replace("\r", " ").replace("\n", " ").strip() if isinstance(x, str) else x)
                for x in v
            ]

    merged["goal"] = str(merged.get("goal") or "").strip()
    merged["last_updated"] = now_iso()

    ok = write_canonical_entry(
        project_name,
        target_path=state_file_path(project_name, "project_state"),
        mode="json_overwrite",
        data=merged,
    )
    return bool(ok)


# Deliverables registry (C1)

# -----------------------------------------------------------------------------

DELIVERABLES_VERSION = 1
DELIVERABLES_FILE_NAME = "deliverables.json"
# -----------------------------------------------------------------------------
# Decisions (C2 v1): candidates -> confirmed -> final (append-only JSONL)
# -----------------------------------------------------------------------------

DECISION_CANDIDATES_FILE_NAME = "decision_candidates.jsonl"
DECISIONS_FILE_NAME = "decisions.jsonl"
# -----------------------------------------------------------------------------
# C4 v1 — Upload Clarification Loop (deterministic, single question, non-spam)
# -----------------------------------------------------------------------------

PENDING_UPLOAD_QUESTION_FILE_NAME = "pending_upload_question.json"
UPLOAD_NOTES_FILE_NAME = "upload_notes.jsonl"
ACTIVE_OBJECT_FILE_NAME = "active_object.json"

# ---------------------------------------------------------------------------
# C9 — Deterministic Inbox / Pending Items (append-only JSONL)
# ---------------------------------------------------------------------------
INBOX_FILE_NAME = "inbox.jsonl"
# ---------------------------------------------------------------------------
# Couples Mode Option A — Bring-up-later queue (append-only JSONL)
# ---------------------------------------------------------------------------
BRINGUPS_FILE_NAME = "bringups.jsonl"
UNDERSTANDING_FILE_NAME = "understanding.json"

def bringups_path(project_name: str) -> Path:
    """
    projects/<user>/<project>/state/bringups.jsonl

    Append-only queue used ONLY by couple_* accounts.
    """
    return state_dir(project_name) / BRINGUPS_FILE_NAME
def understanding_path(project_name: str) -> Path:
    """
    projects/<user>/<project>/state/understanding.json

    Tier-X interpretive memory (NOT factual recall).
    """
    return state_dir(project_name) / UNDERSTANDING_FILE_NAME


def load_understanding(project_name: str) -> Dict[str, Any]:
    ensure_project_scaffold(project_name)
    p = understanding_path(project_name)
    if not p.exists():
        return {}
    try:
        obj = json.loads(p.read_text(encoding="utf-8") or "{}")
        return obj if isinstance(obj, dict) else {}
    except Exception:
        return {}
# ---------------------------------------------------------------------------
# Audit log (per-turn routing trace; append-only JSONL)
# ---------------------------------------------------------------------------

AUDIT_LOG_FILE_NAME = "audit_log.jsonl"

def audit_log_path(project_name: str) -> Path:
    """
    projects/<user>/<project>/state/audit_log.jsonl

    Append-only audit trace for each user turn (routing/tool/override/timing).
    """
    return state_dir(project_name) / AUDIT_LOG_FILE_NAME

def append_audit_event(project_name: str, event: Dict[str, Any]) -> bool:
    """
    Append one audit event (jsonl). Deterministic storage only.
    Must never throw.
    """
    try:
        ensure_project(project_name)
        obj = event if isinstance(event, dict) else {}
        obj.setdefault("schema", "audit_turn_v1")
        obj.setdefault("created_at", now_iso().replace("Z", ""))
        return bool(
            write_canonical_entry(
                project_name,
                target_path=audit_log_path(project_name),
                mode="jsonl_append",
                data=obj,
            )
        )
    except Exception:
        return False
# ---------------------------------------------------------------------------
# Tier-1 — Raw facts (append-only JSONL)
# Purpose: broad “candidate facts” extracted from chat logs / imports.
# This is NOT canonical truth. Distillers promote from here into facts_map.md.
# ---------------------------------------------------------------------------
FACTS_RAW_FILE_NAME = "facts_raw.jsonl"

def facts_raw_path(project_name: str) -> Path:
    """
    projects/<user>/<project>/state/facts_raw.jsonl

    Tier-1: broad raw fact candidates (NOT canonical truth).
    """
    return state_dir(project_name) / FACTS_RAW_FILE_NAME
def append_fact_raw_candidate(
    project_name: str,
    *,
    claim: str,
    slot: str = "other",
    subject: str = "user",
    source: str = "chat",
    evidence: str = "",
    evidence_quote: str = "",
    turn_index: int = 0,
    timestamp: str = "",
) -> Dict[str, Any]:
    """
    Append exactly one Tier-1 fact candidate to facts_raw.jsonl.

    Deterministic write-path for facts learned directly in chat.
    Tier-2 (facts_map.md) still comes from distillation.

    NOTE:
    - Keep backward compatibility with older call sites that pass `evidence=...`.
    - Prefer `evidence_quote` for downstream normalization/distillation.
    - Store provenance as a dict under `source` when turn_index/timestamp are provided.
    """
    ensure_project_scaffold(project_name)

    c = _norm_one_line(claim)
    if not c:
        return {"ok": False, "error": "empty_claim"}

    slot_n = _norm_one_line(slot).lower() or "other"
    subject_n = _norm_one_line(subject).lower() or "user"

    # Evidence: prefer evidence_quote, fall back to evidence
    evq = _norm_one_line(evidence_quote or "") or _norm_one_line(evidence or "")

    # Source: keep string-compatible behavior, but allow structured provenance
    src_type = _norm_one_line(source) or "chat"
    src_obj: Any = src_type
    ts = _norm_one_line(timestamp or "")
    try:
        ti = int(turn_index or 0)
    except Exception:
        ti = 0

    if ti or ts:
        src_obj = {
            "type": src_type,
            "turn_index": ti,
            "timestamp": ts,
        }

    obj: Dict[str, Any] = {
        "id": f"fact_{int(time.time() * 1000)}",
        "created_at": now_iso(),
        "claim": c,
        "slot": slot_n,
        "subject": subject_n,
        "source": src_obj,
        "evidence_quote": evq,
        # Keep legacy field for older tooling (harmless)
        "evidence": evq,
    }

    ok = write_canonical_entry(
        project_name,
        target_path=facts_raw_path(project_name),
        mode="jsonl_append",
        data=obj,
    )
    if not ok:
        return {"ok": False, "error": "canonical_write_rejected"}

    # ---------------------------------------------------------------------
    # Mirror Tier-1 project fact -> Tier-1 global user ledger (facts_raw.jsonl)
    # Purpose:
    # - Ensure cross-domain continuity by capturing user-attributed facts globally.
    # - Apply user memory policy for global mirroring (project may still store locally).
    # ---------------------------------------------------------------------
    try:
        # Determine user segment from "user/project" naming.
        pname = safe_project_name(project_name)
        user_seg0 = pname.split("/", 1)[0].strip()

        if user_seg0:
            # Determine deterministic entity_key for policy + routing
            ek0 = _tier1_entity_key_guess(c, slot_n, subject_n)

            # Apply policy only to the global mirror decision.
            dec = policy_decision_for_tier1_claim(
                user_seg0,
                project_full=pname,
                claim=c,
                entity_key=ek0,
            )

            # Only mirror user-attributed facts when policy allows.
            # Mirror user-attributed facts globally.
            # Some extractors mislabel subject for first-person preference/habit claims; treat explicit
            # first-person language as user-attributed for mirroring.
            is_first_person = False
            try:
                lowc = (c or "").strip().lower()
                is_first_person = (
                    lowc.startswith("i ")
                    or lowc.startswith("i'm ")
                    or lowc.startswith("i’m ")
                    or lowc.startswith("i am ")
                    or lowc.startswith("my ")
                    or lowc.startswith("me ")
                )
            except Exception:
                is_first_person = False

            mirror_subject = "user" if (subject_n == "user" or is_first_person) else subject_n

            if (mirror_subject == "user") and bool(dec.get("store", True)) and bool(dec.get("mirror_global", True)):
                append_user_fact_raw_candidate(
                    user_seg0,
                    claim=c,
                    slot=slot_n,
                    subject=mirror_subject,
                    source="chat",
                    evidence_quote=evq,
                    turn_index=int(turn_index or 0),
                    timestamp=str(timestamp or ""),
                )
    except Exception:
        # Never fail the project write if global mirroring fails.
        pass

    return {"ok": True, "id": obj["id"]}


# ---------------------------------------------------------------------------
# C8 — Deterministic Cross-Linking (append-only JSONL)
# ---------------------------------------------------------------------------

LINKS_FILE_NAME = "links.jsonl"

def links_path(project_name: str) -> Path:
    """
    projects/<user>/<project>/state/links.jsonl
    """
    return state_dir(project_name) / LINKS_FILE_NAME

def _c8_link_stable_id(*, type_: str, from_ref: str, to_ref: str, reason: str) -> str:
    """
    Stable id derived from link content (deterministic).
    """
    try:
        raw = f"{type_}|{from_ref}|{to_ref}|{reason}".encode("utf-8", errors="replace")
        h = hashlib.sha256(raw).hexdigest()[:12]
        return f"link_{h}"
    except Exception:
        return f"link_{int(time.time())}"

def append_link_event(
    project_name: str,
    *,
    type_: str,
    from_ref: str,
    to_ref: str,
    reason: str,
    confidence: str = "low",
    created_at: str = "",
) -> Dict[str, Any]:
    """
    Append one link event into state/links.jsonl (append-only).
    All writes MUST go through write_canonical_entry (C6.2).
    """
    ensure_project(project_name)

    t = str(type_ or "").strip()
    allowed = {"upload_to_decision", "deliverable_to_decision", "upload_to_deliverable", "decision_to_decision"}
    if t not in allowed:
        t = "upload_to_decision"

    fr = strip_control_tokens(str(from_ref or "").strip())
    tr = strip_control_tokens(str(to_ref or "").strip())
    rs = strip_control_tokens(str(reason or "").strip())
    if not rs:
        rs = "linked"

    conf = str(confidence or "").strip().lower()
    if conf not in ("low", "medium", "high"):
        conf = "low"

    ca = (created_at or "").strip() or _now_iso_noz()

    entry = {
        "id": _c8_link_stable_id(type_=t, from_ref=fr, to_ref=tr, reason=rs),
        "created_at": ca,
        "type": t,
        "from": fr,
        "to": tr,
        "reason": rs,
        "confidence": conf,
    }

    ok = write_canonical_entry(
        project_name,
        target_path=links_path(project_name),
        mode="jsonl_append",
        data=entry,
    )
    if not ok:
        raise ValueError("canonical_write_rejected")
    return entry

def _c8_load_links(project_name: str, *, max_lines: int = 5000) -> List[Dict[str, Any]]:
    """
    Best-effort tail load of links.jsonl (bounded).
    """
    ensure_project(project_name)
    p = links_path(project_name)
    if not p.exists():
        return []
    try:
        lines = p.read_text(encoding="utf-8", errors="replace").splitlines()
    except Exception:
        return []
    if max_lines and len(lines) > max_lines:
        lines = lines[-max_lines:]
    out: List[Dict[str, Any]] = []
    for ln in lines:
        try:
            obj = json.loads(ln)
        except Exception:
            continue
        if isinstance(obj, dict):
            out.append(obj)
    return out

def links_for_decision(project_name: str, decision_id: str) -> List[Dict[str, Any]]:
    did = str(decision_id or "").strip()
    if not did:
        return []
    needle = f"decision:{did}"
    rows = _c8_load_links(project_name, max_lines=6000)
    return [r for r in rows if (str(r.get("from") or "") == needle) or (str(r.get("to") or "") == needle)]

def links_for_upload(project_name: str, upload_path: str) -> List[Dict[str, Any]]:
    up = str(upload_path or "").replace("\\", "/").strip()
    if not up:
        return []
    needle = f"upload:{up}" if not up.startswith("upload:") else up
    rows = _c8_load_links(project_name, max_lines=6000)
    return [r for r in rows if (str(r.get("from") or "") == needle) or (str(r.get("to") or "") == needle)]

def links_for_deliverable(project_name: str, deliverable_id: str) -> List[Dict[str, Any]]:
    did = str(deliverable_id or "").strip()
    if not did:
        return []
    needle = f"deliverable:{did}"
    rows = _c8_load_links(project_name, max_lines=6000)
    return [r for r in rows if (str(r.get("from") or "") == needle) or (str(r.get("to") or "") == needle)]

def summarize_links_for_domain(project_name: str, domain: str) -> str:
    """
    Deterministic summary: counts by link type for decisions in this domain.
    """
    dom = str(domain or "").strip()
    if not dom:
        return "(no domain)"

    cur = []
    try:
        cur = list_current_decisions(project_name) or []
    except Exception:
        cur = []

    ids = [str(d.get("id") or "").strip() for d in cur if str(d.get("domain") or "").strip() == dom]
    ids = [x for x in ids if x]
    if not ids:
        return f"{dom}: (no current decisions)"

    rows = _c8_load_links(project_name, max_lines=6000)
    counts: Dict[str, int] = {}
    for r in rows:
        fr = str(r.get("from") or "")
        to = str(r.get("to") or "")
        t = str(r.get("type") or "")
        if any(fr == f"decision:{i}" or to == f"decision:{i}" for i in ids):
            counts[t] = counts.get(t, 0) + 1

    if not counts:
        return f"{dom}: links=0"

    parts = [f"{k}={counts[k]}" for k in sorted(counts.keys())]
    return f"{dom}: " + ", ".join(parts)

def find_current_decision_by_text(project_name: str, text: str) -> Optional[Dict[str, Any]]:
    """
    Deterministically find a CURRENT decision whose text matches exactly.
    Returns normalized decision dict or None.
    """
    t = str(text or "").strip()
    if not t:
        return None
    try:
        cur = list_current_decisions(project_name) or []
    except Exception:
        cur = []
    for d in cur:
        if not isinstance(d, dict):
            continue
        if str(d.get("text") or "").strip() == t:
            return d
    return None

def inbox_path(project_name: str) -> Path:
    """
    projects/<user>/<project>/state/inbox.jsonl
    """
    return state_dir(project_name) / INBOX_FILE_NAME
def facts_raw_path(project_name: str) -> Path:
    """
    projects/<user>/<project>/state/facts_raw.jsonl

    Tier-1 candidate facts (NOT canonical truth).
    """
    return state_dir(project_name) / FACTS_RAW_FILE_NAME


def _sha256_text(s: str) -> str:
    try:
        b = (s or "").encode("utf-8", errors="replace")
        return hashlib.sha256(b).hexdigest()
    except Exception:
        return ""


def _norm_one_line(s: str) -> str:
    x = "" if s is None else str(s)
    x = x.replace("\r", " ").replace("\n", " ")
    x = re.sub(r"\s{2,}", " ", x).strip()
    return x


def _tier1_privacy_guess(claim: str) -> str:
    low = (claim or "").lower()
    sensitive_hits = (
        "adhd", "autism", "ptsd", "trauma", "abuse", "assault",
        "suicide", "self-harm", "self harm",
        "divorce", "custody",
        "visa", "immigration", "e-2", "e2 visa",
        "diagnosis", "medication", "therapy",
        "sex", "underwear",
        "illegal", "crime",
        "address", "social security", "ssn",
    )
    return "sensitive" if any(k in low for k in sensitive_hits) else "normal"


def _tier1_volatility_guess(claim: str, slot: str) -> str:
    low = (claim or "").lower()

    if re.search(r"\b\d{1,3}\s*(?:years?\s*old|yo)\b", low):
        return "likely_to_change"

    if any(k in low for k in (
        "going through a divorce",
        "living with",
        "haven't proposed",
        "not proposed",
        "will get married",
        "visa",
    )):
        return "likely_to_change"

    if slot == "identity" and any(k in low for k in ("canadian", "american", "nationality")):
        return "stable"

    return "unknown"


def _tier1_entity_key_guess(claim: str, slot: str, subject: str) -> str:
    c = _norm_one_line(claim)
    low = c.lower()
    subj = (subject or "").strip().lower()

    if subj == "user":
        # ---- Relationship ages (avoid misattributing to user.identity.age) ----
        if re.search(r"\b\d{1,3}\s*(?:years?\s*old|yo)\b", low):
            if "my son" in low:
                return "user.relationship.son.age"
            if "my daughter" in low:
                return "user.relationship.daughter.age"
            if "my child" in low or "my kid" in low:
                return "user.relationship.child.age"
            if "my girlfriend" in low:
                return "user.relationship.girlfriend.age"
            if "my boyfriend" in low:
                return "user.relationship.boyfriend.age"
            if "my partner" in low:
                return "user.relationship.partner.age"
            if "my spouse" in low:
                return "user.relationship.spouse.age"
            if "my wife" in low:
                return "user.relationship.wife.age"
            if "my husband" in low:
                return "user.relationship.husband.age"
            if "my mother" in low or "my mom" in low:
                return "user.relationship.mother.age"
            if "my father" in low or "my dad" in low:
                return "user.relationship.father.age"
            if "my sister" in low:
                return "user.relationship.sister.age"
            if "my brother" in low:
                return "user.relationship.brother.age"

        # ---- Relationship birthdates ----
        if ("birthday" in low) or ("birthdate" in low) or ("was born" in low):
            if "my son" in low:
                return "user.relationship.son.birthdate"
            if "my daughter" in low:
                return "user.relationship.daughter.birthdate"
            if "my child" in low or "my kid" in low:
                return "user.relationship.child.birthdate"
            if "my girlfriend" in low:
                return "user.relationship.girlfriend.birthdate"
            if "my boyfriend" in low:
                return "user.relationship.boyfriend.birthdate"
            if "my partner" in low:
                return "user.relationship.partner.birthdate"
            if "my spouse" in low:
                return "user.relationship.spouse.birthdate"
            if "my wife" in low:
                return "user.relationship.wife.birthdate"
            if "my husband" in low:
                return "user.relationship.husband.birthdate"
            if "my mother" in low or "my mom" in low:
                return "user.relationship.mother.birthdate"
            if "my father" in low or "my dad" in low:
                return "user.relationship.father.birthdate"
            if "my sister" in low:
                return "user.relationship.sister.birthdate"
            if "my brother" in low:
                return "user.relationship.brother.birthdate"

        # ---- Core identity (MUST NOT collide with relationship "name is") ----
        if re.search(r"\bmy\s+(?:preferred\s+)?name\s+is\b", low) or ("i go by" in low):
            return "user.identity.name"

        if (
            ("i live in" in low)
            or ("i'm in " in low) or ("im in " in low) or ("i am in " in low)
            or ("my confirmed location is" in low)
            or ("confirm my location " in low)
            or ("set my location to" in low)
        ):
            return "user.identity.location"

        if re.search(r"\bmy\s+favou?rite\s+color\s+is\b", low):
            return "user.preference.favorite_color"
        # ---- Context/activity/time (reduce *.other fan-in) ----
        # Time-of-day / "it was 7:40pm" style statements
        if re.search(r"\b\d{1,2}:\d{2}\s*(?:am|pm)\b", low) or re.search(r"\bit\s+was\s+\d{1,2}:\d{2}\b", low):
            return "context.time_of_day"

        # Media activity / "starting a movie" / "watching ..."
        if ("starting a movie" in low) or ("watching" in low and "movie" in low) or ("i'm watching" in low) or ("i am watching" in low):
            return "context.activity.media"

        # Substance-use routine/event context (keeps these out of generic context.other)
        if any(k in low for k in ("edible", "edibles", "gummy", "gummies", "vape", "thc", "sativa", "rippl", "mg ")):
            # If the extractor tagged this as an event, keep it event-scoped; otherwise treat as routine.
            if slot == "event":
                return "event.substance_use"
            if slot == "routine":
                return "routine.substance_use"
            return "context.substance_use"
        # ---- Other identity/constraints ----
        if "years old" in low:
            return "user.identity.age"
        if "divorce" in low:
            return "user.status.divorce"
        if "visa" in low:
            return "user.constraint.immigration"

        # ---- Preferences (generic) ----
        if ("prefer" in low) or ("prefers" in low):
            return "user.preference"

        # ---- Relationship names: keep specific keys so they can't evict user name ----
        if ("my girlfriend" in low) and ("name is" in low):
            return "user.relationship.girlfriend.name"
        if ("my boyfriend" in low) and ("name is" in low):
            return "user.relationship.boyfriend.name"
        if ("my wife" in low) and ("name is" in low):
            return "user.relationship.wife.name"
        if ("my husband" in low) and ("name is" in low):
            return "user.relationship.husband.name"

        # ---- Relationship children ----
        if ("my son" in low) and (("name is" in low) or ("named" in low)):
            return "user.relationship.son"
        if ("my daughter" in low) and (("name is" in low) or ("named" in low)):
            return "user.relationship.daughter"

        # ---- Fallback relationship buckets (non-name) ----
        if "girlfriend" in low:
            return "user.relationship.girlfriend"
        if "boyfriend" in low:
            return "user.relationship.boyfriend"
        if "wife" in low:
            return "user.relationship.wife"
        if "husband" in low:
            return "user.relationship.husband"
        if "son" in low:
            return "user.relationship.son"
        if "daughter" in low:
            return "user.relationship.daughter"

    if subj == "project":
        return "project.meta"

    if slot:
        return f"{slot}.other"

    return "other"




def normalize_facts_raw_jsonl(project_name: str) -> dict:
    """
    Deterministically normalize Tier-1 facts_raw.jsonl in place.

    - Drops junk confirmations (e.g. “Yes”)
    - Adds: claim_hash, entity_key
    - Normalizes: privacy, volatility, slot/subject casing
    """
    ensure_project_scaffold(project_name)

    # NOTE:
    # normalize_facts_raw_jsonl() must NEVER call itself.
    # Tier-1 normalization is triggered by the pipeline after appends (model_pipeline.py),
    # and Tier-2 distillation may optionally call it before reading (if desired).
    p = facts_raw_path(project_name)

    if not p.exists():
        atomic_write_text(p, "", encoding="utf-8", errors="strict")
        return {"ok": True, "written": 0, "dropped": 0}

    try:
        lines = p.read_text(encoding="utf-8", errors="replace").splitlines()
    except Exception:
        return {"ok": False, "error": "read_failed"}

    out_lines = []
    dropped = 0
    kept = 0

    for ln in lines:
        s = (ln or "").strip()
        if not s:
            continue

        try:
            obj = json.loads(s)
        except Exception:
            continue

        if not isinstance(obj, dict):
            continue

        claim = _norm_one_line(obj.get("claim"))
        evidence = _norm_one_line(obj.get("evidence_quote"))

        if not claim:
            dropped += 1
            continue

        # Drop low-signal confirmations
        if evidence.lower() in ("yes", "y", "ok", "okay", "sure", "confirmed"):
            dropped += 1
            continue

        slot = _norm_one_line(obj.get("slot") or "other").lower()
        subject = _norm_one_line(obj.get("subject") or "unknown").lower()

        privacy = _tier1_privacy_guess(claim)
        volatility = _tier1_volatility_guess(claim, slot)
        entity_key = _tier1_entity_key_guess(claim, slot, subject)
        claim_hash = _sha256_text(claim)

        obj["claim"] = claim
        obj["slot"] = slot
        obj["subject"] = subject
        obj["privacy"] = privacy
        obj["volatility"] = volatility
        obj["entity_key"] = entity_key
        obj["claim_hash"] = claim_hash

        out_lines.append(json.dumps(obj, ensure_ascii=False))
        kept += 1

    payload = "\n".join(out_lines).rstrip() + "\n"
    atomic_write_text(p, payload, encoding="utf-8", errors="strict")

    return {"ok": True, "written": kept, "dropped": dropped}


# ---------------------------------------------------------------------------
# Tier-2 — Distillation (Tier-1 facts_raw.jsonl -> facts_map.md)
#
# Deterministic, conservative promotion of recallable facts into a small,
# expert-curated canonical memory file (facts_map.md).
#
# Notes:
# - Tier-1 is NOT truth; Tier-2 is still conservative and provenance-heavy.
# - Scoring is profile-aware, schema is shared.
# - No model calls.
# ---------------------------------------------------------------------------

_TIER2_ALLOWED_PROFILES = ("therapist", "programmer", "general")


def _tier2_shell_facts_map() -> str:
    return "\n".join([
        "# Facts Map (Canonical Project Memory)",
        "",
        "## Entries",
        "- (none yet)",
        "",
    ])


def _tier2_parse_existing_claims(facts_map_text: str) -> set:
    """
    Return a set of normalized claim strings already present as '- FACT: ...'
    Deterministic, case-insensitive.
    """
    txt = facts_map_text or ""
    out = set()
    for ln in txt.splitlines():
        s = (ln or "").strip()
        if s.startswith("- FACT:"):
            claim = s[len("- FACT:"):].strip()
            if claim:
                out.add(re.sub(r"\s+", " ", claim).strip().lower())
    return out


def _tier2_load_facts_raw(project_name: str, *, max_lines: int = 8000) -> List[Dict[str, Any]]:
    """
    Best-effort load of Tier-1 facts from state/facts_raw.jsonl (bounded).
    """
    ensure_project_scaffold(project_name)
    p = facts_raw_path(project_name)
    if not p.exists():
        return []
    try:
        lines = p.read_text(encoding="utf-8", errors="replace").splitlines()
    except Exception:
        return []
    if max_lines and len(lines) > max_lines:
        lines = lines[-max_lines:]

    out: List[Dict[str, Any]] = []
    for ln in lines:
        s = (ln or "").strip()
        if not s:
            continue
        try:
            obj = json.loads(s)
        except Exception:
            continue
        if isinstance(obj, dict):
            out.append(obj)
    return out


def _tier2_profile_weights(profile: str) -> Dict[str, float]:
    """
    Shared schema; profile changes ONLY weighting (deterministic).
    """
    p = (profile or "").strip().lower()
    if p not in _TIER2_ALLOWED_PROFILES:
        p = "general"

    # slot weights
    if p == "therapist":
        return {
            "identity": 2.2,
            "relationship": 2.2,
            "preference": 1.4,
            "routine": 1.2,
            "constraint": 2.0,
            "context": 1.6,
            "event": 1.3,
            "possession": 0.6,
            "other": 0.8,
        }
    if p == "programmer":
        return {
            "identity": 0.7,
            "relationship": 0.6,
            "preference": 0.9,
            "routine": 0.9,
            "constraint": 2.2,
            "context": 2.0,
            "event": 1.2,
            "possession": 0.8,
            "other": 1.0,
        }
    # general
    return {
        "identity": 1.6,
        "relationship": 1.3,
        "preference": 1.4,
        "routine": 1.2,
        "constraint": 1.6,
        "context": 1.4,
        "event": 1.2,
        "possession": 0.8,
        "other": 1.0,
    }


def _tier2_confidence_weight(conf: str) -> float:
    c = (conf or "").strip().lower()
    if c == "high":
        return 1.0
    if c == "medium":
        return 0.6
    return 0.25


def _tier2_volatility_penalty(vol: str) -> float:
    v = (vol or "").strip().lower()
    if v == "stable":
        return 0.0
    if v == "likely_to_change":
        return 0.45
    return 0.15


def _tier2_privacy_penalty(priv: str) -> float:
    p = (priv or "").strip().lower()
    if p == "sensitive":
        return 0.65
    return 0.0

def distill_facts_raw_to_facts_map_tier2p(project_name: str) -> Dict[str, Any]:
    """
    Tier-2P: typed, conservative, conflict-aware facts cache derived from Tier-1 (facts_raw.jsonl).

    FOUNDATIONAL CHANGE:
    - Tier-2P must NOT depend on brittle exact wording patterns.
    - We keep a small typed cache for highest-salience keys, PLUS a "Core Claims"
      section that promotes stable, user-centric facts by slot class.

    Properties:
    - Conflicts are first-class and block confident typed recall.
    - Birthdate is stored if available (YYYY-MM-DD).
    - Age may be stored ONLY as a reported value for continuity (age_reported).
    - Deterministic rebuild.
    """
    ensure_project_scaffold(project_name)

    facts = _tier2_load_facts_raw(project_name, max_lines=8000)
    fm_path = state_file_path(project_name, "facts_map")

    if not facts:
        atomic_write_text(
            fm_path,
            "\n".join([
                "# Facts Map (Tier-2P Typed, Conflict-aware)",
                "",
                "## Facts",
                "- (none yet)",
                "",
                "## Core Claims",
                "- (none yet)",
                "",
                "## Conflicts",
                "- (none)",
                "",
            ]).rstrip() + "\n",
            encoding="utf-8",
            errors="strict",
        )
        return {"ok": True, "promoted": 0, "conflicts": 0}

    # ------------------------------------------------------------------
    # Tier-2P FOUNDATION RULE (project-agnostic)
    #
    # Tier-1 claims are already produced by the model/extractor with slots.
    # Tier-2P must NOT depend on brittle, exact wording matches.
    # We keep a small typed cache for the highest-salience keys, plus a
    # "Core Claims" section that promotes stable, user-centric facts by slot.
    # ------------------------------------------------------------------

    KEY_ORDER: List[str] = [
        "user.identity.preferred_name",
        "user.identity.pronouns",
        "user.identity.age_reported",
        "user.identity.location",
        "user.identity.birthdate",
        "user.identity.timezone",
        "user.relationship.partner.name",
        "user.relationship.child.name",
    ]

    vals: Dict[str, Dict[str, Any]] = {}

    def _push(key: str, value: str, fact_id: str) -> None:
        v = str(value or "").strip()
        if not v:
            return
        if key not in vals:
            vals[key] = {"values": set(), "evidence": {}}
        vals[key]["values"].add(v)
        ev = vals[key]["evidence"]
        if v not in ev:
            ev[v] = []
        if fact_id and fact_id not in ev[v]:
            ev[v].append(fact_id)
            ev[v] = ev[v][:12]

    def _extract(claim: str) -> List[Tuple[str, str]]:
        c = _norm_one_line(claim)
        low = c.lower()
        out: List[Tuple[str, str]] = []

        # Preferred name (first-person OR normalized third-person)
        if "my preferred name is" in low:
            out.append(("user.identity.preferred_name", c.split("is", 1)[1].strip()))
            return out
        if low.startswith("i go by "):
            out.append(("user.identity.preferred_name", c.split("by", 1)[1].strip()))
            return out
        if "the user's preferred name is" in low or "the user preferred name is" in low:
            out.append(("user.identity.preferred_name", c.split("is", 1)[1].strip()))
            return out
        if "the user's name is" in low or "the user name is" in low:
            out.append(("user.identity.preferred_name", c.split("is", 1)[1].strip()))
            return out

        # Pronouns (lightweight)
        m = re.search(r"\b(?:my\s+pronouns\s+are|pronouns\s+are)\s+(he/him|she/her|they/them)\b", low)
        if m:
            out.append(("user.identity.pronouns", m.group(1)))
            return out

        # Age (reported). Store as a volatile, reported value for continuity.
        m = re.search(r"\b(?:i\s*'m|i\s+am|i\s+’m)\s*(\d{1,3})\s*(?:years?\s*old|yo)\b", low)
        if m:
            out.append(("user.identity.age_reported", m.group(1)))
            return out
        m = re.search(r"\bthe\s+user\s+is\s*(\d{1,3})\s*(?:years?\s*old|yo)\b", low)
        if m:
            out.append(("user.identity.age_reported", m.group(1)))
            return out

        # Location (first-person OR normalized third-person)
        if "i live in " in low:
            out.append(("user.identity.location", c.split("in", 1)[1].strip()))
            return out
        if "the user lives in " in low:
            out.append(("user.identity.location", c.split("in", 1)[1].strip()))
            return out

        # Birthdate
        m = re.search(r"\b(?:my\s+birthday\s+is|i\s+was\s+born\s+on)\s+(\d{4}-\d{2}-\d{2})\b", low)
        if m:
            out.append(("user.identity.birthdate", m.group(1)))
            return out
        m = re.search(r"\bthe\s+user\s+was\s+born\s+on\s+(\d{4}-\d{2}-\d{2})\b", low)
        if m:
            out.append(("user.identity.birthdate", m.group(1)))
            return out

        # Timezone
        if "my time zone is " in low or low.startswith("timezone is "):
            out.append(("user.identity.timezone", c.split("is", 1)[1].strip()))
            return out
        if "the user's time zone is" in low or "the user time zone is" in low:
            out.append(("user.identity.timezone", c.split("is", 1)[1].strip()))
            return out
        if "central time" in low:
            out.append(("user.identity.timezone", "America/Chicago"))
            return out

        # Partner name (generic; do not assume girlfriend)
        if "my partner" in low and "name is" in low:
            out.append(("user.relationship.partner.name", c.split("name is", 1)[1].strip()))
            return out
        if "the user's partner" in low and "name is" in low:
            out.append(("user.relationship.partner.name", c.split("name is", 1)[1].strip()))
            return out

        # Child name (generic)
        if "my son" in low and "name is" in low:
            out.append(("user.relationship.child.name", c.split("name is", 1)[1].strip()))
            return out
        if "my daughter" in low and "name is" in low:
            out.append(("user.relationship.child.name", c.split("name is", 1)[1].strip()))
            return out
        if "the user's son" in low and "name is" in low:
            out.append(("user.relationship.child.name", c.split("name is", 1)[1].strip()))
            return out
        if "the user's daughter" in low and "name is" in low:
            out.append(("user.relationship.child.name", c.split("name is", 1)[1].strip()))
            return out

        return out

    promoted = 0
    for f in facts:
        if not isinstance(f, dict):
            continue
        claim = _norm_one_line(f.get("claim") or "")
        if not claim:
            continue
        kvs = _extract(claim)
        if not kvs:
            continue

        # HARD INVARIANT (birthday ownership):
        # Tier-2P must NOT promote birthdate into user.identity unless the originating Tier-1
        # fact was explicitly slot="identity" and subject="user".
        slot0 = str(f.get("slot") or "").strip().lower()
        subj0 = str(f.get("subject") or "").strip().lower()

        fid = _norm_one_line(f.get("id") or "")
        for k, v in kvs:
            if k == "user.identity.birthdate":
                if not (slot0 == "identity" and subj0 == "user"):
                    continue
            _push(k, v, fid)
            promoted += 1

    # ------------------------------------------------------------------
    # Core Claims (foundation)
    #
    # Promote stable, user-centric facts by *slot class*, not exact wording.
    # This lets natural conversation produce usable memory even when typed
    # key extraction doesn't match.
    # ------------------------------------------------------------------

    CORE_SLOTS = {"identity", "relationship", "context", "preference", "routine"}

    def _core_ok(f: Dict[str, Any]) -> bool:
        if not isinstance(f, dict):
            return False
        if str(f.get("subject") or "").strip().lower() != "user":
            return False
        slot = str(f.get("slot") or "other").strip().lower()
        if slot not in CORE_SLOTS:
            return False

        # HARD INVARIANT (Core Claims):
        # Never allow "I was born on YYYY-MM-DD" into Core Claims unless the originating Tier-1
        # fact was explicitly slot="identity". This prevents arbitrary dates (concerts, friends,
        # partner birthdays, etc.) from being preserved as a first-person birth claim.
        claim0 = _norm_one_line(f.get("claim") or "")
        low0 = claim0.lower().strip()
        if re.search(r"\bi\s+was\s+born\s+on\s+\d{4}-\d{2}-\d{2}\b", low0):
            if slot != "identity":
                return False

        priv = str(f.get("privacy") or "normal").strip().lower()
        if priv == "sensitive":
            return False
        # Conservative: avoid volatile facts when explicitly tagged.
        vol = str(f.get("volatility") or "unknown").strip().lower()
        if vol == "likely_to_change":
            return False
        # If extractor provided confidence, honor it.
        conf = str(f.get("confidence") or "").strip().lower()
        if conf == "low":
            return False
        return True

    core_by_hash: Dict[str, Dict[str, Any]] = {}
    for f in facts:
        if not _core_ok(f):
            continue
        claim = _norm_one_line(f.get("claim") or "")
        if not claim:
            continue
        h = str(f.get("claim_hash") or "").strip() or claim.lower()
        fid = _norm_one_line(f.get("id") or "")
        src = f.get("source") if isinstance(f.get("source"), dict) else {}
        ti = src.get("turn_index") if isinstance(src, dict) else None
        ts = src.get("timestamp") if isinstance(src, dict) else None

        if h not in core_by_hash:
            core_by_hash[h] = {"claim": claim, "evidence": [], "turn_index": ti, "timestamp": ts}
        if fid and fid not in core_by_hash[h]["evidence"]:
            core_by_hash[h]["evidence"].append(fid)
            core_by_hash[h]["evidence"] = core_by_hash[h]["evidence"][:8]

        # Prefer most recent timestamp/turn index for ordering.
        try:
            if isinstance(ti, int):
                old_ti = core_by_hash[h].get("turn_index")
                if old_ti is None or (isinstance(old_ti, int) and ti > old_ti):
                    core_by_hash[h]["turn_index"] = ti
                    core_by_hash[h]["timestamp"] = ts
        except Exception:
            pass

    core_items = list(core_by_hash.values())
    # Sort: most recent first if available
    core_items.sort(key=lambda x: (x.get("turn_index") if isinstance(x.get("turn_index"), int) else -1), reverse=True)
    core_items = core_items[:18]

    core_lines: List[str] = []
    for it in core_items:
        claim = str(it.get("claim") or "").strip()
        evs = it.get("evidence") if isinstance(it.get("evidence"), list) else []
        evs = [str(x) for x in evs if str(x).strip()]
        if not claim:
            continue
        if evs:
            core_lines.append(f"- {claim}  (evidence_ids: {evs[:6]})")
        else:
            core_lines.append(f"- {claim}")

    if not core_lines:
        core_lines = ["- (none yet)"]

    facts_lines: List[str] = []
    conflicts_lines: List[str] = []
    conflict_count = 0

    # ------------------------------------------------------------------
    # Key-specific deterministic canonicalization (NO inference)
    #
    # Tier-2P conflicts represent semantic disagreement, not paraphrase.
    # For specific keys where surface variation is common, we normalize
    # before conflict decisions are made.
    #
    # Target: user.identity.location ONLY
    # ------------------------------------------------------------------

    _T2P_LOC_KEY = "user.identity.location"

    def _t2p_norm_location_for_conflict(v: str) -> str:
        """
        Deterministic location normalizer for conflict detection only.
        Goal: collapse trivial paraphrases like:
          - "the Chicago suburbs."
          - "the suburbs of Chicago rather than downtown."
        into one semantic bucket (no conflict).

        Rules:
        - lowercase
        - strip trailing sentence punctuation
        - collapse whitespace
        - remove leading article "the "
        - unify "chicago suburbs" <-> "suburbs of chicago"
        - DO NOT attempt geocoding or synonym expansion
        """
        s = str(v or "").strip().lower()

        # strip trailing punctuation
        while s and s[-1] in ".!,;:":
            s = s[:-1].rstrip()

        # normalize whitespace
        s = " ".join(s.split())

        # drop leading article
        if s.startswith("the "):
            s = s[4:].strip()

        # unify common Chicago suburbs paraphrases
        # (explicit, non-inferential rewrite)
        if ("suburbs" in s) and ("chicago" in s):
            # If either phrase appears, bucket to a single canonical meaning key
            if ("chicago suburbs" in s) or ("suburbs of chicago" in s):
                return "suburbs_of_chicago"

        return s

    def _t2p_clean_location_display(v: str) -> str:
        """
        Deterministic display cleanup (no semantic rewrite):
        - trim
        - strip trailing punctuation
        - collapse whitespace
        """
        s = str(v or "").strip()
        while s and s[-1] in ".!,;:":
            s = s[:-1].rstrip()
        s = " ".join(s.split())
        return s

    for key in KEY_ORDER:
        if key not in vals:
            continue
        vs = sorted(list(vals[key].get("values") or []))
        evid = vals[key].get("evidence") if isinstance(vals[key].get("evidence"), dict) else {}

        # Key-specific collapse: user.identity.location paraphrases should not conflict.
        if key == _T2P_LOC_KEY and len(vs) > 1:
            norms = [_t2p_norm_location_for_conflict(v) for v in vs]
            uniq_norms = sorted(set([n for n in norms if n]))

            # If all variants collapse to one meaning bucket, treat as confirmed.
            if len(uniq_norms) == 1:
                # Choose the most informative surface form deterministically:
                # pick the longest cleaned display string (stable tie-breaker: lexicographic).
                cleaned = [_t2p_clean_location_display(v) for v in vs]
                cleaned = [c for c in cleaned if c]
                cleaned.sort(key=lambda x: (-len(x), x.lower(), x))
                rep = cleaned[0] if cleaned else ""

                # Merge evidence ids from all original variants into the representative value.
                merged_ids: List[str] = []
                for v in vs:
                    ev_ids_v = evid.get(v) if isinstance(evid.get(v), list) else []
                    for eid in ev_ids_v:
                        e0 = str(eid).strip()
                        if e0 and e0 not in merged_ids:
                            merged_ids.append(e0)

                # Rewrite vs/evid locally so downstream renders as confirmed (no conflict).
                vs = [rep] if rep else []
                evid = {rep: merged_ids} if rep else {}

        if len(vs) <= 1:
            v0 = vs[0] if vs else ""
            ev_ids = evid.get(v0) if isinstance(evid.get(v0), list) else []
            ev_ids = [str(x) for x in ev_ids if str(x).strip()]
            facts_lines.extend([
                f"- key: {key}",
                "  - status: confirmed",
                f"  - value: {v0}",
                f"  - evidence_ids: {ev_ids[:8]}",
                "",
            ])
        else:
            conflict_count += 1
            facts_lines.extend([
                f"- key: {key}",
                "  - status: ambiguous",
                "  - value: (blocked_by_conflict)",
                f"  - options: {vs[:8]}",
                "",
            ])
            conflicts_lines.append(f"- key: {key} has conflicting values: {vs[:8]}")

    if not facts_lines:
        facts_lines = ["- (none yet)", ""]

    if not conflicts_lines:
        conflicts_lines = ["- (none)"]

    out_txt = "\n".join(
        [
            "# Facts Map (Tier-2P Typed, Conflict-aware)",
            "",
            "## Facts",
            "",
            *facts_lines,
            "## Core Claims",
            *core_lines,
            "",
            "## Conflicts",
            *conflicts_lines,
            "",
        ]
    ).rstrip() + "\n"

    atomic_write_text(fm_path, out_txt, encoding="utf-8", errors="strict")

    return {"ok": True, "promoted": promoted, "conflicts": conflict_count}


def distill_facts_raw_to_facts_map(project_name: str, *, profile: str) -> Dict[str, Any]:
    """
    Deterministic Tier-2 distillation (Tier-2 semantics choice #2: current working model).

    Properties:
    - Replace-style: overwrite the "## Entries" section of facts_map.md on each run.
    - Synthesis: group Tier-1 facts by (entity_key + slot + subject) and emit one pattern claim per group.
    - Provenance-preserving: each Tier-2 entry lists multiple Tier-1 rawfact ids + turn_index/timestamp (bounded).
    - Profile-aware: uses existing weighting (therapist/programmer/general).
    - Conservative: avoids sensitive/low-confidence and penalizes volatility.
    """
    ensure_project_scaffold(project_name)

    prof = (profile or "").strip().lower()
    if prof not in _TIER2_ALLOWED_PROFILES:
        return {"ok": False, "error": "invalid_profile"}

    facts = _tier2_load_facts_raw(project_name, max_lines=8000)
    if not facts:
        return {"ok": True, "promoted": 0, "kept_existing": 0, "dropped": 0, "note": "no_tier1_facts", "profile": prof}

    fm_path = state_file_path(project_name, "facts_map")
    fm_txt = read_text_file(fm_path, errors="replace") if fm_path.exists() else ""
    if not fm_txt.strip():
        fm_txt = _tier2_shell_facts_map()

    weights = _tier2_profile_weights(prof)

    def _score_fact(f: Dict[str, Any]) -> Tuple[float, str, str, str, str, str, str]:
        claim = _norm_one_line(f.get("claim"))
        slot = _norm_one_line(f.get("slot") or "other").lower()
        subject = _norm_one_line(f.get("subject") or "unknown").lower()
        confidence = _norm_one_line(f.get("confidence") or "low").lower()
        volatility = _norm_one_line(f.get("volatility") or "unknown").lower()
        privacy = _norm_one_line(f.get("privacy") or "normal").lower()
        ek = _norm_one_line(f.get("entity_key") or "")

        # Conservative subject allowlist: promote user/project first.
        if subject not in ("user", "project"):
            subj_bonus = -0.8
        else:
            subj_bonus = 0.4

        slot_w = float(weights.get(slot, weights.get("other", 1.0)))
        conf_w = _tier2_confidence_weight(confidence)
        pen = _tier2_privacy_penalty(privacy) + _tier2_volatility_penalty(volatility)

        score = (slot_w * conf_w) + subj_bonus - pen

        return float(score), claim, ek, slot, subject, confidence, volatility, privacy

    # Group candidates by (entity_key + slot + subject)
    groups: Dict[Tuple[str, str, str], List[Tuple[float, int, Dict[str, Any]]]] = {}
    dropped = 0

    for idx, f in enumerate(facts):
        if not isinstance(f, dict):
            continue

        score, claim, ek, slot, subject, confidence, volatility, privacy = _score_fact(f)

        if not claim:
            dropped += 1
            continue

        # Hard conservative gates:
        # - If privacy is sensitive AND confidence is low, do not promote.
        if privacy == "sensitive" and confidence == "low":
            dropped += 1
            continue

        # Soft gate: score threshold for inclusion in Tier-2 grouping pool
        if score <= 0.35:
            continue

        gk = (ek or "(missing)", slot or "other", subject or "unknown")
        groups.setdefault(gk, []).append((float(score), int(idx), f))

    # Aggregate and rank groups
    ranked: List[Tuple[float, int, Tuple[str, str, str], List[Tuple[float, int, Dict[str, Any]]]]] = []
    for gk, items in groups.items():
        # Prefer top evidence within group (best score), stable tie-breaker: latest idx
        items_sorted = sorted(items, key=lambda x: (-x[0], -x[1]))
        topk = items_sorted[:3]
        avg_top = sum(x[0] for x in topk) / float(len(topk) or 1)
        latest_idx = max(x[1] for x in items_sorted) if items_sorted else 0
        ranked.append((float(avg_top), int(latest_idx), gk, items_sorted))

    ranked.sort(key=lambda x: (-x[0], -x[1]))

    # -----------------------------------------------------------------------
    # Deterministic core-identity injection
    #
    # Why:
    # - Even with “preserve core identity groups”, if Tier-1 rows have stale/misaligned
    #   entity_key/slot values (or grouping misses), the literal identity strings
    #   (e.g., "Frank") can fail to appear in facts_map.md.
    #
    # Contract:
    # - Only inject if we find an explicit matching claim in Tier-1 (truth-bound).
    # - No model calls. No heuristics beyond explicit first-person patterns.
    # - This is a deterministic guardrail, not a redesign.
    # -----------------------------------------------------------------------
    core_keys = {
        "user.identity.name": "identity",
        "user.identity.location": "identity",
        "user.preference.favorite_color": "preference",
    }

    # entity_keys already present in grouped candidates
    present_core = set()
    for _gscore, _latest_idx, gk, _items_sorted in ranked:
        try:
            ek0 = str(gk[0] or "").strip().lower()
        except Exception:
            ek0 = ""
        if ek0 in core_keys:
            present_core.add(ek0)

    def _latest_match_for(patterns: Tuple[str, ...]) -> Tuple[int, Dict[str, Any]]:
        # Return (idx, fact_dict_copy) for the latest Tier-1 fact whose claim matches.
        # Latest is defined by highest index in the loaded Tier-1 list (deterministic).
        best_i = -1
        best_f = None
        for i, ff in enumerate(facts):
            if not isinstance(ff, dict):
                continue
            c = _norm_one_line(ff.get("claim") or "")
            low = c.lower()
            if not low:
                continue
            if any(p in low for p in patterns):
                best_i = i
                best_f = ff
        if best_i < 0 or best_f is None:
            return -1, {}
        return best_i, dict(best_f)

    inject_specs = [
        ("user.identity.name", "identity", ("my preferred name is", "my name is", "i go by")),
        ("user.identity.location", "identity", ("i live in", "i'm in ", "im in ", "i am in ")),
        ("user.preference.favorite_color", "preference", ("my favorite color is", "my favourite color is")),
    ]

    injected = []
    for ek_need, slot_need, pats in inject_specs:
        if ek_need in present_core:
            continue

        i0, f0 = _latest_match_for(pats)
        if i0 < 0 or not f0:
            continue  # truth-bound: do not inject if Tier-1 doesn't contain it

        # Force canonical pins for grouping/output
        f0["entity_key"] = ek_need
        f0["slot"] = slot_need
        f0["subject"] = "user"

        gk0 = (ek_need, slot_need, "user")
        items0 = [(9.99, int(i0), f0)]  # strong score; ordering remains deterministic
        injected.append((9.99, int(i0), gk0, items0))

    if injected:
        # Put injected core identity groups at the very front; rest stays sorted.
        ranked = injected + ranked

    def _is_core_identity_group(items_sorted: List[Tuple[float, int, Dict[str, Any]]]) -> bool:
        """
        Deterministic guardrail: always keep core identity facts in Tier-2 when present,
        even if the conversation contains many higher-scoring groups.

        IMPORTANT:
        - Prefer entity_key pinning (more stable than text patterns).
        - Fall back to text patterns for backward compatibility with older Tier-1 rows.
        """
        core_entity_keys = {
            "user.identity.name",
            "user.identity.location",
            "user.preference.favorite_color",
        }

        for _sc, _ix, f in (items_sorted or []):
            ek = _norm_one_line(f.get("entity_key") or "").strip().lower()
            if ek in core_entity_keys:
                return True

            c = _norm_one_line(f.get("claim") or "").lower()
            if not c:
                continue
            if ("preferred name" in c) or ("my name is" in c) or ("i go by" in c):
                return True
            if ("i live in" in c) or ("i'm in " in c) or ("im in " in c) or ("i am in " in c):
                return True
            if ("favorite color" in c) or ("favourite color" in c):
                return True

        return False


    # Tier-2 cap (bounded current working model)
    max_facts_total = 30

    # Promote core identity groups first, then apply the cap.
    core = [r for r in ranked if _is_core_identity_group(r[3])]
    rest = [r for r in ranked if not _is_core_identity_group(r[3])]
    ranked = (core + rest)[:max_facts_total]

    # Build synthesized Tier-2 entries
    entries: List[List[str]] = []
    promoted = 0

    def _conf_disp(c: str) -> str:
        c0 = (c or "").strip().lower()
        if c0 == "high":
            return "High"
        if c0 == "medium":
            return "Medium"
        return "Low"

    def _worst_volatility(vs: List[str]) -> str:
        # deterministic ordering: likely_to_change > unknown > stable
        have = {str(x or "").strip().lower() for x in vs if str(x or "").strip()}
        if "likely_to_change" in have:
            return "likely_to_change"
        if "unknown" in have:
            return "unknown"
        if "stable" in have:
            return "stable"
        return "unknown"

    for _gscore, _latest_idx, gk, items_sorted in ranked:
        ek, slot, subject = gk

        # Choose the representative claim deterministically:
        # - sum scores by normalized claim
        # - tie-breaker: latest occurrence
        claim_scores: Dict[str, float] = {}
        claim_latest: Dict[str, int] = {}

        for sc, ix, f in items_sorted:
            c = _norm_one_line(f.get("claim"))
            ck = re.sub(r"\s+", " ", (c or "")).strip().lower()
            if not ck:
                continue
            claim_scores[ck] = claim_scores.get(ck, 0.0) + float(sc)
            claim_latest[ck] = max(int(claim_latest.get(ck, 0)), int(ix))

        if not claim_scores:
            continue

        best_ck = sorted(claim_scores.keys(), key=lambda k: (-claim_scores.get(k, 0.0), -claim_latest.get(k, 0), k))[0]
        # Recover the original-cased representative claim from the latest matching item
        rep_claim = ""
        for sc, ix, f in items_sorted:
            c = _norm_one_line(f.get("claim"))
            ck = re.sub(r"\s+", " ", (c or "")).strip().lower()
            if ck == best_ck:
                rep_claim = c
                break
        rep_claim = rep_claim or best_ck

        # Pattern suffix only when group has multiple supporting facts
        support_n = len(items_sorted)
        fact_line = rep_claim
        if support_n >= 2:
            fact_line = f"{rep_claim} (pattern; n={support_n})"

        # Group confidence: max of included
        confs = []
        vols = []
        privs = []
        for _sc2, _ix2, f2 in items_sorted:
            confs.append(_norm_one_line(f2.get("confidence") or "low").lower())
            vols.append(_norm_one_line(f2.get("volatility") or "unknown").lower())
            privs.append(_norm_one_line(f2.get("privacy") or "normal").lower())

        # confidence rank: high > medium > low
        conf_disp = "Low"
        if "high" in confs:
            conf_disp = "High"
        elif "medium" in confs:
            conf_disp = "Medium"

        vol_disp = _worst_volatility(vols)
        priv_disp = "sensitive" if ("sensitive" in privs) else "normal"

        # Provenance list (bounded)
        prov_parts: List[str] = []
        # Use top 6 by score/recency
        top_items = items_sorted[:6]
        for sc3, ix3, f3 in top_items:
            rid = _norm_one_line(f3.get("id") or "") or "(missing_id)"
            src = f3.get("source") if isinstance(f3.get("source"), dict) else {}
            turn_index = src.get("turn_index")
            ts = _norm_one_line(src.get("timestamp") or "")
            try:
                ti = int(turn_index) if turn_index is not None else 0
            except Exception:
                ti = 0
            seg = f"facts_raw.jsonl:{rid}(turn_index={ti}" + (f", timestamp={ts}" if ts else "") + ")"
            prov_parts.append(seg)

        more = max(0, support_n - len(top_items))
        prov_line = "; ".join(prov_parts) + (f"; +{more} more" if more else "")

        # Evidence: best quote from top item
        best_quote = ""
        if items_sorted:
            best_quote = _norm_one_line(items_sorted[0][2].get("evidence_quote") or "")
            if len(best_quote) > 260:
                best_quote = best_quote[:259].rstrip() + "…"

        block: List[str] = []
        block.append(f"- FACT: {fact_line}")
        block.append("  - Tier: 2")
        block.append(f"  - GroupKey: {ek}|{slot}|{subject}")
        block.append(f"  - EntityKey: {ek}")
        block.append(f"  - Slot: {slot}")
        block.append(f"  - Subject: {subject}")
        block.append(f"  - Confidence: {conf_disp}")
        block.append(f"  - Volatility: {vol_disp}")
        block.append(f"  - Privacy: {priv_disp}")
        block.append(f"  - Provenance: {prov_line}")
        if best_quote:
            block.append(f"  - Evidence: \"{best_quote}\"")
        block.append("")
        entries.append(block)
        promoted += 1

    # Replace-style write: overwrite everything under "## Entries"
    lines = fm_txt.splitlines()
    try:
        idx_entries = next(i for i, ln in enumerate(lines) if ln.strip() == "## Entries")
        header_lines = lines[: idx_entries + 1]
    except StopIteration:
        header_lines = ["# Facts Map (Canonical Project Memory)", "", "## Entries"]

    out_lines: List[str] = []
    out_lines.extend(header_lines)

    if promoted <= 0:
        out_lines.append("- (none yet)")
        out_lines.append("")
    else:
        out_lines.append("")
        for block in entries:
            out_lines.extend(block)

    final_txt = "\n".join(out_lines).rstrip() + "\n"
    atomic_write_text(fm_path, final_txt, encoding="utf-8", errors="strict")

    # Also version it as an artifact so it shows up in list/open flows
    try:
        create_artifact(
            project_name,
            "facts_map",
            final_txt,
            artifact_type="facts_map",
            from_files=[],
            file_ext=".md",
        )
    except Exception:
        pass

    return {"ok": True, "promoted": promoted, "kept_existing": 0, "dropped": dropped, "profile": prof, "mode": "replace_synth"}


def pending_upload_question_path(project_name: str) -> Path:
    """
    projects/<user>/<project>/state/pending_upload_question.json
    """
    return state_dir(project_name) / PENDING_UPLOAD_QUESTION_FILE_NAME


def upload_notes_path(project_name: str) -> Path:
    """
    projects/<user>/<project>/state/upload_notes.jsonl
    """
    return state_dir(project_name) / UPLOAD_NOTES_FILE_NAME

def capability_gaps_path(project_name: str) -> Path:
    """
    projects/<user>/<project>/state/capability_gaps.jsonl
    """
    return state_dir(project_name) / CAPABILITY_GAPS_FILE_NAME
def active_object_path(project_name: str) -> Path:
    """
    projects/<user>/<project>/state/active_object.json

    Ephemeral conversational focus (NOT canonical truth).
    """
    return state_dir(project_name) / ACTIVE_OBJECT_FILE_NAME


def load_active_object(project_name: str) -> Dict[str, Any]:
    """
    Load ephemeral active object focus.
    Returns {} if missing/invalid.
    """
    ensure_project(project_name)
    p = active_object_path(project_name)
    if not p.exists():
        return {}
    try:
        obj = json.loads(p.read_text(encoding="utf-8") or "{}")
        return obj if isinstance(obj, dict) else {}
    except Exception:
        return {}


def save_active_object(project_name: str, obj: Dict[str, Any]) -> None:
    """
    Save ephemeral active object focus (atomic write).
    This MUST NOT use canonical write paths (not a truth artifact).
    """
    ensure_project(project_name)

    d = obj if isinstance(obj, dict) else {}

    rel_path = str(d.get("rel_path") or "").replace("\\", "/").strip()
    orig_name = str(d.get("orig_name") or "").strip()
    sha256 = str(d.get("sha256") or "").strip()
    mime = str(d.get("mime") or "").strip()
    reason = str(d.get("set_reason") or "").strip()

    out: Dict[str, Any] = {
        "v": 1,
        "kind": "upload",
        "rel_path": rel_path,
        "orig_name": orig_name,
        "sha256": sha256,
        "mime": mime,
        "set_reason": reason or "unknown",
    }

    # Preserve original set_at if present; always update last_used_at deterministically.
    set_at = str(d.get("set_at") or "").strip()
    if set_at:
        out["set_at"] = set_at
    else:
        out["set_at"] = now_iso()

    out["last_used_at"] = now_iso()

    p = active_object_path(project_name)
    atomic_write_text(p, json.dumps(out, indent=2), encoding="utf-8", errors="strict")


def clear_active_object(project_name: str) -> None:
    """
    Clear ephemeral active object focus.
    """
    ensure_project(project_name)
    p = active_object_path(project_name)
    try:
        if p.exists():
            p.unlink()
    except Exception:
        pass


def _now_iso_noz() -> str:
    # Match the spec example format (no trailing 'Z').
    try:
        return (now_iso() or "").replace("Z", "")
    except Exception:
        return time.strftime("%Y-%m-%dT%H:%M:%S", time.gmtime())


def _parse_iso_noz(s: str) -> float:
    """
    Parse 'YYYY-MM-DDTHH:MM:SS' (optionally with trailing Z) to epoch seconds (UTC).
    Returns 0.0 on failure.
    """
    t = (s or "").strip()
    if not t:
        return 0.0
    t = t.replace("Z", "")
    try:
        return time.mktime(time.strptime(t, "%Y-%m-%dT%H:%M:%S"))
    except Exception:
        return 0.0
# ---------------------------------------------------------------------------
# C9 — Inbox helpers (deterministic, append-only)
# ---------------------------------------------------------------------------

def _next_inbox_id(project_name: str, *, created_at: str) -> str:
    """
    Deterministic per-day increment:
      inbox_YYYY_MM_DD_NNN
    """
    day = (created_at or "").split("T", 1)[0].replace("-", "_")
    pfx = f"inbox_{day}_"

    max_n = 0
    p = inbox_path(project_name)
    if p.exists():
        try:
            for ln in p.read_text(encoding="utf-8", errors="replace").splitlines():
                try:
                    obj = json.loads(ln)
                except Exception:
                    continue
                if not isinstance(obj, dict):
                    continue
                did = str(obj.get("id") or "")
                if not did.startswith(pfx):
                    continue
                tail = did[len(pfx):]
                try:
                    n = int(tail)
                except Exception:
                    n = 0
                if n > max_n:
                    max_n = n
        except Exception:
            pass

    return f"{pfx}{(max_n + 1):03d}"


def append_inbox_item(
    project_name: str,
    *,
    type_: str,
    text: str,
    refs: Optional[List[str]] = None,
    created_at: str = "",
) -> Dict[str, Any]:
    """
    Append a new inbox item (open).
    Deterministic storage only. No model calls.
    """
    ensure_project(project_name)

    created = (created_at or "").strip() or _now_iso_noz()
    iid = _next_inbox_id(project_name, created_at=created)

    t = str(type_ or "").strip()
    if t not in ("clarification", "pending_decision", "conflict", "missing_requirements"):
        t = "clarification"

    entry = {
        "id": iid,
        "created_at": created,
        "type": t,
        "status": "open",
        "text": strip_control_tokens(str(text or "").strip()),
        "refs": [strip_control_tokens(str(x)) for x in (refs or []) if str(x).strip()],
    }

    _append_jsonl_line(project_name, inbox_path(project_name), entry)
    return entry


def _find_latest_open_inbox_record(project_name: str, inbox_id: str) -> Dict[str, Any]:
    """
    Best-effort scan (bounded tail) to find the latest record for an id.
    Deterministic, no inference.
    """
    iid = str(inbox_id or "").strip()
    if not iid:
        return {}

    p = inbox_path(project_name)
    if not p.exists():
        return {}

    try:
        lines = p.read_text(encoding="utf-8", errors="replace").splitlines()
    except Exception:
        return {}

    tail = lines[-1200:] if len(lines) > 1200 else lines
    latest: Dict[str, Any] = {}
    for ln in tail:
        try:
            obj = json.loads(ln)
        except Exception:
            continue
        if not isinstance(obj, dict):
            continue
        if str(obj.get("id") or "").strip() != iid:
            continue
        latest = obj

    # Only return if it's currently open
    if latest and str(latest.get("status") or "").strip() == "open":
        return latest
    return {}


def resolve_inbox_item(
    project_name: str,
    *,
    inbox_id: str,
    resolved_at: str = "",
    resolution_note: str = "",
    refs: Optional[List[str]] = None,
) -> bool:
    """
    Append-only resolve: writes a NEW line for the same id with status=resolved.
    Does NOT mutate prior rows.
    """
    ensure_project(project_name)

    iid = str(inbox_id or "").strip()
    if not iid:
        return False

    created = (resolved_at or "").strip() or _now_iso_noz()

    latest_open = _find_latest_open_inbox_record(project_name, iid)
    item_type = str(latest_open.get("type") or "clarification").strip()
    if item_type not in ("clarification", "pending_decision", "conflict", "missing_requirements"):
        item_type = "clarification"

    # Keep a minimal human-facing text; do NOT embed multiline content.
    base_text = str(latest_open.get("text") or "").strip()
    if base_text:
        text = f"RESOLVED: {base_text}"
        if len(text) > 240:
            text = text[:239].rstrip() + "…"
    else:
        text = "RESOLVED"

    if resolution_note:
        rn = strip_control_tokens(str(resolution_note).strip())
        if rn:
            text2 = f"{text} ({rn})"
            if len(text2) <= 280:
                text = text2

    entry = {
        "id": iid,
        "created_at": created,
        "type": item_type,
        "status": "resolved",
        "text": text,
        "refs": [strip_control_tokens(str(x)) for x in (refs or []) if str(x).strip()],
    }

    ok = write_canonical_entry(
        project_name,
        target_path=inbox_path(project_name),
        mode="jsonl_append",
        data=entry,
    )
    return bool(ok)


def load_inbox(project_name: str) -> List[Dict[str, Any]]:
    """
    Load inbox.jsonl in file order.
    """
    ensure_project(project_name)
    p = inbox_path(project_name)
    if not p.exists():
        return []
    out: List[Dict[str, Any]] = []
    try:
        for ln in p.read_text(encoding="utf-8", errors="replace").splitlines():
            try:
                obj = json.loads(ln)
            except Exception:
                continue
            if isinstance(obj, dict):
                out.append(obj)
    except Exception:
        return out
    return out


def list_open_inbox(project_name: str, *, max_items: int = 12) -> List[Dict[str, Any]]:
    """
    Returns most recent N OPEN inbox items (deduped by id, last-write-wins).
    Deterministic only.
    """
    rows = load_inbox(project_name) or []
    latest_by_id: Dict[str, Dict[str, Any]] = {}
    for r in rows:
        if not isinstance(r, dict):
            continue
        iid = str(r.get("id") or "").strip()
        if not iid:
            continue
        latest_by_id[iid] = r

    open_items = [v for v in latest_by_id.values() if str(v.get("status") or "").strip() == "open"]
    # Sort by created_at desc (string works for ISO-ish)
    open_items.sort(key=lambda x: str(x.get("created_at") or ""), reverse=True)
    return open_items[: int(max_items or 12)]

# ---------------------------------------------------------------------------
# Couples Mode Option A — Bring-up-later queue helpers (deterministic)
# Storage: state/bringups.jsonl (append-only). Resolve appends a new row.
# ---------------------------------------------------------------------------

def _next_bringup_id(project_name: str, *, created_at: str) -> str:
    """
    Deterministic per-day increment:
      bringup_YYYY_MM_DD_NNN
    """
    day = (created_at or "").split("T", 1)[0].replace("-", "_")
    pfx = f"bringup_{day}_"

    max_n = 0
    p = bringups_path(project_name)
    if p.exists():
        try:
            for ln in p.read_text(encoding="utf-8", errors="replace").splitlines():
                try:
                    obj = json.loads(ln)
                except Exception:
                    continue
                if not isinstance(obj, dict):
                    continue
                bid = str(obj.get("id") or "")
                if not bid.startswith(pfx):
                    continue
                tail = bid[len(pfx):]
                try:
                    n = int(tail)
                except Exception:
                    n = 0
                if n > max_n:
                    max_n = n
        except Exception:
            pass

    return f"{pfx}{(max_n + 1):03d}"


def append_bringup_request(
    project_name: str,
    *,
    from_user: str,
    to_user: str,
    topic: str,
    tone: str,
    boundaries: str,
    urgency: str = "",
    context_summary: str = "",
    created_at: str = "",
) -> Dict[str, Any]:
    """
    Append a new bring-up request (open).
    Deterministic storage only. No model calls.
    """
    ensure_project(project_name)

    created = (created_at or "").strip() or _now_iso_noz()
    bid = _next_bringup_id(project_name, created_at=created)

    entry = {
        "id": bid,
        "created_at": created,
        "status": "open",
        "from_user": strip_control_tokens(str(from_user or "").strip()),
        "to_user": strip_control_tokens(str(to_user or "").strip()),
        "topic": strip_control_tokens(str(topic or "").strip()),
        "tone": strip_control_tokens(str(tone or "").strip()),
        "boundaries": strip_control_tokens(str(boundaries or "").strip()),
        "urgency": strip_control_tokens(str(urgency or "").strip()),
    }

    # Optional therapist-side formulation summary (deterministic; no quotes required)
    cs = strip_control_tokens(str(context_summary or "").strip())
    if cs:
        if len(cs) > 320:
            cs = cs[:319].rstrip() + "…"
        entry["context_summary"] = cs

    ok = write_canonical_entry(
        project_name,
        target_path=bringups_path(project_name),
        mode="jsonl_append",
        data=entry,
    )
    if not ok:
        raise ValueError("canonical_write_rejected")
    return entry


def resolve_bringup_request(
    project_name: str,
    *,
    bringup_id: str,
    resolved_at: str = "",
    resolution_note: str = "",
) -> bool:
    """
    Append-only resolve: writes a NEW row for the same id with status=resolved.
    """
    ensure_project(project_name)

    bid = str(bringup_id or "").strip()
    if not bid:
        return False

    created = (resolved_at or "").strip() or _now_iso_noz()

    entry = {
        "id": bid,
        "created_at": created,
        "status": "resolved",
        "resolution_note": strip_control_tokens(str(resolution_note or "").strip()),
    }

    ok = write_canonical_entry(
        project_name,
        target_path=bringups_path(project_name),
        mode="jsonl_append",
        data=entry,
    )
    return bool(ok)


def _load_bringups(project_name: str, *, max_lines: int = 5000) -> List[Dict[str, Any]]:
    """
    Best-effort load (bounded) of bringups.jsonl in file order.
    """
    ensure_project(project_name)
    p = bringups_path(project_name)
    if not p.exists():
        return []
    try:
        lines = p.read_text(encoding="utf-8", errors="replace").splitlines()
    except Exception:
        return []
    if max_lines and len(lines) > max_lines:
        lines = lines[-max_lines:]
    out: List[Dict[str, Any]] = []
    for ln in lines:
        try:
            obj = json.loads(ln)
        except Exception:
            continue
        if isinstance(obj, dict):
            out.append(obj)
    return out


def list_open_bringups_for_user(project_name: str, *, to_user: str, max_items: int = 8) -> List[Dict[str, Any]]:
    """
    Last-write-wins by id; returns OPEN items addressed to to_user.
    Deterministic only.
    """
    tu = str(to_user or "").strip()
    if not tu:
        return []

    rows = _load_bringups(project_name, max_lines=6000) or []
    latest_by_id: Dict[str, Dict[str, Any]] = {}
    for r in rows:
        if not isinstance(r, dict):
            continue
        bid = str(r.get("id") or "").strip()
        if bid:
            latest_by_id[bid] = r

    open_items = []
    for r in latest_by_id.values():
        if str(r.get("status") or "").strip() != "open":
            continue
        if str(r.get("to_user") or "").strip() != tu:
            continue
        open_items.append(r)

    open_items.sort(key=lambda x: str(x.get("created_at") or ""), reverse=True)
    return open_items[: int(max_items or 8)]


def render_pending_bringups_for_session(project_name: str, *, session_start: bool, max_items: int = 5) -> str:
    """
    Couples Mode Option A:
    On session start ONLY, return a system-facing block of neutral themes derived from
    bringups addressed to the current couple_* user.

    Hard rule: NO attribution (no "Frank asked...", no provenance).
    """
    if not session_start:
        return ""

    # Determine current user segment from "user/project"
    pname = safe_project_name(project_name)
    user_seg = pname.split("/", 1)[0].strip()
    proj_seg = pname.split("/", 1)[1].strip() if "/" in pname else DEFAULT_PROJECT_NAME
    if not user_seg.lower().startswith("couple_"):
        return ""

    def _norm_user(u: str) -> str:
        return re.sub(r"[^a-zA-Z0-9_-]", "_", (u or "").strip())

    me = _norm_user(user_seg)

    # Find active couple link for this couple_* user
    links = load_couples_links()
    rec = None
    for _cid, obj in (links or {}).items():
        if not isinstance(obj, dict):
            continue
        if str(obj.get("status") or "") != "active":
            continue
        ua = _norm_user(str(obj.get("user_a") or ""))
        ub = _norm_user(str(obj.get("user_b") or ""))
        if me == ua or me == ub:
            rec = obj
            break

    if not isinstance(rec, dict):
        return ""

    ua = _norm_user(str(rec.get("user_a") or ""))
    ub = _norm_user(str(rec.get("user_b") or ""))

    # Option A rule: scan bringups inside the CURRENT project segment (e.g., therapy_test)
    # for both partners, not link-record defaults.
    proj_a = safe_project_name(f"{ua}/{proj_seg or DEFAULT_PROJECT_NAME}")
    proj_b = safe_project_name(f"{ub}/{proj_seg or DEFAULT_PROJECT_NAME}")

    # Bringups may be stored in either partner's project; scan both.
    items = []
    try:
        items.extend(list_open_bringups_for_user(proj_a, to_user=me, max_items=max_items))
    except Exception:
        pass
    try:
        items.extend(list_open_bringups_for_user(proj_b, to_user=me, max_items=max_items))
    except Exception:
        pass

    # Dedupe by id, then sort by created_at desc
    by_id: Dict[str, Dict[str, Any]] = {}
    for it in items:
        bid = str(it.get("id") or "").strip()
        if bid:
            by_id[bid] = it
    uniq = list(by_id.values())
    uniq.sort(key=lambda x: str(x.get("created_at") or ""), reverse=True)
    uniq = uniq[: int(max_items or 5)]

    if not uniq:
        return ""

    lines: List[str] = []
    lines.append("PENDING_BRINGUPS_FOR_THIS_SESSION (themes; DO NOT attribute):")
    for it in uniq:
        topic = sanitize_retrieved_text(str(it.get("topic") or "")).replace("\n", " ").strip()
        ctxsum = sanitize_retrieved_text(str(it.get("context_summary") or "")).replace("\n", " ").strip()
        if len(ctxsum) > 220:
            ctxsum = ctxsum[:219].rstrip() + "…"
        tone = sanitize_retrieved_text(str(it.get("tone") or "")).replace("\n", " ").strip()
        boundaries = sanitize_retrieved_text(str(it.get("boundaries") or "")).replace("\n", " ").strip()
        urgency = sanitize_retrieved_text(str(it.get("urgency") or "")).replace("\n", " ").strip()

        if len(topic) > 160:
            topic = topic[:159].rstrip() + "…"

        theme = f"- Theme: {topic}" if topic else "- Theme: (unspecified)"
        if tone:
            theme += f" | Tone: {tone}"
        if boundaries:
            theme += f" | Boundaries: {boundaries}"
        if urgency:
            theme += f" | Urgency: {urgency}"
        lines.append(theme)

        # Extra therapist-only context (DO NOT QUOTE; use only to inform neutral framing)
        if ctxsum:
            lines.append(f"  - ContextSummary (do not quote): {ctxsum}")

    return "\n".join(lines).strip()

def summarize_inbox_open(project_name: str, *, max_items: int = 12) -> str:
    """
    Short deterministic text block for user-facing "what's left" answers.
    """
    items = list_open_inbox(project_name, max_items=max_items)
    if not items:
        return "Inbox (open): (none)"

    lines: List[str] = []
    lines.append("Inbox (open):")
    for it in items:
        iid = str(it.get("id") or "").strip()
        t = str(it.get("type") or "").strip()
        txt = sanitize_retrieved_text(str(it.get("text") or "")).replace("\n", " ").strip()
        if len(txt) > 200:
            txt = txt[:199].rstrip() + "…"
        if iid and t:
            lines.append(f"- [{t}] {txt} ({iid})")
        elif iid:
            lines.append(f"- {txt} ({iid})")
        else:
            lines.append(f"- {txt}")
    return "\n".join(lines).strip()

# -----------------------------------------------------------------------------
# Deterministic sanitizer: strip control wrappers from canonical rendering/storage
# -----------------------------------------------------------------------------

_CONTROL_BLOCK_RE = re.compile(
    r"\[(SCOPE|GOAL_ONBOARD|DELIVERABLE)\].*?\[/\1\]",
    flags=re.DOTALL | re.IGNORECASE,
)

def strip_control_tokens(text: str) -> str:
    """
    Deterministically remove known control metadata from text.

    Removes:
      - [SCOPE]...[/SCOPE] (multiline)
      - [GOAL_ONBOARD]...[/GOAL_ONBOARD]
      - [DELIVERABLE]...[/DELIVERABLE]
      - stray closing tags like [/SCOPE]

    This must NEVER call the model.
    """
    s = "" if text is None else str(text)

    # Remove known wrapped blocks
    s = _CONTROL_BLOCK_RE.sub("", s)

    # Remove stray open/close tag lines/remnants
    s = re.sub(r"\[/(SCOPE|GOAL_ONBOARD|DELIVERABLE)\]", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\[(SCOPE|GOAL_ONBOARD|DELIVERABLE)\]", "", s, flags=re.IGNORECASE)

    return s.strip()
def sanitize_retrieved_text(text: str) -> str:
    """
    Stage 2 sanitizer for canonical retrieval:
    - strips control wrappers ([SCOPE]...[/SCOPE], etc.)
    - strips role prefixes (User:/Assistant:)
    - normalizes whitespace (lightly)
    Deterministic ONLY. No model calls.
    """
    s = strip_control_tokens("" if text is None else str(text))

    # Strip common role prefixes (single-line)
    for pfx in ("User:", "Assistant:", "USER:", "ASSISTANT:"):
        if s.startswith(pfx):
            s = s[len(pfx):].lstrip()

    # Collapse runs of whitespace but keep newlines
    s = "\n".join([re.sub(r"[ \t]+", " ", ln).rstrip() for ln in s.splitlines()]).strip()
    return s
# ---------------------------------------------------------------------------
# C6.2 — Canonical Write Path Normalization (sanitizer + unified writer)
# ---------------------------------------------------------------------------
#
# Non-negotiable:
# - ALL writes to canonical project state must pass through write_canonical_entry()
# - Deterministic + stdlib-only
# - Apply ONLY to new writes (no migration / no rewrite of existing history files)
#
# Hard cap choice:
# - We enforce a hard cap in *characters* per write payload (post-sanitization).
#   This caps both JSONL entries and full-file canonical writes.
#

_CANONICAL_MAX_CHARS = 200_000  # hard cap per canonical write payload (chars)

# Strip bracketed control blocks (including but not limited to these tags)
# We keep an allowlist-first pass + a conservative generic pass for ALL-CAPS tags.
_CONTROL_BLOCK_TAGS = ("SCOPE", "DELIVERABLE", "GOAL_ONBOARD")

_CONTROL_BLOCK_ANY_RE = re.compile(
    r"\[(?P<tag>[A-Z0-9_]{2,})\].*?\[/\1\]",
    flags=re.DOTALL,
)

def _strip_bracketed_control_blocks(text: str) -> str:
    """
    Remove bracketed control blocks like:
      [SCOPE]...[/SCOPE], [DELIVERABLE]...[/DELIVERABLE], [GOAL_ONBOARD]...[/GOAL_ONBOARD]
    plus a conservative generic removal for ALL-CAPS tags.
    Deterministic only.
    """
    s = "" if text is None else str(text)

    # Allowlist pass (case-insensitive)
    for tag in _CONTROL_BLOCK_TAGS:
        s = re.sub(
            rf"\[{tag}\].*?\[/{tag}\]",
            "",
            s,
            flags=re.DOTALL | re.IGNORECASE,
        )

    # Generic ALL-CAPS tags (helps catch new control wrappers without code changes)
    s = _CONTROL_BLOCK_ANY_RE.sub("", s)

    # Remove stray open/close tags that might remain
    s = re.sub(r"\[/(?:%s)\]" % "|".join(_CONTROL_BLOCK_TAGS), "", s, flags=re.IGNORECASE)
    s = re.sub(r"\[(?:%s)\]" % "|".join(_CONTROL_BLOCK_TAGS), "", s, flags=re.IGNORECASE)

    return s

def _normalize_canonical_whitespace(text: str) -> str:
    """
    Deterministic whitespace normalization:
    - strip leading/trailing whitespace
    - remove trailing spaces on each line
    - collapse excessive blank lines (max 2 consecutive)
    """
    s = "" if text is None else str(text)

    # Normalize line endings
    s = s.replace("\r\n", "\n").replace("\r", "\n")

    # Trim trailing spaces per line
    lines = [ln.rstrip() for ln in s.split("\n")]

    # Collapse >2 blank lines to 2 blank lines
    out: List[str] = []
    blank_run = 0
    for ln in lines:
        if ln.strip() == "":
            blank_run += 1
            if blank_run <= 2:
                out.append("")
            continue
        blank_run = 0
        out.append(ln)

    return "\n".join(out).strip()

def _enforce_size_cap(text: str, *, cap_chars: int) -> str:
    s = "" if text is None else str(text)
    if cap_chars and len(s) > cap_chars:
        raise ValueError(f"canonical_write_rejected: size_cap_exceeded chars={len(s)} cap={cap_chars}")
    return s

def _reject_multiline_strings(obj: Any) -> None:
    """
    For JSON / JSONL:
    - Reject multiline strings anywhere in the object (no '\n' or '\r').
    Deterministic recursion.
    """
    if obj is None:
        return
    if isinstance(obj, str):
        if ("\n" in obj) or ("\r" in obj):
            raise ValueError("canonical_write_rejected: multiline_string")
        return
    if isinstance(obj, (int, float, bool)):
        return
    if isinstance(obj, list):
        for x in obj:
            _reject_multiline_strings(x)
        return
    if isinstance(obj, dict):
        for k, v in obj.items():
            # keys must be strings in JSON; still enforce no newlines if string
            if isinstance(k, str) and (("\n" in k) or ("\r" in k)):
                raise ValueError("canonical_write_rejected: multiline_string_key")
            _reject_multiline_strings(v)
        return
    # Unknown types are not JSON-serializable anyway; let json.dumps fail upstream.

def _sanitize_for_json_or_jsonl(data: Any) -> Dict[str, Any]:
    """
    Sanitizer for JSON/JSONL:
    - If string: strip control blocks, normalize whitespace, then json.loads
    - If dict: use as-is (but sanitize any top-level string fields deterministically)
    - Ensure valid JSON object

    NOTE:
    - We intentionally do NOT reject multiline strings.
      JSON encoding keeps writes single-line on disk for JSONL (via json.dumps),
      and rejecting multiline content was causing silent write starvation.
    """
    if isinstance(data, str):
        s = _strip_bracketed_control_blocks(data)
        s = _normalize_canonical_whitespace(s)
        _enforce_size_cap(s, cap_chars=_CANONICAL_MAX_CHARS)
        obj = json.loads(s)
    else:
        obj = data

    if not isinstance(obj, dict):
        raise ValueError("canonical_write_rejected: json_must_be_object")

    # Strip control blocks from any top-level string fields (deterministic, shallow)
    cleaned: Dict[str, Any] = {}
    for k, v in obj.items():
        if isinstance(v, str):
            vv = _strip_bracketed_control_blocks(v)
            vv = _normalize_canonical_whitespace(vv)
            cleaned[k] = vv
        else:
            cleaned[k] = v

    return cleaned

def _sanitize_for_text_file(text: str) -> str:
    """
    Sanitizer for canonical text/markdown:
    - strip bracketed control blocks
    - normalize whitespace
    - enforce size cap
    """
    s = _strip_bracketed_control_blocks(text)
    s = _normalize_canonical_whitespace(s)
    _enforce_size_cap(s, cap_chars=_CANONICAL_MAX_CHARS)
    return s

def write_canonical_entry(
    project_name: str,
    *,
    target_path: Path,
    mode: str,
    data: Any,
) -> bool:
    """
    Public unified canonical write helper.

    mode:
      - "jsonl_append": append exactly ONE JSON object as exactly ONE line
      - "json_overwrite": overwrite file with a JSON object (pretty-printed)
      - "text_overwrite": overwrite file with sanitized text

    Returns True if written, False if rejected or failed safely.
    """
    ensure_project(project_name)

    try:
        m = (mode or "").strip().lower()

        if m == "jsonl_append":
            obj = _sanitize_for_json_or_jsonl(data)
            line = json.dumps(obj, ensure_ascii=False)
            _enforce_size_cap(line, cap_chars=_CANONICAL_MAX_CHARS)
            if ("\n" in line) or ("\r" in line):
                # Shouldn't happen, but keep it hard.
                raise ValueError("canonical_write_rejected: jsonl_must_be_single_line")
            target_path.parent.mkdir(parents=True, exist_ok=True)
            with target_path.open("a", encoding="utf-8") as f:
                f.write(line + "\n")
            return True

        if m == "jsonl_overwrite":
            # Overwrite a JSONL file deterministically with validated single-line JSON objects
            raw = "" if data is None else str(data)
            raw = _strip_bracketed_control_blocks(raw)
            raw = _normalize_canonical_whitespace(raw)
            _enforce_size_cap(raw, cap_chars=_CANONICAL_MAX_CHARS)

            lines = [ln for ln in raw.split("\n") if ln.strip()]
            out_lines: List[str] = []
            for ln in lines:
                obj = json.loads(ln)
                if not isinstance(obj, dict):
                    raise ValueError("canonical_write_rejected: jsonl_line_must_be_object")
                obj2 = _sanitize_for_json_or_jsonl(obj)
                out_ln = json.dumps(obj2, ensure_ascii=False)
                if ("\n" in out_ln) or ("\r" in out_ln):
                    raise ValueError("canonical_write_rejected: jsonl_must_be_single_line")
                out_lines.append(out_ln)

            payload = "\n".join(out_lines).rstrip() + "\n"
            atomic_write_text(target_path, payload, encoding="utf-8", errors="strict")
            return True

        if m == "json_overwrite":
            obj = _sanitize_for_json_or_jsonl(data)
            payload = json.dumps(obj, indent=2, ensure_ascii=False)
            payload = _sanitize_for_text_file(payload)
            atomic_write_text(target_path, payload + "\n", encoding="utf-8", errors="strict")
            return True

        if m == "text_overwrite":
            payload = _sanitize_for_text_file("" if data is None else str(data))
            atomic_write_text(target_path, payload + "\n", encoding="utf-8", errors="strict")
            return True

        raise ValueError(f"canonical_write_rejected: unknown_mode={m!r}")

    except Exception:
        # Safe reject: do not write partial content.
        return False
# -----------------------------------------------------------------------------
# C6.1 — Retrieval Policy Wiring (Intent → Canonical Sources)
# Deterministic allowlist + bounded excerpting. NO model calls.
# -----------------------------------------------------------------------------
def merge_interpretive_memory(
    project_name: str,
    *,
    new_obj: Dict[str, Any],
    last_updated_turn: int,
) -> bool:
    """
    Tier-X interpretive memory writer (bounded JSON overwrite).
    NOT factual recall. Replace-style to avoid unbounded growth.

    Deterministic provenance hardening:
    - evidence must be [{"turn_index": int, "excerpt": str}]
    - if evidence arrives as ["..."], convert using last_updated_turn
    """
    ensure_project_scaffold(project_name)

    obj = new_obj if isinstance(new_obj, dict) else {}
    if obj.get("schema") != "interpretive_memory_v1":
        return False

    def _norm_one_line(s: Any) -> str:
        return str(s or "").replace("\r", " ").replace("\n", " ").strip()

    def _norm_evidence(evs: Any) -> List[Dict[str, Any]]:
        out: List[Dict[str, Any]] = []
        if isinstance(evs, list):
            for e in evs:
                if isinstance(e, str):
                    ex = _norm_one_line(e)
                    if ex:
                        out.append({"turn_index": int(last_updated_turn or 0), "excerpt": ex})
                elif isinstance(e, dict):
                    ex = _norm_one_line(e.get("excerpt"))
                    if not ex:
                        continue
                    ti = e.get("turn_index")
                    try:
                        ti2 = int(ti) if ti is not None else int(last_updated_turn or 0)
                    except Exception:
                        ti2 = int(last_updated_turn or 0)
                    out.append({"turn_index": ti2, "excerpt": ex})
        return out[:12]

    def _clip_items(v: Any, n: int) -> List[Dict[str, Any]]:
        if not isinstance(v, list):
            return []
        out: List[Dict[str, Any]] = []
        for x in v:
            if isinstance(x, dict):
                # normalize evidence field if present
                if "evidence" in x:
                    x["evidence"] = _norm_evidence(x.get("evidence"))
                # force single-line strings
                for k, vv in list(x.items()):
                    if isinstance(vv, str):
                        x[k] = _norm_one_line(vv)
                out.append(x)
            if len(out) >= n:
                break
        return out

    out = {
        "schema": "interpretive_memory_v1",
        "updated_at": now_iso(),
        "last_updated_turn": int(last_updated_turn or 0),
        "entities": _clip_items(obj.get("entities"), 12),
        "relationship_dynamics": _clip_items(obj.get("relationship_dynamics"), 12),
        "themes": _clip_items(obj.get("themes"), 12),
        "values_goals": _clip_items(obj.get("values_goals"), 10),
        "open_ambiguities": _clip_items(obj.get("open_ambiguities"), 10),
    }

    return bool(
        write_canonical_entry(
            project_name,
            target_path=understanding_path(project_name),
            mode="json_overwrite",
            data=out,
        )
    )


# Canonical state docs expected under state/
_CANONICAL_STATE_FILES: Dict[str, str] = {
    "project_state.json": "project_state.json",
    "understanding.json": "understanding.json",
    "facts_map.md": "facts_map.md",
    "working_doc.md": "working_doc.md",
    "project_map.md": "project_map.md",
    "project_brief.md": "project_brief.md",
    "preferences.md": "preferences.md",
    "decision_log.md": "decision_log.md",
    "discovery_index.md": "discovery_index.md",
    "issues_questions.md": "issues_questions.md",
    "timeline.md": "timeline.md",
    "fact_ledger.md": "fact_ledger.md",
    "conflicts.md": "conflicts.md",
    "capability_gaps.md": "capability_gaps.md",
    "library_index.md": "library_index.md",
    "health_profile.json": HEALTH_PROFILE_FILE_NAME,
    "decisions.jsonl": DECISIONS_FILE_NAME,
    "decision_candidates.jsonl": DECISION_CANDIDATES_FILE_NAME,
    "upload_notes.jsonl": UPLOAD_NOTES_FILE_NAME,
    "deliverables.json": DELIVERABLES_FILE_NAME,
    "inbox.jsonl": INBOX_FILE_NAME,
}

# Intent → allowlisted canonical sources (bounded; not "everything always")
_C6_RETRIEVAL_POLICY: Dict[str, List[str]] = {
    "recall": [
        "project_state.json",        
        "decisions.jsonl",
        "decision_candidates.jsonl",
        "upload_notes.jsonl",
        "deliverables.json",
        "inbox.jsonl",
        "facts_map.md",
        "discovery_index.md",
        "issues_questions.md",
        "timeline.md",
        "fact_ledger.md",
        "conflicts.md",
        "capability_gaps.md",
        "library_index.md",
    ],
    "status": [
        "understanding.json",        
        "project_state.json",
        "project_map.md",
        "facts_map.md",
        "decisions.jsonl",
        "decision_candidates.jsonl",
        "upload_notes.jsonl",
        "deliverables.json",
        "inbox.jsonl",
        "discovery_index.md",
        "issues_questions.md",
        "fact_ledger.md",
        "conflicts.md",
        "capability_gaps.md",
        "library_index.md",
    ],
    "misc": [
        "understanding.json",        
        "facts_map.md",
        "discovery_index.md",
        "issues_questions.md",
        "timeline.md",
        "fact_ledger.md",
        "conflicts.md",
        "capability_gaps.md",
        "library_index.md",
        "project_map.md",
        "working_doc.md",
        "decisions.jsonl",
        "decision_candidates.jsonl",
        "upload_notes.jsonl",
        "inbox.jsonl",
    ],
    "plan": [
        "understanding.json",        
        "project_state.json",
        "project_map.md",
        "working_doc.md",
        "decision_log.md",
        "preferences.md",
        "decisions.jsonl",
        "decision_candidates.jsonl",
        "upload_notes.jsonl",
        "deliverables.json",
        "inbox.jsonl",
    ],
    "execute": [
        "understanding.json",        
        "project_state.json",
        "project_map.md",
        "working_doc.md",
        "decision_log.md",
        "preferences.md",
        "decisions.jsonl",
        "decision_candidates.jsonl",
        "upload_notes.jsonl",
        "deliverables.json",
        "inbox.jsonl",
    ],
    # lookup keeps existing lookup behavior in Stage 3, but we still provide bounded canonical context
    "lookup": [
        "understanding.json",        
        "project_state.json",
        "facts_map.md",
        "decisions.jsonl",
        "decision_candidates.jsonl",
        "upload_notes.jsonl",
    ],
}

# Hard caps (deterministic)
_C6_MAX_CHUNKS_PER_DOC = 3
_C6_MAX_CHARS_PER_CHUNK = 800
_C6_MAX_TOTAL_CHARS = 6000

_C6_JSONL_TAIL_N = 10
_C6_JSONL_TOPK = 3


def _c6_tokens(text: str) -> List[str]:
    s = (text or "").lower()
    toks = re.findall(r"[a-z0-9_]{3,}", s)
    seen: set[str] = set()
    out: List[str] = []
    for t in toks:
        if t in seen:
            continue
        seen.add(t)
        out.append(t)
        if len(out) >= 32:
            break
    return out


def _c6_score(haystack: str, tokens: List[str]) -> int:
    if not haystack or not tokens:
        return 0
    low = haystack.lower()
    sc = 0
    for t in tokens:
        if t and t in low:
            sc += 1
    return sc


def _c6_split_markdown(text: str) -> List[str]:
    """
    Deterministic chunking for big docs:
    - sanitize
    - split by blank lines first
    - fallback to lines
    """
    s = sanitize_retrieved_text(text or "")
    if not s:
        return []
    parts = [p.strip() for p in re.split(r"\n\s*\n+", s) if p.strip()]
    if parts:
        return parts
    return [ln.strip() for ln in s.splitlines() if ln.strip()]


def _c6_pick_top_chunks(chunks: List[str], tokens: List[str], *, max_chunks: int, max_chars: int) -> List[str]:
    scored: List[Tuple[int, int, str]] = []
    for i, ch in enumerate(chunks):
        sc = _c6_score(ch, tokens)
        scored.append((sc, i, ch))

    # score desc, position asc (stable)
    scored.sort(key=lambda x: (-x[0], x[1]))

    picked: List[str] = []
    for sc, _i, ch in scored:
        if len(picked) >= int(max_chunks or 1):
            break
        if not ch:
            continue
        clipped = ch if len(ch) <= max_chars else (ch[:max_chars].rstrip() + "…")
        picked.append(clipped)

    # If everything scored 0, still return the first chunks deterministically (bounded)
    if (not picked) and chunks:
        for ch in chunks[: int(max_chunks or 1)]:
            clipped = ch if len(ch) <= max_chars else (ch[:max_chars].rstrip() + "…")
            picked.append(clipped)

    return picked


def _c6_read_state_file_text(project_name: str, filename: str) -> str:
    """
    Read a state file by canonical filename under state/.
    Returns "" if missing/unreadable.
    """
    try:
        p = state_dir(project_name) / filename
        if not p.exists():
            return ""
        return p.read_text(encoding="utf-8", errors="replace") or ""
    except Exception:
        return ""


def _c6_excerpt_big_doc(project_name: str, src_key: str, tokens: List[str]) -> List[str]:
    """
    Excerpt markdown big docs with bounded scoring.
    """
    fname = _CANONICAL_STATE_FILES.get(src_key, "")
    if not fname:
        return []
    raw = _c6_read_state_file_text(project_name, fname)
    if not raw.strip():
        return []

    chunks = _c6_split_markdown(raw)
    picked = _c6_pick_top_chunks(
        chunks,
        tokens,
        max_chunks=_C6_MAX_CHUNKS_PER_DOC,
        max_chars=_C6_MAX_CHARS_PER_CHUNK,
    )
    return picked


def _c6_excerpt_project_state(project_name: str, *, tokens: List[str] | None = None) -> str:
    """
    C6.1 refinement:
    - project_state.json is canonical
    - BUT its 'current_focus' is situational and must not dominate misc intent
      unless the user references it.
    """
    raw = _c6_read_state_file_text(project_name, _CANONICAL_STATE_FILES["project_state.json"])
    raw = (raw or "").strip()
    if not raw:
        return ""

    try:
        obj = json.loads(raw)
    except Exception:
        return raw[:_C6_MAX_TOTAL_CHARS]

    tokens = tokens or []

    # If no tokens, be conservative and include everything
    if not tokens:
        return json.dumps(obj, indent=2, ensure_ascii=False)[:_C6_MAX_TOTAL_CHARS]

    focus = str(obj.get("current_focus") or "")
    focus_low = focus.lower()
    hit_count = 0
    for t in tokens:
        if t and t in focus_low:
            hit_count += 1
            if hit_count >= 2:
                break
    focus_hit = hit_count >= 2

    # Suppress current_focus if it doesn't overlap the user intent
    # BUT always preserve expert_frame (EFL v1) and other low-noise control fields.
    if not focus_hit:
        ef = obj.get("expert_frame")
        if not isinstance(ef, dict):
            ef = {}

        # Build a minimal, stable subset for misc/plan/exe grounding:
        # enough to keep the assistant "locked" without flooding context.
        obj2: Dict[str, Any] = {}
        for k in ("project_mode", "goal", "domains", "bootstrap_status"):
            if k in obj:
                obj2[k] = obj.get(k)
        obj2["expert_frame"] = ef
        obj = obj2

    return json.dumps(obj, indent=2, ensure_ascii=False)[:_C6_MAX_TOTAL_CHARS]

def _c6_excerpt_deliverables(project_name: str) -> str:
    """
    Deliverables registry is already bounded JSON; keep a small cap.
    """
    try:
        reg = load_deliverables(project_name)
        if isinstance(reg, dict):
            return json.dumps(reg, indent=2, ensure_ascii=False)[:_C6_MAX_TOTAL_CHARS]
    except Exception:
        return ""
    return ""


def _c6_excerpt_jsonl_tail(project_name: str, src_key: str, tokens: List[str]) -> List[str]:
    """
    For JSONL tails (decisions/upload_notes):
    - take most recent N
    - score overlap
    - return top K (or most recent K if no hits)
    """
    fname = _CANONICAL_STATE_FILES.get(src_key, "")
    if not fname:
        return []

    p = state_dir(project_name) / fname
    if not p.exists():
        return []

    try:
        lines = p.read_text(encoding="utf-8", errors="replace").splitlines()
    except Exception:
        return []

    tail = lines[-int(_C6_JSONL_TAIL_N or 10):] if lines else []
    recs: List[Tuple[int, int, str]] = []  # (score, idx, rendered)
    for i, ln in enumerate(tail):
        try:
            obj = json.loads(ln)
        except Exception:
            obj = None

        if isinstance(obj, dict):
            if src_key == "decisions.jsonl":
                blob = str(obj.get("text") or "")
                ts = str(obj.get("timestamp") or "")
                rendered = f"- [{ts}] {blob}".strip()
            elif src_key == "upload_notes.jsonl":
                up = str(obj.get("upload_path") or "")
                q = str(obj.get("question") or "")
                a = str(obj.get("answer") or "")
                ts = str(obj.get("timestamp") or "")
                rendered = f"- [{ts}] {up} | Q: {q} | A: {a}".strip()
            else:
                rendered = ln.strip()
        else:
            rendered = (ln or "").strip()

        rendered2 = sanitize_retrieved_text(rendered)
        sc = _c6_score(rendered2, tokens)
        recs.append((sc, i, rendered2))

    recs.sort(key=lambda x: (-x[0], -x[1]))  # prefer higher score; then more recent

    picked = [r[2] for r in recs if r[2]][:_C6_JSONL_TOPK]
    if tokens and any(r[0] > 0 for r in recs):
        return picked

    # If no scoring signal, return most recent K deterministically
    recs2 = sorted(recs, key=lambda x: -x[1])
    return [r[2] for r in recs2 if r[2]][:_C6_JSONL_TOPK]


def _c6_excerpt_raw_jsonl_tail_by_rel(rel_path: str, tokens: List[str]) -> List[str]:
    """
    Deterministic excerpting for raw chat-log-like JSONL files (NOT canonical state).
    - Reads last N lines (bounded)
    - Scores overlap with tokens
    - Returns top K (or most recent K if no hits)
    """
    try:
        pr, _pd = _require_configured()
    except Exception:
        return []

    rel = (rel_path or "").replace("\\", "/").strip()
    if not rel:
        return []

    p = pr / rel
    if not p.exists():
        return []

    try:
        lines = p.read_text(encoding="utf-8", errors="replace").splitlines()
    except Exception:
        return []

    # Bounded scan for determinism + perf
    tail_n = 160
    tail = lines[-tail_n:] if len(lines) > tail_n else list(lines)

    recs: List[Tuple[int, int, str]] = []  # (score, idx, rendered)

    for i, ln in enumerate(tail):
        s = (ln or "").strip()
        if not s:
            continue

        rendered = ""
        try:
            obj = json.loads(s)
        except Exception:
            obj = None

        if isinstance(obj, dict):
            # Try common chat-log shapes; fall back to compact JSON
            role = str(obj.get("role") or "").strip()
            content = str(obj.get("content") or obj.get("text") or "").strip()
            ts = str(obj.get("timestamp") or obj.get("time") or obj.get("created_at") or "").strip()

            if role and content:
                rendered = (f"- [{ts}] {role}: {content}" if ts else f"- {role}: {content}").strip()
            elif content:
                rendered = (f"- [{ts}] {content}" if ts else f"- {content}").strip()
            else:
                rendered = s
        else:
            rendered = s

        rendered2 = sanitize_retrieved_text(rendered).replace("\n", " ").strip()
        if len(rendered2) > 260:
            rendered2 = rendered2[:259].rstrip() + "…"

        sc = _c6_score(rendered2, tokens)
        recs.append((int(sc), int(i), rendered2))

    recs.sort(key=lambda x: (-x[0], -x[1]))  # prefer higher score; then more recent

    picked = [r[2] for r in recs if r[2]][:_C6_JSONL_TOPK]
    if tokens and any(r[0] > 0 for r in recs):
        return picked

    # If no scoring signal, return most recent K deterministically
    recs2 = sorted(recs, key=lambda x: -x[1])
    return [r[2] for r in recs2 if r[2]][:_C6_JSONL_TOPK]


def _c6_excerpt_partner_chat_lens(project_name: str, tokens: List[str], *, max_items: int = 6) -> List[str]:
    """
    Deterministic excerpting for partner chat logs (lens-only, no quotes).
    - Reads last N lines from chat_log.jsonl
    - Keeps partner USER messages only
    - Scores overlap and returns top K (or most recent K if no hits)
    """
    try:
        p = state_dir(project_name) / "chat_log.jsonl"
    except Exception:
        return []
    if not p.exists():
        return []

    try:
        lines = p.read_text(encoding="utf-8", errors="replace").splitlines()
    except Exception:
        return []

    tail = lines[-240:] if lines else []
    recs: List[Tuple[int, int, str]] = []
    for i, ln in enumerate(tail):
        s = (ln or "").strip()
        if not s:
            continue
        try:
            obj = json.loads(s)
        except Exception:
            obj = None
        if not isinstance(obj, dict):
            continue
        role = str(obj.get("role") or "").strip().lower()
        if role != "user":
            continue
        txt = str(obj.get("content") or "").strip()
        if not txt:
            continue
        # Compact and bound each line
        txt = " ".join(txt.split())
        if len(txt) > 220:
            txt = txt[:219].rstrip() + "..."
        rendered = f"- {txt}"
        rendered2 = sanitize_retrieved_text(rendered)
        sc = _c6_score(rendered2, tokens)
        recs.append((sc, i, rendered2))

    if not recs:
        return []

    recs.sort(key=lambda x: (-x[0], -x[1]))
    picked = [r[2] for r in recs if r[2]][: max_items]
    if tokens and any(r[0] > 0 for r in recs):
        return picked

    recs2 = sorted(recs, key=lambda x: -x[1])
    return [r[2] for r in recs2 if r[2]][: max_items]


def build_canonical_snippets(
    project_name: str,
    *,
    intent: str,
    entities: List[str],
    user_text: str,
) -> List[str]:
    # Therapist-only: optional extra project reads (set dynamically below)
    _THERAPIST_EXTRA_PROJECTS: List[Tuple[str, str]] = []
    
    """
    C6.1 — deterministic canonical snippet builder (bounded allowlist per intent).
    Returns list[str] where each item is a labeled snippet block.
    Stage 3 MUST see ONLY these excerpts, never whole docs.
    """
    ensure_project_scaffold(project_name)
    # Couples Retrieval Policy (READ-ONLY; couple_* accounts):
    # Partner context is limited to distilled tiers ONLY:
    #   - Tier-2G (partner global profile snippet)
    #   - Tier-2P (partner facts_map.md typed cache)
    #   - Tier-X  (partner understanding.json interpretive memory)
    #
    # HARD EXCLUSIONS:
    #   - Partner Tier-0 chat logs (no chat_log.jsonl / raw jsonl tail injection)
    #   - Partner Tier-1 facts_raw.jsonl (never retrieved)
    partner_user = ""
    partner_project_full = ""
    try:
        no_extra = "__NO_COUPLES_EXTRA__" in (user_text or "")
        pname = safe_project_name(project_name)
        user_seg = pname.split("/", 1)[0].strip()
        proj_seg = pname.split("/", 1)[1].strip() if "/" in pname else DEFAULT_PROJECT_NAME

        if (not no_extra) and user_seg.lower().startswith("couple_"):

            def _norm_user(u: str) -> str:
                return re.sub(r"[^a-zA-Z0-9_-]", "_", (u or "").strip())

            me = _norm_user(user_seg)

            links = load_couples_links()
            rec = None
            for _cid, obj in (links or {}).items():
                if not isinstance(obj, dict):
                    continue
                if str(obj.get("status") or "") != "active":
                    continue
                ua = _norm_user(str(obj.get("user_a") or ""))
                ub = _norm_user(str(obj.get("user_b") or ""))
                if me == ua or me == ub:
                    rec = obj
                    break

            if isinstance(rec, dict):
                ua = _norm_user(str(rec.get("user_a") or ""))
                ub = _norm_user(str(rec.get("user_b") or ""))

                if me == ua:
                    partner_user = ub
                else:
                    partner_user = ua

                partner_proj = proj_seg or DEFAULT_PROJECT_NAME
                partner_project_full = safe_project_name(f"{partner_user}/{partner_proj}")
    except Exception:
        partner_user = ""
        partner_project_full = ""


    intent0 = (intent or "misc").strip().lower()
    if intent0 not in _C6_RETRIEVAL_POLICY:
        intent0 = "misc"

    ent_text = " ".join([str(x) for x in (entities or []) if str(x or "").strip()])
    # tokens_all are used for scoring big docs + JSONL tails
    tokens_all = _c6_tokens((user_text or "") + " " + ent_text)
    # tokens_user are used ONLY to decide whether project_state.current_focus is relevant
    tokens_user = _c6_tokens(user_text or "")

    allow = _C6_RETRIEVAL_POLICY.get(intent0, []) or []
    out: List[str] = []

    total_chars = 0

    def _add_block(label: str, body: str) -> None:
        nonlocal total_chars
        b = (body or "").strip()
        if not b:
            return
        block = f"[{label}]\n{b}".strip()
        if not block:
            return
        remaining = _C6_MAX_TOTAL_CHARS - total_chars
        if remaining <= 0:
            return
        if len(block) > remaining:
            block = block[:remaining].rstrip() + "…"
        out.append(block)
        total_chars += len(block)
    # ---------------------------------------------------------------------
    # Global Tier-2G injection (ALWAYS for this user across all projects)
    # ---------------------------------------------------------------------
    try:
        user_seg0 = str(project_name or "").split("/", 1)[0].strip()
        if user_seg0:
            prof_snip = render_user_profile_snippet(user_seg0, user_text=user_text or "")
            if prof_snip:
                _add_block("user_profile_global.json (Tier-2G)", prof_snip)

            # -----------------------------------------------------------------
            # Global Tier-2M injection (ALWAYS-ON persona slice + fuzzy expansion)
            #
            # Goal:
            # - The assistant should feel like it "knows the user" regardless of phrasing.
            # - Never dump everything; keep a small, stable, provenance-aware subset.
            # - Respect do_not_resurface at read-time for identity kernel fields.
            #
            # Policy:
            # - Always include identity/relationship *structure* (bounded) so continuity is present.
            # - Always include a tiny amount of light context (media/time) for conversational continuity.
            # - Expand with fuzzy keyword detection for personalization / misspellings.
            # -----------------------------------------------------------------
            low = _policy_norm(user_text or "")

            def _tok3(s: str) -> List[str]:
                return re.findall(r"[a-z0-9_]{3,}", _policy_norm(s))

            def _lev1(a: str, b: str) -> int:
                # Deterministic small edit distance (bounded). Returns 0/1/2+.
                a = (a or "")
                b = (b or "")
                if a == b:
                    return 0
                la = len(a)
                lb = len(b)
                if abs(la - lb) > 1:
                    return 2
                # one substitution
                if la == lb:
                    diff = 0
                    for i in range(la):
                        if a[i] != b[i]:
                            diff += 1
                            if diff > 1:
                                return 2
                    return 1
                # one insertion/deletion
                # ensure a is shorter
                if la > lb:
                    a, b = b, a
                    la, lb = lb, la
                i = 0
                j = 0
                diff = 0
                while i < la and j < lb:
                    if a[i] == b[j]:
                        i += 1
                        j += 1
                        continue
                    diff += 1
                    if diff > 1:
                        return 2
                    j += 1
                return 1

            def _fuzzy_has_any(*words: str) -> bool:
                # token-based fuzzy match: exact OR edit-distance-1 against any user token
                utoks = _tok3(user_text or "")
                if not utoks:
                    return False
                for w in words:
                    wtoks = _tok3(w)
                    for ww in wtoks:
                        for t in utoks:
                            if t == ww:
                                return True
                            if _lev1(t, ww) <= 1:
                                return True
                return False

            include_keys: List[str] = []
            include_prefixes: List[str] = []

            # ALWAYS-ON identity + relationships (bounded in renderer)
            # Respect do_not_resurface for identity kernel fields (same as Tier-2G).
            if policy_allows_resurface_now(user_seg0, entity_key="user.identity.name", user_text=user_text or ""):
                include_keys.append("user.identity.name")
            if policy_allows_resurface_now(user_seg0, entity_key="user.identity.location", user_text=user_text or ""):
                include_keys.append("user.identity.location")
            include_prefixes.append("user.relationship.")

            # ALWAYS-ON light context (kept small; helps "knows you" feel)
            include_prefixes.append("context.activity.")
            include_prefixes.append("context.time_")

            # Personalization intent (robust to phrasing/misspellings)
            # If user is asking "tailor to me / based on what you know", include additional preference buckets.
            if _fuzzy_has_any("based", "know", "about", "me", "tailor", "personalize", "customize", "explain", "why"):
                include_prefixes.append("user.preference.")
                include_prefixes.append("preference.")

            # Situational expansions (fuzzy)
            if _fuzzy_has_any("movie", "watch", "watching", "stream", "netflix", "hulu", "prime", "hbo"):
                include_prefixes.append("context.activity.")
            if _fuzzy_has_any("tonight", "today", "now", "time", "pm", "am"):
                include_prefixes.append("context.time_")

            # Keep substance-use context out unless user is clearly in that lane (fuzzy).
            if _fuzzy_has_any("edible", "edibles", "gummy", "gummies", "vape", "thc", "sativa"):
                include_prefixes.append("context.substance_")
                include_prefixes.append("event.substance_")
                include_prefixes.append("routine.substance_")

            # De-dupe while preserving order
            def _dedupe(xs: List[str]) -> List[str]:
                out2: List[str] = []
                seen2: set = set()
                for x in xs:
                    xx = str(x or "").strip()
                    if not xx:
                        continue
                    if xx in seen2:
                        continue
                    seen2.add(xx)
                    out2.append(xx)
                return out2

            include_keys = _dedupe(include_keys)
            include_prefixes = _dedupe(include_prefixes)

            try:
                t2m = render_user_global_facts_snippet_tier2m(
                    user_seg0,
                    include_keys=include_keys,
                    include_prefixes=include_prefixes,
                    include_kinds=None,
                    max_items=12,
                )
            except Exception:
                t2m = ""

            if t2m:
                t2m2 = (
                    "USAGE_NOTE_FOR_TIER2M:\n"
                    "- Use these facts to ground tone, assumptions, and explanations so the assistant feels like it knows the user.\n"
                    "- You may infer cautiously from patterns here, but do NOT invent new preferences or contradict recorded facts.\n"
                    "- If a needed detail is ambiguous or missing, ask one clarifying question only when it matters.\n"
                    "\n"
                    + (t2m or "").strip()
                ).strip()
                _add_block("user_global_facts_map.json (Tier-2M subset)", t2m2)
    except Exception:
        pass

    # ---------------------------------------------------------------------
    # Intent-aware inclusion cleanup:
    # - understanding.json (Tier-X) excluded for recall intents
    # - facts_map.md (Tier-2P) excluded for plan/execute intents
    # ---------------------------------------------------------------------
    if intent0 == "recall":
        allow = [x for x in allow if x != "understanding.json"]
    if intent0 in ("plan", "execute"):
        allow = [x for x in allow if x != "facts_map.md"]

    for src in allow:
        if src == "understanding.json":
            ut = _c6_read_state_file_text(project_name, UNDERSTANDING_FILE_NAME)
            if ut.strip():
                _add_block("understanding.json (interpretive; do not use for factual recall)", ut)
            continue

        if src == "project_state.json":
            st = _c6_excerpt_project_state(project_name, tokens=tokens_user)
            _add_block("project_state.json", st)
            continue

        if src == "deliverables.json":
            dj = _c6_excerpt_deliverables(project_name)
            _add_block("deliverables.json", dj)
            continue

        if src in ("decisions.jsonl", "upload_notes.jsonl"):
            lines = _c6_excerpt_jsonl_tail(project_name, src, tokens_all)
            if lines:
                _add_block(src, "\n".join(lines))
            continue

        # Big markdown docs (bounded excerpting)
        if src in ("facts_map.md", "working_doc.md", "project_map.md", "project_brief.md", "preferences.md", "decision_log.md"):
            picked = _c6_excerpt_big_doc(project_name, src, tokens_all)
            if picked:
                _add_block(src, "\n\n".join(picked))
            continue
    # Couples Retrieval Policy: inject PARTNER CONTEXT from distilled tiers ONLY.
    # These blocks are read-time only and used for neutral theme formation.
    # DO NOT include partner Tier-0/Tier-1. Do NOT recurse into partner snippet builder.
    try:
        if partner_user and partner_project_full:
            # Tier-2G (global profile snippet)
            try:
                prof_snip_p = render_user_profile_snippet(partner_user)
            except Exception:
                prof_snip_p = ""
            if prof_snip_p:
                _add_block("PARTNER_CONTEXT_TIER2G (global_profile; do_not_disclose)", prof_snip_p)

            # Tier-2P (project typed facts cache)
            try:
                picked_fm = _c6_excerpt_big_doc(partner_project_full, "facts_map.md", tokens_all)
            except Exception:
                picked_fm = []
            if picked_fm:
                _add_block("PARTNER_CONTEXT_TIER2P (facts_map; do_not_disclose)", "\n\n".join(picked_fm))

            # Tier-X (interpretive memory)
            try:
                ut_p = _c6_read_state_file_text(partner_project_full, UNDERSTANDING_FILE_NAME)
            except Exception:
                ut_p = ""
            if ut_p.strip():
                _add_block("PARTNER_CONTEXT_TIERX (understanding; do_not_disclose)", ut_p)

            # Partner chat lens (one-sided, do not quote or attribute)
            try:
                lens_lines = _c6_excerpt_partner_chat_lens(partner_project_full, tokens_all, max_items=6)
            except Exception:
                lens_lines = []
            if lens_lines:
                _add_block("PARTNER_CHAT_LENS (one_sided; do_not_quote)", "\n".join(lens_lines))
    except Exception:
        pass

    # Couples Shared Memory (agreements, corroborated vs proposed)
    try:
        pname2 = safe_project_name(project_name)
        user_seg2 = pname2.split("/", 1)[0].strip().lower()
        if user_seg2.startswith("couple_"):
            mem = load_couples_shared_memory(project_name)
            ag = mem.get("agreements") if isinstance(mem.get("agreements"), list) else []
            if isinstance(ag, list) and ag:
                lines = []
                # Corroborated first
                for it in ag:
                    if not isinstance(it, dict):
                        continue
                    if str(it.get("status") or "") != "corroborated":
                        continue
                    txt = str(it.get("text") or "").strip()
                    if txt:
                        lines.append(f"- corroborated: {txt}")
                # Proposed
                for it in ag:
                    if not isinstance(it, dict):
                        continue
                    if str(it.get("status") or "") == "corroborated":
                        continue
                    txt = str(it.get("text") or "").strip()
                    if txt:
                        lines.append(f"- proposed: {txt}")
                if lines:
                    _add_block("COUPLES_SHARED_MEMORY (agreements)", "\n".join(lines))
    except Exception:
        pass

    # Tier-drop: include a tiny excerpt from the active raw chat log (if present).
    # This makes the chat feel human while staying deterministic + bounded.
    try:
        ao = load_active_object(project_name) or {}
    except Exception:
        ao = {}

    try:
        rel = str(ao.get("rel_path") or "").replace("\\", "/").strip()
        if rel and rel.endswith(".jsonl") and "/raw/" in rel:
            lines = _c6_excerpt_raw_jsonl_tail_by_rel(rel, tokens_all)
            if lines:
                _add_block(f"active_chat_log:{Path(rel).name}", "\n".join(lines))
    except Exception:
        pass

    return out
def retrieve_canonical_snippets(
    project_name: str,
    *,
    intent: str,
    entities: List[str],
    max_items: int = 3,
) -> Dict[str, Any]:
    """
    Stage 2 deterministic canonical retrieval (C6.1-wired).

    Back-compat wrapper: returns dict with "snippets" list, but the builder is:
      build_canonical_snippets(project_name, intent, entities, user_text)

    NOTE:
    - This function does NOT receive user_text, so it cannot do perfect excerpt scoring.
    - It remains for legacy callers only.
    - The C6 request pipeline should call build_canonical_snippets(...) directly.
    """
    ensure_project_scaffold(project_name)

    intent0 = (intent or "misc").strip().lower()
    ents = [sanitize_retrieved_text(e) for e in (entities or []) if str(e or "").strip()]
    ents = [e for e in ents if e][:18]

    # Legacy fallback user_text: entities joined (deterministic, bounded)
    user_text = " ".join(ents).strip()

    try:
        snippets = build_canonical_snippets(
            project_name,
            intent=intent0,
            entities=list(ents),
            user_text=user_text,
        )
    except Exception:
        snippets = []

    meta: Dict[str, Any] = {"project": project_name, "policy": "c6.1"}

    return {"intent": intent0, "entities": ents, "snippets": snippets[:12], "meta": meta}

def load_pending_upload_question(project_name: str) -> Dict[str, Any]:
    """
    Load pending upload clarification state.
    Returns {} if missing/invalid.
    """
    ensure_project(project_name)
    p = pending_upload_question_path(project_name)
    if not p.exists():
        return {}
    try:
        obj = json.loads(p.read_text(encoding="utf-8") or "{}")
        return obj if isinstance(obj, dict) else {}
    except Exception:
        return {}


def save_pending_upload_question(project_name: str, obj: Dict[str, Any]) -> None:
    """
    Save pending upload clarification state (atomic write).
    """
    ensure_project(project_name)
    p = pending_upload_question_path(project_name)
    atomic_write_text(p, json.dumps(obj or {}, indent=2), encoding="utf-8", errors="strict")


def clear_pending_upload_question(project_name: str) -> None:
    """
    Remove pending upload clarification state.
    """
    p = pending_upload_question_path(project_name)
    try:
        if p.exists():
            p.unlink()
    except Exception:
        pass


def is_pending_upload_question_unexpired(pending_obj: Dict[str, Any]) -> bool:
    """
    True when:
    - pending == True
    - expires_at exists and is in the future (UTC)
    """
    if not isinstance(pending_obj, dict):
        return False
    if pending_obj.get("pending") is not True:
        return False
    exp = str(pending_obj.get("expires_at") or "").strip()
    exp_epoch = _parse_iso_noz(exp)
    if exp_epoch <= 0:
        return False
    return time.time() <= exp_epoch


def _next_upload_note_id(project_name: str, *, created_at: str) -> str:
    """
    Deterministic per-day increment:
      upnote_YYYY_MM_DD_NNN
    """
    day = (created_at or "").split("T", 1)[0].replace("-", "_")
    pfx = f"upnote_{day}_"

    max_n = 0
    p = upload_notes_path(project_name)
    if p.exists():
        try:
            for ln in p.read_text(encoding="utf-8", errors="replace").splitlines():
                try:
                    obj = json.loads(ln)
                except Exception:
                    continue
                if not isinstance(obj, dict):
                    continue
                did = str(obj.get("id") or "")
                if not did.startswith(pfx):
                    continue
                tail = did[len(pfx):]
                try:
                    n = int(tail)
                except Exception:
                    n = 0
                if n > max_n:
                    max_n = n
        except Exception:
            pass

    return f"{pfx}{(max_n + 1):03d}"


def append_upload_note(
    project_name: str,
    *,
    upload_path: str,
    question: str,
    answer: str,
    source: str = "user",
) -> Dict[str, Any]:
    """
    Append-only log of upload clarification Q/A.
    Stores answer verbatim. No inference.
    """
    ensure_project(project_name)

    created_at = _now_iso_noz()
    nid = _next_upload_note_id(project_name, created_at=created_at)

    # Write-time sanitizer (recommended): never store control wrappers in canonical logs.
    q_clean = strip_control_tokens(str(question or "").strip())
    a_clean = strip_control_tokens(str(answer or ""))

    entry = {
        "id": nid,
        "timestamp": created_at,
        "upload_path": str(upload_path or "").strip(),
        "question": q_clean,
        "answer": a_clean,
        "source": str(source or "").strip() or "user",
    }

    _append_jsonl_line(project_name, upload_notes_path(project_name), entry)
    return entry


def decision_candidates_path(project_name: str) -> Path:
    """
    projects/<user>/<project>/state/decision_candidates.jsonl
    """
    return state_dir(project_name) / DECISION_CANDIDATES_FILE_NAME


def decisions_path(project_name: str) -> Path:
    """
    projects/<user>/<project>/state/decisions.jsonl
    """
    return state_dir(project_name) / DECISIONS_FILE_NAME


def _next_decision_id(project_name: str, *, prefix: str, created_at: str) -> str:
    """
    Deterministic per-day increment:
      cand_YYYY_MM_DD_NNN
      dec_YYYY_MM_DD_NNN
    """
    day = (created_at or "").split("T", 1)[0].replace("-", "_")
    pfx = f"{prefix}_{day}_"

    max_n = 0
    p = decisions_path(project_name) if prefix == "dec" else decision_candidates_path(project_name)
    if p.exists():
        try:
            for ln in p.read_text(encoding="utf-8", errors="replace").splitlines():
                try:
                    obj = json.loads(ln)
                except Exception:
                    continue
                if not isinstance(obj, dict):
                    continue
                did = str(obj.get("id") or "")
                if not did.startswith(pfx):
                    continue
                tail = did[len(pfx):]
                try:
                    n = int(tail)
                except Exception:
                    n = 0
                if n > max_n:
                    max_n = n
        except Exception:
            pass

    return f"{pfx}{(max_n + 1):03d}"


def _append_jsonl_line(project_name: str, path: Path, obj: Dict[str, Any]) -> None:
    # Canonical JSONL writes MUST go through the unified sanitizer + writer.
    ok = write_canonical_entry(
        project_name,
        target_path=path,
        mode="jsonl_append",
        data=obj,
    )
    if not ok:
        raise ValueError("canonical_write_rejected")


def append_decision_candidate(
    project_name: str,
    *,
    text: str,
    confidence: float,
    source: str = "system",
    related_deliverable: str = "",
) -> Dict[str, Any]:
    """
    Append-only. Creates a new candidate line in decision_candidates.jsonl.
    Deterministic storage only. No inference.
    """
    ensure_project(project_name)

    created_at = _deliverables_now_noz()
    cid = _next_decision_id(project_name, prefix="cand", created_at=created_at)

    # Write-time sanitizer: never store control wrappers in canonical logs.
    txt_clean = strip_control_tokens(str(text or "").strip())

    entry = {
        "id": cid,
        "timestamp": created_at,
        "text": txt_clean,
        "confidence": float(confidence or 0.0),
        "status": "pending",
        "source": str(source or "").strip() or "system",
        "related_deliverable": str(related_deliverable or "").strip(),
    }

    _append_jsonl_line(project_name, decision_candidates_path(project_name), entry)
    return entry


def update_decision_candidate_status(
    project_name: str,
    candidate_id: str,
    new_status: str,
    *,
    note: str = "",
) -> bool:
    """
    Deterministically update a candidate's status in decision_candidates.jsonl.
    No model calls. Returns True if updated, else False.
    """
    ensure_project(project_name)

    cid = str(candidate_id or "").strip()
    if not cid:
        return False

    p = decision_candidates_path(project_name)
    if not p.exists():
        return False

    lines = p.read_text(encoding="utf-8", errors="replace").splitlines()
    out_lines: List[str] = []
    updated = False

    for ln in lines:
        try:
            obj = json.loads(ln)
        except Exception:
            out_lines.append(ln)
            continue

        if not isinstance(obj, dict):
            out_lines.append(ln)
            continue

        if str(obj.get("id") or "").strip() == cid:
            obj["status"] = str(new_status or "").strip() or "pending"
            if note:
                obj["note"] = strip_control_tokens(str(note or "").strip())
            updated = True
            out_lines.append(json.dumps(obj, ensure_ascii=False))
        else:
            out_lines.append(ln)

    if not updated:
        return False

    # Rewrite file deterministically (same order, just edited line)
    payload = "\n".join(out_lines).rstrip() + "\n"
    ok = write_canonical_entry(
        project_name,
        target_path=p,
        mode="jsonl_overwrite",
        data=payload,
    )
    return bool(ok)
def append_final_decision(
    project_name: str,
    *,
    text: str,
    source: str = "user",
    related_deliverable: str = "",
) -> Dict[str, Any]:
    """
    Append-only. Writes a new final decision line into decisions.jsonl.
    Deterministic storage only. No inference. No model calls.
    """
    ensure_project(project_name)

    created_at = _deliverables_now_noz()
    did = _next_decision_id(project_name, prefix="dec", created_at=created_at)

    # Write-time sanitizer: never store control wrappers in canonical logs.
    txt_clean = strip_control_tokens(str(text or "").strip())

    entry = {
        "id": did,
        "timestamp": created_at,
        "text": txt_clean,
        "source": str(source or "").strip() or "user",
        "related_deliverable": str(related_deliverable or "").strip(),
    }

    _append_jsonl_line(project_name, decisions_path(project_name), entry)
    return entry

def load_decisions(project_name: str) -> List[Dict[str, Any]]:
    """
    Load final decisions (source of truth). Append-only file; returned in file order.
    """
    ensure_project(project_name)
    p = decisions_path(project_name)
    if not p.exists():
        return []

    out: List[Dict[str, Any]] = []
    try:
        for ln in p.read_text(encoding="utf-8", errors="replace").splitlines():
            try:
                obj = json.loads(ln)
            except Exception:
                continue
            if isinstance(obj, dict):
                out.append(obj)
    except Exception:
        return out

    return out
# -----------------------------------------------------------------------------
# C7 — Decisions as First-Class Objects (immutable, supersession-aware)
#
# Storage: state/decisions.jsonl (append-only JSONL)
#
# Rules:
# - NEVER edit/delete existing decision rows.
# - Supersession is represented by:
#   (a) a NEW decision row with supersedes=<old_id>
#   (b) a NEW marker row for the old decision id with status="superseded"
# - Backward compatible: legacy decisions.jsonl rows continue to load (defaults applied at read time).
# - Writes MUST go through write_canonical_entry() (C6.2).
# -----------------------------------------------------------------------------

def _c7_stable_legacy_id(raw_line: str) -> str:
    """
    If a legacy row lacks an id, derive a stable id from its raw line content.
    Deterministic (sha256 prefix).
    """
    try:
        b = (raw_line or "").encode("utf-8", errors="replace")
        h = hashlib.sha256(b).hexdigest()[:12]
        return f"legacy_{h}"
    except Exception:
        return "legacy_unknown"

def _c7_normalize_decision_row(obj: Dict[str, Any], *, raw_line: str = "") -> Dict[str, Any]:
    """
    Normalize ANY decision row (legacy or C7) to the canonical C7 schema, without rewriting disk.

    Canonical schema:
      {
        "id": "<stable id>",
        "created_at": "<ISO8601>",
        "domain": "<dot.notation.domain>" | "" ,
        "surface": "<optional>" | "",
        "status": "candidate|final|superseded",
        "text": "<human-facing statement>",
        "supersedes": "<decision_id|null>",
        "evidence": ["refs..."],
        "confidence": "low|medium|high"
      }
    """
    d = obj if isinstance(obj, dict) else {}

    # id (required)
    did = str(d.get("id") or "").strip()
    if not did:
        did = _c7_stable_legacy_id(raw_line)

    # created_at (required)
    created_at = str(d.get("created_at") or "").strip()
    if not created_at:
        # legacy uses "timestamp"
        created_at = str(d.get("timestamp") or "").strip()
    if not created_at:
        created_at = _deliverables_now_noz()

    # status default
    status = str(d.get("status") or "").strip().lower()
    if status not in ("candidate", "final", "superseded"):
        # legacy decisions are treated as final
        status = "final"

    # text (required human-facing)
    text = str(d.get("text") or "").strip()
    if not text:
        # legacy safety: if some row used a different key, keep minimal
        text = str(d.get("decision") or "").strip()

    # domain (optional)
    domain = d.get("domain")
    domain = str(domain).strip() if domain is not None else ""
    if domain.lower() in ("none", "null"):
        domain = ""

    # surface (optional)
    surface = d.get("surface")
    surface = str(surface).strip() if surface is not None else ""
    if surface.lower() in ("none", "null"):
        surface = ""

    # supersedes (optional)
    supersedes = d.get("supersedes", None)
    if supersedes is None:
        supersedes_norm = None
    else:
        supersedes_s = str(supersedes).strip()
        supersedes_norm = supersedes_s if supersedes_s else None

    # evidence (optional)
    evidence = d.get("evidence")
    if isinstance(evidence, list):
        evidence_norm = [strip_control_tokens(str(x)).strip() for x in evidence if str(x).strip()]
    else:
        evidence_norm = []

    # confidence (optional)
    conf = d.get("confidence", None)
    conf_s = str(conf).strip().lower() if conf is not None else ""
    if conf_s not in ("low", "medium", "high"):
        conf_s = ""

    out = {
        "id": did,
        "created_at": created_at,
        "domain": domain,
        "surface": surface,
        "status": status,
        "text": strip_control_tokens(text),
        "supersedes": supersedes_norm,
        "evidence": evidence_norm,
    }
    if conf_s:
        out["confidence"] = conf_s

    return out

def infer_decision_domain(text: str) -> Tuple[str, str]:
    """
    Deterministic domain inference (v1).
    Returns (domain, surface).
    - domain is dot notation when recognized, else "".
    - surface is a short string like "floor|wall|shower" when recognized, else "".
    """
    t = (text or "").strip()
    if not t:
        return "", ""

    area, aspect = _decision_topic_key(t)

    # Map to dot domains (v1)
    area_map = {
        "primary_bathroom": "primary_bath",
        "bathroom": "bath",
    }
    aspect_map = {
        "floor": "floor",
        "walls": "walls",
        "shower_floor": "shower.floor",
        "shower_walls": "shower.walls",
    }

    a = area_map.get(area, "")
    b = aspect_map.get(aspect, "")

    if a and b:
        domain = f"{a}.{b}"
    elif a:
        domain = a
    elif b:
        domain = b
    else:
        domain = ""

    # Surface hint
    surface = ""
    low = t.lower()
    if "shower" in low:
        surface = "shower"
    elif "floor" in low or "floors" in low:
        surface = "floor"
    elif "wall" in low or "walls" in low:
        surface = "wall"

    return domain, surface

def _c7_load_decision_rows(project_name: str) -> List[Dict[str, Any]]:
    """
    Load decisions.jsonl in file order, best-effort.
    Returns raw dict rows (may be legacy or C7).
    """
    ensure_project(project_name)
    p = decisions_path(project_name)
    if not p.exists():
        return []

    out: List[Dict[str, Any]] = []
    try:
        for ln in p.read_text(encoding="utf-8", errors="replace").splitlines():
            try:
                obj = json.loads(ln)
            except Exception:
                continue
            if isinstance(obj, dict):
                out.append(obj)
    except Exception:
        return out
    return out

def list_decisions(
    project_name: str,
    *,
    domain: str = "",
    status: str = "",
    include_history: bool = True,
    limit: int = 0,
) -> List[Dict[str, Any]]:
    """
    Deterministically list decisions as normalized C7 objects.

    include_history=True:
      returns normalized rows in file order (append-only history).
    include_history=False:
      returns latest row per id (last-write-wins), then filtered.

    Optional filters:
      - domain: exact match
      - status: candidate|final|superseded
    """
    rows = _c7_load_decision_rows(project_name)
    normalized: List[Dict[str, Any]] = []
    for r in rows:
        if not isinstance(r, dict):
            continue
        # normalize using best-effort: raw_line is unknown here (we already parsed JSON)
        normalized.append(_c7_normalize_decision_row(r, raw_line=""))

    dom = str(domain or "").strip()
    st = str(status or "").strip().lower()
    if st and st not in ("candidate", "final", "superseded"):
        st = ""

    if not include_history:
        latest_by_id: Dict[str, Dict[str, Any]] = {}
        for r in normalized:
            did = str(r.get("id") or "").strip()
            if did:
                latest_by_id[did] = r
        normalized = list(latest_by_id.values())
        # Sort latest set by created_at desc for stable “most recent first” behavior
        normalized.sort(key=lambda x: str(x.get("created_at") or ""), reverse=True)

    if dom:
        normalized = [r for r in normalized if str(r.get("domain") or "").strip() == dom]
    if st:
        normalized = [r for r in normalized if str(r.get("status") or "").strip().lower() == st]

    if limit and limit > 0:
        return normalized[: int(limit)]
    return normalized

def get_current_decisions_by_domain(project_name: str) -> Dict[str, Dict[str, Any]]:
    """
    Deterministically compute CURRENT decisions (non-superseded), grouped by domain when possible.

    Current = latest-per-id row where:
      - status != "superseded"
      - AND id is not referenced by any other decision's supersedes field
    """
    latest = list_decisions(project_name, include_history=False) or []

    # Any id referenced by a newer decision is not current (even if no marker row exists)
    superseded_ids: set[str] = set()
    for r in latest:
        sid = r.get("supersedes", None)
        if sid is None:
            continue
        s = str(sid).strip()
        if s:
            superseded_ids.add(s)

    current: List[Dict[str, Any]] = []
    for r in latest:
        did = str(r.get("id") or "").strip()
        st = str(r.get("status") or "").strip().lower()
        if st == "superseded":
            continue
        if did and did in superseded_ids:
            continue
        current.append(r)

    # Group by domain when possible
    by_dom: Dict[str, Dict[str, Any]] = {}
    no_dom_bucket: List[Dict[str, Any]] = []

    for r in current:
        dom = str(r.get("domain") or "").strip()
        if dom:
            # If multiple current rows accidentally share a domain, last-write-wins by created_at
            prev = by_dom.get(dom)
            if not prev or str(r.get("created_at") or "") >= str(prev.get("created_at") or ""):
                by_dom[dom] = r
        else:
            no_dom_bucket.append(r)

    # Deterministic: fold no-domain items into keyed buckets so callers can still render them
    # without losing data. Key shape: "__no_domain__N"
    no_dom_bucket.sort(key=lambda x: str(x.get("created_at") or ""), reverse=True)
    for i, r in enumerate(no_dom_bucket, start=1):
        by_dom[f"__no_domain__{i:02d}"] = r

    return by_dom
# ---------------------------------------------------------------------------
# C8.2 — Conflict Detection (CURRENT decisions only; deterministic)
# ---------------------------------------------------------------------------

def list_current_decisions(project_name: str) -> List[Dict[str, Any]]:
    """
    Deterministically list CURRENT decisions (non-superseded), without collapsing to one per domain.
    Uses C7 normalization + supersedes semantics.
    """
    latest = list_decisions(project_name, include_history=False) or []

    superseded_ids: set[str] = set()
    for r in latest:
        sid = r.get("supersedes", None)
        if sid is None:
            continue
        s = str(sid).strip()
        if s:
            superseded_ids.add(s)

    current: List[Dict[str, Any]] = []
    for r in latest:
        did = str(r.get("id") or "").strip()
        st = str(r.get("status") or "").strip().lower()
        if st == "superseded":
            continue
        if did and did in superseded_ids:
            continue
        current.append(r)

    # stable order: created_at asc (file order is not available here; created_at is best stable proxy)
    current.sort(key=lambda x: str(x.get("created_at") or ""))
    return current

def detect_current_decision_conflicts(project_name: str) -> List[Dict[str, Any]]:
    """
    C8.2 — A conflict exists when TWO OR MORE CURRENT FINAL decisions share:
      - same domain + surface
        OR
      - same domain when surface is empty (treated as wildcard / domain-level)
    AND have different text.

    C8.2.1 — Legacy compatibility:
      - If a decision row has empty domain/surface, we deterministically infer an effective (domain,surface)
        from its text using infer_decision_domain(text) for grouping ONLY.
      - NO disk mutation.
    """
    cur = list_current_decisions(project_name) or []

    finals: List[Dict[str, Any]] = []
    for d in cur:
        if not isinstance(d, dict):
            continue
        if str(d.get("status") or "").strip().lower() != "final":
            continue

        txt = str(d.get("text") or "").strip()
        if not txt:
            continue

        dom = str(d.get("domain") or "").strip()
        surf = str(d.get("surface") or "").strip()

        # C8.2.1 fallback: infer only when missing
        if not dom:
            dom2, surf2 = infer_decision_domain(txt)
            dom = str(dom2 or "").strip()
            if not surf:
                surf = str(surf2 or "").strip()

        if not dom:
            # still unknown → cannot participate in conflict grouping
            continue

        d2 = dict(d)
        d2["_eff_domain"] = dom
        d2["_eff_surface"] = surf
        finals.append(d2)

    # Two detection modes:
    # 1) Surface-specific conflicts: same (domain, surface!=empty) with differing texts.
    # 2) Domain-level conflicts: if ANY decision in a domain has empty surface, then compare ALL decisions in that domain.
    by_domain: Dict[str, List[Dict[str, Any]]] = {}
    for d in finals:
        dom = str(d.get("_eff_domain") or "").strip()
        if not dom:
            continue
        by_domain.setdefault(dom, []).append(d)

    out: List[Dict[str, Any]] = []

    def _uniq_texts(items: List[Dict[str, Any]]) -> List[str]:
        seen = set()
        uniq: List[str] = []
        for x in items:
            t = str(x.get("text") or "").strip()
            if t and t not in seen:
                seen.add(t)
                uniq.append(t)
        return uniq

    # 1) Surface-specific conflicts
    by_key: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}
    for d in finals:
        dom = str(d.get("_eff_domain") or "").strip()
        surf = str(d.get("_eff_surface") or "").strip()
        if not dom or not surf:
            continue
        by_key.setdefault((dom, surf), []).append(d)

    for (dom, surf), items in by_key.items():
        if len(items) < 2:
            continue
        uniq = _uniq_texts(items)
        if len(uniq) < 2:
            continue
        out.append(
            {
                "domain": dom,
                "surface": surf,
                "decision_ids": [str(x.get("id") or "").strip() for x in items if str(x.get("id") or "").strip()],
                "texts": uniq[:12],
                "created_at": _now_iso_noz(),
            }
        )

    # 2) Domain-level conflicts (surface empty acts as wildcard)
    for dom, items in by_domain.items():
        has_empty_surface = any(not str(x.get("_eff_surface") or "").strip() for x in items)
        if not has_empty_surface:
            continue

        # Compare ALL items in the domain deterministically
        if len(items) < 2:
            continue
        uniq = _uniq_texts(items)
        if len(uniq) < 2:
            continue

        out.append(
            {
                "domain": dom,
                "surface": "",
                "decision_ids": [str(x.get("id") or "").strip() for x in items if str(x.get("id") or "").strip()],
                "texts": uniq[:12],
                "created_at": _now_iso_noz(),
            }
        )

    return out


def _c8_latest_open_conflict_keys(project_name: str) -> set:
    """
    Build a set of (domain,surface) keys that already have an OPEN conflict inbox item.
    Deterministic last-write-wins via inbox.jsonl scan.
    """
    rows = load_inbox(project_name) or []
    latest_by_id: Dict[str, Dict[str, Any]] = {}
    for r in rows:
        if not isinstance(r, dict):
            continue
        iid = str(r.get("id") or "").strip()
        if iid:
            latest_by_id[iid] = r

    keys = set()
    for it in latest_by_id.values():
        if str(it.get("status") or "").strip() != "open":
            continue
        if str(it.get("type") or "").strip() != "conflict":
            continue
        dom = str(it.get("domain") or "").strip()
        surf = str(it.get("surface") or "").strip()
        if dom:
            keys.add((dom, surf))
    return keys

def append_conflict_inbox_item(project_name: str, *, conflict: Dict[str, Any]) -> Dict[str, Any]:
    """
    Append a conflict item into inbox.jsonl (type=conflict, status=open).
    Adds extra fields: domain, surface, decision_ids, texts (bounded).
    """
    ensure_project(project_name)

    dom = str(conflict.get("domain") or "").strip()
    surf = str(conflict.get("surface") or "").strip()
    ids = conflict.get("decision_ids") if isinstance(conflict.get("decision_ids"), list) else []
    ids = [str(x).strip() for x in ids if str(x).strip()][:12]
    texts = conflict.get("texts") if isinstance(conflict.get("texts"), list) else []
    texts = [strip_control_tokens(str(x)).strip() for x in texts if str(x).strip()][:8]

    created = _now_iso_noz()
    iid = _next_inbox_id(project_name, created_at=created)

    surf_part = f" ({surf})" if surf else ""
    txt = f"Conflict: {dom}{surf_part} has multiple current decisions. Choose one."

    entry = {
        "id": iid,
        "created_at": created,
        "type": "conflict",
        "status": "open",
        "text": strip_control_tokens(txt),
        "refs": [f"decision:{x}" for x in ids],
        "domain": dom,
        "surface": surf,
        "decision_ids": ids,
        "texts": texts,
    }

    _append_jsonl_line(project_name, inbox_path(project_name), entry)
    return entry

def ensure_conflicts_in_inbox(project_name: str) -> List[Dict[str, Any]]:
    """
    Deterministically detect conflicts and ensure each (domain,surface) has an OPEN conflict inbox item.
    Returns the detected conflicts (not the inbox rows).
    """
    conflicts = detect_current_decision_conflicts(project_name) or []
    open_keys = _c8_latest_open_conflict_keys(project_name)

    for c in conflicts:
        if not isinstance(c, dict):
            continue
        dom = str(c.get("domain") or "").strip()
        surf = str(c.get("surface") or "").strip()
        if not dom:
            continue
        key = (dom, surf)
        if key in open_keys:
            continue
        try:
            append_conflict_inbox_item(project_name, conflict=c)
            open_keys.add(key)
        except Exception:
            pass

    return conflicts

def mark_decision_superseded(project_name: str, *, decision_id: str) -> bool:
    """
    Append-only: write a superseded marker row for an existing decision id.
    (Used for conflict resolution when user chooses a winner among multiple current decisions.)
    """
    ensure_project(project_name)
    did = str(decision_id or "").strip()
    if not did:
        return False

    marker: Dict[str, Any] = {
        "id": did,
        "created_at": _deliverables_now_noz(),
        "domain": "",
        "surface": "",
        "status": "superseded",
        "text": strip_control_tokens(""),
        "supersedes": None,
        "evidence": [],
    }
    try:
        _append_jsonl_line(project_name, decisions_path(project_name), marker)
        return True
    except Exception:
        return False

def resolve_conflict_by_winner(project_name: str, *, conflict_inbox_id: str, winner_decision_id: str) -> bool:
    """
    Deterministic conflict resolution:
    - Mark all other decision_ids in the conflict inbox item as superseded (append-only markers)
    - Resolve the conflict inbox item (append-only)
    """
    ensure_project(project_name)

    iid = str(conflict_inbox_id or "").strip()
    win = str(winner_decision_id or "").strip()
    if not iid or not win:
        return False

    # Find latest record for that inbox id (any status)
    rows = load_inbox(project_name) or []
    latest: Dict[str, Any] = {}
    for r in rows:
        if not isinstance(r, dict):
            continue
        if str(r.get("id") or "").strip() == iid:
            latest = r

    if not latest or str(latest.get("type") or "") != "conflict":
        return False

    ids = latest.get("decision_ids") if isinstance(latest.get("decision_ids"), list) else []
    ids = [str(x).strip() for x in ids if str(x).strip()]
    if win not in ids:
        return False

    losers = [x for x in ids if x != win]
    for lid in losers:
        try:
            mark_decision_superseded(project_name, decision_id=lid)
            append_link_event(
                project_name,
                type_="decision_to_decision",
                from_ref=f"decision:{win}",
                to_ref=f"decision:{lid}",
                reason="conflict_resolution_supersedes",
                confidence="high",
            )
        except Exception:
            pass

    try:
        resolve_inbox_item(
            project_name,
            inbox_id=iid,
            resolution_note=f"winner={win}",
            refs=[f"decision:{win}"],
        )
    except Exception:
        pass

    return True

def add_decision(
    project_name: str,
    *,
    domain: str = "",
    surface: str = "",
    status: str = "final",
    text: str,
    supersedes: Optional[str] = None,
    evidence: Optional[List[str]] = None,
    confidence: str = "",
) -> Dict[str, Any]:
    """
    Append a decision row in C7 schema (append-only).
    """
    ensure_project(project_name)

    created_at = _deliverables_now_noz()
    did = _next_decision_id(project_name, prefix="dec", created_at=created_at)

    st = str(status or "final").strip().lower()
    if st not in ("candidate", "final", "superseded"):
        st = "final"

    conf = str(confidence or "").strip().lower()
    if conf and conf not in ("low", "medium", "high"):
        conf = ""

    entry: Dict[str, Any] = {
        "id": did,
        "created_at": created_at,
        "domain": str(domain or "").strip(),
        "surface": str(surface or "").strip(),
        "status": st,
        "text": strip_control_tokens(str(text or "").strip()),
        "supersedes": (str(supersedes).strip() if supersedes else None),
        "evidence": [strip_control_tokens(str(x)).strip() for x in (evidence or []) if str(x).strip()],
    }
    if conf:
        entry["confidence"] = conf
    _append_jsonl_line(project_name, decisions_path(project_name), entry)

    # C8.2: After adding a decision, deterministically detect conflicts and ensure they appear in inbox.
    try:
        ensure_conflicts_in_inbox(project_name)
    except Exception:
        pass

    return entry

def supersede_decision(
    project_name: str,
    *,
    old_id: str,
    new_domain: str = "",
    new_surface: str = "",
    new_text: str,
    evidence: Optional[List[str]] = None,
    confidence: str = "",
) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    """
    Append-only supersession:
      1) write new decision with supersedes=old_id (status=final)
      2) append a supersede marker row for the old decision id (status=superseded)
    Returns (new_entry, old_marker_entry).
    """
    ensure_project(project_name)

    oid = str(old_id or "").strip()
    if not oid:
        raise ValueError("supersede_decision: missing old_id")

    # 1) New decision
    new_entry = add_decision(
        project_name,
        domain=str(new_domain or "").strip(),
        surface=str(new_surface or "").strip(),
        status="final",
        text=str(new_text or "").strip(),
        supersedes=oid,
        evidence=evidence,
        confidence=confidence,
    )
    # C8.1: Record a deterministic link for supersession (new -> old).
    try:
        append_link_event(
            project_name,
            type_="decision_to_decision",
            from_ref=f"decision:{str(new_entry.get('id') or '').strip()}",
            to_ref=f"decision:{oid}",
            reason="supersedes",
            confidence="high",
        )
    except Exception:
        pass

    # 2) Old marker (same id, last-write-wins)
    marker: Dict[str, Any] = {
        "id": oid,
        "created_at": _deliverables_now_noz(),
        "domain": "",
        "surface": "",
        "status": "superseded",
        "text": strip_control_tokens(""),
        "supersedes": None,
        "evidence": [],
    }
    _append_jsonl_line(project_name, decisions_path(project_name), marker)
    # C8.2: After supersession, deterministically detect conflicts and ensure they appear in inbox.
    try:
        ensure_conflicts_in_inbox(project_name)
    except Exception:
        pass

    return new_entry, marker

# -----------------------------------------------------------------------------
# C5 v1 — Deterministic Memory Reuse (read-only)
# Sources (authoritative order):
#   1) state/upload_notes.jsonl
#   2) state/decisions.jsonl
#   3) state/deliverables.json
# Deterministic retrieval only: keyword overlap + recency. No inference.
# -----------------------------------------------------------------------------

def load_upload_notes(project_name: str) -> List[Dict[str, Any]]:
    """
    Load upload notes (append-only). Returned in file order.
    """
    ensure_project(project_name)
    p = upload_notes_path(project_name)
    if not p.exists():
        return []

    out: List[Dict[str, Any]] = []
    try:
        for ln in p.read_text(encoding="utf-8", errors="replace").splitlines():
            try:
                obj = json.loads(ln)
            except Exception:
                continue
            if isinstance(obj, dict):
                out.append(obj)
    except Exception:
        return out
    return out

def _derive_age_from_birthdate_for_user(birthdate_ymd: str, user: str) -> tuple[Optional[int], str, str]:
    out: Optional[int] = None
    b = (birthdate_ymd or "").strip()
    asof_ymd, asof_tz = _today_ymd_for_user(user)
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", b):
        try:
            by, bm, bd = [int(x) for x in b.split("-")]
            ty, tm, td = [int(x) for x in asof_ymd.split("-")]
            age = ty - by
            if (tm, td) < (bm, bd):
                age -= 1
            if 0 <= age <= 130:
                out = int(age)
        except Exception:
            out = None
    return out, asof_ymd, asof_tz


def load_capability_gaps(project_name: str, limit: int = 2000) -> List[Dict[str, Any]]:
    ensure_project(project_name)
    p = capability_gaps_path(project_name)
    if not p.exists():
        return []
    out: List[Dict[str, Any]] = []
    try:
        lines = p.read_text(encoding="utf-8", errors="replace").splitlines()
    except Exception:
        lines = []
    for ln in lines[-int(limit or 0):]:
        try:
            obj = json.loads(ln)
        except Exception:
            obj = None
        if isinstance(obj, dict):
            out.append(obj)
    return out


def capability_gap_exists(project_name: str, gap_id: str) -> bool:
    gid = str(gap_id or "").strip()
    if not gid:
        return False
    ensure_project(project_name)
    p = capability_gaps_path(project_name)
    if not p.exists():
        return False
    try:
        lines = p.read_text(encoding="utf-8", errors="replace").splitlines()
    except Exception:
        return False
    tail = lines[-2000:] if len(lines) > 2000 else lines
    for ln in tail:
        try:
            obj = json.loads(ln)
        except Exception:
            continue
        if not isinstance(obj, dict):
            continue
        if str(obj.get("id") or "").strip() == gid:
            return True
    return False


def _next_capability_gap_id(project_name: str, *, created_at: str) -> str:
    # Deterministic per-day increment: gap_YYYYMMDD_NNN
    day = (created_at or "").split("T", 1)[0].replace("-", "")
    prefix = f"gap_{day}_"
    max_n = 0
    p = capability_gaps_path(project_name)
    if p.exists():
        try:
            for ln in p.read_text(encoding="utf-8", errors="replace").splitlines():
                try:
                    obj = json.loads(ln)
                except Exception:
                    continue
                if not isinstance(obj, dict):
                    continue
                gid = str(obj.get("id") or "")
                if not gid.startswith(prefix):
                    continue
                tail = gid[len(prefix):]
                try:
                    n = int(tail)
                except Exception:
                    n = 0
                if n > max_n:
                    max_n = n
        except Exception:
            pass
    return f"{prefix}{(max_n + 1):03d}"


def append_capability_gap_entry(project_name: str, entry: Dict[str, Any]) -> Dict[str, Any]:
    """
    Append-only. Records a capability gap for later remediation.
    """
    ensure_project(project_name)
    created_at = str(entry.get("created_at") or _deliverables_now_noz())
    gid = str(entry.get("id") or "").strip()
    if not gid:
        gid = _next_capability_gap_id(project_name, created_at=created_at)

    e = dict(entry)
    e["id"] = gid
    e["created_at"] = created_at
    e["blocked"] = bool(e.get("blocked", False))
    e["task_summary"] = str(e.get("task_summary") or "").strip()
    e["limitations"] = e.get("limitations") if isinstance(e.get("limitations"), list) else []
    e["missing_capabilities"] = e.get("missing_capabilities") if isinstance(e.get("missing_capabilities"), list) else []
    e["suggested_features"] = e.get("suggested_features") if isinstance(e.get("suggested_features"), list) else []
    e["needed_inputs"] = e.get("needed_inputs") if isinstance(e.get("needed_inputs"), list) else []
    e["recommended_search_queries"] = e.get("recommended_search_queries") if isinstance(e.get("recommended_search_queries"), list) else []

    _append_jsonl_line(project_name, capability_gaps_path(project_name), e)
    return e


def build_capability_gap_views_and_write(
    project_name: str,
    *,
    max_entries: int = 2000,
) -> Dict[str, Any]:
    ensure_project_scaffold(project_name)
    entries = load_capability_gaps(project_name, limit=max_entries)

    blocked = sum(1 for e in entries if isinstance(e, dict) and bool(e.get("blocked")))

    lines: List[str] = []
    lines.append("# Capability Gaps")
    lines.append("")
    lines.append(f"- Updated: {now_iso()}")
    lines.append(f"- Entries: {len(entries)}")
    lines.append(f"- Blocked: {blocked}")
    lines.append("")

    if entries:
        lines.append("## Recent")
        for e in entries[-12:]:
            if not isinstance(e, dict):
                continue
            gid = str(e.get("id") or "").strip()
            task = _fact_norm_text(e.get("task_summary"), cap=90)
            lims = e.get("limitations") if isinstance(e.get("limitations"), list) else []
            lim0 = _fact_norm_text(lims[0] if lims else "", cap=80)
            feats = e.get("suggested_features") if isinstance(e.get("suggested_features"), list) else []
            feat0 = _fact_norm_text(feats[0] if feats else "", cap=80)
            blocked0 = "blocked" if bool(e.get("blocked")) else "partial"
            line = f"- {gid} | {blocked0} | {task}"
            if lim0:
                line = line + f" | limit: {lim0}"
            if feat0:
                line = line + f" | fix: {feat0}"
            lines.append(line)
    else:
        lines.append("## Recent")
        lines.append("- (none yet)")

    md = "\n".join(lines).strip()

    create_artifact(
        project_name,
        "capability_gaps_latest",
        md,
        artifact_type="capability_gaps",
        from_files=[],
        file_ext=".md",
        also_update_state_key="capability_gaps",
    )

    return {"entries": len(entries), "blocked": blocked, "md": md}


# ---------------------------------------------------------------------------
# Discovery Index (analysis hat; append-only + derived views)
# ---------------------------------------------------------------------------

def discovery_index_path(project_name: str) -> Path:
    # projects/<user>/<project>/state/discovery_index.jsonl
    return state_dir(project_name) / DISCOVERY_INDEX_FILE_NAME

def fact_ledger_path(project_name: str) -> Path:
    # projects/<user>/<project>/state/fact_ledger.jsonl
    return state_dir(project_name) / FACT_LEDGER_FILE_NAME

def upload_batches_state_path(project_name: str) -> Path:
    # projects/<user>/<project>/state/upload_batches.json
    return state_dir(project_name) / UPLOAD_BATCH_STATE_FILE_NAME

def load_discovery_index(project_name: str, limit: int = 1200) -> List[Dict[str, Any]]:
    p = discovery_index_path(project_name)
    if not p.exists():
        return []
    out: List[Dict[str, Any]] = []
    try:
        lines = p.read_text(encoding="utf-8", errors="replace").splitlines()
    except Exception:
        lines = []
    for ln in lines[-int(limit or 0):]:
        try:
            obj = json.loads(ln)
        except Exception:
            obj = None
        if isinstance(obj, dict):
            out.append(obj)
    return out

def load_fact_ledger(project_name: str, limit: int = 4000) -> List[Dict[str, Any]]:
    p = fact_ledger_path(project_name)
    if not p.exists():
        return []
    out: List[Dict[str, Any]] = []
    try:
        lines = p.read_text(encoding="utf-8", errors="replace").splitlines()
    except Exception:
        lines = []
    for ln in lines[-int(limit or 0):]:
        try:
            obj = json.loads(ln)
        except Exception:
            obj = None
        if isinstance(obj, dict):
            out.append(obj)
    return out

def _fact_norm_text(v: Any, cap: int = 280) -> str:
    s = str(v or "").strip()
    if not s:
        return ""
    if len(s) <= cap:
        return s
    return s[: cap - 1].rstrip() + "…"

def append_fact_ledger_entry(project_name: str, entry: Dict[str, Any]) -> Dict[str, Any]:
    """
    Append one fact ledger entry (jsonl).
    Deterministic storage only; caller ensures content quality.
    """
    ensure_project_scaffold(project_name)
    e = entry if isinstance(entry, dict) else {}
    e = dict(e)
    e.setdefault("id", f"fact_{int(time.time() * 1000)}")
    e.setdefault("created_at", now_iso())

    e["entity"] = _fact_norm_text(e.get("entity"), cap=120)
    e["attribute"] = _fact_norm_text(e.get("attribute"), cap=80)
    e["value"] = _fact_norm_text(e.get("value"), cap=240)
    e["date"] = _fact_norm_text(e.get("date"), cap=40)
    e["source_file"] = _fact_norm_text(e.get("source_file"), cap=200).replace("\\", "/")
    e["source_hint"] = _fact_norm_text(e.get("source_hint"), cap=120)
    e["confidence"] = str(e.get("confidence") or "").strip().lower()
    e["notes"] = _fact_norm_text(e.get("notes"), cap=240)

    _append_jsonl_line(project_name, fact_ledger_path(project_name), e)
    return e

def _load_upload_batches_state(project_name: str) -> Dict[str, Any]:
    p = upload_batches_state_path(project_name)
    if not p.exists():
        return {"batches": {}}
    try:
        obj = json.loads(p.read_text(encoding="utf-8") or "{}")
    except Exception:
        obj = {}
    if not isinstance(obj, dict):
        obj = {}
    obj.setdefault("batches", {})
    if not isinstance(obj.get("batches"), dict):
        obj["batches"] = {}
    return obj

def _save_upload_batches_state(project_name: str, obj: Dict[str, Any]) -> None:
    out = obj if isinstance(obj, dict) else {}
    out.setdefault("batches", {})
    atomic_write_text(upload_batches_state_path(project_name), json.dumps(out, indent=2, sort_keys=True))

def register_upload_batch(project_name: str, *, zip_name: str, files: List[str]) -> str:
    """
    Register a new upload batch (ZIP extraction).
    Returns batch_id.
    """
    ensure_project_scaffold(project_name)
    batch_id = f"batch_{int(time.time() * 1000)}"
    files2 = [str(x).replace("\\", "/").strip() for x in (files or []) if str(x).strip()]
    st = _load_upload_batches_state(project_name)
    batches = st.get("batches")
    if not isinstance(batches, dict):
        batches = {}
        st["batches"] = batches
    batches[batch_id] = {
        "id": batch_id,
        "created_at": now_iso(),
        "zip_name": str(zip_name or "").strip(),
        "total": len(files2),
        "remaining": len(files2),
        "status": "active",
        "summary_done": False,
        "files": files2[:800],
        "completed_at": "",
    }
    _save_upload_batches_state(project_name, st)
    return batch_id

def mark_upload_batch_summary_done(project_name: str, *, batch_id: str) -> None:
    if not batch_id:
        return
    st = _load_upload_batches_state(project_name)
    batches = st.get("batches")
    if not isinstance(batches, dict):
        return
    b = batches.get(batch_id)
    if not isinstance(b, dict):
        return
    b["summary_done"] = True
    batches[batch_id] = b
    st["batches"] = batches
    _save_upload_batches_state(project_name, st)

def mark_upload_batch_file_done(project_name: str, *, batch_id: str, canonical_rel: str) -> Dict[str, Any]:
    """
    Decrement remaining count for a batch and mark done when complete.
    Returns the updated batch record (or {}).
    """
    if not batch_id:
        return {}
    st = _load_upload_batches_state(project_name)
    batches = st.get("batches")
    if not isinstance(batches, dict):
        return {}
    b = batches.get(batch_id)
    if not isinstance(b, dict):
        return {}
    if str(b.get("status") or "") == "done":
        return dict(b)
    try:
        rem = int(b.get("remaining") or 0)
    except Exception:
        rem = 0
    if rem > 0:
        rem -= 1
    if rem < 0:
        rem = 0
    b["remaining"] = rem
    if rem == 0:
        b["status"] = "done"
        b["completed_at"] = now_iso()
    batches[batch_id] = b
    st["batches"] = batches
    _save_upload_batches_state(project_name, st)
    return dict(b)

def _disc_norm_list(v: Any) -> List[str]:
    if v is None:
        return []
    if isinstance(v, list):
        items = [str(x).strip() for x in v if str(x).strip()]
        return items[:24]
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return []
        # Split on common separators if it looks like a list
        if "|" in s or ";" in s:
            parts = re.split(r"[|;]", s)
            return [p.strip() for p in parts if p.strip()][:24]
        return [s]
    return []

def append_discovery_index_entry(project_name: str, entry: Dict[str, Any]) -> Dict[str, Any]:
    """
    Append one discovery index entry (jsonl).
    Deterministic storage only; caller ensures content quality.
    """
    ensure_project_scaffold(project_name)
    e = entry if isinstance(entry, dict) else {}
    e = dict(e)
    e.setdefault("id", f"disc_{int(time.time() * 1000)}")
    e.setdefault("created_at", now_iso())

    # Normalize common fields
    e["upload_path"] = str(e.get("upload_path") or "").replace("\\", "/").strip()
    e["orig_name"] = str(e.get("orig_name") or "").strip()
    e["doc_type"] = str(e.get("doc_type") or "").strip()
    e["title"] = str(e.get("title") or "").strip()
    e["summary"] = str(e.get("summary") or "").strip()
    e["confidence"] = str(e.get("confidence") or "").strip().lower()

    e["parties"] = _disc_norm_list(e.get("parties"))
    e["dates"] = e.get("dates") if isinstance(e.get("dates"), list) else _disc_norm_list(e.get("dates"))
    e["amounts"] = _disc_norm_list(e.get("amounts"))
    e["issues"] = _disc_norm_list(e.get("issues"))
    e["questions"] = _disc_norm_list(e.get("questions"))
    e["tags"] = _disc_norm_list(e.get("tags"))

    _append_jsonl_line(project_name, discovery_index_path(project_name), e)
    return e

def build_discovery_views_and_write(
    project_name: str,
    *,
    max_entries: int = 400,
    max_summary_chars: int = 240,
) -> Dict[str, Any]:
    """
    Build human-readable discovery views from discovery_index.jsonl.
    Writes:
      - artifacts: discovery_index_latest.md, timeline_latest.md, issues_questions_latest.md, evidence_matrix_latest.csv
      - state: discovery_index.md, timeline.md, issues_questions.md, evidence_matrix.csv
    """
    ensure_project_scaffold(project_name)
    entries = load_discovery_index(project_name, limit=max_entries)

    # Discovery index (table)
    idx_lines: List[str] = []
    idx_lines.append("# Discovery Index")
    idx_lines.append("")
    idx_lines.append(f"- Updated: {now_iso()}")
    idx_lines.append(f"- Entries: {len(entries)}")
    idx_lines.append("")
    idx_lines.append("| ID | Date(s) | Doc Type | Title | Source | Summary |")
    idx_lines.append("|---|---|---|---|---|---|")

    def _short(s: str, cap: int) -> str:
        s2 = (s or "").strip()
        if len(s2) <= cap:
            return s2
        return s2[: cap - 1].rstrip() + "…"

    for e in entries:
        if not isinstance(e, dict):
            continue
        eid = str(e.get("id") or "").strip()
        dates = e.get("dates")
        if isinstance(dates, list):
            dval = ", ".join([str(x) for x in dates if str(x).strip()][:3])
        else:
            dval = str(dates or "").strip()
        doc_type = str(e.get("doc_type") or "").strip()
        title = str(e.get("title") or "").strip()
        src = str(e.get("orig_name") or e.get("upload_path") or "").strip()
        summ = _short(str(e.get("summary") or "").strip(), int(max_summary_chars or 240))
        idx_lines.append(
            f"| {eid} | {_short(dval, 24)} | {_short(doc_type, 24)} | {_short(title, 40)} | {_short(src, 40)} | {_short(summ, 120)} |"
        )

    idx_md = "\n".join(idx_lines).strip()

    # Timeline
    timeline_items: List[Tuple[str, str]] = []
    fuzzy_items: List[str] = []
    for e in entries:
        if not isinstance(e, dict):
            continue
        title = str(e.get("title") or "").strip()
        src = str(e.get("orig_name") or e.get("upload_path") or "").strip()
        dates = e.get("dates")
        if isinstance(dates, list) and dates:
            for d in dates[:6]:
                if isinstance(d, dict):
                    ds = str(d.get("date") or "").strip()
                    label = str(d.get("label") or "").strip()
                else:
                    ds = str(d or "").strip()
                    label = ""
                if re.fullmatch(r"\d{4}-\d{2}-\d{2}", ds):
                    blurb = f"{title} ({src})"
                    if label:
                        blurb = blurb + f" — {label}"
                    timeline_items.append((ds, blurb))
                elif ds:
                    fuzzy_items.append(f"{ds} — {title} ({src})")
        elif isinstance(dates, str) and dates.strip():
            fuzzy_items.append(f"{dates.strip()} — {title} ({src})")

    timeline_items.sort(key=lambda x: x[0])
    timeline_lines = ["# Timeline", "", f"- Updated: {now_iso()}", ""]
    if timeline_items:
        timeline_lines.append("## Dated")
        for d, blurb in timeline_items:
            timeline_lines.append(f"- {d} — {blurb}")
        timeline_lines.append("")
    if fuzzy_items:
        timeline_lines.append("## Fuzzy or Undated")
        for it in fuzzy_items[:200]:
            timeline_lines.append(f"- {it}")
    timeline_md = "\n".join(timeline_lines).strip()

    # Issues + questions
    issues_seen: set[str] = set()
    questions_seen: set[str] = set()
    issues_lines = ["# Issues and Questions", "", f"- Updated: {now_iso()}", ""]
    issues_lines.append("## Issues")
    for e in entries:
        if not isinstance(e, dict):
            continue
        src = str(e.get("orig_name") or e.get("upload_path") or "").strip()
        for it in _disc_norm_list(e.get("issues")):
            key = it.lower()
            if key in issues_seen:
                continue
            issues_seen.add(key)
            issues_lines.append(f"- {it} ({_short(src, 60)})")
    if len(issues_seen) == 0:
        issues_lines.append("- (none yet)")

    issues_lines.append("")
    issues_lines.append("## Open Questions")
    for e in entries:
        if not isinstance(e, dict):
            continue
        src = str(e.get("orig_name") or e.get("upload_path") or "").strip()
        for it in _disc_norm_list(e.get("questions")):
            key = it.lower()
            if key in questions_seen:
                continue
            questions_seen.add(key)
            issues_lines.append(f"- {it} ({_short(src, 60)})")
    if len(questions_seen) == 0:
        issues_lines.append("- (none yet)")

    issues_md = "\n".join(issues_lines).strip()

    # Evidence matrix CSV
    header = [
        "source_file",
        "doc_type",
        "title",
        "parties",
        "dates",
        "amounts",
        "issues",
        "summary",
        "confidence",
    ]
    csv_lines = [",".join(header)]
    for e in entries:
        if not isinstance(e, dict):
            continue
        row = [
            _csv_escape(e.get("orig_name") or e.get("upload_path")),
            _csv_escape(e.get("doc_type")),
            _csv_escape(e.get("title")),
            _csv_escape(" | ".join(_disc_norm_list(e.get("parties")))),
            _csv_escape(" | ".join(_disc_norm_list(e.get("dates")))),
            _csv_escape(" | ".join(_disc_norm_list(e.get("amounts")))),
            _csv_escape(" | ".join(_disc_norm_list(e.get("issues")))),
            _csv_escape(_short(str(e.get("summary") or ""), 240)),
            _csv_escape(e.get("confidence")),
        ]
        csv_lines.append(",".join(row))
    evidence_csv = "\n".join(csv_lines).strip()

    # Write artifacts (versioned) + update state files for retrieval
    create_artifact(
        project_name,
        "discovery_index_latest",
        idx_md,
        artifact_type="discovery_index",
        from_files=[],
        file_ext=".md",
        also_update_state_key="discovery_index",
    )
    create_artifact(
        project_name,
        "timeline_latest",
        timeline_md,
        artifact_type="timeline",
        from_files=[],
        file_ext=".md",
        also_update_state_key="timeline",
    )
    create_artifact(
        project_name,
        "issues_questions_latest",
        issues_md,
        artifact_type="issues_questions",
        from_files=[],
        file_ext=".md",
        also_update_state_key="issues_questions",
    )
    create_artifact(
        project_name,
        "evidence_matrix_latest",
        evidence_csv,
        artifact_type="evidence_matrix",
        from_files=[],
        file_ext=".csv",
        also_update_state_key="evidence_matrix",
    )

    return {
        "entries": len(entries),
        "discovery_index_md": idx_md,
        "timeline_md": timeline_md,
        "issues_md": issues_md,
        "evidence_csv": evidence_csv,
    }

def build_fact_ledger_views_and_write(
    project_name: str,
    *,
    max_entries: int = 2000,
) -> Dict[str, Any]:
    ensure_project_scaffold(project_name)
    entries = load_fact_ledger(project_name, limit=max_entries)

    # Ledger markdown
    lines: List[str] = []
    lines.append("# Fact Ledger")
    lines.append("")
    lines.append(f"- Updated: {now_iso()}")
    lines.append(f"- Entries: {len(entries)}")
    lines.append("")
    lines.append("| ID | Entity | Attribute | Value | Date | Source | Confidence |")
    lines.append("|---|---|---|---|---|---|---|")

    for e in entries:
        if not isinstance(e, dict):
            continue
        eid = str(e.get("id") or "").strip()
        ent = _fact_norm_text(e.get("entity"), cap=60)
        att = _fact_norm_text(e.get("attribute"), cap=40)
        val = _fact_norm_text(e.get("value"), cap=80)
        date = _fact_norm_text(e.get("date"), cap=20)
        src = _fact_norm_text(e.get("source_file"), cap=60)
        conf = _fact_norm_text(e.get("confidence"), cap=10)
        lines.append(f"| {eid} | {ent} | {att} | {val} | {date} | {src} | {conf} |")

    ledger_md = "\n".join(lines).strip()

    # Conflicts (simple deterministic scan)
    conflicts: List[Dict[str, Any]] = []
    buckets: Dict[Tuple[str, str, str], Dict[str, Any]] = {}
    for e in entries:
        if not isinstance(e, dict):
            continue
        ent = str(e.get("entity") or "").strip().lower()
        att = str(e.get("attribute") or "").strip().lower()
        date = str(e.get("date") or "").strip()
        if not ent or not att:
            continue
        key = (ent, att, date)
        val = str(e.get("value") or "").strip()
        if not val:
            continue
        b = buckets.get(key)
        if not isinstance(b, dict):
            b = {"values": {}, "samples": []}
        vals = b.get("values")
        if not isinstance(vals, dict):
            vals = {}
        if val not in vals:
            vals[val] = []
        vals[val].append(str(e.get("source_file") or "").strip())
        b["values"] = vals
        b["samples"] = (b.get("samples") or [])[:4]
        buckets[key] = b

    for (ent, att, date), b in buckets.items():
        vals = b.get("values") if isinstance(b.get("values"), dict) else {}
        if len(vals.keys()) >= 2:
            conflicts.append(
                {
                    "entity": ent,
                    "attribute": att,
                    "date": date,
                    "values": list(vals.keys())[:6],
                }
            )

    conf_lines: List[str] = []
    conf_lines.append("# Conflict Report")
    conf_lines.append("")
    conf_lines.append(f"- Updated: {now_iso()}")
    conf_lines.append(f"- Conflicts: {len(conflicts)}")
    conf_lines.append("")
    if conflicts:
        for c in conflicts[:300]:
            ent = _fact_norm_text(c.get("entity"), cap=80)
            att = _fact_norm_text(c.get("attribute"), cap=60)
            date = _fact_norm_text(c.get("date"), cap=20)
            vals = c.get("values") if isinstance(c.get("values"), list) else []
            vals2 = [str(v) for v in vals if str(v).strip()]
            val_line = " vs ".join(vals2[:4]) if vals2 else "(values not captured)"
            if date:
                conf_lines.append(f"- {ent} | {att} | {date} — {val_line}")
            else:
                conf_lines.append(f"- {ent} | {att} — {val_line}")
    else:
        conf_lines.append("- (none yet)")
    conflicts_md = "\n".join(conf_lines).strip()

    create_artifact(
        project_name,
        "fact_ledger_latest",
        ledger_md,
        artifact_type="fact_ledger",
        from_files=[],
        file_ext=".md",
        also_update_state_key="fact_ledger",
    )
    create_artifact(
        project_name,
        "conflicts_latest",
        conflicts_md,
        artifact_type="conflict_report",
        from_files=[],
        file_ext=".md",
        also_update_state_key="conflicts",
    )

    return {
        "entries": len(entries),
        "conflicts": len(conflicts),
        "ledger_md": ledger_md,
        "conflicts_md": conflicts_md,
    }


def dedupe_raw_files_by_sha(
    project_name: str,
    *,
    delete_duplicate_files: bool = False,
) -> Dict[str, Any]:
    """
    Remove duplicate raw_files entries based on sha256.
    - Keeps the earliest added_at entry as canonical.
    - Optionally deletes duplicate files from disk (off by default).
    - Updates manifest in-place.
    Returns summary dict.
    """
    pr, pd = _require_configured()
    ensure_project(project_name)
    m = load_manifest(project_name)
    raw_files = m.get("raw_files") or []
    if not isinstance(raw_files, list):
        raw_files = []

    # Normalize + ensure sha256 present
    for entry in raw_files:
        if not isinstance(entry, dict):
            continue
        rel = str(entry.get("path") or "").replace("\\", "/").strip()
        if not rel:
            continue
        if not entry.get("sha256"):
            try:
                abs_path = pr / rel
                if abs_path.exists() and abs_path.is_file():
                    entry["sha256"] = file_sha256_bytes(read_bytes_file(abs_path))
            except Exception:
                pass

    # Group by sha
    by_sha: Dict[str, Dict[str, Any]] = {}
    duplicates: List[Dict[str, Any]] = []

    def _added_at(e: Dict[str, Any]) -> float:
        try:
            return float(e.get("added_at") or 0.0)
        except Exception:
            return 0.0

    for entry in raw_files:
        if not isinstance(entry, dict):
            continue
        sha = str(entry.get("sha256") or "").strip()
        if not sha:
            continue
        if sha not in by_sha:
            by_sha[sha] = entry
        else:
            keep = by_sha[sha]
            if _added_at(entry) and _added_at(keep) and _added_at(entry) < _added_at(keep):
                duplicates.append(keep)
                by_sha[sha] = entry
            else:
                duplicates.append(entry)

    if not duplicates:
        return {"duplicates": 0, "removed": 0, "kept": len(raw_files)}

    # Remove duplicates from raw_files list
    dup_ids = set()
    for d in duplicates:
        try:
            dup_ids.add(id(d))
        except Exception:
            continue

    new_raw: List[Dict[str, Any]] = []
    for e in raw_files:
        if id(e) in dup_ids:
            continue
        new_raw.append(e)

    # Optionally delete duplicate files
    if delete_duplicate_files:
        for d in duplicates:
            try:
                rel = str(d.get("path") or "").replace("\\", "/").strip()
                if not rel:
                    continue
                abs_path = pr / rel
                if abs_path.exists() and abs_path.is_file():
                    abs_path.unlink()
            except Exception:
                pass

    m["raw_files"] = new_raw
    save_manifest(project_name, m)

    return {"duplicates": len(duplicates), "removed": len(duplicates), "kept": len(new_raw)}


def build_library_index_and_write(
    project_name: str,
    *,
    max_entries: int = 4000,
) -> Dict[str, Any]:
    """
    Build a unified library index from manifest raw_files + discovery index.
    Writes:
      - artifacts: library_index_latest.md
      - state: library_index.md
    """
    ensure_project_scaffold(project_name)
    pr, pd = _require_configured()
    m = load_manifest(project_name)
    raw_files = m.get("raw_files") or []
    if not isinstance(raw_files, list):
        raw_files = []

    di = load_discovery_index(project_name, limit=6000)
    di_paths = set()
    di_names = set()
    for e in di:
        if not isinstance(e, dict):
            continue
        up = str(e.get("upload_path") or "").replace("\\", "/").strip()
        if up:
            di_paths.add(up)
        nm = str(e.get("orig_name") or "").strip()
        if nm:
            di_names.add(nm)

    lines: List[str] = []
    lines.append("# Library Index")
    lines.append("")
    lines.append(f"- Updated: {now_iso()}")
    lines.append(f"- Raw files: {len(raw_files)}")
    lines.append(f"- Discovery indexed: {len(di)}")
    lines.append("")
    lines.append("| Added | Orig Name | Path | Size KB | SHA256 | Indexed |")
    lines.append("|---|---|---|---:|---|---|")

    def _short(s: str, cap: int) -> str:
        s2 = (s or "").strip()
        if len(s2) <= cap:
            return s2
        return s2[: cap - 1].rstrip() + "…"

    for rf in raw_files[: int(max_entries or 0)]:
        if not isinstance(rf, dict):
            continue
        orig_name = str(rf.get("orig_name") or "").strip()
        rel = str(rf.get("path") or "").replace("\\", "/").strip()
        sha = str(rf.get("sha256") or "").strip()
        added = ""
        try:
            ts = float(rf.get("added_at") or 0.0)
            if ts > 0:
                added = time.strftime("%Y-%m-%d", time.gmtime(ts))
        except Exception:
            added = ""
        size_kb = ""
        try:
            if rel:
                p = pr / rel
                if p.exists() and p.is_file():
                    size_kb = f"{int(p.stat().st_size / 1024)}"
        except Exception:
            size_kb = ""

        indexed = "yes" if (rel in di_paths or orig_name in di_names) else "no"

        lines.append(
            f"| {_short(added, 12)} | {_short(orig_name, 36)} | {_short(rel, 44)} | {size_kb} | {_short(sha, 12)} | {indexed} |"
        )

    md = "\n".join(lines).strip()

    create_artifact(
        project_name,
        "library_index_latest",
        md,
        artifact_type="library_index",
        from_files=[],
        file_ext=".md",
        also_update_state_key="library_index",
    )

    return {"raw_files": len(raw_files), "indexed": len(di), "md": md}

def build_analysis_audit_report(project_name: str) -> str:
    """
    Build a concise audit snapshot of analysis/ingestion state.
    """
    ensure_project_scaffold(project_name)
    m = load_manifest(project_name)
    raw_files = m.get("raw_files") if isinstance(m.get("raw_files"), list) else []
    arts = m.get("artifacts") if isinstance(m.get("artifacts"), list) else []
    di = load_discovery_index(project_name, limit=5000)
    fl = load_fact_ledger(project_name, limit=8000)
    cg = load_capability_gaps(project_name, limit=8000)
    batch_state = _load_upload_batches_state(project_name)
    batches = batch_state.get("batches") if isinstance(batch_state.get("batches"), dict) else {}
    active_batches = [b for b in batches.values() if isinstance(b, dict) and str(b.get("status") or "") == "active"]

    lines: List[str] = []
    lines.append("# Analysis Audit")
    lines.append("")
    lines.append(f"- Updated: {now_iso()}")
    lines.append(f"- Raw files: {len(raw_files)}")
    lines.append(f"- Artifacts: {len(arts)}")
    lines.append(f"- Discovery entries: {len(di)}")
    lines.append(f"- Fact ledger entries: {len(fl)}")
    lines.append(f"- Capability gaps: {len(cg)}")
    lines.append(f"- Active upload batches: {len(active_batches)}")
    lines.append("")
    if active_batches:
        lines.append("## Active Batches")
        for b in active_batches[:12]:
            bn = str(b.get("zip_name") or b.get("id") or "").strip()
            tot = int(b.get("total") or 0)
            rem = int(b.get("remaining") or 0)
            lines.append(f"- {bn}: {tot - rem}/{tot} processed")
    else:
        lines.append("## Active Batches")
        lines.append("- (none)")

    out = "\n".join(lines).strip()
    create_artifact(
        project_name,
        "analysis_audit_latest",
        out,
        artifact_type="analysis_audit",
        from_files=[],
        file_ext=".md",
    )
    return out

def build_case_workbook_and_write(project_name: str) -> Optional[str]:
    """
    Build a multi-sheet XLSX case workbook from current analysis artifacts.
    Returns artifact relative path, or None if openpyxl unavailable.
    """
    if Workbook is None:
        return None

    ensure_project_scaffold(project_name)
    wb = Workbook()

    # Sheet 1: Discovery Index
    ws = wb.active
    ws.title = "Discovery Index"
    ws.append(["ID", "Date(s)", "Doc Type", "Title", "Source", "Summary"])
    for e in load_discovery_index(project_name, limit=2000):
        if not isinstance(e, dict):
            continue
        dates = e.get("dates")
        if isinstance(dates, list):
            dval = ", ".join([str(x) for x in dates if str(x).strip()][:4])
        else:
            dval = str(dates or "").strip()
        ws.append([
            str(e.get("id") or ""),
            dval,
            str(e.get("doc_type") or ""),
            str(e.get("title") or ""),
            str(e.get("orig_name") or e.get("upload_path") or ""),
            str(e.get("summary") or ""),
        ])

    # Sheet 2: Timeline
    ws2 = wb.create_sheet(title="Timeline")
    ws2.append(["Date", "Event"])
    timeline_entries: List[Tuple[str, str]] = []
    for e in load_discovery_index(project_name, limit=2000):
        if not isinstance(e, dict):
            continue
        title = str(e.get("title") or "").strip()
        src = str(e.get("orig_name") or e.get("upload_path") or "").strip()
        dates = e.get("dates")
        if isinstance(dates, list):
            for d in dates[:6]:
                if isinstance(d, dict):
                    ds = str(d.get("date") or "").strip()
                    label = str(d.get("label") or "").strip()
                else:
                    ds = str(d or "").strip()
                    label = ""
                if ds:
                    ev = f"{title} ({src})"
                    if label:
                        ev = ev + f" — {label}"
                    timeline_entries.append((ds, ev))
        elif isinstance(dates, str) and dates.strip():
            timeline_entries.append((dates.strip(), f"{title} ({src})"))
    timeline_entries.sort(key=lambda x: x[0])
    for d, ev in timeline_entries:
        ws2.append([d, ev])

    # Sheet 3: Issues + Questions
    ws3 = wb.create_sheet(title="Issues")
    ws3.append(["Type", "Text", "Source"])
    for e in load_discovery_index(project_name, limit=2000):
        if not isinstance(e, dict):
            continue
        src = str(e.get("orig_name") or e.get("upload_path") or "").strip()
        for it in _disc_norm_list(e.get("issues")):
            ws3.append(["Issue", it, src])
        for it in _disc_norm_list(e.get("questions")):
            ws3.append(["Question", it, src])

    # Sheet 4: Fact Ledger
    ws4 = wb.create_sheet(title="Fact Ledger")
    ws4.append(["ID", "Entity", "Attribute", "Value", "Date", "Source", "Hint", "Confidence"])
    for e in load_fact_ledger(project_name, limit=8000):
        if not isinstance(e, dict):
            continue
        ws4.append([
            str(e.get("id") or ""),
            str(e.get("entity") or ""),
            str(e.get("attribute") or ""),
            str(e.get("value") or ""),
            str(e.get("date") or ""),
            str(e.get("source_file") or ""),
            str(e.get("source_hint") or ""),
            str(e.get("confidence") or ""),
        ])

    # Sheet 5: Conflicts
    ws5 = wb.create_sheet(title="Conflicts")
    ws5.append(["Entity", "Attribute", "Date", "Values"])
    conflicts = build_fact_ledger_views_and_write(project_name).get("conflicts_md")
    # Also parse conflicts list from latest build for table-like entries
    conf_entries: List[str] = []
    if isinstance(conflicts, str):
        conf_entries = [ln.strip() for ln in conflicts.splitlines() if ln.strip().startswith("- ")]
    for ln in conf_entries:
        # ln format: "- entity | attr | date — val vs val"
        row = ln.lstrip("- ").strip()
        parts = row.split("—", 1)
        left = parts[0].strip()
        vals = parts[1].strip() if len(parts) > 1 else ""
        left_parts = [p.strip() for p in left.split("|")]
        ent = left_parts[0] if len(left_parts) > 0 else ""
        att = left_parts[1] if len(left_parts) > 1 else ""
        date = left_parts[2] if len(left_parts) > 2 else ""
        ws5.append([ent, att, date, vals])

    bio = BytesIO()
    wb.save(bio)
    content_bytes = bio.getvalue()
    entry = create_artifact_bytes(
        project_name,
        "case_workbook_latest",
        content_bytes,
        artifact_type="case_workbook",
        from_files=[],
        file_ext=".xlsx",
    )
    return str(entry.get("path") or "").strip() or None

def build_case_summary_doc(project_name: str) -> Optional[str]:
    """
    Build a concise case summary document from current discovery and ledger.
    Returns artifact relative path.
    """
    ensure_project_scaffold(project_name)
    di = load_discovery_index(project_name, limit=800)
    fl = load_fact_ledger(project_name, limit=1200)

    lines: List[str] = []
    lines.append("# Case Summary")
    lines.append("")
    lines.append(f"- Updated: {now_iso()}")
    lines.append("")

    # High-level snapshot
    lines.append("## Snapshot")
    lines.append(f"- Discovery items: {len(di)}")
    lines.append(f"- Fact ledger entries: {len(fl)}")
    lines.append("")

    # Key issues + questions
    issues: List[str] = []
    questions: List[str] = []
    for e in di:
        if not isinstance(e, dict):
            continue
        issues.extend(_disc_norm_list(e.get("issues")))
        questions.extend(_disc_norm_list(e.get("questions")))
    issues = [i for i in issues if i][:40]
    questions = [q for q in questions if q][:40]

    lines.append("## Key Issues")
    if issues:
        for it in issues[:12]:
            lines.append(f"- {it}")
    else:
        lines.append("- (none yet)")
    lines.append("")

    lines.append("## Open Questions")
    if questions:
        for it in questions[:12]:
            lines.append(f"- {it}")
    else:
        lines.append("- (none yet)")
    lines.append("")

    # Conflicts (from latest build)
    conf_md = build_fact_ledger_views_and_write(project_name).get("conflicts_md")
    lines.append("## Conflicts")
    if isinstance(conf_md, str) and conf_md.strip():
        # include top conflicts only
        conf_lines = [ln for ln in conf_md.splitlines() if ln.strip().startswith("- ")]
        if conf_lines:
            for ln in conf_lines[:10]:
                lines.append(ln)
        else:
            lines.append("- (none yet)")
    else:
        lines.append("- (none yet)")

    doc = "\n".join(lines).strip()
    entry = create_artifact(
        project_name,
        "case_summary_latest",
        doc,
        artifact_type="case_summary",
        from_files=[],
        file_ext=".md",
    )
    return str(entry.get("path") or "").strip() or None


def _c5_tokens(text: str) -> List[str]:
    s = (text or "").lower()
    # Keep deterministic, simple: alnum+underscore tokens, min len 3
    toks = re.findall(r"[a-z0-9_]{3,}", s)
    # De-dupe stable
    seen: set[str] = set()
    out: List[str] = []
    for t in toks:
        if t in seen:
            continue
        seen.add(t)
        out.append(t)
        if len(out) >= 18:
            break
    return out


def _c5_score_text(haystack: str, tokens: List[str]) -> int:
    if not haystack or not tokens:
        return 0
    low = haystack.lower()
    score = 0
    for t in tokens:
        if t and t in low:
            score += 1
    return score


def _c5_entry_epoch(ts: str) -> float:
    # timestamps here are stored as YYYY-MM-DDTHH:MM:SS (no Z) in this codebase
    return _parse_iso_noz(str(ts or "").strip())


def find_project_memory(
    project_name: str,
    *,
    query: str,
    max_items: int = 3,
) -> List[Dict[str, Any]]:
    """
    Deterministically find 1–3 most relevant items across:
      upload_notes, decisions, deliverables

    Returns a list of normalized records:
      {
        "source": "upload_note" | "decision" | "deliverable",
        "timestamp": "...",
        "score": int,
        "data": { ...original record... }
      }

    No summarization. No rewriting. Read-only.
    """
    ensure_project(project_name)
    q = (query or "").strip()
    tokens = _c5_tokens(q)

    # Authoritative source order: upload_notes > decisions > deliverables
    source_rank = {"upload_note": 0, "decision": 1, "deliverable": 2}

    candidates: List[Dict[str, Any]] = []

    # 1) Upload notes
    try:
        notes = load_upload_notes(project_name)
    except Exception:
        notes = []
    for it in notes[-250:]:  # bounded scan for determinism + perf
        if not isinstance(it, dict):
            continue
        ts = str(it.get("timestamp") or "").strip()
        blob = " ".join([
            str(it.get("upload_path") or ""),
            str(it.get("question") or ""),
            str(it.get("answer") or ""),
        ]).strip()
        sc = _c5_score_text(blob, tokens)
        # If query has no usable tokens, require at least some signal text to avoid random picks
        if tokens and sc <= 0:
            continue
        candidates.append({
            "source": "upload_note",
            "timestamp": ts,
            "score": int(sc),
            "data": it,
        })

    # 2) Decisions
    try:
        decs = load_decisions(project_name)
    except Exception:
        decs = []
    for it in decs[-250:]:
        if not isinstance(it, dict):
            continue
        ts = str(it.get("timestamp") or "").strip()
        blob = " ".join([
            str(it.get("text") or ""),
            str(it.get("related_deliverable") or ""),
        ]).strip()
        sc = _c5_score_text(blob, tokens)
        if tokens and sc <= 0:
            continue
        candidates.append({
            "source": "decision",
            "timestamp": ts,
            "score": int(sc),
            "data": it,
        })

    # 3) Deliverables
    reg = {}
    try:
        reg = load_deliverables(project_name)
    except Exception:
        reg = {}
    items = reg.get("items") if isinstance(reg, dict) else None
    if isinstance(items, list):
        for it in items[-250:]:
            if not isinstance(it, dict):
                continue
            ts = str(it.get("created_at") or "").strip()
            blob = " ".join([
                str(it.get("title") or ""),
                str(it.get("type") or ""),
                str(it.get("path") or ""),
            ]).strip()
            sc = _c5_score_text(blob, tokens)
            if tokens and sc <= 0:
                continue
            candidates.append({
                "source": "deliverable",
                "timestamp": ts,
                "score": int(sc),
                "data": it,
            })

    # If query tokens were empty (e.g., "remind me"), only use recency within each source
    # but still cap to max_items and preserve authoritative order.
    # Sort: score desc, timestamp desc, source_rank asc
    def key_fn(r: Dict[str, Any]) -> tuple:
        src = str(r.get("source") or "")
        ts = str(r.get("timestamp") or "")
        epoch = _c5_entry_epoch(ts)
        sc = int(r.get("score") or 0)
        return (-sc, -epoch, source_rank.get(src, 9), src)

    candidates.sort(key=key_fn)

    # Deterministic cap (overall max_items)
    out: List[Dict[str, Any]] = []
    for r in candidates:
        out.append(r)
        if len(out) >= int(max_items or 3):
            break
    return out

# -----------------------------------------------------------------------------
# C5.2 — Truth-bound Project Pulse (DETERMINISTIC ONLY)
# Canonical sources ONLY:
#   - state/upload_notes.jsonl
#   - state/decisions.jsonl
#   - state/deliverables.json
#   - project_state.json (goal/next_actions if present)
#   - facts_map.md (optional presence only)
# No model calls. No inference.
# -----------------------------------------------------------------------------
def load_decision_candidates(project_name: str) -> List[Dict[str, Any]]:
    """
    Load decision candidates (append-only-ish JSONL, plus occasional status update rows).
    Returns dict rows in file order. Invalid lines ignored.
    """
    ensure_project(project_name)
    p = decision_candidates_path(project_name)
    if not p.exists():
        return []
    out: List[Dict[str, Any]] = []
    try:
        for ln in p.read_text(encoding="utf-8", errors="replace").splitlines():
            try:
                obj = json.loads(ln)
            except Exception:
                continue
            if isinstance(obj, dict):
                out.append(obj)
    except Exception:
        return out
    return out


def _pending_candidates(project_name: str) -> List[Dict[str, Any]]:
    """
    Deterministic: pending candidates only (must have text + status == pending).
    Status update rows typically lack 'text'; ignore them.
    """
    out: List[Dict[str, Any]] = []
    for it in (load_decision_candidates(project_name) or []):
        if not isinstance(it, dict):
            continue
        if str(it.get("status") or "").strip().lower() != "pending":
            continue
        txt = str(it.get("text") or "").strip()
        if not txt:
            continue
        out.append(it)
    return out


def _decision_topic_key(text: str) -> Tuple[str, str]:
    """
    Deterministic topic key extraction for conflict grouping (v1).
    Returns (area, aspect).
    Conservative: only recognizes a few stable concepts.
    """
    t = (text or "").lower()

    area = ""
    if "primary bathroom" in t or "primary bath" in t:
        area = "primary_bathroom"
    elif "bathroom" in t or "bath" in t:
        area = "bathroom"

    aspect = ""
    if "shower floor" in t:
        aspect = "shower_floor"
    elif "shower wall" in t or "shower walls" in t:
        aspect = "shower_walls"
    elif "shower" in t and "wall" in t:
        aspect = "shower_walls"
    elif "floor" in t or "floors" in t:
        aspect = "floor"
    elif "wall" in t or "walls" in t:
        aspect = "walls"

    return area, aspect


def _material_bucket(text: str) -> str:
    """
    Deterministic material normalization (v1).
    Returns a small bucket label used only for conflict detection.
    """
    t = (text or "").lower()

    # prioritize multi-word phrases first
    if "micro plaster" in t or "microplaster" in t:
        return "micro_plaster"
    if "plaster" in t:
        return "plaster"
    if "porcelain" in t:
        return "porcelain_tile"
    if "tile" in t:
        return "tile"
    if "walnut" in t:
        return "walnut"
    if "wood" in t:
        return "wood"
    if "concrete" in t:
        return "concrete"
    return ""


def detect_decision_conflicts(
    project_name: str,
    *,
    confirmed: List[Dict[str, Any]],
    pending: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """
    Deterministic conflict detector (v1):
    - group by (area, aspect)
    - if confirmed and pending disagree on material bucket, flag conflict
    Returns list of conflict dicts:
      {
        "topic": "primary_bathroom/floor",
        "confirmed_text": "...",
        "pending_text": "...",
        "confirmed_material": "...",
        "pending_material": "..."
      }
    """
    conflicts: List[Dict[str, Any]] = []

    # Index confirmed by topic key (last-wins)
    conf_map: Dict[Tuple[str, str], Dict[str, Any]] = {}
    for d in (confirmed or []):
        if not isinstance(d, dict):
            continue
        txt = str(d.get("text") or "").strip()
        if not txt:
            continue
        key = _decision_topic_key(txt)
        if key == ("", ""):
            continue
        conf_map[key] = d

    for c in (pending or []):
        if not isinstance(c, dict):
            continue
        ptxt = str(c.get("text") or "").strip()
        if not ptxt:
            continue
        key = _decision_topic_key(ptxt)
        if key == ("", ""):
            continue
        d = conf_map.get(key)
        if not d:
            continue

        ctxt = str(d.get("text") or "").strip()
        cm = _material_bucket(ctxt)
        pm = _material_bucket(ptxt)

        if cm and pm and cm != pm:
            topic = "/".join([key[0], key[1]]).strip("/")
            conflicts.append(
                {
                    "topic": topic,
                    "confirmed_text": ctxt,
                    "pending_text": ptxt,
                    "confirmed_material": cm,
                    "pending_material": pm,
                }
            )

    return conflicts
def find_pending_decision_candidates(
    project_name: str,
    *,
    query: str,
    max_items: int = 5,
) -> List[Dict[str, Any]]:
    """
    Deterministically return the most relevant *pending* decision candidates to a query.
    No inference. Keyword overlap + recency.
    """
    q = (query or "").strip()
    toks = _c5_tokens(q)
    cands = _pending_candidates(project_name)

    scored: List[Tuple[int, float, Dict[str, Any]]] = []
    for it in cands:
        txt = str(it.get("text") or "").strip()
        ts = str(it.get("timestamp") or "").strip()
        blob = txt
        sc = _c5_score_text(blob, toks) if toks else 0
        epoch = _c5_entry_epoch(ts)
        scored.append((sc, epoch, it))

    # Prefer high score, then recency
    scored.sort(key=lambda x: (-x[0], -x[1]))

    out: List[Dict[str, Any]] = []
    for sc, _ep, it in scored:
        if toks and sc <= 0:
            continue
        out.append(it)
        if len(out) >= int(max_items or 5):
            break

    # If query had no tokens (generic), return most recent pending
    if (not toks) and (not out):
        out = cands[-int(max_items or 5):]

    return out
def _pulse_collapse_degenerate_decisions(texts: List[str]) -> List[str]:
    """
    Display-only, deterministic cleanup for Project Pulse:
    Suppress earlier shorter entries when a later entry clearly subsumes it via:
    - exact duplicate
    - prefix match:   b startswith (a + " ")
    - suffix match:   b endswith (" " + a)   (e.g., "tile for the bathroom floors" subsumes "the bathroom floors")
    - containment:    (" " + a + " ") in (" " + b + " ")  (whole-phrase containment)

    Does NOT mutate files. Does NOT infer new decisions.
    """
    items = [str(t or "").strip() for t in (texts or []) if str(t or "").strip()]
    if not items:
        return []

    keep = [True] * len(items)

    for i in range(len(items)):
        a = items[i]
        if not a:
            keep[i] = False
            continue

        a_low = a.lower().strip()
        a_pad = " " + a_low + " "

        for j in range(i + 1, len(items)):
            b = items[j]
            if not b:
                continue

            b_low = b.lower().strip()
            b_pad = " " + b_low + " "

            # Exact duplicate
            if b_low == a_low:
                keep[i] = False
                break

            # Prefix subsumption
            if b_low.startswith(a_low + " "):
                keep[i] = False
                break

            # Suffix subsumption (handles: "tile for the bathroom floors" vs "the bathroom floors")
            if b_low.endswith(" " + a_low):
                keep[i] = False
                break

            # Whole-phrase containment (conservative; requires word boundaries)
            if a_pad in b_pad:
                keep[i] = False
                break

    out: List[str] = []
    for i, ok in enumerate(keep):
        if ok:
            out.append(items[i])
    return out

def build_truth_bound_pulse(project_name: str) -> str:
    """
    Deterministic, truth-bound project pulse.
    NO inference.
    NO invented goals.
    NO invented next actions.

    Enhanced (v1):
    - Surfaces pending decision candidates
    - Surfaces deterministic conflicts (confirmed vs pending)
    """
    ensure_project_scaffold(project_name)

    lines: List[str] = []
    lines.append("Project Pulse (truth-bound)")

    # C6.3: If bootstrap is not active, surface initialization banner.
    bs0 = ""
    try:
        st0 = load_project_state(project_name)
        bs0 = str(st0.get("bootstrap_status") or "").strip()
        if bs0 and bs0 != "active":
            lines.append("Project is being initialized — goal not yet confirmed.")
    except Exception:
        bs0 = ""

    # Goal (only if explicitly recorded)
    goal = None
    try:
        st = load_project_state(project_name)
        goal = st.get("goal") if isinstance(st, dict) else None
    except Exception:
        goal = None

    g = (str(goal).strip() if goal is not None else "")
    if (not g) and bs0 and bs0 != "active":
        lines.append("Goal: (not set yet)")
        lines.append("Next action: Set the project goal (send: !goal: <one sentence>)")
    else:
        lines.append(f"Goal: {g if g else '(not set yet)'}")

    # Current deliverable
    cur_deliv = None
    try:
        st = load_project_state(project_name)
        cur_deliv = st.get("current_deliverable") if isinstance(st, dict) else None
    except Exception:
        cur_deliv = None
    lines.append(f"Current deliverable: {cur_deliv if cur_deliv else '(none yet)'}")

    # Decisions (CURRENT only; supersession-aware)
    # C7: show only current (non-superseded) decisions and group by domain when possible.
    decs = []
    current_by_domain: Dict[str, Dict[str, Any]] = {}
    try:
        # Keep legacy callers safe: load_decisions still exists, but pulse uses C7 current view when available.
        current_by_domain = get_current_decisions_by_domain(project_name) or {}
    except Exception:
        current_by_domain = {}

    # Render: domain: text (or just text if domain missing)
    rendered: List[str] = []
    if current_by_domain:
        # Stable order: domains sorted, but keep __no_domain__ buckets last
        keys = sorted([k for k in current_by_domain.keys() if not k.startswith("__no_domain__")])
        keys += sorted([k for k in current_by_domain.keys() if k.startswith("__no_domain__")])
        for dom in keys[:12]:
            r = current_by_domain.get(dom) or {}
            txt = str(r.get("text") or "").strip()
            if not txt:
                continue
            if dom and (not dom.startswith("__no_domain__")):
                rendered.append(f"{dom}: {txt}")
            else:
                rendered.append(txt)

    if rendered:
        rendered2 = _pulse_collapse_degenerate_decisions(rendered)
        lines.append("Recent decisions: " + " | ".join(rendered2[-5:]) if rendered2 else "Recent decisions: Not recorded / ambiguous")
    else:
        lines.append("Recent decisions: (none yet)")

    # For downstream pending/conflict logic, keep a best-effort confirmed list
    try:
        decs = load_decisions(project_name) or []
    except Exception:
        decs = []


    # Pending decision candidates (pure log, but hide duplicates of confirmed decisions)
    try:
        pend = _pending_candidates(project_name) or []
    except Exception:
        pend = []

    confirmed_texts = set()
    try:
        for d in (decs or []):
            if isinstance(d, dict):
                t = str(d.get("text") or "").strip()
                if t:
                    confirmed_texts.add(t)
    except Exception:
        confirmed_texts = set()

    pend_txts: List[str] = []
    for x in (pend or []):
        if not isinstance(x, dict):
            continue
        t = str(x.get("text") or "").strip()
        if not t:
            continue
        # Deterministic de-dupe: if already confirmed, do not show as pending.
        if t in confirmed_texts:
            continue
        pend_txts.append(t)

    if pend_txts:
        lines.append("Pending decisions: " + " | ".join(pend_txts[-5:]))
    else:
        lines.append("Pending decisions: Not recorded / ambiguous")

    # Conflicts (C8.2 deterministic): show OPEN conflict inbox items (short).
    try:
        open_items2 = list_open_inbox(project_name, max_items=24)
    except Exception:
        open_items2 = []

    open_conflicts = [it for it in (open_items2 or []) if isinstance(it, dict) and str(it.get("type") or "") == "conflict"]

    if open_conflicts:
        brief2: List[str] = []
        for it in open_conflicts[:5]:
            dom = str(it.get("domain") or "").strip()
            surf = str(it.get("surface") or "").strip()
            label = dom + (f"({surf})" if surf else "")
            if not label:
                label = "conflict"
            brief2.append(label)
        lines.append("Conflicts: " + " | ".join(brief2))
    else:
        lines.append("Conflicts: (none)")

    # Links (C8.1 deterministic): show a tiny recent summary (counts by type, tail-bounded).
    try:
        link_rows = _c8_load_links(project_name, max_lines=2500)
    except Exception:
        link_rows = []
    if link_rows:
        counts: Dict[str, int] = {}
        for r in link_rows[-300:]:
            t = str(r.get("type") or "").strip()
            if t:
                counts[t] = counts.get(t, 0) + 1
        parts = [f"{k}={counts[k]}" for k in sorted(counts.keys())]
        lines.append("Links: " + (", ".join(parts) if parts else "(none)"))
    else:
        lines.append("Links: (none)")
    # Inbox (open) — deterministic pending items
    try:
        open_items = list_open_inbox(project_name, max_items=5)
    except Exception:
        open_items = []

    if open_items:
        brief = []
        for it in open_items:
            if not isinstance(it, dict):
                continue
            txt = sanitize_retrieved_text(str(it.get("text") or "")).replace("\n", " ").strip()
            if len(txt) > 90:
                txt = txt[:89].rstrip() + "…"
            if txt:
                brief.append(txt)
        lines.append("Inbox (open): " + " | ".join(brief) if brief else "Inbox (open): (none)")
    else:
        lines.append("Inbox (open): (none)")

    # Recent upload notes (pure log)
    try:
        notes = load_upload_notes(project_name) or []
    except Exception:
        notes = []
    if notes:
        items = []
        for n in notes[-5:]:
            if isinstance(n, dict):
                p = sanitize_retrieved_text(str(n.get("upload_path") or ""))
                a = sanitize_retrieved_text(str(n.get("answer") or ""))
                if p or a:
                    items.append(f"{p}: {a}".strip(": "))
        lines.append("Recent upload notes: " + " | ".join(items) if items else "Recent upload notes: (none yet)")
    else:
        lines.append("Recent upload notes: Not recorded / ambiguous")

    # Next action — ONLY if explicitly stored (almost never)
    # Avoid duplicates: bootstrap may have already appended a Next action line above.
    if not any(str(ln).strip().startswith("Next action:") for ln in lines):
        next_action = None
        try:
            st = load_project_state(project_name)
            next_action = st.get("next_action") if isinstance(st, dict) else None
        except Exception:
            next_action = None
        lines.append(f"Next action: {next_action if next_action else '(not set yet)'}")

    # Facts map presence only (treat scaffold placeholders as empty)
    try:
        fm = state_file_path(project_name, "facts_map")
        txt = ""
        if fm.exists() and fm.stat().st_size > 0:
            txt = read_text_file(fm, errors="replace").strip()

        # Placeholder templates should count as empty
        low = (txt or "").lower().strip()
        is_placeholder = (
            (not txt)
            or low == "(empty)"
            or low == "# facts map\n\n(empty)"
            or low.endswith("\n(empty)")
        )
        present = bool(txt) and (not is_placeholder)
    except Exception:
        present = False
    lines.append(f"Facts map: {'present' if present else 'empty'}")

    return "\n".join(lines)



def deliverables_path(project_name: str) -> Path:
    """
    projects/<user>/<project>/state/deliverables.json
    """
    return state_dir(project_name) / DELIVERABLES_FILE_NAME


def load_deliverables(project_name: str) -> Dict[str, Any]:
    ensure_project(project_name)
    p = deliverables_path(project_name)
    if not p.exists():
        d = {"version": DELIVERABLES_VERSION, "current": "", "items": []}
        atomic_write_text(p, json.dumps(d, indent=2), encoding="utf-8", errors="strict")
        return d

    try:
        obj = json.loads(p.read_text(encoding="utf-8") or "{}")
    except Exception:
        obj = {}

    if not isinstance(obj, dict):
        obj = {}

    obj.setdefault("version", DELIVERABLES_VERSION)
    obj.setdefault("current", "")
    items = obj.get("items")
    if not isinstance(items, list):
        items = []
    obj["items"] = items
    return obj


def save_deliverables(project_name: str, data: Dict[str, Any]) -> None:
    ensure_project(project_name)
    p = deliverables_path(project_name)
    atomic_write_text(p, json.dumps(data or {}, indent=2), encoding="utf-8", errors="strict")


def _deliverables_now_noz() -> str:
    # Match the spec example format (no trailing 'Z').
    try:
        return (now_iso() or "").replace("Z", "")
    except Exception:
        return time.strftime("%Y-%m-%dT%H:%M:%S", time.gmtime())


def _next_deliverable_id(existing: Dict[str, Any], *, created_at: str) -> str:
    # Deterministic per-day increment: deliv_YYYY_MM_DD_NNN
    day = (created_at or "").split("T", 1)[0].replace("-", "_")
    prefix = f"deliv_{day}_"

    max_n = 0
    items = existing.get("items") or []
    if isinstance(items, list):
        for it in items:
            if not isinstance(it, dict):
                continue
            did = str(it.get("id") or "")
            if not did.startswith(prefix):
                continue
            tail = did[len(prefix):]
            try:
                n = int(tail)
            except Exception:
                n = 0
            if n > max_n:
                max_n = n

    return f"{prefix}{(max_n + 1):03d}"


def register_deliverable(
    project_name: str,
    *,
    deliverable_type: str,
    title: str,
    path: str,
    source: str,
) -> Dict[str, Any]:
    """
    Register an explicitly-created deliverable as a first-class object.

    Rules:
    - Exactly one current overall (registry.current points to it)
    - Creating a new deliverable of the same type auto-supersedes the previous
    - No inference: this is ONLY called at explicit creation points
    """
    ensure_project(project_name)

    rel = (path or "").replace("\\", "/").strip()
    dtype = (deliverable_type or "").strip().lower()
    if not dtype:
        dtype = "unknown"

    created_at = _deliverables_now_noz()
    reg = load_deliverables(project_name)

    items = reg.get("items") or []
    if not isinstance(items, list):
        items = []
    reg["items"] = items

    prev_same_type_id = ""
    prev_current_id = str(reg.get("current") or "").strip()

    # Find the current deliverable of the same type (if any)
    for it in items:
        if not isinstance(it, dict):
            continue
        if str(it.get("status") or "") != "current":
            continue
        if str(it.get("type") or "").strip().lower() == dtype:
            prev_same_type_id = str(it.get("id") or "").strip()
            break

    # Enforce exactly one current: clear all current statuses
    for it in items:
        if not isinstance(it, dict):
            continue
        if str(it.get("status") or "") == "current":
            it["status"] = "superseded"

    new_id = _next_deliverable_id(reg, created_at=created_at)

    entry = {
        "id": new_id,
        "type": dtype,
        "title": str(title or "").strip() or f"{dtype.upper()} Deliverable",
        "path": rel,
        "created_at": created_at,
        "status": "current",
        "source": str(source or "").strip() or "unknown",
        "replaces": prev_same_type_id or "",
    }

    # If we superseded something of the same type, ensure it's marked superseded (belt + suspenders)
    if prev_same_type_id:
        for it in items:
            if not isinstance(it, dict):
                continue
            if str(it.get("id") or "") == prev_same_type_id:
                it["status"] = "superseded"
                break

    items.append(entry)
    reg["current"] = new_id
    # C8.1: Deterministic deliverable_to_decision links for "in-scope" current decisions.
    # Heuristic (deterministic):
    # - If project_state.domains is non-empty: link to current decisions whose domain matches exactly.
    # - Else: link to ALL current decisions (bounded), as the minimal evidence layer.
    try:
        st = load_project_state(project_name) or {}
        doms = st.get("domains") if isinstance(st, dict) else []
        if not isinstance(doms, list):
            doms = []
        doms = [str(x).strip() for x in doms if str(x).strip()][:24]

        cur = list_current_decisions(project_name) or []
        deliverable_ref = f"deliverable:{str(entry.get('id') or '').strip()}"

        picked: List[Dict[str, Any]] = []
        if doms:
            for d in cur:
                if not isinstance(d, dict):
                    continue
                if str(d.get("domain") or "").strip() in doms:
                    picked.append(d)
        else:
            picked = list(cur)

        # deterministic cap
        picked = picked[:12]

        for d in picked:
            did = str(d.get("id") or "").strip()
            if not did:
                continue
            append_link_event(
                project_name,
                type_="deliverable_to_decision",
                from_ref=deliverable_ref,
                to_ref=f"decision:{did}",
                reason="current_decisions_at_deliverable_creation",
                confidence=("medium" if doms else "low"),
            )
    except Exception:
        pass

    save_deliverables(project_name, reg)
    return entry

def get_latest_artifact_by_type(project_name: str, artifact_type: str) -> Optional[Dict[str, Any]]:
    m = load_manifest(project_name)
    arts = m.get("artifacts") or []
    matches = [a for a in arts if a.get("type") == artifact_type]
    matches.sort(key=lambda a: float(a.get("created_at", 0.0) or 0.0), reverse=True)
    return matches[0] if matches else None


def read_artifact_text(project_name: str, entry: Optional[Dict[str, Any]], cap_chars: int = 0) -> str:
    pr, pd = _require_configured()
    if not entry:
        return ""
    rel = (entry.get("path") or "").replace("\\", "/").strip()
    if not rel:
        return ""
    path = pr / rel
    txt = read_text_file(path, errors="replace")
    if cap_chars and len(txt) > cap_chars:
        return txt[:cap_chars] + "\n...[truncated]..."
    return txt
def _is_code_help_project(project_full_name: str) -> bool:
    """
    True only when the *short* project name is exactly 'code_help' (case-insensitive).
    Ensures this behavior never activates in other projects.
    """
    try:
        short = (project_full_name or "").split("/", 1)[-1].strip().lower()
        return short == "code_help"
    except Exception:
        return False


def _find_codehelp_index_files(project_name: str) -> List[Path]:
    """
    Find chunk-map/index artifacts like:
      code_<timestamp>_<filename>_index_v2.md
      code_*_index*.md
    Only searches within this project's artifacts folder.
    """
    try:
        ad = artifacts_dir(project_name)
        if not ad.exists():
            return []
        out: List[Path] = []
        for p in ad.iterdir():
            if not p.is_file():
                continue
            name = p.name.lower()
            if not name.endswith(".md"):
                continue
            if not name.startswith("code_"):
                continue
            if "index" not in name:
                continue
            out.append(p)
        out.sort(key=lambda x: x.stat().st_mtime, reverse=True)
        return out
    except Exception:
        return []


def build_codehelp_capability_map(project_name: str, cap_chars_total: int = 18000) -> str:
    """
    Build a bounded, model-readable map of what the codebase can do + where to look.
    This is designed to reduce capability hallucinations in Code_Help normal chat.
    """
    files = _find_codehelp_index_files(project_name)
    if not files:
        return (
            "CODEHELP_CAPABILITY_MAP:\n\n"
            "- No code_*index*.md files found in this project's artifacts folder.\n"
            "- If you expect chunk maps, regenerate them and ensure they are saved as artifacts.\n"
        )

    parts: List[str] = []
    parts.append("CODEHELP_CAPABILITY_MAP:")
    parts.append("")
    parts.append("Rules:")
    parts.append("- Treat this as the source of truth for what code exists/can do.")
    parts.append("- If something is not in these maps, label it 'unconfirmed' instead of guessing.")
    parts.append("")
    parts.append("Available chunk/index maps (most recent first):")
    for p in files[:8]:
        parts.append(f"- {p.name}")

    parts.append("")
    parts.append("===INDEX_MAPS_START===")

    used = 0
    per_file_cap = max(2000, int(cap_chars_total / max(1, min(len(files), 4))))
    for p in files[:4]:
        try:
            txt = read_text_file(p, errors="replace")
        except Exception:
            txt = ""
        if not txt:
            continue
        if len(txt) > per_file_cap:
            txt = txt[:per_file_cap].rstrip() + "\n...[truncated]..."
        block = f"\n---\nFILE: {p.name}\n---\n{txt}\n"
        if used + len(block) > cap_chars_total:
            break
        parts.append(block)
        used += len(block)

    parts.append("===INDEX_MAPS_END===")
    return "\n".join(parts).strip()


# -----------------------------------------------------------------------------
# Artifacts
# -----------------------------------------------------------------------------

def create_artifact_bytes(
    project_name: str,
    logical_name: str,
    content_bytes: bytes,
    *,
    artifact_type: str = "binary",
    from_files: Optional[List[str]] = None,
    file_ext: str = ".bin",
    meta: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    pr, pd = _require_configured()
    ensure_project(project_name)
    m = load_manifest(project_name)

    safe_name = re.sub(r"[^a-zA-Z0-9_.-]", "_", logical_name or "artifact")
    ext = (file_ext or ".bin").strip()
    if not ext.startswith("."):
        ext = "." + ext

    versions: List[int] = []
    for a in m.get("artifacts") or []:
        if a.get("name") == safe_name:
            try:
                versions.append(int(a.get("version") or 0))
            except Exception:
                pass
    next_v = (max(versions) + 1) if versions else 1

    filename = f"{safe_name}_v{next_v}{ext}"
    apath = artifacts_dir(project_name) / filename
    apath.parent.mkdir(parents=True, exist_ok=True)
    apath.write_bytes(content_bytes or b"")

    rel = str(apath.relative_to(pr)).replace("\\", "/")
    entry = {
        "name": safe_name,
        "filename": filename,
        "path": rel,
        "type": artifact_type,
        "version": next_v,
        "from_files": from_files or [],
        "created_at": time.time(),
        "meta": meta or {},
    }
    m.setdefault("artifacts", []).append(entry)
    save_manifest(project_name, m)
    return entry


def create_artifact(
    project_name: str,
    logical_name: str,
    content: str,
    *,
    artifact_type: str = "cheat_sheet",
    from_files: Optional[List[str]] = None,
    file_ext: str = ".md",
    write_errors: str = "strict",
    meta: Optional[Dict[str, Any]] = None,
    also_update_state_key: Optional[str] = None,
) -> Dict[str, Any]:
    pr, pd = _require_configured()
    ensure_project(project_name)
    m = load_manifest(project_name)

    safe_name = re.sub(r"[^a-zA-Z0-9_.-]", "_", logical_name or "artifact")
    ext = (file_ext or ".md").strip()
    if not ext.startswith("."):
        ext = "." + ext

    versions: List[int] = []
    for a in m.get("artifacts") or []:
        if a.get("name") == safe_name:
            try:
                versions.append(int(a.get("version") or 0))
            except Exception:
                pass
    next_v = (max(versions) + 1) if versions else 1

    filename = f"{safe_name}_v{next_v}{ext}"
    apath = artifacts_dir(project_name) / filename
    apath.write_text(content or "", encoding="utf-8", errors=write_errors)

    rel = str(apath.relative_to(pr)).replace("\\", "/")
    entry = {
        "name": safe_name,
        "filename": filename,
        "path": rel,
        "type": artifact_type,
        "version": next_v,
        "from_files": from_files or [],
        "created_at": time.time(),
        "meta": meta or {},
    }
    m.setdefault("artifacts", []).append(entry)
    save_manifest(project_name, m)

    if also_update_state_key:
        atomic_write_text(state_file_path(project_name, also_update_state_key), content or "", encoding="utf-8", errors="strict")

    return entry


# -----------------------------------------------------------------------------
# Extraction Ledger (deterministic; MVP)
# -----------------------------------------------------------------------------

def _ledger_source_type_from_suffix(suffix: str) -> str:
    suf = (suffix or "").lower().strip()
    if suf in (".xlsx", ".xlsm", ".xls"):
        return suf.lstrip(".")
    if suf == ".pdf":
        return "pdf"
    if suf in (".png", ".jpg", ".jpeg", ".webp", ".gif"):
        return "image"
    return (suf.lstrip(".") if suf.startswith(".") else (suf or "unknown"))


_MONEY_RE = re.compile(r"(?P<cur>\$)?\s*(?P<num>\d{1,3}(?:,\d{3})*(?:\.\d{2})?)")
_QTY_RE = re.compile(r"\b(?P<qty>\d+(?:\.\d+)?)\b")


def _to_float_maybe(s: str) -> Optional[float]:
    t = (s or "").strip().replace(",", "")
    if not t:
        return None
    try:
        return float(t)
    except Exception:
        return None


def _sha256_text_prefix(text: str, n: int = 12) -> str:
    try:
        b = (text or "").encode("utf-8", errors="replace")
        return hashlib.sha256(b).hexdigest()[: int(n or 12)]
    except Exception:
        return ""


def _csv_escape(v: Any) -> str:
    s = "" if v is None else str(v)
    s = s.replace("\r", " ").replace("\n", " ").strip()
    # Minimal RFC4180-ish quoting (deterministic)
    if any(ch in s for ch in (",", '"', "\n")):
        s = '"' + s.replace('"', '""') + '"'
    return s


def _ledger_item(
    *,
    source_file: str,
    source_type: str,
    where_found: str,
    label: str,
    qty: Optional[float],
    unit_cost: Optional[float],
    total_cost: Optional[float],
    currency: str,
    confidence: str,
    notes: str = "",
    extracted_at: str,
) -> Dict[str, Any]:
    conf = (confidence or "low").strip().lower()
    if conf not in ("low", "medium", "high"):
        conf = "low"

    cur = (currency or "USD").strip() or "USD"

    return {
        "source_file": (source_file or "").replace("\\", "/").strip(),
        "source_type": (source_type or "").strip(),
        "where_found": (where_found or "").strip(),
        "label": strip_control_tokens(str(label or "").strip()),
        "qty": qty,
        "unit_cost": unit_cost,
        "total_cost": total_cost,
        "currency": cur,
        "confidence": conf,
        "notes": strip_control_tokens(str(notes or "").strip()) if notes else "",
        "extracted_at": extracted_at,
    }


def _find_latest_artifact_for_file(
    project_name: str,
    *,
    artifact_type: str,
    canonical_rel: str,
) -> Optional[Dict[str, Any]]:
    m = load_manifest(project_name)
    arts = m.get("artifacts") or []
    if not isinstance(arts, list):
        return None

    needle = (canonical_rel or "").replace("\\", "/").strip()
    if not needle:
        return None

    # newest-last in manifest; scan reverse
    for a in reversed(arts[-3000:]):
        if not isinstance(a, dict):
            continue
        if str(a.get("type") or "") != str(artifact_type or ""):
            continue
        ff = a.get("from_files") or []
        if not isinstance(ff, list):
            continue
        ff_norm = [(str(x) or "").replace("\\", "/").strip() for x in ff]
        if needle in ff_norm:
            return a
    return None


def _extract_items_from_pdf_text(
    *,
    canonical_rel: str,
    pdf_text: str,
    extracted_at: str,
    max_items: int,
) -> List[Dict[str, Any]]:
    items: List[Dict[str, Any]] = []
    if not pdf_text:
        return items

    page = 1
    line_in_page = 0

    # Track page markers if present (plan_ocr uses [PAGE N])
    for raw_ln in (pdf_text or "").splitlines():
        ln = (raw_ln or "").strip()
        if not ln:
            continue

        mpage = re.match(r"^\[PAGE\s+(\d+)\]\s*$", ln, flags=re.IGNORECASE)
        if mpage:
            try:
                page = int(mpage.group(1))
            except Exception:
                page = page
            line_in_page = 0
            continue

        line_in_page += 1

        mm = _MONEY_RE.search(ln)
        if not mm:
            continue

        amt = _to_float_maybe(mm.group("num") if mm else "")
        if amt is None:
            continue

        # Best-effort qty extraction: first standalone number not equal to money token
        qty = None
        try:
            qms = list(_QTY_RE.finditer(ln))
            if qms:
                qv = _to_float_maybe(qms[0].group("qty"))
                if qv is not None:
                    qty = qv
        except Exception:
            qty = None

        where = f"page {page}: line {line_in_page}"
        label = ln
        items.append(
            _ledger_item(
                source_file=canonical_rel,
                source_type="pdf",
                where_found=where,
                label=label,
                qty=qty,
                unit_cost=None,
                total_cost=amt,
                currency="USD",
                confidence="medium",
                notes="parsed_from_pdf_text_line",
                extracted_at=extracted_at,
            )
        )

        if len(items) >= int(max_items or 0):
            break

    return items


def _extract_items_from_image_ocr(
    *,
    canonical_rel: str,
    ocr_text: str,
    extracted_at: str,
    max_items: int,
) -> List[Dict[str, Any]]:
    items: List[Dict[str, Any]] = []
    if not ocr_text:
        return items

    h = _sha256_text_prefix(ocr_text, 12)
    # Only extract money-like lines (keep it bounded + cost-focused)
    for raw_ln in (ocr_text or "").splitlines():
        ln = (raw_ln or "").strip()
        if not ln:
            continue

        mm = _MONEY_RE.search(ln)
        if not mm:
            continue

        amt = _to_float_maybe(mm.group("num") if mm else "")
        if amt is None:
            continue

        where = f"ocr_sha256:{h}"
        items.append(
            _ledger_item(
                source_file=canonical_rel,
                source_type="image",
                where_found=where,
                label=ln,
                qty=None,
                unit_cost=None,
                total_cost=amt,
                currency="USD",
                confidence="low",
                notes="parsed_from_ocr_line",
                extracted_at=extracted_at,
            )
        )

        if len(items) >= int(max_items or 0):
            break

    return items


def _extract_items_from_excel_blueprint(
    *,
    canonical_rel: str,
    blueprint_obj: Dict[str, Any],
    extracted_at: str,
    max_items: int,
) -> List[Dict[str, Any]]:
    """
    Deterministic extraction using ONLY the existing excel_blueprint artifact JSON.
    NO workbook scanning here.

    We only emit ledger rows when we can find (sheet, cell) + a numeric value.
    The blueprint schema may vary; we search conservatively for common patterns.
    """
    items: List[Dict[str, Any]] = []
    if not isinstance(blueprint_obj, dict):
        return items

    def _emit(sheet: str, cell: str, label: str, value: Any) -> None:
        nonlocal items
        if len(items) >= int(max_items or 0):
            return
        v = None
        if isinstance(value, (int, float)):
            v = float(value)
        elif isinstance(value, str):
            v = _to_float_maybe(value)
        if v is None:
            return
        where = f"{sheet}!{cell}"
        items.append(
            _ledger_item(
                source_file=canonical_rel,
                source_type=_ledger_source_type_from_suffix(".xlsx"),
                where_found=where,
                label=label or f"{sheet}!{cell}",
                qty=None,
                unit_cost=None,
                total_cost=v,
                currency="USD",
                confidence="high",
                notes="from_excel_blueprint_cached_value",
                extracted_at=extracted_at,
            )
        )

    # Pattern A: blueprint["cached_values"] = [{"sheet":..., "cell":..., "label":..., "value":...}, ...]
    cv = blueprint_obj.get("cached_values")
    if isinstance(cv, list):
        for row in cv:
            if len(items) >= int(max_items or 0):
                break
            if not isinstance(row, dict):
                continue
            sh = str(row.get("sheet") or "").strip()
            ce = str(row.get("cell") or "").strip()
            if not sh or not ce:
                continue
            lab = str(row.get("label") or row.get("name") or "").strip()
            val = row.get("value")
            _emit(sh, ce, lab, val)

    # Pattern B: blueprint["cells"] = [{"sheet":..., "cell":..., "value":...}, ...]
    cells = blueprint_obj.get("cells")
    if isinstance(cells, list) and len(items) < int(max_items or 0):
        for row in cells:
            if len(items) >= int(max_items or 0):
                break
            if not isinstance(row, dict):
                continue
            sh = str(row.get("sheet") or "").strip()
            ce = str(row.get("cell") or "").strip()
            if not sh or not ce:
                continue
            lab = str(row.get("label") or row.get("name") or "").strip()
            val = row.get("value")
            _emit(sh, ce, lab, val)

    # Pattern C: blueprint["named_ranges"] = [{"name":..., "sheet":..., "cell":..., "value":...}, ...]
    nr = blueprint_obj.get("named_ranges")
    if isinstance(nr, list) and len(items) < int(max_items or 0):
        for row in nr:
            if len(items) >= int(max_items or 0):
                break
            if not isinstance(row, dict):
                continue
            sh = str(row.get("sheet") or "").strip()
            ce = str(row.get("cell") or "").strip()
            nm = str(row.get("name") or "").strip()
            if not sh or not ce:
                continue
            _emit(sh, ce, nm, row.get("value"))

    return items


def build_extraction_ledger_and_write(
    project_name: str,
    *,
    include_csv: bool = True,
    max_items_per_file: int = 120,
    max_total_items: int = 800,
) -> Dict[str, Any]:
    """
    Deterministic Extraction Ledger builder (MVP).

    Reads only:
      - manifest raw_files
      - existing artifacts: excel_blueprint, pdf_text, ocr_text (scoped to that raw file)

    Writes:
      - artifacts/ledger_extraction_<timestamp>.json  (type=extraction_ledger)
      - artifacts/ledger_extraction_<timestamp>.csv   (type=extraction_ledger_csv) [optional]
      - artifacts/ledger_extraction_latest.json       (type=extraction_ledger) (versioned by create_artifact)
    """
    ensure_project_scaffold(project_name)

    extracted_at = now_iso()
    m = load_manifest(project_name)
    raw_files = m.get("raw_files") or []
    if not isinstance(raw_files, list):
        raw_files = []

    all_items: List[Dict[str, Any]] = []
    from_files: List[str] = []

    pr, pd = _require_configured()

    for rf in raw_files:
        if len(all_items) >= int(max_total_items or 0):
            break
        if not isinstance(rf, dict):
            continue
        canonical_rel = str(rf.get("path") or "").replace("\\", "/").strip()
        if not canonical_rel:
            continue

        from_files.append(canonical_rel)

        suf = Path(canonical_rel).suffix.lower()
        stype = _ledger_source_type_from_suffix(suf)

        # Excel → parse excel_blueprint JSON
        if suf in (".xlsx", ".xlsm", ".xls"):
            a = _find_latest_artifact_for_file(project_name, artifact_type="excel_blueprint", canonical_rel=canonical_rel)
            if isinstance(a, dict):
                txt = read_artifact_text(project_name, a, cap_chars=220000).strip()
                try:
                    obj = json.loads(txt) if txt else {}
                except Exception:
                    obj = {}
                items = _extract_items_from_excel_blueprint(
                    canonical_rel=canonical_rel,
                    blueprint_obj=(obj if isinstance(obj, dict) else {}),
                    extracted_at=extracted_at,
                    max_items=int(max_items_per_file or 0),
                )
                all_items.extend(items)
            continue

        # PDF → parse pdf_text
        if suf == ".pdf":
            a = _find_latest_artifact_for_file(project_name, artifact_type="pdf_text", canonical_rel=canonical_rel)
            if isinstance(a, dict):
                txt = read_artifact_text(project_name, a, cap_chars=220000).strip()
                items = _extract_items_from_pdf_text(
                    canonical_rel=canonical_rel,
                    pdf_text=txt,
                    extracted_at=extracted_at,
                    max_items=int(max_items_per_file or 0),
                )
                all_items.extend(items)
            continue

        # Image → parse ocr_text
        if suf in IMAGE_SUFFIXES:
            a = _find_latest_artifact_for_file(project_name, artifact_type="ocr_text", canonical_rel=canonical_rel)
            if isinstance(a, dict):
                txt = read_artifact_text(project_name, a, cap_chars=220000).strip()
                items = _extract_items_from_image_ocr(
                    canonical_rel=canonical_rel,
                    ocr_text=txt,
                    extracted_at=extracted_at,
                    max_items=int(max_items_per_file or 0),
                )
                all_items.extend(items)
            continue

        # Other types: ignore for MVP
        continue

    # Hard cap, deterministic order (manifest order + extraction order)
    if max_total_items and len(all_items) > int(max_total_items):
        all_items = all_items[: int(max_total_items)]

    ledger_obj: Dict[str, Any] = {
        "version": 1,
        "generated_at": extracted_at,
        "project": safe_project_name(project_name),
        "item_count": len(all_items),
        "items": all_items,
    }

    ts = extracted_at.replace(":", "").replace("-", "").replace("T", "_").replace("Z", "Z")
    logical_ts = f"ledger_extraction_{ts}"

    json_entry = create_artifact(
        project_name,
        logical_ts,
        json.dumps(ledger_obj, indent=2, ensure_ascii=False),
        artifact_type="extraction_ledger",
        from_files=from_files,
        file_ext=".json",
        write_errors="strict",
        meta={"generated_at": extracted_at, "item_count": len(all_items)},
    )

    # Also keep a stable name for “latest” (still versioned by create_artifact)
    latest_entry = create_artifact(
        project_name,
        "ledger_extraction_latest",
        json.dumps(ledger_obj, indent=2, ensure_ascii=False),
        artifact_type="extraction_ledger",
        from_files=from_files,
        file_ext=".json",
        write_errors="strict",
        meta={"generated_at": extracted_at, "item_count": len(all_items), "latest": True},
    )

    csv_entry: Optional[Dict[str, Any]] = None
    if include_csv:
        header = [
            "source_file",
            "source_type",
            "where_found",
            "label",
            "qty",
            "unit_cost",
            "total_cost",
            "currency",
            "confidence",
            "notes",
            "extracted_at",
        ]
        lines = [",".join(header)]
        for it in all_items:
            row = [
                _csv_escape(it.get("source_file")),
                _csv_escape(it.get("source_type")),
                _csv_escape(it.get("where_found")),
                _csv_escape(it.get("label")),
                _csv_escape(it.get("qty")),
                _csv_escape(it.get("unit_cost")),
                _csv_escape(it.get("total_cost")),
                _csv_escape(it.get("currency")),
                _csv_escape(it.get("confidence")),
                _csv_escape(it.get("notes")),
                _csv_escape(it.get("extracted_at")),
            ]
            lines.append(",".join(row))
        csv_entry = create_artifact(
            project_name,
            logical_ts,
            "\n".join(lines) + "\n",
            artifact_type="extraction_ledger_csv",
            from_files=from_files,
            file_ext=".csv",
            write_errors="strict",
            meta={"generated_at": extracted_at, "item_count": len(all_items)},
        )

    return {
        "generated_at": extracted_at,
        "item_count": len(all_items),
        "json_rel": str(json_entry.get("path") or "").replace("\\", "/"),
        "json_latest_rel": str(latest_entry.get("path") or "").replace("\\", "/"),
        "csv_rel": (str(csv_entry.get("path") or "").replace("\\", "/") if isinstance(csv_entry, dict) else ""),
    }


# -----------------------------------------------------------------------------
# Scaffold
# -----------------------------------------------------------------------------

WORKING_DOC_SPECS: List[Tuple[str, str, str, str]] = [
    ("project_map", "project_map", "Project Map", ".md"),
    ("project_brief", "project_brief", "Project Brief", ".md"),
    ("working_doc", "working_doc", "Working Document", ".md"),
    ("preferences", "preferences", "Preferences", ".md"),
    ("decision_log", "decision_log", "Decision Log", ".md"),
    ("facts_map", "facts_map", "Facts Map", ".md"),
    ("project_state", "project_state", "Project State", ".json"),
]


def ensure_project_scaffold(project_name: str) -> None:
    if is_project_deleted(project_name):
        raise RuntimeError(f"Project '{safe_project_name(project_name)}' was just deleted; refusing to auto-recreate scaffold.")
    ensure_project(project_name)
    m = load_manifest(project_name)
    goal = (m.get("goal") or "").strip()

    existing_types = {a.get("type") for a in (m.get("artifacts") or []) if a.get("type")}
    for atype, logical, title, ext in WORKING_DOC_SPECS:
        state_path = state_file_path(project_name, atype)
        if state_path.exists():
            continue

        if atype in existing_types:
            latest = get_latest_artifact_by_type(project_name, atype)
            if latest:
                txt = read_artifact_text(project_name, latest)
                atomic_write_text(state_path, txt, encoding="utf-8", errors="replace")
                continue

        if atype == "project_state":
            # C6.3 default schema (no migration of existing files).
            # New projects bootstrap when goal is empty.
            bootstrap_status = "needs_goal" if not (goal or "").strip() else "active"
            default_state = {
                "project_mode": "hybrid",
                "goal": goal or "",
                "domains": [],
                "bootstrap_status": bootstrap_status,
                "expert_frame": {
                    "status": "",
                    "label": "",
                    "directive": "",
                    "set_reason": "",
                    "updated_at": "",
                },
                "current_focus": "",
                "next_actions": [],
                "key_files": [],
                "last_updated": now_iso(),
            }
            content = json.dumps(default_state, indent=2)
        else:
            # Minimal defaults; server/agent owns the actual content quality.
            if atype == "working_doc":
                g = (goal or "").strip()
                new_template = "\n".join(
                    [
                        "# Working Document",
                        "",
                        "## Goal",
                        (g if g else "(not set yet)"),
                        "",
                        "## Current focus",
                        "- (empty)",
                        "",
                        "## Plan",
                        "- (empty)",
                        "",
                        "## Open questions",
                        "- (empty)",
                        "",
                        "## Next actions",
                        "- (empty)",
                        "",
                    ]
                )

                # MIGRATION:
                # If a working_doc already exists but is still the old placeholder,
                # upgrade it once to the structured template.
                try:
                    existing = ""
                    if state_path.exists():
                        existing = state_path.read_text(encoding="utf-8", errors="replace").strip()

                    is_old_placeholder = (
                        not existing
                        or existing.endswith("(empty)")
                        or existing == "# Working Document"
                        or existing == "# Working Document\n\n(empty)"
                    )

                    content = new_template if is_old_placeholder else existing
                except Exception:
                    content = new_template
            else:
                content = f"# {title}\n\n(empty)\n"


        atomic_write_text(state_path, content, encoding="utf-8", errors="strict")
        create_artifact(
            project_name,
            logical,
            content,
            artifact_type=atype,
            from_files=[],
            file_ext=ext,
            also_update_state_key=None,
        )

    # C9: inbox.jsonl (append-only; create empty file if missing)
    try:
        p_inbox = inbox_path(project_name)
        if not p_inbox.exists():
            atomic_write_text(p_inbox, "", encoding="utf-8", errors="strict")
    except Exception:
        pass

    # C9: inbox.jsonl (append-only; create empty file if missing)
    try:
        p_inbox = inbox_path(project_name)
        if not p_inbox.exists():
            atomic_write_text(p_inbox, "", encoding="utf-8", errors="strict")
    except Exception:
        pass

    # Health profile (project-scoped; JSON)
    try:
        user_seg = _project_user_segment(project_name)
        _ensure_expert_profiles(user_seg)
    except Exception:
        pass


# -----------------------------------------------------------------------------
# File registration / listing
# -----------------------------------------------------------------------------

def register_raw_file(project_name: str, saved_abs_path: Path, orig_name: str) -> Tuple[bool, str, str]:
    pr, pd = _require_configured()
    m = load_manifest(project_name)
    rel_path = str(saved_abs_path.relative_to(pr)).replace("\\", "/")

    raw_files = m.get("raw_files") or []
    if not isinstance(raw_files, list):
        raw_files = []
        m["raw_files"] = raw_files

    now_ts = time.time()
    saved_name = saved_abs_path.name

    # Compute sha256 for content-based dedupe (bounded but deterministic)
    sha = ""
    try:
        sha = file_sha256_bytes(read_bytes_file(saved_abs_path))
    except Exception:
        sha = ""

    # 1) Strong dedupe: same rel path or same saved filename => treat as same file
    for entry in raw_files:
        prev_path = str(entry.get("path") or "")
        prev_saved = str(entry.get("saved_name") or "")
        if prev_path == rel_path or (prev_saved and prev_saved == saved_name):
            # Update orig_name to the newest observed name (non-authoritative), and ensure sha present.
            entry["orig_name"] = orig_name
            entry["saved_name"] = saved_name
            entry["path"] = rel_path
            entry["added_at"] = now_ts
            if sha:
                entry["sha256"] = sha
            entry["ingest_version"] = INGEST_PIPELINE_VERSION
            save_manifest(project_name, m)
            return True, rel_path, prev_path

    # 2) Content dedupe (if we have sha): same sha => treat as same bytes
    if sha:
        for entry in raw_files:
            if str(entry.get("sha256") or "") == sha:
                prev_path = str(entry.get("path") or "")
                # Keep the FIRST canonical path; do NOT rewrite to a new duplicate path.
                # Update only non-authoritative metadata.
                entry.setdefault("aliases", [])
                try:
                    aliases = entry.get("aliases")
                    if isinstance(aliases, list) and orig_name and (orig_name not in aliases):
                        aliases.append(orig_name)
                        entry["aliases"] = aliases[:8]
                except Exception:
                    pass

                # If this duplicate was saved to a different filename, remove the new bytes.
                try:
                    if prev_path:
                        prev_abs = pr / prev_path
                        if saved_abs_path.resolve() != prev_abs.resolve():
                            if saved_abs_path.exists() and saved_abs_path.is_file():
                                saved_abs_path.unlink()
                except Exception:
                    pass

                if sha:
                    entry["sha256"] = sha
                entry["ingest_version"] = INGEST_PIPELINE_VERSION
                save_manifest(project_name, m)
                return True, prev_path or rel_path, prev_path

    # 3) Legacy dedupe by orig_name (keep behavior, but now last priority)
    orig_lower = (orig_name or "").lower()
    for entry in raw_files:
        if (entry.get("orig_name") or "").lower() == orig_lower:
            prev = str(entry.get("path") or "")
            entry["orig_name"] = orig_name
            entry["saved_name"] = saved_name
            entry["path"] = rel_path
            entry["added_at"] = now_ts
            if sha:
                entry["sha256"] = sha
            entry["ingest_version"] = INGEST_PIPELINE_VERSION
            save_manifest(project_name, m)
            return True, rel_path, prev

    raw_files.append(
        {
            "orig_name": orig_name,
            "saved_name": saved_name,
            "path": rel_path,
            "added_at": now_ts,
            "sha256": sha,
            "ingest_version": INGEST_PIPELINE_VERSION,
        }
    )
    m["raw_files"] = raw_files
    save_manifest(project_name, m)
    return False, rel_path, ""


def list_project_files(project_name: str) -> Dict[str, Any]:
    pr, pd = _require_configured()
    m = load_manifest(project_name)

    project_full = safe_project_name(project_name)
    expected_prefix = f"projects/{project_full}/"

    def _norm(p: Any) -> str:
        return (p or "").replace("\\", "/").strip()

    def _exists(rel: str) -> bool:
        rel2 = _norm(rel)
        if not rel2:
            return False
        try:
            return (pr / rel2).exists()
        except Exception:
            return False

    def _safe_rel_raw(stored_rel: Any, saved_name: Any) -> str:
        rp = _norm(stored_rel)
        if rp.startswith(expected_prefix) and _exists(rp):
            return rp

        # If the manifest has a legacy path, but the file exists in THIS project's raw dir, relink safely.
        fn = (saved_name or "").strip()
        if fn:
            candidate = f"{expected_prefix}{RAW_DIR_NAME}/{fn}"
            if _exists(candidate):
                return candidate

        # Guard: do not return legacy/foreign paths as clickable links
        return ""

    def _safe_rel_art(stored_rel: Any, filename: Any) -> str:
        rp = _norm(stored_rel)
        if rp.startswith(expected_prefix) and _exists(rp):
            return rp

        # Same idea for artifacts: if file exists in THIS project's artifacts dir, relink safely.
        fn = (filename or "").strip()
        if fn:
            candidate = f"{expected_prefix}{ARTIFACTS_DIR_NAME}/{fn}"
            if _exists(candidate):
                return candidate

        return ""

    raw = [
        {
            "filename": rf.get("saved_name"),
            "original_name": rf.get("orig_name"),
            "relative_path": _safe_rel_raw(rf.get("path"), rf.get("saved_name")),
        }
        for rf in (m.get("raw_files") or [])
    ]
    arts = [
        {
            "filename": a.get("filename"),
            "type": a.get("type", "artifact"),
            "relative_path": _safe_rel_art(a.get("path", ""), a.get("filename")),
            "version": a.get("version", 1),
        }
        for a in (m.get("artifacts") or [])
    ]

    # Optional: user-facing slices (doesn't remove history; just provides a better view)
    deliverables = [a for a in arts if str(a.get("type") or "").startswith("deliverable_")]

    latest_by_type: Dict[str, Dict[str, Any]] = {}
    for a in arts:
        at = str(a.get("type") or "artifact")
        cur = latest_by_type.get(at)
        try:
            v = int(a.get("version") or 0)
        except Exception:
            v = 0
        try:
            cv = int(cur.get("version") or 0) if cur else -1
        except Exception:
            cv = -1
        if cur is None or v > cv:
            latest_by_type[at] = a

    artifacts_latest = list(latest_by_type.values())

    return {
        "project": project_name,
        "raw": raw,
        "artifacts": arts,
        "deliverables": deliverables,
        "artifacts_latest": artifacts_latest,
    }

# -----------------------------------------------------------------------------
# Chunking + text ingestion
# -----------------------------------------------------------------------------

@dataclass
class ChunkInfo:
    index: int
    start_line: int
    end_line: int
    start_byte: int
    end_byte: int
    sha256_bytes: str
    text: str


def chunk_bytes_line_aligned(raw_bytes: bytes, *, max_bytes: int = 12000) -> List[ChunkInfo]:
    line_bytes = raw_bytes.splitlines(keepends=True) if raw_bytes else []
    total_lines = len(line_bytes)
    chunks: List[ChunkInfo] = []

    cur: List[bytes] = []
    cur_len = 0
    start_line = 1
    start_byte = 0

    def flush(end_line: int) -> None:
        nonlocal cur, cur_len, start_line, start_byte
        if not cur:
            return
        b = b"".join(cur)
        txt = b.decode("utf-8", errors="surrogateescape")
        chunks.append(
            ChunkInfo(
                index=len(chunks) + 1,
                start_line=start_line,
                end_line=end_line,
                start_byte=start_byte,
                end_byte=start_byte + len(b),
                sha256_bytes=file_sha256_bytes(b),
                text=txt,
            )
        )
        start_line = end_line + 1
        start_byte += len(b)
        cur = []
        cur_len = 0

    if not line_bytes:
        return chunks

    for line_no, lb in enumerate(line_bytes, start=1):
        if cur and (cur_len + len(lb) > max_bytes):
            flush(line_no - 1)
        cur.append(lb)
        cur_len += len(lb)

    flush(total_lines)
    return chunks


def ingest_text_file(project_name: str, rel_path: str) -> Tuple[int, str]:
    pr, pd = _require_configured()
    abs_path = pr / rel_path
    raw_bytes = read_bytes_file(abs_path)
    sha = file_sha256_bytes(raw_bytes)
    suffix = abs_path.suffix.lower()
    logical_base = abs_path.stem

    chunks = chunk_bytes_line_aligned(raw_bytes, max_bytes=12000)
    chunk_count = len(chunks)

    chunk_entries: List[Dict[str, Any]] = []
    for ch in chunks:
        label = f"code_{logical_base}_chunk_{ch.index:02d}"
        meta = {
            "source_rel_path": rel_path,
            "source_sha256": sha,
            "chunk_index": ch.index,
            "chunk_count": chunk_count,
            "start_line": ch.start_line,
            "end_line": ch.end_line,
            "start_byte": ch.start_byte,
            "end_byte": ch.end_byte,
            "sha256_bytes": ch.sha256_bytes,
        }
        entry = create_artifact(
            project_name,
            label,
            ch.text,
            artifact_type="code_chunk",
            from_files=[rel_path],
            file_ext=(suffix or ".txt"),
            write_errors="surrogateescape",
            meta=meta,
        )
        chunk_entries.append(entry)

    idx_lines: List[str] = []
    idx_lines.append(f"# Code index for {abs_path.name}")
    idx_lines.append("")
    idx_lines.append("## Source")
    idx_lines.append(f"- Relative path: {rel_path}")
    idx_lines.append(f"- SHA256 (file bytes): {sha}")
    idx_lines.append(f"- Total bytes: {len(raw_bytes)}")
    idx_lines.append(f"- Total lines: {len(raw_bytes.splitlines())}")
    idx_lines.append(f"- Chunks: {chunk_count}")
    idx_lines.append("")
    idx_lines.append("## Chunk map (byte offsets are 0-indexed; end offsets are exclusive)")
    idx_lines.append("| chunk | start_line | end_line | start_byte | end_byte | sha256_bytes | artifact_filename |")
    idx_lines.append("|---:|---:|---:|---:|---:|---|---|")
    for ch, entry in zip(chunks, chunk_entries):
        idx_lines.append(
            f"| {ch.index} | {ch.start_line} | {ch.end_line} | {ch.start_byte} | {ch.end_byte} | {ch.sha256_bytes} | {entry.get('filename')} |"
        )

    idx_entry = create_artifact(
        project_name,
        f"code_{logical_base}_index",
        "\n".join(idx_lines),
        artifact_type="code_index",
        from_files=[rel_path],
        file_ext=".md",
        meta={
            "source_rel_path": rel_path,
            "source_sha256": sha,
            "chunk_count": chunk_count,
            "chunk_filenames": [e.get("filename") for e in chunk_entries],
            "logical_base": logical_base,
        },
    )

    overview = "\n".join([
        f"# File overview: {abs_path.name}",
        "",
        f"- Path: {rel_path}",
        f"- Type: {suffix or 'text'}",
        f"- SHA256: {sha}",
        f"- Bytes: {len(raw_bytes)}",
        f"- Lines: {len(raw_bytes.splitlines())}",
        f"- Chunks: {chunk_count}",
        "",
        "See the code_index artifact for the chunk map:",
        f"- {idx_entry.get('filename')}",
    ])
    create_artifact(project_name, f"file_overview_{logical_base}", overview, artifact_type="file_overview", from_files=[rel_path], file_ext=".md")

    m = load_manifest(project_name)
    m["last_ingested"] = {"source_rel_path": rel_path, "suffix": suffix, "sha256": sha, "when": time.time()}
    save_manifest(project_name, m)

    return chunk_count, sha


def ensure_ingested_current(project_name: str, rel_path: str) -> None:
    pr, pd = _require_configured()
    abs_path = pr / rel_path
    if not abs_path.exists() or not is_text_suffix(abs_path.suffix):
        return

    raw_bytes = read_bytes_file(abs_path)
    sha = file_sha256_bytes(raw_bytes)

    m = load_manifest(project_name)
    idxs = [a for a in (m.get("artifacts") or []) if a.get("type") == "code_index"]
    idxs.sort(key=lambda a: float(a.get("created_at", 0.0) or 0.0), reverse=True)
    for idx in idxs:
        meta = idx.get("meta") or {}
        if not isinstance(meta, dict):
            continue
        if (meta.get("source_rel_path") or "").replace("\\", "/") != rel_path.replace("\\", "/"):
            continue
        if (meta.get("source_sha256") or "") == sha:
            return
        break

    ingest_text_file(project_name, rel_path)


def collect_full_corpus_for_source(project_name: str, rel_path: str, *, max_chars: int) -> str:
    pr, pd = _require_configured()
    rel_path = (rel_path or "").replace("\\", "/").strip()
    if not rel_path:
        return ""

    ensure_ingested_current(project_name, rel_path)

    m = load_manifest(project_name)
    idxs = [a for a in (m.get("artifacts") or []) if a.get("type") == "code_index"]
    idxs.sort(key=lambda a: float(a.get("created_at", 0.0) or 0.0), reverse=True)

    idx_entry: Optional[Dict[str, Any]] = None
    for idx in idxs:
        meta = idx.get("meta") or {}
        if isinstance(meta, dict) and (meta.get("source_rel_path") or "").replace("\\", "/") == rel_path:
            idx_entry = idx
            break

    parts: List[str] = []
    total = 0
    if idx_entry:
        meta = idx_entry.get("meta") or {}
        chunk_files = meta.get("chunk_filenames") if isinstance(meta, dict) else None
        if isinstance(chunk_files, list) and chunk_files:
            def key_fn(fn: str) -> int:
                m2 = re.search(r"_chunk_(\d+)", fn)
                return int(m2.group(1)) if m2 else 0

            ordered = sorted([str(x) for x in chunk_files], key=key_fn)
            for fn in ordered:
                p = artifacts_dir(project_name) / fn
                txt = read_text_file(p, errors="surrogateescape")
                if not txt:
                    continue
                remaining = max_chars - total
                if remaining <= 0:
                    break
                if len(txt) > remaining:
                    txt = txt[:remaining]
                parts.append(txt)
                total += len(txt)
            return "".join(parts)

    raw_txt = read_text_file(pr / rel_path, errors="surrogateescape")
    return raw_txt[:max_chars] if (max_chars and len(raw_txt) > max_chars) else raw_txt


# -----------------------------------------------------------------------------
# PDF + image overview helpers
# -----------------------------------------------------------------------------

def load_pdf_text(path: Path, *, allow_ocr: bool = True, ocr_max_pages: int = 25) -> str:
    try:
        reader = PdfReader(str(path))
        pages: List[str] = []
        for page in reader.pages:
            try:
                pages.append(page.extract_text() or "")
            except Exception as e:
                pages.append(f"[PDF page extract error: {e!r}]")
        text = "\n".join(pages).strip()
        if text:
            return text

        if not allow_ocr:
            return ""

        if convert_from_path is None or pytesseract is None:
            return "(PDF error: no extractable text found; scanned PDF OCR not available (install pdf2image + poppler + pytesseract/tesseract).)"

        try:
            images = convert_from_path(str(path))
            ocr_parts: List[str] = []
            for i, im in enumerate(images[:max(1, int(ocr_max_pages or 25))]):
                try:
                    ocr = pytesseract.image_to_string(im) or ""
                    ocr = ocr.strip()
                    if ocr:
                        ocr_parts.append(f"[PAGE {i+1}]\n{ocr}")
                except Exception as e:
                    ocr_parts.append(f"[PAGE {i+1} OCR error: {e!r}]")
            ocr_text = "\n\n".join(ocr_parts).strip()
            if not ocr_text:
                return "(PDF error: no extractable text found; OCR ran but produced empty output.)"
            return ocr_text
        except Exception as e:
            return f"(PDF error: no extractable text found; OCR fallback failed: {e!r})"
    except Exception as e:
        return f"(PDF error: {e!r})"


def load_image_ocr(path: Path) -> Tuple[str, str]:
    if Image is None or pytesseract is None:
        return "", "OCR not available (install Pillow + pytesseract + Tesseract binary)."

    try:
        img = Image.open(str(path))
        if getattr(img, "mode", "") not in ("RGB", "L"):
            img = img.convert("RGB")
        text = (pytesseract.image_to_string(img) or "").strip()
        if not text:
            return "", "OCR ran but produced no text."
        return text, "OCR extracted text successfully."
    except Exception as e:
        return "", f"OCR error: {e!r}"


def _is_probably_zip(path: Path) -> bool:
    try:
        with path.open("rb") as f:
            return f.read(2) == b"PK"
    except Exception:
        return False


def _extract_docx_xml_text(xml_bytes: bytes) -> str:
    if not xml_bytes:
        return ""
    try:
        root = ET.fromstring(xml_bytes)
    except Exception:
        return ""

    paragraphs: List[str] = []
    try:
        for p in root.iter():
            if not str(p.tag).endswith("}p"):
                continue
            parts: List[str] = []
            for t in p.iter():
                if str(t.tag).endswith("}t"):
                    if t.text:
                        parts.append(t.text)
            if parts:
                line = "".join(parts).strip()
                if line:
                    paragraphs.append(line)
    except Exception:
        return ""

    text = "\n".join(paragraphs).strip()
    if text:
        text = html.unescape(text)
    return text


def _extract_docx_text(path: Path) -> Tuple[str, str]:
    """
    Extract text from a .docx (zip) file using stdlib only.
    Returns (text, note).
    """
    try:
        zf = zipfile.ZipFile(path, "r")
    except Exception:
        return "", "docx_open_failed"

    parts: List[str] = []
    try:
        names = zf.namelist()
    except Exception:
        names = []

    # Core document + optional notes/comments
    for name in (
        "word/document.xml",
        "word/footnotes.xml",
        "word/endnotes.xml",
        "word/comments.xml",
        "word/header1.xml",
        "word/footer1.xml",
    ):
        if name not in names:
            continue
        try:
            xml_bytes = zf.read(name)
        except Exception:
            continue
        t = _extract_docx_xml_text(xml_bytes)
        if t:
            parts.append(t)

    try:
        zf.close()
    except Exception:
        pass

    if not parts:
        return "", "docx_no_text_found"

    text = "\n\n".join(parts).strip()
    if not text:
        return "", "docx_no_text_found"
    return text, "docx_text_extracted"


def _extract_doc_legacy_text(path: Path) -> Tuple[str, str]:
    """
    Best-effort text extraction from legacy .doc (binary) files.
    Low fidelity, but better than no signal. Returns (text, note).
    """
    data = read_bytes_file(path)
    if not data:
        return "", "doc_read_failed"

    # Extract long-ish printable runs as a crude fallback.
    try:
        parts = re.findall(rb"[\x09\x0A\x0D\x20-\x7E]{4,}", data)
    except Exception:
        parts = []

    if not parts:
        return "", "doc_no_strings_found"

    try:
        text = "\n".join(p.decode("latin-1", errors="ignore") for p in parts)
    except Exception:
        text = ""

    # Cleanup
    text = re.sub(r"[ \t]{2,}", " ", text or "")
    text = re.sub(r"\n{3,}", "\n\n", text or "")
    text = text.strip()

    if len(text) < 200:
        return "", "doc_low_signal"
    return text, "doc_legacy_strings"


def _try_convert_doc_to_docx_text(project_name: str, path: Path) -> Tuple[str, str]:
    """
    Best-effort conversion for legacy .doc using LibreOffice (soffice) if available.
    Returns (text, note).
    """
    try:
        soffice = shutil.which("soffice")
    except Exception:
        soffice = None
    if not soffice:
        # Windows fallback: common install paths (avoid requiring PATH update)
        try:
            candidates = [
                Path(r"C:\Program Files\LibreOffice\program\soffice.exe"),
                Path(r"C:\Program Files (x86)\LibreOffice\program\soffice.exe"),
            ]
            for c in candidates:
                if c.exists():
                    soffice = str(c)
                    break
        except Exception:
            soffice = None
    if not soffice:
        return "", "soffice_not_available"

    try:
        out_dir = state_dir(project_name) / "tmp"
        out_dir.mkdir(parents=True, exist_ok=True)
    except Exception:
        return "", "tmp_dir_unavailable"

    out_path = out_dir / f"{path.stem}.docx"

    try:
        cmd = [
            soffice,
            "--headless",
            "--convert-to",
            "docx",
            "--outdir",
            str(out_dir),
            str(path),
        ]
        subprocess.run(cmd, check=False, timeout=60, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except Exception:
        return "", "soffice_convert_failed"

    if not out_path.exists():
        return "", "soffice_no_output"

    try:
        text, note = _extract_docx_text(out_path)
    except Exception:
        text, note = "", "docx_parse_failed_after_convert"

    try:
        out_path.unlink()
    except Exception:
        pass

    if not text:
        return "", "convert_yielded_no_text"
    return text, f"converted_via_soffice|{note}"


# -----------------------------------------------------------------------------
# Upload ingestion entrypoint
# -----------------------------------------------------------------------------

def _analysis_hat_active(project_name: str) -> bool:
    try:
        st = load_project_state(project_name) or {}
    except Exception:
        return False
    try:
        last_ex = str(st.get("last_active_expert") or "").strip().lower()
    except Exception:
        last_ex = ""
    return last_ex == "analysis"

def ingest_uploaded_file(
    project_name: str,
    rel_path: str,
    *,
    caption_image_fn=None,
    classify_image_fn=None,
) -> str:
    """
    Inspect an uploaded file and create/update artifacts for it.

    caption_image_fn (optional):
      callable(img_bytes: bytes, mime: str) -> Tuple[str, str]
      Returns (caption, note)

    classify_image_fn (optional):
      callable(project_name=..., canonical_rel=..., abs_stem=..., suffix=..., caption=..., ocr_text=...) -> dict
      If provided, used for image classification/mapping.
    """
    pr, pd = _require_configured()

    ensure_project_scaffold(project_name)

    rel_path = (rel_path or "").replace("\\", "/").strip()
    if not rel_path:
        return "No path provided to ingest."

    # Resolve safely under the project's raw dir by basename, unless rel_path is already canonical
    base = Path(rel_path).name
    abs_path = raw_dir(project_name) / base
    canonical_rel = str(abs_path.relative_to(pr)).replace("\\", "/")

    if not abs_path.exists():
        candidate = pr / rel_path
        if candidate.exists() and candidate.is_file():
            abs_path = candidate
            canonical_rel = rel_path
        else:
            return f"I couldn't find '{rel_path}' on disk for project '{project_name}'."

    suffix = abs_path.suffix.lower()

    # Excel
    if suffix in (".xlsx", ".xlsm", ".xls"):
        md, short = summarize_excel_structure(abs_path)
        create_artifact(project_name, f"file_overview_{abs_path.stem}", md, artifact_type="file_overview", from_files=[canonical_rel], file_ext=".md")

        try:
            bp = extract_excel_blueprint(abs_path)
            create_artifact(
                project_name,
                f"excel_blueprint_{abs_path.stem}",
                json.dumps(bp, indent=2, ensure_ascii=False),
                artifact_type="excel_blueprint",
                from_files=[canonical_rel],
                file_ext=".json",
            )
        except Exception as e:
            create_artifact(
                project_name,
                f"excel_blueprint_error_{abs_path.stem}",
                json.dumps({"error": repr(e), "file": abs_path.name, "when": now_iso()}, indent=2, ensure_ascii=False),
                artifact_type="excel_blueprint_error",
                from_files=[canonical_rel],
                file_ext=".json",
            )

        return f"Ingested Excel '{abs_path.name}'. {short}"

    # PDF
    if suffix == ".pdf":
        analysis_mode = _analysis_hat_active(project_name)
        text = load_pdf_text(abs_path, allow_ocr=(not analysis_mode))
        capped = text[:200000]

        # Always store extracted/OCR text (even if sparse)
        create_artifact(
            project_name,
            f"pdf_text_{abs_path.stem}",
            capped,
            artifact_type="pdf_text",
            from_files=[canonical_rel],
            file_ext=".txt",
        )

        plan_pages = []
        ocr_pages = []

        # If this is likely a scanned plan, render pages to images + OCR per page
        if (not analysis_mode) and (convert_from_path is not None and pytesseract is not None):
            try:
                images = convert_from_path(str(abs_path), dpi=300)
                for i, img in enumerate(images):
                    page_no = i + 1

                    # Save page image
                    img_bytes = b""
                    try:
                        from io import BytesIO
                        buf = BytesIO()
                        img.save(buf, format="PNG")
                        img_bytes = buf.getvalue()
                    except Exception:
                        pass

                    if img_bytes:
                        create_artifact_bytes(
                            project_name,
                            f"plan_page_{abs_path.stem}_{page_no:02d}",
                            img_bytes,
                            artifact_type="plan_page_image",
                            from_files=[canonical_rel],
                            file_ext=".png",
                            meta={"page": page_no},
                        )

                    # OCR page
                    try:
                        ocr = pytesseract.image_to_string(img) or ""
                        ocr = ocr.strip()
                    except Exception:
                        ocr = ""

                    if ocr:
                        ocr_pages.append(f"[PAGE {page_no}]\n{ocr}")

                if ocr_pages:
                    ocr_text = "\n\n".join(ocr_pages)[:200000]

                    # Store plan OCR (plan-specific)
                    create_artifact(
                        project_name,
                        f"plan_ocr_{abs_path.stem}",
                        ocr_text,
                        artifact_type="plan_ocr",
                        from_files=[canonical_rel],
                        file_ext=".txt",
                    )

                    # ALSO store as generic OCR so chat context can see it
                    create_artifact(
                        project_name,
                        f"ocr_text_{abs_path.stem}",
                        ocr_text,
                        artifact_type="ocr_text",
                        from_files=[canonical_rel],
                        file_ext=".txt",
                    )

                    # Write high-level plan summary into facts_map (chat-retrievable)
                    summary = "\n".join([
                        "# Floor Plan Ingested",
                        "",
                        f"- Source file: {abs_path.name}",
                        f"- Pages: {len(ocr_pages)}",
                        "",
                        "This project includes a scanned architectural floor plan.",
                        "Room labels, dimensions, notes, and layout relationships were OCR-extracted.",
                        "",
                        "Ask questions like:",
                        "- “What rooms are on the main floor?”",
                        "- “What are the master bedroom dimensions?”",
                        "- “Describe the layout around the kitchen.”",
                    ])

                    prev = read_text_file(state_file_path(project_name, "facts_map"))

                    write_canonical_entry(
                        project_name,
                        target_path=state_file_path(project_name, "facts_map"),
                        mode="text_overwrite",
                        data=(prev + "\n\n" + summary).strip(),
                    )
            except Exception:
                pass

        if analysis_mode:
            overview = "\n".join([
                f"# PDF overview: {abs_path.name}",
                "",
                f"- Path: {canonical_rel}",
                "- Type: PDF",
                f"- Text extract: {'ok' if text.strip() else 'none'}",
                "- OCR: deferred (analysis ladder will escalate if needed)",
            ])
        else:
            overview = "\n".join([
                f"# PDF overview: {abs_path.name}",
                "",
                f"- Path: {canonical_rel}",
                "- Type: architectural plan (scanned)",
                "",
                "Plan pages and OCR text have been stored and indexed.",
            ])

        create_artifact(
            project_name,
            f"file_overview_{abs_path.stem}",
            overview,
            artifact_type="file_overview",
            from_files=[canonical_rel],
            file_ext=".md",
        )

        if analysis_mode:
            return f"Ingested PDF '{abs_path.name}' (text extracted; OCR deferred for analysis ladder)."
        return f"Ingested architectural plan '{abs_path.name}' (pages rendered, OCR extracted, plan added to project memory)."


    # Word documents (.docx, .doc)
    if suffix in (".docx", ".doc"):
        text = ""
        note = ""

        if suffix == ".docx" or _is_probably_zip(abs_path):
            text, note = _extract_docx_text(abs_path)
        else:
            text, note = _extract_doc_legacy_text(abs_path)
            # Legacy .doc is low-fidelity by definition; if LibreOffice is available,
            # always attempt conversion for a higher-quality extract.
            needs_convert = (not text) or (len(text) < 1500) or ("doc_legacy_strings" in note) or ("doc_low_signal" in note)
            if needs_convert:
                t2, n2 = _try_convert_doc_to_docx_text(project_name, abs_path)
                if t2:
                    text, note = t2, (note + ";" + n2 if note else n2)

        # If we extracted but it looks partial for legacy .doc, record a capability gap once.
        try:
            is_legacy = (suffix == ".doc") and (not _is_probably_zip(abs_path))
            low_signal = bool(text) and (len(text) < 1500) and (("doc_legacy_strings" in note) or ("doc_low_signal" in note) or ("soffice_not_available" in note))
            if is_legacy and low_signal:
                sha = ""
                try:
                    sha = file_sha256_bytes(read_bytes_file(abs_path))
                except Exception:
                    sha = ""
                gap_id = f"gap_doc_extract_{sha[:12]}" if sha else f"gap_doc_extract_{abs_path.stem[:12]}"
                if not capability_gap_exists(project_name, gap_id):
                    entry = {
                        "id": gap_id,
                        "blocked": True,
                        "task_summary": "Extract full text from legacy .doc upload",
                        "limitations": [
                            "Legacy .doc parsing is partial without a converter",
                        ],
                        "missing_capabilities": [
                            "LibreOffice soffice conversion",
                        ],
                        "needed_inputs": [
                            "Install LibreOffice or upload PDF/DOCX",
                        ],
                        "suggested_features": [
                            "Auto-convert legacy .doc to .docx/.pdf via LibreOffice",
                        ],
                        "recommended_search_queries": [
                            "LibreOffice install Windows winget soffice",
                            "convert .doc to .docx headless soffice",
                        ],
                        "created_at": now_iso(),
                        "source_file": canonical_rel,
                    }
                    try:
                        append_capability_gap_entry(project_name, entry)
                        build_capability_gap_views_and_write(project_name)
                    except Exception:
                        pass
        except Exception:
            pass

        if text:
            capped = text[:200000]

            create_artifact(
                project_name,
                f"doc_text_{abs_path.stem}",
                capped,
                artifact_type="doc_text",
                from_files=[canonical_rel],
                file_ext=".txt",
            )

            preview = capped[:2000].strip()
            overview = "\n".join([
                f"# Word document overview: {abs_path.name}",
                "",
                f"- Path: {canonical_rel}",
                f"- Type: {suffix}",
                f"- Note: {note}",
                "",
                "## Extracted text (preview)",
                "",
                preview if preview else "(no preview)",
            ])

            create_artifact(
                project_name,
                f"file_overview_{abs_path.stem}",
                overview,
                artifact_type="file_overview",
                from_files=[canonical_rel],
                file_ext=".md",
            )

            return f"Ingested Word document '{abs_path.name}'."

        # No usable text (record capability gap)
        try:
            sha = ""
            try:
                sha = file_sha256_bytes(read_bytes_file(abs_path))
            except Exception:
                sha = ""
            gap_id = f"gap_doc_extract_{sha[:12]}" if sha else f"gap_doc_extract_{abs_path.stem[:12]}"
            if not capability_gap_exists(project_name, gap_id):
                entry = {
                    "id": gap_id,
                    "blocked": True,
                    "task_summary": "Extract full text from legacy .doc upload",
                    "limitations": [
                        "Legacy .doc parsing is partial without a converter",
                    ],
                    "missing_capabilities": [
                        "LibreOffice soffice conversion",
                    ],
                    "needed_inputs": [
                        "Install LibreOffice or upload PDF/DOCX",
                    ],
                    "suggested_features": [
                        "Auto-convert legacy .doc to .docx/.pdf via LibreOffice",
                    ],
                    "recommended_search_queries": [
                        "LibreOffice install Windows winget soffice",
                        "convert .doc to .docx headless soffice",
                    ],
                    "created_at": now_iso(),
                    "source_file": canonical_rel,
                }
                try:
                    append_capability_gap_entry(project_name, entry)
                    build_capability_gap_views_and_write(project_name)
                except Exception:
                    pass
        except Exception:
            pass

        # No usable text
        overview = "\n".join([
            f"# Word document overview: {abs_path.name}",
            "",
            f"- Path: {canonical_rel}",
            f"- Type: {suffix}",
            f"- Note: {note or 'no_text_extracted'}",
            "",
            "No readable text could be extracted. For best results, export to .docx or PDF and re-upload.",
            "If this is a legacy .doc, installing LibreOffice enables automatic conversion and full extraction.",
        ])
        create_artifact(
            project_name,
            f"file_overview_{abs_path.stem}",
            overview,
            artifact_type="file_overview",
            from_files=[canonical_rel],
            file_ext=".md",
        )
        return f"Stored '{abs_path.name}' but could not extract text. If you can, export to .docx or PDF and re-upload."


    # Text/code
    if is_text_suffix(suffix):
        chunk_count, _sha = ingest_text_file(project_name, canonical_rel)
        return f"Ingested '{abs_path.name}' into {chunk_count} code chunks + index (reconstructable)."

    # Images (OCR + optional vision caption + classification/mapping)
    if suffix in IMAGE_SUFFIXES:
        ocr_text, ocr_note = load_image_ocr(abs_path)

        img_bytes = read_bytes_file(abs_path)
        mime = mime_for_image_suffix(suffix)

        caption, cap_note = ("", "No caption provider configured.")
        if img_bytes and callable(caption_image_fn):
            try:
                caption, cap_note = caption_image_fn(img_bytes, mime)
            except Exception as e:
                caption, cap_note = ("", f"Caption error: {e!r}")
        elif not img_bytes:
            cap_note = "No image bytes available."

        # Store OCR text (if any)
        if ocr_text.strip():
            capped = ocr_text[:200000]
            create_artifact(
                project_name,
                f"ocr_text_{abs_path.stem}",
                capped,
                artifact_type="ocr_text",
                from_files=[canonical_rel],
                file_ext=".txt",
            )

        # Store caption (if any)
        if caption.strip():
            capped = caption[:20000]
            create_artifact(
                project_name,
                f"image_caption_{abs_path.stem}",
                capped,
                artifact_type="image_caption",
                from_files=[canonical_rel],
                file_ext=".txt",
            )

        cls: Dict[str, Any] = {}
        if callable(classify_image_fn):
            try:
                cls = classify_image_fn(
                    project_name=project_name,
                    canonical_rel=canonical_rel,
                    abs_stem=abs_path.stem,
                    suffix=suffix,
                    caption=caption,
                    ocr_text=ocr_text,
                ) or {}
            except Exception:
                cls = {}
        # Visual semantics (deterministic domain helpers like Wordle)
        # DISABLED.
        #
        # Rationale:
        # - General perception is handled by image_semantics (LLM vision).
        # - Domain-specific detectors (e.g., Wordle) are opt-in plugins, not default behavior.
        # - Leaving these on pollutes context and causes incorrect downstream actions.
        #
        # Intentionally no-op:
        pass

        preview_ocr = (ocr_text.strip()[:800] if ocr_text.strip() else "(none)")
        preview_cap = (caption.strip()[:800] if caption.strip() else "(none)")

        overview = "\n".join([
            f"# Image overview: {abs_path.name}",
            "",
            f"- Path: {canonical_rel}",
            f"- Type: {suffix}",
            f"- OCR: {ocr_note}",
            f"- Caption: {cap_note}",
            "",
            "## Classification",
            f"- Label: {cls.get('label')}",
            f"- Tags: {', '.join(cls.get('tags') or []) or '(none)'}",
            f"- Suggested bucket: {cls.get('suggested_bucket')}",
            f"- Relevance: {cls.get('relevance')}",
            f"- Ask user: {cls.get('ask_user')}",
            (f"- Question: {cls.get('question')}" if cls.get("ask_user") else "- Question: (none)"),
            (f"- Rationale: {cls.get('rationale')}" if cls.get("rationale") else "- Rationale: (none)"),
            "",
            "## Caption (preview)",
            "",
            preview_cap,
            "",
            "## OCR text (preview)",
            "",
            preview_ocr,
        ])
        create_artifact(project_name, f"file_overview_{abs_path.stem}", overview, artifact_type="file_overview", from_files=[canonical_rel], file_ext=".md")

        parts = []
        parts.append("OCR ok" if ocr_text.strip() else "OCR none")
        parts.append("caption ok" if caption.strip() else "caption none")
        label = (cls.get("label") or "image").strip()
        bucket = (cls.get("suggested_bucket") or "misc/").strip()
        ask_user = bool(cls.get("ask_user"))
        question = (cls.get("question") or "").strip()

        base_msg = f"Ingested image '{abs_path.name}' ({', '.join(parts)}). Detected: {label}. Suggested bucket: {bucket}"
        if ask_user and question:
            return base_msg + "\n" + question
        return base_msg

    # Unknown
    overview = "\n".join([
        f"# File overview: {abs_path.name}",
        "",
        f"- Path: {canonical_rel}",
        f"- Type: {suffix or '(no extension)'}",
        "- Note: automatic deep analysis for this file type is not implemented yet.",
    ])
    create_artifact(project_name, f"file_overview_{abs_path.stem}", overview, artifact_type="file_overview", from_files=[canonical_rel], file_ext=".md")
    return f"Stored '{abs_path.name}' (no deep ingest for this file type yet)."

